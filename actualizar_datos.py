"""
PEGSA & BULLTRADE - Actualizador de Datos v6
Columnas calculadas:
  DIAS_EN_FEEDLOT    : dias desde FECHA_INGRESO (techo 365 dias)
  CLASIFICACION      : Macho / Hembra / Vaca / Toro segun CATEGORIA
  ENGORDE_DIARIO_KG  : kg/dia segun clasificacion y KG_INGRESO
  KG_ESTIMADO_HOY    : KG_INGRESO + dias*engorde (techo 650 kg)
  CATEGORIA_FINAL    : categoria segun KG_ESTIMADO_HOY
  NOMBRE_CORRAL      : nombre de establecimiento segun NRO_CORRAL

KPIs agrupados por:
  - Propietario       (HOTELERO)
  - Establecimiento   (NOMBRE_CORRAL)
  - Clasificacion     (Macho/Hembra/Vaca/Toro)
  - Categoria final
"""
import sys, os, json, logging, configparser, warnings, re
from datetime import datetime, timedelta as _td
from pathlib import Path

# v15.4: el bot AUTO corre actualizar_datos.py desde
# OneDrive\PEGSA_Portal\datos\ pero wincampo_source.py vive en
# OneDrive\PEGSA_Portal\ (parent dir). Agregamos el parent al sys.path
# para que el import 'from wincampo_source import WinCampoAPI' funcione
# independientemente de cwd o de quien lance el script.
_THIS = Path(__file__).resolve().parent
_PARENT = _THIS.parent
for _p in (_PARENT, _THIS):
    if str(_p) not in sys.path:
        sys.path.insert(0, str(_p))

# ── Verificar dependencias antes de importar ──────────────
_missing = []
try:
    import pandas as pd
except ImportError:
    _missing.append("pandas")

if _missing:
    # Escribir error en log aunque el logger no este listo aun
    _log_dir = Path(__file__).parent / "logs"
    _log_dir.mkdir(exist_ok=True)
    _err_file = _log_dir / f"log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
    _msg = (f"ERROR CRITICO: faltan dependencias: {', '.join(_missing)}\n"
            f"Python: {sys.executable} (v{sys.version})\n"
            f"Solución: correr  1_INSTALAR.bat  o ejecutar:\n"
            f"  {sys.executable} -m pip install pandas\n")
    _err_file.write_text(_msg, encoding="utf-8")
    print(_msg)
    sys.exit(1)

# ── Logging ───────────────────────────────────────────────
log_dir = Path(__file__).parent / "logs"
log_dir.mkdir(exist_ok=True)
log_file = log_dir / f"log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-8s  %(message)s",
    datefmt="%H:%M:%S",
    handlers=[
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ]
)
log = logging.getLogger(__name__)

def separador(titulo=""):
    log.info("-" * 55)
    if titulo:
        log.info(f"  {titulo}")
        log.info("-" * 55)

def to_num(v):
    """Convierte string numérico argentino a float.
    Soporta: 334 / 293,74 / 1.234,56 / 1234.56
    """
    try:
        s = str(v or 0).strip().replace(" ", "")
        if not s or s in ("-", "None", "nan"):
            return 0.0
        # Si tiene coma: es separador decimal argentino (ej: 293,74 o 1.234,56)
        if "," in s:
            s = s.replace(".", "").replace(",", ".")
        # Si no tiene coma pero sí tiene punto: puede ser decimal anglosajón (334.5)
        # o separador de miles (1.234) — si hay exactamente 3 dígitos tras el punto es miles
        elif "." in s:
            parts = s.split(".")
            if len(parts) == 2 and len(parts[1]) == 3:
                s = s.replace(".", "")  # era separador de miles: 1.234 → 1234
            # si len(parts[1]) != 3, es decimal: 334.5 → 334.5 (no tocar)
        return float(s)
    except:
        return 0.0

def find_col(columnas, keywords):
    cn = [c.lower() for c in columnas]
    for kw in keywords:
        for i, c in enumerate(cn):
            if kw in c:
                return columnas[i]
    return None

# ═══════════════════════════════════════════════════════════
#  TABLAS DE NEGOCIO
# ═══════════════════════════════════════════════════════════

CLASIFICACION_MAP = {
    "TM": "Macho",  "NT": "Macho",  "NV": "Macho",
    "TH": "Hembra", "VQ": "Hembra",
    "VA": "Vaca",
    "TO": "Toro",
}

# Tabla 1: engorde diario en kg/dia segun clasificacion y peso de ingreso
# Fuente: "Aumento Proyectado dieta a fecha (analisis anual).xlsx" — muestra anual real
#   Filtros por categoría: días en feedlot + rango peso de entrada (igual que Excel)
#   Hembra  0-250  → ternero hembra: ADP obs=1.32  (días 100-450, pesoE 0-200,   N=19)
#   Hembra 250+    → vaquillona:      ADP obs=1.35  (días 30-350,  pesoE 200-400, N=685)
#   Macho   0-250  → ternero macho:   ADP obs=1.37  (días 100-450, pesoE 0-200,   N=87)
#   Macho  250-350 → novillito:       ADP obs=1.49  (días 30-350,  pesoE 200-400, N=691)
#   Macho  350-550 → novillo pesado:  ADP obs=1.23  (días 30-350,  pesoE 350-750, N=188)
#   Vaca    0-650  → vacas engorde:   ADP obs=1.40  (días 30-350,  pesoE 0-750,   N=1727)
# DEPRECADO v15.13: ya NO se consulta para feedlot. El path El Haras pasó a
# ADP_CAL_FALLBACK / _ADP_CAL_RUNTIME (abajo). Se deja como comentario muerto.
ENGORDE_DIARIO = [
    ("Hembra", 0,    250,  1.32),
    ("Hembra", 250,  1000, 1.35),
    ("Macho",  0,    250,  1.37),
    ("Macho",  250,  350,  1.49),
    ("Macho",  350,  550,  1.23),
    ("Macho",  550,  1000, 1.10),
    ("Toro",   0,    1000, 1.60),
    ("Vaca",   0,    650,  1.40),
    ("Vaca",   650,  1000, 1.00),
]

# v15.13: ADP calibrado (límite inferior del rango ±25% del teórico).
# v15.13.2: renombrado a ADP_CAL_FALLBACK. Ya NO es el valor usado en runtime:
# main() actualiza _ADP_CAL_RUNTIME con el adp_calibrado dinámico que calcula
# procesar_productivo (clamp ±25% del teórico) en cada tick. Este dict queda
# como FALLBACK para una categoría sin observado suficiente.
# v15.63: el clamp pasó de ±15% a ±25% (decisión Nicolás 2026-08-18) — 5 de 7
# categorías venían clavadas en el piso. Los pisos de acá se recalculan a × 0.75.
ADP_CAL_FALLBACK = {
    'TM': 1.028,  # ternero macho   (= 1.371 × 0.75)
    'TH': 0.993,  # ternera         (= 1.324 × 0.75)
    'NT': 1.117,  # novillito       (= 1.489 × 0.75)
    'NV': 0.923,  # novillo         (= 1.231 × 0.75)
    'VQ': 1.010,  # vaquillona      (= 1.346 × 0.75)
    'VA': 1.049,  # vaca            (= 1.399 × 0.75)
    'TO': 1.200,  # toro            (= 1.60  × 0.75)
}

# v15.13.2: ADP runtime (dinámico). Se inicializa con el fallback y main() lo
# pisa con los valores de procesar_productivo al inicio. Se MUTA in-place (no se
# re-bindea) para que calc_engorde, que lo lee como global, vea los valores frescos.
_ADP_CAL_RUNTIME = dict(ADP_CAL_FALLBACK)

# ── v15.67: teóricos de ADP calculados desde la base histórica ───────────────
# Criterio Nicolás 31/08/2026: 2026 quedó atípico (problema de estructura +
# clima golpearon estadía y engorde), así que el teórico ya NO sale de la tabla
# del Excel sino de los egresos reales de los dos años calendario ANTERIORES al
# de la corrida. Dos teóricos distintos:
#   · ADP Real (90d)  -> misma ventana de 90 días, pero de cada año base
#   · Por categoría   -> año calendario completo de cada año base
# La lista se calcula sola. Para CONGELARLA (p.ej. cuando 2026 empiece a entrar
# en la base y no lo quieras adentro), poner los años a mano:
#     ADP_BASE_ANIOS = [2024, 2025]
ADP_BASE_ANIOS   = None   # None = los dos años calendario anteriores al de hoy
ADP_BASE_MIN_CAB = 30     # cabezas mínimas para aceptar un teórico calculado
# Tolerancia del clamp del observado contra el teórico. El valor clampeado
# (adp_calibrado) es el que usa el módulo Stock para estimar masa histórica.
# OJO: con el teórico calculado sobre años sanos, un año malo como 2026 se va
# más lejos del teórico y toca el límite más seguido — o sea, el clamp le tapa
# parte de la caída a la estimación de masa. Subirlo (0.40) deja pasar más de
# la realidad; bajarlo protege más la estimación.
ADP_CLAMP_TOL = 0.25

def adp_base_anios(hoy=None):
    """Años calendario que sirven de base para los teóricos de ADP."""
    from datetime import date as _date
    hoy = hoy or _date.today()
    a = ADP_BASE_ANIOS or [hoy.year - 2, hoy.year - 1]
    return sorted(set(int(x) for x in a))

# v15.13: techo kg_estimado_hoy por categoría (decisión usuario 2026-06-08).
# Reemplaza TECHO_KG_FEEDLOT=650 único para el path El Haras.
TECHO_KG_POR_CAT = {
    'TM': 550,  # ternero (terminación)
    'TH': 500,  # ternera
    'NT': 550,  # novillito
    'NV': 600,  # v15.13.1: subido de 550 → 600 (decisión usuario 2026-06-08)
    'VQ': 550,  # v15.13.1: subido de 500 → 550 (decisión usuario 2026-06-08)
    'VA': 650,  # vaca
    'TO': 800,  # toro
}

# Tabla 2: categoria final segun clasificacion y kg estimado
CATEGORIA_FINAL_MAP = [
    ("Hembra", 0,    250,  "ternera"),
    ("Hembra", 250,  1000, "vaquillona"),
    ("Macho",  0,    250,  "ternero"),
    ("Macho",  250,  350,  "novillito"),
    ("Macho",  350,  550,  "novillo"),
    ("Macho",  550,  1000, "novillo mayor a 550 kg"),
    ("Toro",   0,    1000, "toro"),
    ("Vaca",   0,    650,  "vaca"),
    ("Vaca",   650,  1000, "vaca mayor a 650 kg"),
]

# Tabla 3: nombre de establecimiento segun NRO_CORRAL
CORRALES = [
    (1,   199, "El Haras"),
    (200, 299, "El Coloradito"),
    (300, 399, "Don Pedro"),
    (400, 499, "El Descanso"),
    (500, 599, "Campo Medel"),
    (600, 699, "El Morrón"),
    (700, 799, "La Panchita"),
    (800, 899, "La Cucuca"),
    (900, 999, "El Durazno"),
    (1000, 1099, "Recepción"),
    (1100, 1199, "San Antonio"),   # v15.56
]

TECHO_KG_FEEDLOT = 650   # DEPRECADO v15.13 (reemplazado por TECHO_KG_POR_CAT); techo único viejo El Haras
TECHO_KG_RECRIA  = 380   # techo para establecimientos de recría
ENGORDE_RECRIA   = 0.5   # kg/día fijo para todos los establecimientos de recría
TECHO_DIAS = 365         # dias maximos en feedlot

# ── Funciones de lookup ───────────────────────────────────
def get_clasificacion(cat):
    if not cat or str(cat).strip() == "":
        return "Sin clasificar"
    return CLASIFICACION_MAP.get(str(cat).strip().upper(), "Sin clasificar")

def get_engorde(clasificacion, kg_ingreso):
    for clas, desde, hasta, kg_dia in ENGORDE_DIARIO:
        if clas.lower() == clasificacion.lower() and desde <= kg_ingreso < hasta:
            return kg_dia
    return 0.0

def get_categoria_final(clasificacion, kg_estimado):
    for clas, desde, hasta, cat_final in CATEGORIA_FINAL_MAP:
        if clas.lower() == clasificacion.lower() and desde <= kg_estimado < hasta:
            return cat_final
    return "Sin clasificar"

def get_nombre_corral(nro):
    try:
        n = int(float(nro or 0))
    except:
        return "Sin asignar"
    for desde, hasta, nombre in CORRALES:
        if desde <= n <= hasta:
            return nombre
    return "Sin asignar"

def limpiar_nan(obj):
    """Elimina NaN/inf recursivamente antes de guardar JSON."""
    if isinstance(obj, dict):
        return {k: limpiar_nan(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [limpiar_nan(v) for v in obj]
    if isinstance(obj, float):
        if obj != obj or obj == float("inf") or obj == float("-inf"):
            return None
    return obj

def esperar_si_interactivo(mensaje="\nPresiona Enter para cerrar..."):
    try:
        if sys.stdin and sys.stdin.isatty():
            input(mensaje)
    except Exception:
        pass

def resolver_carpeta_salida(ruta_cfg):
    script_dir = Path(__file__).resolve().parent
    ruta = str(ruta_cfg or "").strip()
    if (not ruta or ruta.lower() in {"auto", ".", ".\\", "./"} or
        "C:\\Users\\USER\\" in ruta or "C:/Users/USER/" in ruta):
        return str(script_dir)
    p = Path(ruta).expanduser()
    if not p.is_absolute():
        p = (script_dir / p).resolve()
    return str(p)

# ═══════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════
def cargar_config():
    path = Path(__file__).parent / "config.ini"
    if not path.exists():
        log.error(f"No encontre config.ini en: {path}")
        esperar_si_interactivo("\nPresiona Enter para cerrar...")
        sys.exit(1)
    cfg = configparser.ConfigParser()
    cfg.read(path, encoding="utf-8")
    return cfg

# ═══════════════════════════════════════════════════════════
#  SMOKE TEST (v15.14)
# ═══════════════════════════════════════════════════════════
def smoke_test(carpeta, periodo):
    """
    v15.14: Smoke test de los JSONs generados por el pipeline.
    Valida rangos razonables de los KPIs principales. Detecta bugs silenciosos
    (ej. ADP teórico null como pasó en v15.5.1, masa inflada como v15.13).
    Devuelve dict {ok, checks_passed, checks_failed, errors[]}.
    Nunca tira excepción — si un JSON no se puede leer, lo reporta como error.
    """
    from pathlib import Path
    import json as _json
    base = Path(carpeta)
    results = {"ok": True, "checks_passed": 0, "checks_failed": 0, "errors": [], "avisos": []}

    def _check(name, cond, detail):
        if cond:
            results["checks_passed"] += 1
        else:
            results["checks_failed"] += 1
            results["ok"] = False
            results["errors"].append(f"{name}: {detail}")

    def _load(filename):
        try:
            return _json.load(open(base / filename, encoding="utf-8"))
        except Exception as e:
            results["errors"].append(f"no se pudo leer {filename}: {e}")
            results["ok"] = False
            return None

    # --- stock_kpis ---
    sk = _load(f"stock_kpis_{periodo}.json")
    if sk:
        cab = sk.get("kpis", {}).get("total_cabezas", 0)
        kg  = sk.get("kpis", {}).get("total_kg_estimado_hoy", 0)
        _check("stock.cabezas",  5000 < cab < 20000, f"cabezas={cab:,} fuera de [5000, 20000]")
        _check("stock.kg",       2_500_000 < kg < 6_000_000, f"kg={kg:,} fuera de [2.5M, 6M]")
        if cab > 0:
            kg_cab = kg / cab
            _check("stock.kg_por_cab", 300 < kg_cab < 700, f"kg/cab={kg_cab:.0f} fuera de [300, 700]")

    # --- productivo (ADP teórico null detection — el bug de ayer) ---
    p = _load(f"productivo_{periodo}.json")
    if p:
        pc90 = p.get("por_categoria_90d", {})
        cats_con_teo = sum(1 for c in pc90.values() if c.get("adp_teorico") is not None)
        _check("productivo.cats_con_adp_teorico", cats_con_teo >= 6,
               f"solo {cats_con_teo}/7 categorías tienen adp_teorico — posible regresión v15.5.1")
        registros = p.get("meta", {}).get("registros_filtrados", 0)
        _check("productivo.registros", registros > 1000, f"solo {registros} egresos filtrados")
        # v15.67: aviso (NO error) si la mayoría de las categorías toca el clamp.
        # Señal de que el teórico de la base quedó lejos del año corriente y que
        # el clamp le está tapando la caída real a la estimación de masa.
        _con_teo = [c for c, v in pc90.items() if v.get("adp_teorico") is not None]
        _clamp   = [c for c in _con_teo if pc90[c].get("ajustado")]
        if _con_teo and len(_clamp) > len(_con_teo) / 2:
            _tol = p.get("meta", {}).get("clamp_tolerancia", 0.25)
            results["avisos"].append(
                f"productivo.clamp: {len(_clamp)}/{len(_con_teo)} categorías tocan el límite "
                f"+/-{_tol:.0%} ({', '.join(sorted(_clamp))}). El ADP calibrado dejo de seguir al "
                f"observado y con eso se estima la masa historica. Si la caida es real, subir "
                f"ADP_CLAMP_TOL (p.ej. 0.40) en actualizar_datos.py.")
            log.warning("  [!] " + results["avisos"][-1])

    # --- muertes (anual razonable) ---
    m = _load(f"muertes_{periodo}.json")
    if m:
        muertes_anio = m.get("anio", {}).get("total_muertes", 0)
        _check("muertes.anio", 100 < muertes_anio < 600,
               f"muertes año={muertes_anio} fuera de [100, 600]")
        tasa = m.get("mortandad", {}).get("tasa_mensual_pct")
        if tasa is not None:
            _check("muertes.tasa_mensual", 0.3 < tasa < 3.0,
                   f"tasa mensual={tasa}% fuera de [0.3, 3.0]")

    # --- movimientos (ingresos razonables) ---
    mv = _load(f"movimientos_{periodo}.json")
    if mv:
        ing_anio = mv.get("anio", {}).get("ingresos", {})
        cab_ing = ing_anio.get("total_cabezas", 0)
        _check("movimientos.ingresos_anio", 5000 < cab_ing < 30000,
               f"ingresos año={cab_ing:,} fuera de [5000, 30000]")

    # --- insumos (al menos los 7 críticos) ---
    ins = _load(f"stock_insumos_{periodo}.json")
    if ins:
        insumos_list = ins.get("insumos") or ins.get("registros") or []
        _check("insumos.count", len(insumos_list) >= 6,
               f"solo {len(insumos_list)} insumos críticos — esperado >= 6 de los 7")

    # --- resultado por remito: canario del cruce con Datamars (v15.68.2) ---
    # AVISO, no error. El control "sin caravana vs pesadas sin EID de la sesión"
    # se eliminó con el criterio por caravana (ya no hay sesión de referencia).
    # Queda el único síntoma que sí es un cruce roto: Datamars activo pero
    # ninguna caravana confirmada en ningún remito (token, cache o EIDs mal).
    rr = _load("resultado_remitos.json")
    if rr:
        _dm = (rr.get("meta") or {}).get("datamars") or {}
        if _dm.get("activo") and not _dm.get("confirmadas_total"):
            results["avisos"].append(
                "remitos.datamars: la cache tiene lecturas pero NO se confirmó ninguna "
                "caravana en ningún remito — revisar el cruce RFID/EID.")
            log.warning("  [!] " + results["avisos"][-1])

    # --- consumo (anual razonable) ---
    c = _load(f"consumo_{periodo}.json")
    if c:
        anual_kg = c.get("anual", {}).get("total_kg", 0)
        _check("consumo.anual", anual_kg > 10_000_000,
               f"consumo anual={anual_kg:,} kg muy bajo (esperado > 10M)")

    return results

# ═══════════════════════════════════════════════════════════
#  EXTRACCION Y ENRIQUECIMIENTO
# ═══════════════════════════════════════════════════════════
def extraer(tabla, fecha_col=None, dias=730, df_override=None):
    """
    Aplica las transformaciones de enriquecimiento (DIAS_EN_FEEDLOT,
    CLASIFICACION, NOMBRE_CORRAL, ENGORDE_DIARIO_KG, KG_ESTIMADO_HOY,
    CATEGORIA_FINAL) sobre el DataFrame que entrega el adapter WinCampo Web.

    v15.10: el path SQL se eliminó (migración completa). `df_override` es
    obligatorio; sin él la tabla se skipea. `fecha_col`/`dias` se mantienen
    por compatibilidad de firma (el filtro temporal lo hace cada fetch_*).
    """
    try:
        import pandas as pd
    except ImportError:
        log.error("Falta pandas. Ejecuta: pip install pandas")
        sys.exit(1)

    try:
        if df_override is None:
            log.warning(f"  Skipeando {tabla} - sin df_override (v15.10: ya no hay path SQL)")
            return [], []
        log.info(f"  Leyendo {tabla} desde df_override (adapter externo) ...")
        df = df_override.copy() if hasattr(df_override, "copy") else pd.DataFrame(df_override)

        # Limpiar NaN/inf
        df = df.where(pd.notnull(df), None)
        for col in df.columns:
            df[col] = df[col].apply(
                lambda x: None if isinstance(x, float) and (x != x or x in (float("inf"), float("-inf"))) else x
            )
        log.info(f"  {len(df):,} registros  |  {len(df.columns)} columnas originales")

        # 1. Dias en feedlot con techo de TECHO_DIAS
        if "FECHA_INGRESO" in df.columns:
            hoy    = pd.Timestamp.now().normalize()
            fechas = pd.to_datetime(df["FECHA_INGRESO"], errors="coerce")
            df["DIAS_EN_FEEDLOT"] = (hoy - fechas).dt.days.fillna(0).astype(int).clip(upper=TECHO_DIAS)
            log.info(f"  + DIAS_EN_FEEDLOT  (techo {TECHO_DIAS} dias, prom: {df['DIAS_EN_FEEDLOT'].mean():.0f})")

        # 2. Clasificacion Macho / Hembra / Vaca / Toro
        if "CATEGORIA" in df.columns:
            df["CLASIFICACION"] = df["CATEGORIA"].apply(get_clasificacion)
            log.info(f"  + CLASIFICACION    {df['CLASIFICACION'].value_counts().to_dict()}")

        # 2b. Nombre de establecimiento (necesario antes del techo por campo)
        if "NRO_CORRAL" in df.columns:
            df["NOMBRE_CORRAL"] = df["NRO_CORRAL"].apply(get_nombre_corral)

        # 3. Engorde diario y Kg estimado con techo por establecimiento
        # Feedlot (El Haras, corrales 1-199): usa tabla ENGORDE_DIARIO por categoría y peso
        # Recría (resto de establecimientos): 0.5 kg/día fijo, techo 380 kg
        if all(c in df.columns for c in ["CLASIFICACION", "KG_INGRESO", "DIAS_EN_FEEDLOT"]):
            def calc_engorde(row):
                nombre = str(row.get("NOMBRE_CORRAL") or "").strip().lower()
                if "haras" in nombre:
                    # v15.13: ADP calibrado per categoría (código TM/VA/etc.) en vez
                    # de la tabla ENGORDE_DIARIO de teóricos plenos.
                    # v15.13.2: lee _ADP_CAL_RUNTIME (dinámico, lo setea main() desde
                    # procesar_productivo); cae a ADP_CAL_FALLBACK si falta la cat.
                    cat = str(row.get("CATEGORIA") or "").strip().upper()
                    return _ADP_CAL_RUNTIME.get(cat, ADP_CAL_FALLBACK.get(cat, 1.0))
                return ENGORDE_RECRIA  # recría: fijo 0.5 kg/día sin importar categoría

            def calc_kg_est(row):
                kg_ing  = float(row["KG_INGRESO"] or 0)
                dias    = int(row["DIAS_EN_FEEDLOT"] or 0)
                engorde = float(row["ENGORDE_DIARIO_KG"] or 0)
                nombre  = str(row.get("NOMBRE_CORRAL") or "").strip().lower()
                if "haras" in nombre:
                    # v15.13: techo per categoría
                    cat   = str(row.get("CATEGORIA") or "").strip().upper()
                    techo = TECHO_KG_POR_CAT.get(cat, 650)
                else:
                    techo = TECHO_KG_RECRIA
                return round(min(kg_ing + dias * engorde, techo), 1)

            df["ENGORDE_DIARIO_KG"] = df.apply(calc_engorde, axis=1)
            df["KG_ESTIMADO_HOY"]   = df.apply(calc_kg_est,  axis=1)
            log.info(f"  + KG_ESTIMADO_HOY  feedlot (v15.13: ADP calibrado + techo por cat) | recría techo {TECHO_KG_RECRIA} kg @ {ENGORDE_RECRIA} kg/día | prom: {df['KG_ESTIMADO_HOY'].mean():.1f}")

        # 4. Categoria final segun KG_ESTIMADO_HOY
        if all(c in df.columns for c in ["CLASIFICACION", "KG_ESTIMADO_HOY"]):
            df["CATEGORIA_FINAL"] = df.apply(
                lambda r: get_categoria_final(
                    str(r["CLASIFICACION"] or ""),
                    float(r["KG_ESTIMADO_HOY"] or 0)
                ), axis=1
            )
            log.info(f"  + CATEGORIA_FINAL  {df['CATEGORIA_FINAL'].value_counts().to_dict()}")

        # 5. Log NOMBRE_CORRAL
        if "NOMBRE_CORRAL" in df.columns:
            log.info(f"  + NOMBRE_CORRAL    {df['NOMBRE_CORRAL'].value_counts().to_dict()}")

        registros = df.to_dict(orient="records")
        columnas  = list(df.columns)
        return registros, columnas

    except Exception as e:
        log.error(f"  Error leyendo {tabla}: {e}")
        import traceback; traceback.print_exc()
        return [], []

# ═══════════════════════════════════════════════════════════
#  KPIs
# ═══════════════════════════════════════════════════════════
def calcular_kpis(registros, columnas):
    if not registros:
        return {}

    col_cab  = "CANTIDAD"  if "CANTIDAD"  in columnas else find_col(columnas, ["cabez","nro_cab"])
    col_cat  = "CATEGORIA" if "CATEGORIA" in columnas else None
    col_prop = "HOTELERO"  if "HOTELERO"  in columnas else find_col(columnas, ["hotel","propiet"])

    # Totales generales
    total_cab    = sum(to_num(r.get(col_cab, 0)) for r in registros) if col_cab else 0
    col_kg = "KG_ESTIMADO_HOY" if "KG_ESTIMADO_HOY" in registros[0] else ("KG_ESTIMADO" if "KG_ESTIMADO" in registros[0] else None)
    total_kg_est = sum(
        float(r.get(col_kg) or 0) * to_num(r.get(col_cab, 0))
        for r in registros
    ) if col_kg else 0
    prom_kg   = round(total_kg_est / total_cab, 1) if total_cab > 0 else 0
    dias_vals = [float(r.get("DIAS_EN_FEEDLOT") or 0) for r in registros if r.get("DIAS_EN_FEEDLOT")]
    dias_prom = round(sum(dias_vals) / len(dias_vals), 1) if dias_vals else 0

    # Funcion generica de agrupacion por cualquier columna
    def agrupar(col_grupo):
        grupos = {}
        for r in registros:
            g  = str(r.get(col_grupo) or "Sin datos").strip()
            c  = to_num(r.get(col_cab, 0)) if col_cab else 0
            ke = float(r.get(col_kg) or 0) * c if col_kg else 0
            if g not in grupos:
                grupos[g] = {"cabezas": 0, "kg_estimado": 0}
            grupos[g]["cabezas"]     += c
            grupos[g]["kg_estimado"] += ke
        for g in grupos:
            d = grupos[g]
            d["ton_estimado"] = round(d["kg_estimado"] / 1000, 2)
            d["kg_promedio"]  = round(d["kg_estimado"] / d["cabezas"], 1) if d["cabezas"] > 0 else 0
            d["kg_estimado"]  = round(d["kg_estimado"])
        return grupos

    # Agrupacion cruzada: establecimiento -> categoria_final
    def agrupar_est_cat():
        result = {}
        for r in registros:
            est = str(r.get("NOMBRE_CORRAL") or "Sin datos").strip()
            cat = str(r.get("CATEGORIA_FINAL") or r.get("CATEGORIA") or "Sin datos").strip()
            c   = to_num(r.get(col_cab, 0)) if col_cab else 0
            ke  = float(r.get(col_kg) or 0) * c if col_kg else 0
            if est not in result:
                result[est] = {}
            if cat not in result[est]:
                result[est][cat] = {"cabezas": 0, "kg_estimado": 0}
            result[est][cat]["cabezas"]     += c
            result[est][cat]["kg_estimado"] += ke
        for est in result:
            for cat in result[est]:
                d = result[est][cat]
                d["ton_estimado"] = round(d["kg_estimado"] / 1000, 2)
                d["kg_promedio"]  = round(d["kg_estimado"] / d["cabezas"], 1) if d["cabezas"] > 0 else 0
        return result

    # Agrupacion cruzada: categoria_final -> establecimiento + propietario
    def agrupar_cat_desglose():
        result = {}
        for r in registros:
            cat  = str(r.get("CATEGORIA_FINAL") or r.get("CATEGORIA") or "Sin datos").strip()
            est  = str(r.get("NOMBRE_CORRAL") or "Sin datos").strip()
            prop = str(r.get(col_prop) or "Sin datos").strip() if col_prop else "Sin datos"
            c    = to_num(r.get(col_cab, 0)) if col_cab else 0
            ke   = float(r.get(col_kg) or 0) * c if col_kg else 0
            if cat not in result:
                result[cat] = {"por_establecimiento": {}, "por_propietario": {}}
            for grupo, key in [(est, "por_establecimiento"), (prop, "por_propietario")]:
                if grupo not in result[cat][key]:
                    result[cat][key][grupo] = {"cabezas": 0, "kg_estimado": 0}
                result[cat][key][grupo]["cabezas"]     += c
                result[cat][key][grupo]["kg_estimado"] += ke
        for cat in result:
            for key in ["por_establecimiento", "por_propietario"]:
                for g in result[cat][key]:
                    d = result[cat][key][g]
                    d["ton_estimado"] = round(d["kg_estimado"] / 1000, 2)
                    d["kg_promedio"]  = round(d["kg_estimado"] / d["cabezas"], 1) if d["cabezas"] > 0 else 0
        return result

    por_cat             = agrupar(col_cat)            if col_cat  else {}
    por_propietario     = agrupar(col_prop)            if col_prop else {}
    por_establecimiento = agrupar("NOMBRE_CORRAL")    if "NOMBRE_CORRAL"   in registros[0] else {}
    por_clas            = agrupar("CLASIFICACION")    if "CLASIFICACION"   in registros[0] else {}
    por_cat_final       = agrupar("CATEGORIA_FINAL")  if "CATEGORIA_FINAL" in registros[0] else {}
    por_cat_desglose    = agrupar_cat_desglose()      if "CATEGORIA_FINAL" in registros[0] else {}

    return {
        "total_cabezas":          int(total_cab),
        "total_kg_estimado_hoy":  round(total_kg_est),
        "total_ton_estimado_hoy": round(total_kg_est / 1000, 2),
        "kg_promedio_estimado":   prom_kg,
        "dias_promedio_feedlot":  dias_prom,
        "total_establecimientos": len(por_establecimiento),
        "total_propietarios":     len(por_propietario),
        "por_categoria":          por_cat,
        "por_propietario":        por_propietario,
        "por_establecimiento":    por_establecimiento,
        "por_clasificacion":      por_clas,
        "por_categoria_final":    por_cat_final,
        "por_establecimiento_categoria": agrupar_est_cat() if "NOMBRE_CORRAL" in registros[0] else {},
        "por_categoria_desglose":        por_cat_desglose,
    }

# ═══════════════════════════════════════════════════════════
#  MOVIMIENTOS PRODUCTIVOS (v_PB_Ingresos + v_PB_Egresos)
# ═══════════════════════════════════════════════════════════
# Columnas esperadas (se buscan por keyword si el nombre exacto no existe):
#   Ingresos: FECHA, HOTELERO/PROPIETARIO, ESTABLECIMIENTO/CORRAL/NRO_CORRAL,
#             CATEGORIA, CANTIDAD/CABEZAS, KG_TOTAL/KG/PESO_TOTAL
#   Egresos:  FECHA, HOTELERO/PROPIETARIO, ESTABLECIMIENTO/CORRAL/NRO_CORRAL,
#             CATEGORIA, CANTIDAD/CABEZAS, KG_TOTAL/KG/PESO_TOTAL,
#             TIPO_EGRESO/MOTIVO/DESTINO (opcional)
# Si alguna columna no existe el campo queda en 0 / "Sin datos".
# ────────────────────────────────────────────────────────────

def _find(cols, *keywords):
    """Retorna el nombre de columna que matchea algún keyword (case-insensitive)."""
    cl = [c.lower() for c in cols]
    for kw in keywords:
        kw = kw.lower()
        for i, c in enumerate(cl):
            if kw in c:
                return cols[i]
    return None

# Valores de consignataria a EXCLUIR de ingresos (case-insensitive)
CONSIGNATARIA_EXCLUIR = {"destete", "traslado"}

def _es_excluido_consignataria(r, col_cons):
    """True si el registro debe excluirse por tipo de consignataria."""
    if not col_cons:
        return False
    val = str(r.get(col_cons) or "").strip().lower()
    return val in CONSIGNATARIA_EXCLUIR

def _kpis_bloque(regs, col_cab, col_kg):
    """Calcula totales de un conjunto de registros."""
    total_cab = sum(to_num(r.get(col_cab, 0)) for r in regs) if col_cab else 0
    total_kg  = sum(to_num(r.get(col_kg,  0)) for r in regs) if col_kg  else 0
    kg_prom   = round(total_kg / total_cab, 1) if total_cab > 0 else 0
    return {
        "cabezas":    round(total_cab),
        "kg":         round(total_kg, 1),
        "kg_promedio": kg_prom,
    }

def _agrupar_movimientos(regs, col_fecha, col_prop, col_cat, col_cab, col_kg,
                          filtro_mes=None):
    """
    Agrupa registros por propietario, categoría y mes.
    filtro_mes: si se pasa (str "YYYY-MM"), solo acumula ese mes.
    """
    import pandas as pd

    por_prop      = {}
    por_cat       = {}
    por_mes       = {}
    por_mes_det   = {}   # {mes: {cabezas, kg, por_categoria:{}, por_propietario:{}}}
    total_cab = 0
    total_kg  = 0

    for r in regs:
        cab  = to_num(r.get(col_cab, 0)) if col_cab else 0
        kg   = to_num(r.get(col_kg,  0)) if col_kg  else 0
        prop = str(r.get(col_prop, "Sin datos") or "Sin datos").strip() if col_prop else "Sin datos"
        cat  = str(r.get(col_cat,  "Sin datos") or "Sin datos").strip() if col_cat  else "Sin datos"

        mes = "Sin fecha"
        if col_fecha:
            try:
                f = pd.to_datetime(r.get(col_fecha), errors="coerce")
                if f is not None and not pd.isnull(f):
                    mes = f.strftime("%Y-%m")
            except:
                pass

        # Si filtramos por mes específico, saltear los que no corresponden
        if filtro_mes and mes != filtro_mes:
            continue

        total_cab += cab
        total_kg  += kg

        for grupo, key in [(prop, por_prop), (cat, por_cat), (mes, por_mes)]:
            if grupo not in key:
                key[grupo] = {"cabezas": 0, "kg": 0}
            key[grupo]["cabezas"] += cab
            key[grupo]["kg"]      += kg

        # Detalle por mes: desglose por categoría y propietario dentro de cada mes
        if mes not in por_mes_det:
            por_mes_det[mes] = {"cabezas": 0, "kg": 0, "por_categoria": {}, "por_propietario": {}}
        por_mes_det[mes]["cabezas"] += cab
        por_mes_det[mes]["kg"]      += kg
        for grp, slot in [(cat, "por_categoria"), (prop, "por_propietario")]:
            if grp not in por_mes_det[mes][slot]:
                por_mes_det[mes][slot][grp] = {"cabezas": 0, "kg": 0}
            por_mes_det[mes][slot][grp]["cabezas"] += cab
            por_mes_det[mes][slot][grp]["kg"]      += kg

    # Redondear y calcular kg_promedio
    for d in [por_prop, por_cat, por_mes]:
        for k in d:
            d[k]["cabezas"]    = round(d[k]["cabezas"])
            d[k]["kg"]         = round(d[k]["kg"], 1)
            d[k]["kg_promedio"] = round(d[k]["kg"] / d[k]["cabezas"], 1) if d[k]["cabezas"] > 0 else 0

    for mes_k, mv in por_mes_det.items():
        mv["cabezas"] = round(mv["cabezas"])
        mv["kg"]      = round(mv["kg"], 1)
        mv["kg_promedio"] = round(mv["kg"] / mv["cabezas"], 1) if mv["cabezas"] > 0 else 0
        for slot in ("por_categoria", "por_propietario"):
            for grp in mv[slot]:
                g = mv[slot][grp]
                g["cabezas"]    = round(g["cabezas"])
                g["kg"]         = round(g["kg"], 1)
                g["kg_promedio"] = round(g["kg"] / g["cabezas"], 1) if g["cabezas"] > 0 else 0

    return {
        "total_cabezas":    round(total_cab),
        "total_kg":         round(total_kg, 1),
        "kg_promedio":      round(total_kg / total_cab, 1) if total_cab > 0 else 0,
        "por_propietario":  por_prop,
        "por_categoria":    por_cat,
        "por_mes":          por_mes,
        "por_mes_detalle":  dict(sorted(por_mes_det.items())),  # ordenado cronológico
    }


def procesar_movimientos(regs_ing, cols_ing, regs_egr, cols_egr, periodo):
    """
    Procesa ingresos y egresos.
    Genera dos cortes:
      - anio: últimos 365 días desde hoy
      - ultimo_mes: mes calendario actual
    Ingresos filtrados: excluye registros con consignataria DESTETE o TRASLADO.
    """
    import pandas as pd
    from datetime import timedelta

    hoy          = datetime.now()
    hace_un_anio = hoy - timedelta(days=365)
    mes_actual   = hoy.strftime("%Y-%m")

    # Mes anterior (completo, no el mes en curso)
    primer_dia_mes_actual = hoy.replace(day=1)
    ultimo_mes_dt  = primer_dia_mes_actual - timedelta(days=1)
    mes_anterior   = ultimo_mes_dt.strftime("%Y-%m")

    MESES_ES = {
        "January":"Enero","February":"Febrero","March":"Marzo","April":"Abril",
        "May":"Mayo","June":"Junio","July":"Julio","August":"Agosto",
        "September":"Septiembre","October":"Octubre","November":"Noviembre","December":"Diciembre"
    }
    nombre_mes_en  = hoy.strftime("%B %Y")
    nombre_mes     = MESES_ES.get(hoy.strftime("%B"), hoy.strftime("%B")) + " " + hoy.strftime("%Y")
    nombre_mes_ant = MESES_ES.get(ultimo_mes_dt.strftime("%B"), ultimo_mes_dt.strftime("%B")) + " " + ultimo_mes_dt.strftime("%Y")

    # ── Detectar columnas — primero intenta nombres exactos de WinCampo, luego keywords ──
    def cols_ing_det(cols, tipo):
        col_fecha = _find(cols, "FechaIngreso", "fechaingreso", "fecha_ingreso", "fecha")
        col_prop  = _find(cols, "hotelero", "propietario", "hotel")
        col_cat   = _find(cols, "categoria", "category", "cat")
        col_cab   = _find(cols, "Cantidad",   "cantidad", "cabezas", "nro_cab", "cant")
        # Hardcodear KgIngreso exacto — evitar KgEgreso u otras columnas similares
        col_kg = "KgIngreso" if regs_ing and "KgIngreso" in regs_ing[0] else                  "kgingreso" if regs_ing and "kgingreso" in regs_ing[0] else                  next((c for c in cols if c.lower() == "kgingreso"), None)
        col_cons   = _find(cols, "Consignatario", "consignatario", "consignataria", "consignat")
        col_origen = "Proveedor" if regs_ing and "Proveedor" in regs_ing[0] else                      next((c for c in cols if c.lower() == "proveedor"), None)
        log.info(f"  {tipo}: {len(regs_ing):,} regs | fecha={col_fecha} prop={col_prop} "
                 f"cat={col_cat} cab={col_cab} kg={col_kg} cons={col_cons} origen={col_origen}")
        log.info(f"  Todas las columnas de INGRESOS: {cols}")
        if col_kg and regs_ing:
            sample_kg = [r.get(col_kg) for r in regs_ing[:5]]
            log.info(f"  Primeros 5 valores KgIngreso ({col_kg}): {sample_kg}")
        if col_origen and regs_ing:
            vals_orig = sorted(set(str(r.get(col_origen) or "").strip() for r in regs_ing[:500] if r.get(col_origen)))
            log.info(f"  Valores únicos origen ({col_origen}) muestra: {vals_orig[:15]}")
        return col_fecha, col_prop, col_cat, col_cab, col_kg, col_cons, col_origen

    def cols_egr_det(cols, tipo):
        col_fecha   = _find(cols, "FechaSalida",  "fechasalida",  "fecha_salida",  "fecha_egreso", "fecha")
        col_prop    = _find(cols, "hotelero", "propietario", "hotel")
        col_cat     = _find(cols, "categoria", "category", "cat")
        col_cab     = _find(cols, "Cantidad",  "cantidad", "cabezas", "nro_cab", "cant")
        # Hardcodear KgEgreso exacto — la vista tiene KgIngreso, KgEgreso y KgGanado
        col_kg = "KgEgreso" if regs_egr and "KgEgreso" in regs_egr[0] else                  "kgegreso" if regs_egr and "kgegreso" in regs_egr[0] else                  next((c for c in cols if c.lower() == "kgegreso"), None)
        col_motivo  = _find(cols, "MotivoSalida", "motivosalida", "motivo_salida", "motivo", "tipo_egreso", "tipo")
        col_destino = "DestinoVenta" if regs_egr and "DestinoVenta" in regs_egr[0] else                       next((c for c in cols if c.lower() == "destinoventa"), None)
        log.info(f"  {tipo}: {len(regs_egr):,} regs | fecha={col_fecha} prop={col_prop} "
                 f"cat={col_cat} cab={col_cab} kg={col_kg} motivo={col_motivo} destino={col_destino}")
        log.info(f"  Todas las columnas de EGRESOS: {cols}")
        if col_kg and regs_egr:
            sample_kg_e = [r.get(col_kg) for r in regs_egr[:5]]
            log.info(f"  Primeros 5 valores KgEgreso ({col_kg}): {sample_kg_e}")
        if col_destino and regs_egr:
            vals_dest = sorted(set(str(r.get(col_destino) or "").strip() for r in regs_egr[:500] if r.get(col_destino)))
            log.info(f"  Valores únicos destino ({col_destino}) muestra: {vals_dest[:15]}")
        if col_motivo and regs_egr:
            vals_unicos = list({str(r.get(col_motivo) or "").strip() for r in regs_egr[:500]})
            vals_unicos.sort()
            log.info(f"  Valores únicos en '{col_motivo}' (muestra 500 regs): {vals_unicos}")
        return col_fecha, col_prop, col_cat, col_cab, col_kg, col_motivo, col_destino

    # ── Filtrar por fecha (último año) ──
    def filtrar_anio(regs, col_fecha):
        if not col_fecha:
            return regs
        filtrados = []
        for r in regs:
            try:
                f = pd.to_datetime(r.get(col_fecha), errors="coerce")
                if f is not None and not pd.isnull(f) and f >= pd.Timestamp(hace_un_anio):
                    filtrados.append(r)
            except:
                pass
        log.info(f"    Filtro último año: {len(regs):,} → {len(filtrados):,} registros")
        return filtrados

    # ── Filtrar consignataria (solo ingresos) ──
    def filtrar_consignataria(regs, col_cons):
        if not col_cons:
            return regs
        filtrados = [r for r in regs if not _es_excluido_consignataria(r, col_cons)]
        excluidos = len(regs) - len(filtrados)
        if excluidos:
            log.info(f"    Excluidos por consignataria (DESTETE/TRASLADO): {excluidos:,}")
        return filtrados

    # ── Filtrar solo VENTA en egresos ──
    # v15.5: tolera tanto el codigo 1-letra "V" que devuelve la API WinCampo Web
    # como el string largo "VENTA" / "VENTA con destino X" del SQL viejo. Asi
    # funciona con ambas fuentes durante la migracion progresiva. Cuando v15.10
    # haga cleanup del path SQL, esto se simplifica a `m == "V"`.
    def filtrar_solo_venta(regs, col_motivo):
        if not col_motivo:
            log.warning("    ⚠ No se encontró columna MotivoSalida — se usan todos los egresos")
            return regs
        filtrados = []
        for r in regs:
            m = str(r.get(col_motivo) or "").strip().upper()
            if m == "V" or "VENTA" in m:
                filtrados.append(r)
        log.info(f"    Filtro MotivoSalida=VENTA: {len(regs):,} → {len(filtrados):,} registros")
        return filtrados

    # ── Agregar por_tipo_egreso (sobre todos los egresos del año, sin filtro VENTA) ──
    def calc_por_tipo(regs, col_motivo, col_cab, col_kg):
        if not col_motivo:
            return {}
        d = {}
        for r in regs:
            t   = str(r.get(col_motivo) or "Sin datos").strip()
            cab = to_num(r.get(col_cab, 0)) if col_cab else 0
            kg  = to_num(r.get(col_kg,  0)) if col_kg  else 0
            if t not in d:
                d[t] = {"cabezas": 0, "kg": 0}
            d[t]["cabezas"] += cab
            d[t]["kg"]      += kg
        for t in d:
            d[t]["cabezas"]    = round(d[t]["cabezas"])
            d[t]["kg"]         = round(d[t]["kg"], 1)
            d[t]["kg_promedio"] = round(d[t]["kg"] / d[t]["cabezas"], 1) if d[t]["cabezas"] > 0 else 0
        return d

    # ── v15.38: detalle granular últimos 30 días (era 15d en v15.37).
    # Agrupa por NRO_TROPA (ingresos) / NRO_TRANSACCION (egresos) = 1 fila
    # por evento documental. Cada fila incluye categorias_detalle =
    # [{categoria, cabezas, kg_prom}] para el desplegable del frontend, más
    # kg_total y kg_prom (promedio ponderado) del evento.
    from datetime import date as _date30, timedelta as _td30
    corte_30d = _date30.today() - _td30(days=30)

    def _detalle_30d(regs, col_fecha, col_cab, col_kg, id_keys,
                     lugar_keys, vendedor_keys, consig_keys):
        grupos = {}
        for r in regs:
            fv = r.get(col_fecha) if col_fecha else None
            try:
                f = pd.to_datetime(fv, errors="coerce")
                if f is None or pd.isnull(f) or f.date() < corte_30d:
                    continue
                fecha_iso = f.strftime("%Y-%m-%d")
            except Exception:
                continue
            doc = next((str(r.get(k)).strip() for k in id_keys
                        if r.get(k) not in (None, "")), "—")
            key = (fecha_iso, doc)
            g = grupos.get(key)
            if g is None:
                g = {"fecha": fecha_iso, "doc": doc, "cabezas": 0, "kg_total": 0.0,
                     "lugar": "—", "vendedor": "—", "consignatario": "—",
                     "_por_cat": {}}
                grupos[key] = g
            try:
                cab_reg = int(round(float(r.get(col_cab) or 0))) if col_cab else 1
            except (TypeError, ValueError):
                cab_reg = 1
            g["cabezas"] += cab_reg
            try:
                kg_reg = float(r.get(col_kg) or 0) if col_kg else 0.0
            except (TypeError, ValueError):
                kg_reg = 0.0
            g["kg_total"] += kg_reg
            cat = str(r.get("categoria") or r.get("Categoria") or "?").strip() or "?"
            pc = g["_por_cat"].setdefault(cat, {"cabezas": 0, "kg": 0.0})
            pc["cabezas"] += cab_reg
            pc["kg"]      += kg_reg
            if g["lugar"] == "—":
                lv = next((str(r.get(k)).strip() for k in lugar_keys
                           if r.get(k) not in (None, "")), None)
                if lv:
                    g["lugar"] = lv
            if g["vendedor"] == "—":
                vv = next((str(r.get(k)).strip() for k in vendedor_keys
                           if r.get(k) not in (None, "")), None)
                if vv:
                    g["vendedor"] = vv
            if g["consignatario"] == "—":
                cv = next((str(r.get(k)).strip() for k in consig_keys
                           if r.get(k) not in (None, "")), None)
                if cv:
                    g["consignatario"] = cv
        items = []
        for g in grupos.values():
            cats = []
            for cat, d in sorted(g["_por_cat"].items()):
                cats.append({
                    "categoria": cat,
                    "cabezas":   d["cabezas"],
                    "kg_prom":   round(d["kg"] / d["cabezas"], 1) if d["cabezas"] else 0,
                })
            items.append({
                "fecha":              g["fecha"],
                "doc":                g["doc"],
                "cabezas":            g["cabezas"],
                "kg_total":           round(g["kg_total"], 1),
                "kg_prom":            round(g["kg_total"] / g["cabezas"], 1) if g["cabezas"] else 0,
                "lugar":              g["lugar"],
                "vendedor":           g["vendedor"],
                "consignatario":      g["consignatario"],
                "categorias_detalle": cats,
            })
        items.sort(key=lambda x: (x["fecha"], x["doc"]), reverse=True)
        return items

    ingresos_30d = []
    egresos_30d  = []

    # ────────────────────────────────────────────────────────
    # INGRESOS
    # ────────────────────────────────────────────────────────
    EMPTY_ING = {"total_cabezas": 0, "total_kg": 0, "kg_promedio": 0,
                 "por_propietario": {}, "por_categoria": {}, "por_mes": {}}

    if regs_ing:
        ci = cols_ing_det(cols_ing, "INGRESOS")
        col_fecha_i, col_prop_i, col_cat_i, col_cab_i, col_kg_i, col_cons_i, col_origen_i = ci

        # Filtrar: último año + consignataria
        ing_anio = filtrar_anio(regs_ing, col_fecha_i)
        ing_anio = filtrar_consignataria(ing_anio, col_cons_i)

        # v15.38: detalle últimos 30 días (1 fila por tropa) + kg + vendedor + cats
        ingresos_30d = _detalle_30d(
            ing_anio, col_fecha_i, col_cab_i, col_kg_i,
            id_keys=["NRO_TROPA"], lugar_keys=["ORIGEN"],
            vendedor_keys=["Proveedor", "PROVEEDOR"],
            consig_keys=["Consignatario"])

        # Último mes → mes anterior (completo)
        ing_mes_regs = [r for r in ing_anio if _get_mes(r, col_fecha_i) == mes_anterior]
        log.info(f"    Ingresos mes anterior ({mes_anterior}): {len(ing_mes_regs):,} registros")

        ing_anio_data = _agrupar_movimientos(ing_anio,     col_fecha_i, col_prop_i, col_cat_i, col_cab_i, col_kg_i)
        ing_mes_data  = _agrupar_movimientos(ing_mes_regs, col_fecha_i, col_prop_i, col_cat_i, col_cab_i, col_kg_i)

        # Top 10 orígenes por cabezas (Proveedor)
        if col_origen_i:
            orig_cnt  = {}
            orig_cats = {}  # origen -> {categoria: cabezas}
            for r in ing_anio:
                o   = str(r.get(col_origen_i) or "Sin datos").strip()
                cat = str(r.get(col_cat_i)    or "Sin datos").strip() if col_cat_i else "Sin datos"
                cab = round(to_num(r.get(col_cab_i, 1) if col_cab_i else 1))
                orig_cnt[o] = orig_cnt.get(o, 0) + cab
                if o not in orig_cats:
                    orig_cats[o] = {}
                orig_cats[o][cat] = orig_cats[o].get(cat, 0) + cab
            top10_origen = sorted(orig_cnt.items(), key=lambda x: -x[1])[:10]
            ing_anio_data["top10_origen"] = [
                {"nombre": k, "cabezas": v,
                 "por_categoria": dict(sorted(orig_cats[k].items(), key=lambda x: -x[1]))}
                for k, v in top10_origen
            ]
            log.info(f"    Top 3 orígenes: {top10_origen[:3]}")
        else:
            ing_anio_data["top10_origen"] = []
    else:
        log.warning("  ⚠ Sin datos de Ingresos")
        ing_anio_data = EMPTY_ING.copy()
        ing_mes_data  = EMPTY_ING.copy()

    # ────────────────────────────────────────────────────────
    # EGRESOS  (KgEgreso, FechaSalida, solo MotivoSalida=VENTA para KPIs/tablas)
    # ────────────────────────────────────────────────────────
    EMPTY_EGR = {"total_cabezas": 0, "total_kg": 0, "kg_promedio": 0,
                 "por_propietario": {}, "por_categoria": {}, "por_mes": {}, "por_tipo_egreso": {}}

    if regs_egr:
        ce = cols_egr_det(cols_egr, "EGRESOS")
        col_fecha_e, col_prop_e, col_cat_e, col_cab_e, col_kg_e, col_motivo_e, col_destino_e = ce

        # Filtrar: último año (por FechaSalida)
        egr_anio_todos = filtrar_anio(regs_egr, col_fecha_e)

        # Para KPIs y tablas: solo VENTA
        egr_anio_venta = filtrar_solo_venta(egr_anio_todos, col_motivo_e)

        # v15.38: detalle últimos 30 días (1 fila por VENTA = NRO_TRANSACCION;
        # egreso = 1 cab/reg). Fallback a NRO_TROPA si la transacción viene null.
        # Nota: la API NO expone consignatario en egresos → esa col queda "—".
        egresos_30d = _detalle_30d(
            egr_anio_venta, col_fecha_e, col_cab_e, "KgEgreso",
            id_keys=["NRO_TRANSACCION", "NRO_TROPA"], lugar_keys=["Destino", "DestinoVenta"],
            vendedor_keys=[],
            consig_keys=["Consignatario"])

        # Mes anterior (sobre ventas)
        egr_mes_regs = [r for r in egr_anio_venta if _get_mes(r, col_fecha_e) == mes_anterior]
        log.info(f"    Egresos (VENTA) mes anterior ({mes_anterior}): {len(egr_mes_regs):,} registros")

        egr_anio_data = _agrupar_movimientos(egr_anio_venta, col_fecha_e, col_prop_e, col_cat_e, col_cab_e, col_kg_e)
        egr_anio_data["por_tipo_egreso"] = calc_por_tipo(egr_anio_todos, col_motivo_e, col_cab_e, col_kg_e)

        egr_mes_data = _agrupar_movimientos(egr_mes_regs, col_fecha_e, col_prop_e, col_cat_e, col_cab_e, col_kg_e)
        egr_mes_data["por_tipo_egreso"] = calc_por_tipo(egr_mes_regs, col_motivo_e, col_cab_e, col_kg_e)

        # Top 10 destinos (frigoríficos) por cabezas — solo ventas
        if col_destino_e:
            dest_cnt  = {}
            dest_cats = {}  # destino -> {categoria: cabezas}
            for r in egr_anio_venta:
                d   = str(r.get(col_destino_e) or "Sin datos").strip()
                cat = str(r.get(col_cat_e)     or "Sin datos").strip() if col_cat_e else "Sin datos"
                cab = round(to_num(r.get(col_cab_e, 1) if col_cab_e else 1))
                dest_cnt[d] = dest_cnt.get(d, 0) + cab
                if d not in dest_cats:
                    dest_cats[d] = {}
                dest_cats[d][cat] = dest_cats[d].get(cat, 0) + cab
            top10_destino = sorted(dest_cnt.items(), key=lambda x: -x[1])[:10]
            egr_anio_data["top10_destino"] = [
                {"nombre": k, "cabezas": v,
                 "por_categoria": dict(sorted(dest_cats[k].items(), key=lambda x: -x[1]))}
                for k, v in top10_destino
            ]
            log.info(f"    Top 3 destinos: {top10_destino[:3]}")
        else:
            egr_anio_data["top10_destino"] = []
    else:
        log.warning("  ⚠ Sin datos de Egresos")
        egr_anio_data = EMPTY_EGR.copy()
        egr_mes_data  = EMPTY_EGR.copy()

    # ────────────────────────────────────────────────────────
    # RESÚMENES
    # ────────────────────────────────────────────────────────
    def make_resumen(ing, egr):
        saldo_cab = ing["total_cabezas"] - egr["total_cabezas"]
        saldo_kg  = round(ing["total_kg"] - egr["total_kg"], 1)
        return {
            "cabezas_ingresadas":  ing["total_cabezas"],
            "kg_ingresado":        ing["total_kg"],
            "kg_promedio_ingreso": ing["kg_promedio"],
            "cabezas_egresadas":   egr["total_cabezas"],
            "kg_egresado":         egr["total_kg"],
            "kg_promedio_egreso":  egr["kg_promedio"],
            "saldo_cabezas":       saldo_cab,
            "saldo_kg":            saldo_kg,
        }

    log.info(f"  Último año    → Ing: {ing_anio_data['total_cabezas']:,} cab / {ing_anio_data['total_kg']:,.0f} kg"
             f"  |  Egr: {egr_anio_data['total_cabezas']:,} cab / {egr_anio_data['total_kg']:,.0f} kg")
    log.info(f"  Mes anterior  → Ing: {ing_mes_data['total_cabezas']:,} cab"
             f"  |  Egr: {egr_mes_data['total_cabezas']:,} cab")

    return {
        "meta": {
            "generado":      datetime.now().isoformat(),
            "periodo":       periodo,
            "mes_actual":    mes_actual,
            "nombre_mes":    nombre_mes,
            "mes_anterior":  mes_anterior,
            "nombre_mes_ant": nombre_mes_ant,
            "desde_anio":    hace_un_anio.strftime("%Y-%m-%d"),
            "hasta":         hoy.strftime("%Y-%m-%d"),
            "filtros":       "Ingresos: excluye CONSIGNATARIO en [DESTETE, TRASLADO]. Egresos: solo MotivoSalida=VENTA para KPIs. Por Tipo incluye todos los motivos.",
            "detalle_desde": corte_30d.isoformat(),   # v15.38 (era 15d)
        },
        "anio": {
            "resumen":  make_resumen(ing_anio_data, egr_anio_data),
            "ingresos": ing_anio_data,
            "egresos":  egr_anio_data,
            "ingresos_detalle_30d": ingresos_30d,   # v15.38 (era 15d)
            "egresos_detalle_30d":  egresos_30d,    # v15.38
        },
        "ultimo_mes": {
            "nombre":   nombre_mes_ant,
            "resumen":  make_resumen(ing_mes_data, egr_mes_data),
            "ingresos": ing_mes_data,
            "egresos":  egr_mes_data,
        },
    }


def _get_mes(r, col_fecha):
    """Retorna 'YYYY-MM' del registro, o '' si no hay fecha."""
    if not col_fecha:
        return ""
    try:
        import pandas as pd
        f = pd.to_datetime(r.get(col_fecha), errors="coerce")
        if f is not None and not pd.isnull(f):
            return f.strftime("%Y-%m")
    except:
        pass
    return ""


# ═══════════════════════════════════════════════════════════
#  MUERTES (V_MUERTES)
# ═══════════════════════════════════════════════════════════
# Columnas conocidas:
#   MUERTOS     : cantidad de cabezas muertas
#   ABREVIATURA : categoría del animal
# Columnas buscadas automáticamente:
#   fecha       : para filtro último año y evolución mensual
#   establecimiento: para desglose por campo (si existe)
# ─────────────────────────────────────────────────────────

def procesar_muertes(regs_m, cols_m, regs_ing, cols_ing, regs_stock, cols_stock, periodo):
    """
    Procesa V_MUERTES y calcula tasa de mortandad anual compuesta.

    Tasa mensual = muertes_anio / (ingresos_anio_filtrados + stock_haras_hoy)
    Tasa anual   = (1 + tasa_mensual)^12 - 1

    El Haras = corrales NRO_CORRAL 1–199 en stock.
    Ingresos = todos excepto CONSIGNATARIO en [DESTETE, TRASLADO].
    """
    import pandas as pd
    from datetime import timedelta

    hoy          = datetime.now()
    hace_un_anio = hoy - timedelta(days=365)
    mes_actual   = hoy.strftime("%Y-%m")
    MESES_ES     = {"January":"Enero","February":"Febrero","March":"Marzo","April":"Abril",
                    "May":"Mayo","June":"Junio","July":"Julio","August":"Agosto",
                    "September":"Septiembre","October":"Octubre","November":"Noviembre","December":"Diciembre"}
    nombre_mes   = MESES_ES.get(hoy.strftime("%B"), hoy.strftime("%B")) + " " + hoy.strftime("%Y")

    # ══════════════════════════════════════════════
    # A) COLUMNAS V_MUERTES
    # ══════════════════════════════════════════════
    # FECHA_MUERTE es la columna correcta — buscarla EXACTA primero
    col_muertes = _find(cols_m, "MUERTOS", "muertos", "muertes", "bajas", "baja")
    col_cat     = _find(cols_m, "ABREVIATURA", "abreviatura", "categoria", "cat", "especie")
    col_fecha_m = "FECHA_MUERTE" if regs_m and "FECHA_MUERTE" in regs_m[0] else                   _find(cols_m, "FECHA_MUERTE", "fecha_muerte", "FECHA", "fecha")

    # helpers — definidos ANTES de cualquier uso
    def en_anio(r, col_f):
        try:
            f = pd.to_datetime(r.get(col_f), errors="coerce")
            return f is not None and not pd.isnull(f) and f >= pd.Timestamp(hace_un_anio)
        except:
            return False

    def get_mes(r, col_f):
        try:
            f = pd.to_datetime(r.get(col_f), errors="coerce")
            if f is not None and not pd.isnull(f):
                return f.strftime("%Y-%m")
        except:
            pass
        return "Sin fecha"

    def sumar_col(registros, col):
        return sum(to_num(r.get(col, 0)) for r in registros) if col else 0

    def to_int_directo(v):
        """Convierte valor a entero sin tocar separadores — para columna MUERTOS."""
        try:
            return int(float(str(v or 0)))
        except:
            return 0

    def sumar_muertes(registros):
        """Suma columna MUERTOS usando conversión directa float→int."""
        return sum(to_int_directo(r.get(col_muertes, 0)) for r in registros) if col_muertes else 0

    log.info(f"  V_MUERTES: {len(regs_m):,} regs | muertes={col_muertes} cat={col_cat} fecha={col_fecha_m}")
    log.info(f"  TODAS las columnas: {cols_m}")
    if regs_m and col_muertes:
        vals = [to_int_directo(r.get(col_muertes, 0)) for r in regs_m[:5]]
        log.info(f"  Primeros 5 valores MUERTOS: {vals}")
    if not col_muertes:
        log.error("  ✗ No se encontró columna MUERTOS")

    # ── Grupos de categorías ──────────────────────────────────────────────────
    # Vacas: VA | Machos: TM, NT, NV, TO | Hembras: TH, VQ
    GRUPOS = {
        "Vacas":   {"VA"},
        "Machos":  {"TM", "NT", "NV", "TO"},
        "Hembras": {"TH", "VQ"},
    }
    # Nombres largos tal como vienen en v_PB_Ingresos columna Categoria
    GRUPOS_NOMBRE = {
        "VACA":       "Vacas",
        "TERNERO":    "Machos",
        "NOVILLITO":  "Machos",
        "NOVILLO":    "Machos",
        "TORO":       "Machos",
        "TERNERA":    "Hembras",
        "VAQUILLONA": "Hembras",
    }
    GRUPOS_ORDEN = ["Vacas", "Machos", "Hembras"]

    def get_grupo(abrev):
        """Retorna el grupo según código corto (VA/TM/etc.) o nombre largo (VACA/NOVILLO/etc.)."""
        a = str(abrev or "").strip().upper()
        for g, cats in GRUPOS.items():
            if a in cats:
                return g
        return GRUPOS_NOMBRE.get(a, "Otros")

    def dias_encierre(r):
        """Días de encierre usando columna DIAS_ENCIERRE (ya calculada en V_MUERTES)."""
        try:
            return int(float(r.get("DIAS_ENCIERRE") or 0))
        except:
            return 0

    def agrupar(registros, col_grupo, col_val, label="Sin datos"):
        d = {}
        for r in registros:
            g = str(r.get(col_grupo) or label).strip() if col_grupo else label
            val = to_int_directo(r.get(col_val, 0)) if col_val == col_muertes else                   (to_num(r.get(col_val, 0)) if col_val else 0)
            d[g] = d.get(g, 0) + val
        return {k: round(v) for k, v in sorted(d.items(), key=lambda x: -x[1])}

    def agrupar_mes(registros, col_f, col_val):
        d = {}
        for r in registros:
            mes = get_mes(r, col_f) if col_f else "Sin fecha"
            val = to_int_directo(r.get(col_val, 0)) if col_val == col_muertes else                   (to_num(r.get(col_val, 0)) if col_val else 0)
            d[mes] = d.get(mes, 0) + val
        return {k: round(v) for k, v in sorted(d.items())}

    def sumar_por_grupo_m(registros):
        """Suma muertes agrupadas por Vacas/Machos/Hembras."""
        d = {g: 0 for g in GRUPOS_ORDEN}
        d["Otros"] = 0
        for r in registros:
            g = get_grupo(r.get(col_cat))
            d[g] = d.get(g, 0) + to_int_directo(r.get(col_muertes, 0))
        return {k: round(v) for k, v in d.items() if v > 0}

    # ── Filtrar muertes último año ──────────────────────────────────────────
    if col_fecha_m:
        regs_anio_m = [r for r in regs_m if en_anio(r, col_fecha_m)]
        log.info(f"    Muertes filtro último año: {len(regs_m):,} → {len(regs_anio_m):,}")
        log.info(f"    Suma MUERTOS sin filtro días: {round(sumar_muertes(regs_anio_m)):,}")
    else:
        regs_anio_m = regs_m
        log.warning(f"    ⚠ col_fecha_m=None — sin filtro de fecha")

    # ── Filtro El Haras (corrales 1-199) ──
    # v15.7.1: el SQL viejo V_MUERTES contaba muertes de todos los corrales
    # (incluyendo recría: corrales 200/300/400), inflando el numerador de la
    # tasa de feedlot. Corrección de bug latente: la mortandad del feedlot
    # debe ser solo de El Haras (1-199), igual que el denominador (stock Haras).
    def _es_haras_muerte(r):
        try:
            n = int(float(r.get("NRO_CORRAL") or 0))
            return 1 <= n <= 199
        except (TypeError, ValueError):
            return False

    regs_anio_m_haras = [r for r in regs_anio_m if _es_haras_muerte(r)]
    excluidos_haras = len(regs_anio_m) - len(regs_anio_m_haras)
    log.info(f"    Filtro El Haras (corrales 1-199): {len(regs_anio_m):,} → {len(regs_anio_m_haras):,} ({excluidos_haras} fuera de Haras excluidos)")
    regs_anio_m = regs_anio_m_haras

    # ── Filtro >30 días de encierre (columna DIAS_ENCIERRE) ────────────────
    regs_anio_m30 = [r for r in regs_anio_m if dias_encierre(r) > 30]
    excluidos_30  = len(regs_anio_m) - len(regs_anio_m30)
    log.info(f"    Filtro >30d encierre: {len(regs_anio_m):,} → {len(regs_anio_m30):,} ({excluidos_30} excluidos)")
    log.info(f"    Suma MUERTOS último año (>30d): {round(sumar_muertes(regs_anio_m30)):,}")

    total_anio_m    = round(sumar_muertes(regs_anio_m30))
    por_cat_anio    = agrupar(regs_anio_m30, col_cat, col_muertes)
    por_grupo_anio  = sumar_por_grupo_m(regs_anio_m30)
    por_mes_anio    = agrupar_mes(regs_anio_m30, col_fecha_m, col_muertes)

    regs_mes_m      = [r for r in regs_anio_m30 if get_mes(r, col_fecha_m) == mes_actual] if col_fecha_m else []
    total_mes_m     = round(sumar_muertes(regs_mes_m))
    por_cat_mes     = agrupar(regs_mes_m, col_cat, col_muertes)
    por_grupo_mes   = sumar_por_grupo_m(regs_mes_m)
    por_mes_mes     = agrupar_mes(regs_mes_m, col_fecha_m, col_muertes)

    log.info(f"  Muertes último año (>30d) → {total_anio_m:,}")
    log.info(f"  Muertes mes actual        → {total_mes_m:,}")
    log.info(f"  Por grupo año: {por_grupo_anio}")

    # ══════════════════════════════════════════════
    # B) INGRESOS ÚLTIMO AÑO — total y por grupo
    # ══════════════════════════════════════════════
    col_fecha_i = _find(cols_ing, "FechaIngreso", "fechaingreso", "fecha_ingreso", "fecha")
    col_cab_i   = _find(cols_ing, "CantidadIngreso", "Cantidad", "cantidad", "cabezas", "nro_cab", "cant")
    col_cons_i  = _find(cols_ing, "Consignatario", "consignatario", "consignataria", "consignat")
    col_cat_i   = _find(cols_ing, "Categoria", "categoria", "category", "cat")

    # Diagnóstico: ver valores únicos de categoría en ingresos
    if col_cat_i and regs_ing:
        cats_ing = sorted(set(str(r.get(col_cat_i) or "").strip() for r in regs_ing[:500] if r.get(col_cat_i)))
        log.info(f"  Valores únicos Categoria en ingresos (muestra): {cats_ing[:20]}")

    def es_excluido_cons(r):
        if not col_cons_i:
            return False
        return str(r.get(col_cons_i) or "").strip().upper() in {"DESTETE", "TRASLADO"}

    if col_fecha_i:
        ing_anio = [r for r in regs_ing if en_anio(r, col_fecha_i) and not es_excluido_cons(r)]
    else:
        ing_anio = [r for r in regs_ing if not es_excluido_cons(r)]

    total_cab_ing_anio = round(sumar_col(ing_anio, col_cab_i))

    # Ingresos por grupo (Vacas/Machos/Hembras)
    ing_por_grupo = {g: 0 for g in GRUPOS_ORDEN}
    ing_por_grupo["Otros"] = 0
    for r in ing_anio:
        g   = get_grupo(r.get(col_cat_i) if col_cat_i else "")
        cab = to_num(r.get(col_cab_i, 0)) if col_cab_i else 0
        ing_por_grupo[g] = ing_por_grupo.get(g, 0) + cab
    ing_por_grupo = {k: round(v) for k, v in ing_por_grupo.items() if v > 0}

    log.info(f"  Ingresos último año → {total_cab_ing_anio:,} cab | por grupo: {ing_por_grupo}")

    # ══════════════════════════════════════════════
    # C) STOCK HOY — EL HARAS — total y por grupo
    # ══════════════════════════════════════════════
    col_cab_s    = "CANTIDAD"    if (regs_stock and "CANTIDAD"    in regs_stock[0]) else _find(cols_stock, "cabezas", "cantidad", "cant")
    col_nombre_s = "NOMBRE_CORRAL" if (regs_stock and "NOMBRE_CORRAL" in regs_stock[0]) else _find(cols_stock, "nombre_corral", "establecimiento")
    col_corral_s = "NRO_CORRAL"  if (regs_stock and "NRO_CORRAL"  in regs_stock[0]) else _find(cols_stock, "nro_corral", "corral")
    col_cat_s    = "CATEGORIA_FINAL" if (regs_stock and "CATEGORIA_FINAL" in regs_stock[0]) else                    _find(cols_stock, "ABREVIATURA", "abreviatura", "categoria", "cat")

    def es_haras(r):
        if col_nombre_s:
            nombre = str(r.get(col_nombre_s) or "").strip().lower()
            if nombre == "el haras":
                return True
        if col_corral_s:
            try:
                nro = int(float(r.get(col_corral_s) or 0))
                return 1 <= nro <= 199
            except:
                pass
        return False

    # Mapeo de CATEGORIA_FINAL a grupos
    CAT_FINAL_GRUPO = {
        "vaca":                 "Vacas",
        "vaca mayor a 650 kg":  "Vacas",
        "novillo":              "Machos",
        "novillo mayor a 550 kg": "Machos",
        "novillito":            "Machos",
        "ternero":              "Machos",
        "toro":                 "Machos",
        "vaquillona":           "Hembras",
        "ternera":              "Hembras",
    }

    def get_grupo_stock(r):
        cat = str(r.get(col_cat_s) or "").strip().lower() if col_cat_s else ""
        return CAT_FINAL_GRUPO.get(cat, "Otros")

    stock_haras = [r for r in regs_stock if es_haras(r)]
    total_stock_haras = round(sumar_col(stock_haras, col_cab_s))

    stock_por_grupo = {g: 0 for g in GRUPOS_ORDEN}
    stock_por_grupo["Otros"] = 0
    for r in stock_haras:
        g   = get_grupo_stock(r)
        cab = to_num(r.get(col_cab_s, 0)) if col_cab_s else 0
        stock_por_grupo[g] = stock_por_grupo.get(g, 0) + cab
    stock_por_grupo = {k: round(v) for k, v in stock_por_grupo.items() if v > 0}

    log.info(f"  Stock El Haras → {total_stock_haras:,} cab | por grupo: {stock_por_grupo}")

    # ══════════════════════════════════════════════
    # D) ÚLTIMOS 30 DÍAS
    # ══════════════════════════════════════════════
    hace_30        = hoy - timedelta(days=30)
    nombre_mes_ant = f"Últimos 30 días ({hace_30.strftime('%d/%m/%Y')} – {hoy.strftime('%d/%m/%Y')})"
    mes_anterior   = hace_30.strftime("%Y-%m-%d")

    def en_ultimos_30(r):
        if not col_fecha_m:
            return False
        try:
            import pandas as pd
            f = pd.to_datetime(r.get(col_fecha_m), errors="coerce")
            if f is not None and not pd.isnull(f):
                return f.date() >= hace_30.date()
        except:
            pass
        return False

    regs_mes_ant_m    = [r for r in regs_anio_m30 if en_ultimos_30(r)]
    total_mes_ant_m   = round(sumar_muertes(regs_mes_ant_m))
    por_cat_mes_ant   = agrupar(regs_mes_ant_m, col_cat, col_muertes)
    por_grupo_mes_ant = sumar_por_grupo_m(regs_mes_ant_m)
    log.info(f"  Muertes últimos 30 días → {total_mes_ant_m:,}")

    # ══════════════════════════════════════════════
    # E) TASA DE MORTANDAD — global y por grupo
    # ══════════════════════════════════════════════
    denominador  = total_cab_ing_anio + total_stock_haras
    tasa_mensual = (total_anio_m / denominador) if denominador > 0 else None
    tasa_mens_p  = round(tasa_mensual * 100, 3) if tasa_mensual is not None else None

    # Tasa por grupo: solo mensual
    tasas_grupo = {}
    for g in GRUPOS_ORDEN:
        m_g   = por_grupo_anio.get(g, 0)
        i_g   = ing_por_grupo.get(g, 0)
        s_g   = stock_por_grupo.get(g, 0)
        den_g = i_g + s_g
        if den_g > 0:
            tm_g = m_g / den_g
            tasas_grupo[g] = {
                "muertes":          m_g,
                "ingresos":         i_g,
                "stock":            s_g,
                "denominador":      den_g,
                "tasa_mensual_pct": round(tm_g * 100, 3),
            }
        else:
            tasas_grupo[g] = {"muertes": m_g, "ingresos": i_g, "stock": s_g,
                               "denominador": 0, "tasa_mensual_pct": None}

    log.info(f"  ── Tasa de mortandad ──")
    log.info(f"    Muertes año (>30d) : {total_anio_m:,}")
    log.info(f"    Ingresos año       : {total_cab_ing_anio:,}")
    log.info(f"    Stock El Haras     : {total_stock_haras:,}")
    log.info(f"    Denominador        : {denominador:,}")
    log.info(f"    Tasa mensual       : {tasa_mens_p}%")
    for g, t in tasas_grupo.items():
        log.info(f"    {g:<10} → {t.get('muertes'):>4} muertes / {t.get('denominador'):>6,} den → {t.get('tasa_mensual_pct')}% mensual")

    # ══════════════════════════════════════════════
    # RESULTADO
    # ══════════════════════════════════════════════
    return {
        "meta": {
            "generado":      datetime.now().isoformat(),
            "periodo":       periodo,
            "tabla":         "V_MUERTES",
            "mes_actual":    mes_actual,
            "nombre_mes":    nombre_mes,
            "mes_anterior":  mes_anterior,
            "nombre_mes_ant": nombre_mes_ant,
            "desde_anio":    hace_un_anio.strftime("%Y-%m-%d"),
            "hasta":         hoy.strftime("%Y-%m-%d"),
            "filtro_dias":   ">30 días de encierre",
        },
        "mortandad": {
            "muertes_anio":     total_anio_m,
            "ingresos_anio":    total_cab_ing_anio,
            "stock_haras_hoy":  total_stock_haras,
            "denominador":      denominador,
            "tasa_mensual_pct": tasa_mens_p,
            "formula":          "tasa_mensual = muertes_año / (ingresos_año + stock_hoy)",
            "por_grupo":        tasas_grupo,
        },
        "anio": {
            "total_muertes": total_anio_m,
            "por_categoria": por_cat_anio,
            "por_grupo":     por_grupo_anio,
            "por_mes":       por_mes_anio,
        },
        "mes_anterior": {
            "nombre":        nombre_mes_ant,
            "mes":           mes_anterior,
            "total_muertes": total_mes_ant_m,
            "por_categoria": por_cat_mes_ant,
            "por_grupo":     por_grupo_mes_ant,
        },
    }




# ═══════════════════════════════════════════════════════════
#  MUERTES ÚLTIMOS 30 DÍAS — módulo independiente
#  Mismo análisis que el anual pero ventana = 30 días:
#  muertes (>30d encierre), ingresos y stock por grupo
#  → genera muertes_30d_YYYY.json
# ═══════════════════════════════════════════════════════════
def procesar_muertes_30d(regs_m, cols_m, regs_ing, cols_ing, regs_stock, cols_stock, periodo):
    import pandas as pd
    from datetime import timedelta

    hoy     = datetime.now()
    hace_30 = hoy - timedelta(days=30)
    desde_str = hace_30.strftime("%d/%m/%Y")
    hasta_str = hoy.strftime("%d/%m/%Y")
    label_periodo = f"Últimos 30 días ({desde_str} – {hasta_str})"

    # ── Columnas V_MUERTES ──────────────────────────────────
    col_muertes = _find(cols_m, "MUERTOS", "muertos", "muertes", "bajas", "baja")
    col_cat     = _find(cols_m, "ABREVIATURA", "abreviatura", "categoria", "cat", "especie")
    col_fecha_m = "FECHA_MUERTE" if regs_m and "FECHA_MUERTE" in regs_m[0] else \
                  _find(cols_m, "FECHA_MUERTE", "fecha_muerte", "FECHA", "fecha")

    def to_int_directo(v):
        try: return int(float(str(v or 0)))
        except: return 0

    def to_num_local(v):
        try:
            s = str(v or "0").strip().replace(",", ".")
            return float(s)
        except: return 0.0

    def en_30d(r, col_f):
        try:
            f = pd.to_datetime(r.get(col_f), errors="coerce")
            return f is not None and not pd.isnull(f) and f.date() >= hace_30.date()
        except: return False

    def dias_encierre(r):
        try: return int(float(r.get("DIAS_ENCIERRE") or 0))
        except: return 0

    # Grupos
    GRUPOS = {"Vacas": {"VA"}, "Machos": {"TM","NT","NV","TO"}, "Hembras": {"TH","VQ"}}
    GRUPOS_NOMBRE = {"VACA":"Vacas","TERNERO":"Machos","NOVILLITO":"Machos",
                     "NOVILLO":"Machos","TORO":"Machos","TERNERA":"Hembras","VAQUILLONA":"Hembras"}
    GRUPOS_ORDEN = ["Vacas","Machos","Hembras"]
    CAT_FINAL_GRUPO = {
        "vaca":"Vacas","vaca mayor a 650 kg":"Vacas",
        "novillo":"Machos","novillo mayor a 550 kg":"Machos",
        "novillito":"Machos","ternero":"Machos","toro":"Machos",
        "vaquillona":"Hembras","ternera":"Hembras",
    }

    def get_grupo(abrev):
        a = str(abrev or "").strip().upper()
        for g, cats in GRUPOS.items():
            if a in cats: return g
        return GRUPOS_NOMBRE.get(a, "Otros")

    def get_grupo_stock(r):
        cat = str(r.get(col_cat_s) or "").strip().lower() if col_cat_s else ""
        return CAT_FINAL_GRUPO.get(cat, "Otros")

    # v15.7.1: filtrar El Haras (1-199) — ver procesar_muertes(). El SQL viejo
    # mezclaba muertes de recría en el numerador de la tasa de feedlot.
    def _es_haras_muerte(r):
        try:
            n = int(float(r.get("NRO_CORRAL") or 0))
            return 1 <= n <= 199
        except (TypeError, ValueError):
            return False

    # ── A) MUERTES últimos 30d con >30d encierre ────────────
    if col_fecha_m:
        regs_30d = [r for r in regs_m
                    if en_30d(r, col_fecha_m)
                    and dias_encierre(r) > 30
                    and _es_haras_muerte(r)]
    else:
        regs_30d = []

    total_m = round(sum(to_int_directo(r.get(col_muertes, 0)) for r in regs_30d))

    # Por categoría
    por_cat = {}
    for r in regs_30d:
        c = str(r.get(col_cat) or "Sin datos").strip() if col_cat else "Sin datos"
        por_cat[c] = por_cat.get(c, 0) + to_int_directo(r.get(col_muertes, 0))
    por_cat = {k: round(v) for k, v in sorted(por_cat.items(), key=lambda x: -x[1])}

    # Por grupo
    por_grupo_m = {g: 0 for g in GRUPOS_ORDEN}
    for r in regs_30d:
        g = get_grupo(r.get(col_cat))
        por_grupo_m[g] = por_grupo_m.get(g, 0) + to_int_directo(r.get(col_muertes, 0))
    por_grupo_m = {k: round(v) for k, v in por_grupo_m.items()}

    log.info(f"  [30d] Muertes (>30d encierre): {total_m:,} | por grupo: {por_grupo_m}")

    # ── B) INGRESOS últimos 30d ─────────────────────────────
    col_fecha_i = _find(cols_ing, "FechaIngreso","fechaingreso","fecha_ingreso","fecha")
    col_cab_i   = _find(cols_ing, "CantidadIngreso","Cantidad","cantidad","cabezas","nro_cab","cant")
    col_cons_i  = _find(cols_ing, "Consignatario","consignatario","consignataria","consignat")
    col_cat_i   = _find(cols_ing, "Categoria","categoria","category","cat")

    def es_excluido(r):
        if not col_cons_i: return False
        return str(r.get(col_cons_i) or "").strip().upper() in {"DESTETE","TRASLADO"}

    if col_fecha_i:
        ing_30d = [r for r in regs_ing if en_30d(r, col_fecha_i) and not es_excluido(r)]
    else:
        ing_30d = [r for r in regs_ing if not es_excluido(r)]

    total_ing = round(sum(to_num_local(r.get(col_cab_i, 0)) for r in ing_30d) if col_cab_i else 0)

    ing_por_grupo = {g: 0 for g in GRUPOS_ORDEN}
    for r in ing_30d:
        g   = get_grupo(r.get(col_cat_i) if col_cat_i else "")
        cab = to_num_local(r.get(col_cab_i, 0)) if col_cab_i else 0
        ing_por_grupo[g] = ing_por_grupo.get(g, 0) + cab
    ing_por_grupo = {k: round(v) for k, v in ing_por_grupo.items()}

    log.info(f"  [30d] Ingresos: {total_ing:,} cab | por grupo: {ing_por_grupo}")

    # ── C) STOCK hoy El Haras (igual que el anual — es snapshot) ──
    col_cab_s    = "CANTIDAD"       if (regs_stock and "CANTIDAD"       in regs_stock[0]) else _find(cols_stock, "cabezas","cantidad","cant")
    col_nombre_s = "NOMBRE_CORRAL"  if (regs_stock and "NOMBRE_CORRAL"  in regs_stock[0]) else _find(cols_stock, "nombre_corral","establecimiento")
    col_corral_s = "NRO_CORRAL"     if (regs_stock and "NRO_CORRAL"     in regs_stock[0]) else _find(cols_stock, "nro_corral","corral")
    col_cat_s    = "CATEGORIA_FINAL" if (regs_stock and "CATEGORIA_FINAL" in regs_stock[0]) else \
                   _find(cols_stock, "ABREVIATURA","abreviatura","categoria","cat")

    def es_haras(r):
        if col_nombre_s:
            if str(r.get(col_nombre_s) or "").strip().lower() == "el haras": return True
        if col_corral_s:
            try:
                nro = int(float(r.get(col_corral_s) or 0))
                return 1 <= nro <= 199
            except: pass
        return False

    stock_haras = [r for r in regs_stock if es_haras(r)]
    total_stock = round(sum(to_num_local(r.get(col_cab_s, 0)) for r in stock_haras) if col_cab_s else 0)

    stock_por_grupo = {g: 0 for g in GRUPOS_ORDEN}
    for r in stock_haras:
        g   = get_grupo_stock(r)
        cab = to_num_local(r.get(col_cab_s, 0)) if col_cab_s else 0
        stock_por_grupo[g] = stock_por_grupo.get(g, 0) + cab
    stock_por_grupo = {k: round(v) for k, v in stock_por_grupo.items()}

    log.info(f"  [30d] Stock El Haras: {total_stock:,} cab | por grupo: {stock_por_grupo}")

    # ── D) TASA DE MORTANDAD 30 días ────────────────────────
    denominador  = total_ing + total_stock
    tasa_mens_p  = round(total_m / denominador * 100, 3) if denominador > 0 else None

    tasas_grupo = {}
    for g in GRUPOS_ORDEN:
        m_g   = por_grupo_m.get(g, 0)
        i_g   = ing_por_grupo.get(g, 0)
        s_g   = stock_por_grupo.get(g, 0)
        den_g = i_g + s_g
        tasas_grupo[g] = {
            "muertes":          m_g,
            "ingresos":         i_g,
            "stock":            s_g,
            "denominador":      den_g,
            "tasa_mensual_pct": round(m_g / den_g * 100, 3) if den_g > 0 else None,
        }

    log.info(f"  [30d] Tasa: {tasa_mens_p}% | denom: {denominador:,}")
    for g, t in tasas_grupo.items():
        log.info(f"    {g:<10} → {t['muertes']:>4} muertes / {t['denominador']:>6,} den → {t['tasa_mensual_pct']}%")

    return {
        "meta": {
            "generado":       datetime.now().isoformat(),
            "periodo":        periodo,
            "tabla":          "V_MUERTES",
            "ventana":        "30 días",
            "desde":          hace_30.strftime("%Y-%m-%d"),
            "hasta":          hoy.strftime("%Y-%m-%d"),
            "label_periodo":  label_periodo,
            "filtro_dias":    ">30 días de encierre",
        },
        "mortandad": {
            "muertes_30d":      total_m,
            "ingresos_30d":     total_ing,
            "stock_haras_hoy":  total_stock,
            "denominador":      denominador,
            "tasa_mensual_pct": tasa_mens_p,
            "formula":          "tasa = muertes_30d / (ingresos_30d + stock_hoy)",
            "por_grupo":        tasas_grupo,
        },
        "detalle": {
            "total_muertes": total_m,
            "por_categoria": por_cat,
            "por_grupo":     por_grupo_m,
        },
    }

# ═══════════════════════════════════════════════════════════
#  GUARDAR JSON
# ═══════════════════════════════════════════════════════════
def guardar(datos, carpeta, nombre):
    """Guarda JSON de forma atómica (escribe en temp y renombra) para
    evitar que interrupciones del proceso produzcan archivos truncados."""
    import tempfile, os
    dest = Path(carpeta)
    dest.mkdir(parents=True, exist_ok=True)
    ruta = dest / nombre
    # Escribir en archivo temporal dentro del mismo directorio (mismo filesystem)
    fd, tmp_path = tempfile.mkstemp(dir=dest, suffix='.tmp', prefix=nombre+'_')
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(limpiar_nan(datos), f, ensure_ascii=False, indent=2, default=str)
        # Rename atómico: reemplaza el destino de forma segura
        os.replace(tmp_path, ruta)
    except Exception:
        # Si algo falla, eliminar el temp para no dejar basura
        try: os.unlink(tmp_path)
        except Exception: pass
        raise
    log.info(f"  Guardado: {ruta.name}  ({ruta.stat().st_size // 1024} KB)")
    return str(ruta)

# ═══════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════

def procesar_productivo(regs_egr, cols_egr, periodo):
    """
    Parámetros productivos desde v_PB_Egresos.
    Métricas: AdpSinDebaste (engorde diario) y Estadia (días en feedlot).
    Filtra MotivoSalida = VENTA. Últimos 365 días por FechaSalida.
    """
    from datetime import date, timedelta
    import pandas as pd

    hoy        = date.today()
    hace_anio  = hoy - timedelta(days=365)
    hace_90d   = hoy - timedelta(days=90)

    # v15.67: años base para los teóricos (ver ADP_BASE_ANIOS arriba)
    _base_anios = adp_base_anios(hoy)

    def _shift_anio(d, anio):
        """Misma fecha en otro año; 29/02 cae a 28/02 en año no bisiesto."""
        try:
            return d.replace(year=anio)
        except ValueError:
            return d.replace(year=anio, day=28)

    # Ventana equivalente a "últimos 90 días" en cada año base
    _base_trim = [(_shift_anio(hace_90d, a), _shift_anio(hoy, a)) for a in _base_anios]
    # Fecha más vieja que hace falta conservar para poder calcular las bases
    _corte_viejo = min([date(min(_base_anios), 1, 1)] + [w[0] for w in _base_trim])

    # ADP teórico por categoría — con filtros por estadía Y peso de entrada (igual que Excel)
    # Fuente: "Aumento Proyectado dieta a fecha (analisis anual).xlsx" — hoja resumen
    #   feedlot entero (días 30-400, pesoE 50-500): ADP=1.376
    _ADP_TEO = {
        'TERNERO':    1.371,  # ternero macho:   días 100-450, pesoE 0-200,   N=87
        'TERNERA':    1.324,  # ternero hembra:  días 100-450, pesoE 0-200,   N=19
        'NOVILLITO':  1.489,  # novillito:       días 30-350,  pesoE 200-400, N=691
        'NOVILLO':    1.231,  # novillos pesado: días 30-350,  pesoE 350-750, N=188
        'VAQUILLONA': 1.346,  # vaquillona:      días 30-350,  pesoE 200-400, N=685
        'VACA':       1.399,  # vacas engorde:   días 30-350,  pesoE 0-750,   N=1727
        'TORO':       1.60,   # toro (sin datos propios, referencia anterior)
    }

    # Filtros per-categoría: estadía (días) y peso de entrada (kg)
    # Replica exactamente los rangos de la hoja "resumen" del Excel
    _CAT_FILTROS = {
        'TERNERO':    {'est_min': 100, 'est_max': 450, 'pe_min':   0, 'pe_max': 200},
        'TERNERA':    {'est_min': 100, 'est_max': 450, 'pe_min':   0, 'pe_max': 200},
        'NOVILLITO':  {'est_min':  30, 'est_max': 350, 'pe_min': 200, 'pe_max': 400},
        'NOVILLO':    {'est_min':  30, 'est_max': 350, 'pe_min': 350, 'pe_max': 750},
        'VAQUILLONA': {'est_min':  30, 'est_max': 350, 'pe_min': 200, 'pe_max': 400},
        'VACA':       {'est_min':  30, 'est_max': 350, 'pe_min':   0, 'pe_max': 750},
        'TORO':       {'est_min':  30, 'est_max': 400, 'pe_min':   0, 'pe_max': 1000},
    }

    # v15.5.1 HOTFIX: adapter WinCampo Web (fetch_egresos) devuelve Categoria
    # como código de 2 letras (TM/VA/TH/NV/VQ/NT/TO). _ADP_TEO y _CAT_FILTROS
    # usan nombres largos como claves históricas (TERNERO/VACA/etc.). Mapear
    # código -> largo antes de cada lookup para que clamp ±25% y filtros
    # per-categoría sigan funcionando como pre-v15.5 (commit ac7ad4c).
    CATEGORIA_CODE_TO_LARGO = {
        'TM': 'TERNERO',    'TH': 'TERNERA',
        'NT': 'NOVILLITO',  'NV': 'NOVILLO',
        'VQ': 'VAQUILLONA', 'VA': 'VACA',
        'TO': 'TORO',
    }

    def _cat_largo(cat):
        """Normaliza categoría a nombre largo para lookup en _ADP_TEO/_CAT_FILTROS."""
        k = (cat or '').strip().upper()
        return CATEGORIA_CODE_TO_LARGO.get(k, k)  # si ya viene larga, deja como está

    # Detectar columnas
    def fc(nombres):
        cl = {c.lower(): c for c in cols_egr}
        for n in nombres:
            if n.lower() in cl:
                return cl[n.lower()]
        return None

    col_fecha   = fc(["FechaSalida","fechasalida","fecha_salida","fecha"])
    col_motivo  = fc(["MotivoSalida","motivosalida","motivo_salida","motivo"])
    col_adp     = fc(["AdpSinDebaste","adpsindebaste","adp_sin_debaste","adp"])
    col_estadia = fc(["Estadia","estadia","estadía","dias_estadia","dias_encierre"])
    col_cat     = fc(["Categoria","categoria","category","cat"])
    col_cab     = fc(["Cantidad","cantidad","cabezas"])
    col_rfid    = fc(["RFID","rfid"])
    col_pesoe   = fc(["KgIngreso","kgingreso","PesoEntrada","pesoentrada","peso_entrada","kg_entrada","KgEntrada"])

    log.info(f"  Productivo | fecha={col_fecha} motivo={col_motivo} adp={col_adp} estadia={col_estadia} cat={col_cat} cab={col_cab} rfid={col_rfid} pesoe={col_pesoe}")
    log.info(f"  Todas las columnas de v_PB_Egresos: {cols_egr}")

    # Filtrar con todos los criterios de calidad
    regs = []
    excl = {"motivo": 0, "rfid": 0, "estadia": 0, "adp": 0, "fecha": 0}
    for r in regs_egr:
        # 1) Filtro fecha
        # v15.67: se conservan también los años base (para calcular los teóricos).
        # El recorte a "último año" se hace después, con la marca _en_anio.
        try:
            f = pd.to_datetime(r.get(col_fecha), errors="coerce") if col_fecha else None
            if f is None or pd.isnull(f): excl["fecha"] += 1; continue
            if f.date() < _corte_viejo:   excl["fecha"] += 1; continue
        except: excl["fecha"] += 1; continue

        # 2) MotivoSalida = VENTA
        # v15.5: tolera codigo 1-letra "V" (API WinCampo Web) y string largo
        # "VENTA con destino X" (SQL viejo). Mismo patron que filtrar_solo_venta
        # del modulo movimientos productivos.
        if col_motivo:
            mot = str(r.get(col_motivo) or "").strip().upper()
            if not (mot == "V" or "VENTA" in mot):
                excl["motivo"] += 1; continue

        # 3) RFID solo numérico (si tiene alguna letra, se descarta)
        if col_rfid:
            rfid = str(r.get(col_rfid) or "").strip()
            if rfid and not rfid.isdigit(): excl["rfid"] += 1; continue

        # 4) Estadia: rango amplio 0-450 días (el filtro fino se aplica por categoría)
        est = to_num(r.get(col_estadia)) if col_estadia else None
        if est is None or not (0 < est <= 450): excl["estadia"] += 1; continue

        # 5) AdpSinDebaste entre 0 y 5
        adp = to_num(r.get(col_adp)) if col_adp else None
        if adp is None or not (0 < adp <= 5): excl["adp"] += 1; continue

        # Marcar si cae en los últimos 90 días
        r = dict(r)
        # Guardar peso de entrada para filtros per-categoría
        if col_pesoe:
            r['_pesoe'] = to_num(r.get(col_pesoe))
        try:
            _fd = f.date()
            r['_en_90d']  = (_fd >= hace_90d)
            r['_en_anio'] = (_fd >= hace_anio)
            # v15.67: marcas de base histórica
            r['_base_anual'] = (_fd.year in _base_anios)
            r['_base_trim']  = any(a <= _fd <= b for a, b in _base_trim)
        except:
            r['_en_90d'] = False; r['_en_anio'] = False
            r['_base_anual'] = False; r['_base_trim'] = False
        regs.append(r)

    log.info(f"  Registros totales: {len(regs_egr):,}")
    log.info(f"  Excluidos por fecha:   {excl['fecha']:,}")
    log.info(f"  Excluidos por motivo:  {excl['motivo']:,}")
    log.info(f"  Excluidos por RFID:    {excl['rfid']:,}")
    log.info(f"  Excluidos por estadía: {excl['estadia']:,}")
    log.info(f"  Excluidos por ADP:     {excl['adp']:,}")
    log.info(f"  Registros válidos:     {len(regs):,}")

    if not regs:
        return {
            "meta": {"generado": datetime.now().isoformat(), "periodo": periodo, "tabla": "v_PB_Egresos",
                     "ventana": "365 días",
                     "filtros": "MotivoSalida=VENTA | RFID numérico | Estadía 0-450d | ADP 0-5 | filtros per-cat (est+pesoE)",
                     "registros_filtrados": 0},
            "general": {}, "por_categoria": {}, "por_mes": {}
        }

    # Función para calcular promedios ponderados por cabezas
    # cat: si se provee, aplica los filtros de estadía y peso de entrada de _CAT_FILTROS
    def calc_stats(rows, cat=None):
        filt = _CAT_FILTROS.get(_cat_largo(cat), {}) if cat else {}
        est_min = filt.get('est_min', 0)
        est_max = filt.get('est_max', 9999)
        pe_min  = filt.get('pe_min',  0)
        pe_max  = filt.get('pe_max',  9999)
        adp_vals, est_vals = [], []
        cab_ok = 0
        for r in rows:
            cab  = to_num(r.get(col_cab, 1) if col_cab else 1) or 1
            est  = to_num(r.get(col_estadia) if col_estadia else None)
            adp  = to_num(r.get(col_adp) if col_adp else None)
            peso = r.get('_pesoe') if cat else None
            # Filtro per-categoría: estadía
            if cat and est is not None:
                if not (est_min <= est <= est_max):
                    continue
            # Filtro per-categoría: peso de entrada (solo si columna disponible)
            if cat and col_pesoe and peso is not None:
                if not (pe_min <= peso <= pe_max):
                    continue
            cab_ok += int(round(cab))
            if adp is not None and adp > 0:
                adp_vals.extend([adp] * int(round(cab)))
            if est is not None and est > 0:
                est_vals.extend([est] * int(round(cab)))
        return {
            "cabezas":       cab_ok or len(rows),
            "adp_promedio":  round(sum(adp_vals)/len(adp_vals), 3) if adp_vals else None,
            "adp_min":       round(min(adp_vals), 3) if adp_vals else None,
            "adp_max":       round(max(adp_vals), 3) if adp_vals else None,
            "estadia_promedio": round(sum(est_vals)/len(est_vals), 1) if est_vals else None,
            "estadia_min":      int(min(est_vals)) if est_vals else None,
            "estadia_max":      int(max(est_vals)) if est_vals else None,
        }

    # v15.67: 'regs' ahora arrastra también los años base. El último año —que es
    # lo que muestran General / Por categoría / Por mes— es este subconjunto.
    regs_anio = [r for r in regs if r.get('_en_anio')]
    log.info(f"  Registros último año:  {len(regs_anio):,}  (de {len(regs):,} conservados)")

    # ── v15.67: teóricos calculados sobre los años base ──────────────────────
    def _teoricos(rows_filtro):
        """{cat: {'adp':x, 'cabezas':n}} aplicando los filtros per-categoría."""
        out = {}
        if not col_cat: return out
        agr = {}
        for r in regs:
            if not rows_filtro(r): continue
            c = str(r.get(col_cat) or "Sin datos").strip().upper()
            agr.setdefault(c, []).append(r)
        for c, rows in agr.items():
            st = calc_stats(rows, cat=c)
            out[c] = {"adp": st.get("adp_promedio"), "cabezas": st.get("cabezas") or 0}
        return out

    _teo_anual = _teoricos(lambda r: r.get('_base_anual'))
    _teo_trim  = _teoricos(lambda r: r.get('_base_trim'))
    log.info(f"  Teórico base ANUAL  ({_base_anios}): "
             + ", ".join(f"{c}={v['adp']} (n={v['cabezas']})" for c, v in sorted(_teo_anual.items())))
    log.info(f"  Teórico base 90d    ({_base_anios}): "
             + ", ".join(f"{c}={v['adp']} (n={v['cabezas']})" for c, v in sorted(_teo_trim.items())))

    def _teo_para(cat, preferir_trim):
        """Cadena de fallback: base pedida -> base anual -> tabla del Excel.

        Criterio Nicolás 31/08/2026: si el trimestre no junta cabezas, se usa el
        teórico ANUAL de los mismos años base (no la tabla), para quedarse dentro
        de los datos buenos. La tabla es el último recurso."""
        c = str(cat or "").strip().upper()
        if preferir_trim:
            v = _teo_trim.get(c)
            if v and v["adp"] and v["cabezas"] >= ADP_BASE_MIN_CAB:
                return v["adp"], f"base 90d {_base_anios} (n={v['cabezas']})"
        v = _teo_anual.get(c)
        if v and v["adp"] and v["cabezas"] >= ADP_BASE_MIN_CAB:
            return v["adp"], f"base anual {_base_anios} (n={v['cabezas']})"
        t = _ADP_TEO.get(_cat_largo(c))
        return (round(t, 4) if t else None), "tabla Excel (sin base suficiente)"

    # General
    general = calc_stats(regs_anio)
    log.info(f"  ADP prom: {general.get('adp_promedio')} kg/día | Estadía prom: {general.get('estadia_promedio')} días")

    # Por categoría
    por_cat = {}
    if col_cat:
        cat_regs = {}
        for r in regs_anio:
            cat = str(r.get(col_cat) or "Sin datos").strip().upper()
            cat_regs.setdefault(cat, []).append(r)
        for cat, rows in sorted(cat_regs.items()):
            st  = calc_stats(rows, cat=cat)
            obs = st.get('adp_promedio')
            teo, fuente = _teo_para(cat, preferir_trim=False)
            st["adp_teorico"]    = teo
            st["teorico_fuente"] = fuente
            st["variacion_pct"]  = round((obs - teo) / teo * 100, 2) if (obs is not None and teo) else None
            por_cat[cat] = st

    # Por mes
    por_mes = {}
    if col_fecha:
        mes_regs = {}
        for r in regs_anio:
            try:
                f = pd.to_datetime(r.get(col_fecha), errors="coerce")
                mes = f.strftime("%Y-%m") if f and not pd.isnull(f) else "Sin fecha"
            except: mes = "Sin fecha"
            mes_regs.setdefault(mes, []).append(r)
        for mes in sorted(mes_regs.keys()):
            if mes != "Sin fecha":
                por_mes[mes] = calc_stats(mes_regs[mes])

    # Por categoría — últimos 90 días con comparación vs ADP teórico
    regs_90d    = [r for r in regs_anio if r.get('_en_90d')]
    por_cat_90d = {}
    if col_cat and regs_90d:
        cat_regs_90 = {}
        for r in regs_90d:
            cat = str(r.get(col_cat) or "Sin datos").strip().upper()
            cat_regs_90.setdefault(cat, []).append(r)
        for cat, rows in sorted(cat_regs_90.items()):
            st = calc_stats(rows, cat=cat)
            obs = st.get('adp_promedio')
            # v15.67: teórico = misma ventana de 90d de los años base (fallback
            # a la base anual y, en último caso, a la tabla del Excel)
            teo, _teo_fuente = _teo_para(cat, preferir_trim=True)
            # Variación % entre observado y teórico
            var_pct = round((obs - teo) / teo * 100, 2) if (obs is not None and teo) else None
            # Calibrado: clampear obs a ±ADP_CLAMP_TOL del teórico
            if obs is not None and teo:
                lo, hi = teo * (1 - ADP_CLAMP_TOL), teo * (1 + ADP_CLAMP_TOL)
                cal     = round(max(lo, min(hi, obs)), 4)
                ajust   = (obs < lo or obs > hi)
            else:
                cal   = obs
                ajust = False
            por_cat_90d[cat] = {
                **st,
                "adp_teorico":   round(teo, 4) if teo else None,
                "adp_calibrado": cal,
                "variacion_pct": var_pct,
                "ajustado":      ajust,
                "adp_min_range": round(teo * (1 - ADP_CLAMP_TOL), 4) if teo else None,
                "adp_max_range": round(teo * (1 + ADP_CLAMP_TOL), 4) if teo else None,
                "teorico_fuente": _teo_fuente,
            }
    log.info(f"  por_categoria_90d: {len(por_cat_90d)} categorías ({len(regs_90d)} registros en 90d)")

    return {
        "meta": {
            "generado":            datetime.now().isoformat(),
            "periodo":             periodo,
            "tabla":               "v_PB_Egresos",
            "ventana":             "365 días",
            "filtros":             "MotivoSalida=VENTA | RFID numérico | Estadía 30-365d | ADP 0-5",
            "base_anios":          _base_anios,
            "base_min_cabezas":    ADP_BASE_MIN_CAB,
            "clamp_tolerancia":    ADP_CLAMP_TOL,
            "base_trim_ventanas":  [[a.isoformat(), b.isoformat()] for a, b in _base_trim],
            "teorico_90d":         f"ventana de 90 días de {_base_anios}",
            "teorico_anual":       f"año calendario completo de {_base_anios}",
            "registros_totales":   len(regs_egr),
            "registros_conservados": len(regs),
            "registros_filtrados": len(regs_anio),
            "excluidos":           excl,
            "col_adp":             col_adp,
            "col_estadia":         col_estadia,
            "col_rfid":            col_rfid,
        },
        "general":          general,
        "por_categoria":    por_cat,
        "por_categoria_90d": por_cat_90d,
        "por_mes":          por_mes,
    }


def procesar_consumo(regs, cols, periodo):
    """
    Consumo de alimento desde v_PB_ConsumoDetallado.
    - Total anual por insumo (últimos 365 días): suma KILOS_TC_INSUMO agrupado por DESC_INSUMO
    - Promedio diario últimos 7 días: suma KILOS_TC_INSUMO de los últimos 7 días / 7
    """
    from datetime import date, timedelta
    import pandas as pd

    hoy       = date.today()
    hace_anio = hoy - timedelta(days=365)
    hace_30d  = hoy - timedelta(days=30)   # ventana amplia para detectar días con registros

    # Columnas exactas confirmadas por screenshot
    col_fecha   = "FECHA"      if "FECHA"      in (regs[0] if regs else {}) else next((c for c in cols if c.upper()=="FECHA"), None)
    col_kg      = "KILOS_TC_INSUMO" if "KILOS_TC_INSUMO" in (regs[0] if regs else {}) else next((c for c in cols if c.upper()=="KILOS_TC_INSUMO"), None)
    col_desc    = "DESC_INSUMO" if "DESC_INSUMO" in (regs[0] if regs else {}) else next((c for c in cols if c.upper()=="DESC_INSUMO"), None)
    col_cod     = "COD_INSUMO"  if "COD_INSUMO"  in (regs[0] if regs else {}) else next((c for c in cols if c.upper()=="COD_INSUMO"), None)

    log.info(f"  Consumo | fecha={col_fecha} kg={col_kg} desc={col_desc} cod={col_cod}")
    log.info(f"  Columnas v_PB_ConsumoDetallado: {cols}")

    if not regs or not col_fecha or not col_kg:
        log.warning("  ⚠ Sin datos de consumo o columnas no encontradas")
        return {
            "meta": {"generado": datetime.now().isoformat(), "periodo": periodo,
                     "tabla": "v_PB_ConsumoDetallado", "registros": 0},
            "anual": {"total_kg": 0, "por_insumo": []},
            "semanal": {"desde": str(hoy), "hasta": str(hoy), "total_kg_3d": 0,
                        "promedio_diario_kg": 0, "promedio_diario_kg_ms": 0, "por_insumo": []},
        }

    # Parsear fechas y filtrar último año + últimos 30 días
    regs_anio      = []
    regs_recientes = []   # últimos 30 días (ventana para encontrar los 3 días con datos)
    for r in regs:
        try:
            f = pd.to_datetime(r.get(col_fecha), errors="coerce")
            if f is None or pd.isnull(f): continue
            fd = f.date()
            if fd < hace_anio: continue
            regs_anio.append(r)
            if fd >= hace_30d:
                regs_recientes.append(r)
        except:
            continue

    # ── Detectar los últimos 3 días únicos con registros ──
    dias_recientes_set = set()
    for r in regs_recientes:
        try:
            f = pd.to_datetime(r.get(col_fecha), errors="coerce")
            if f and not pd.isnull(f):
                dias_recientes_set.add(f.date().strftime("%Y-%m-%d"))
        except: pass
    ultimos_3_dias = set(sorted(dias_recientes_set, reverse=True)[:3])
    regs_3d = [r for r in regs_recientes
               if not pd.isnull(pd.to_datetime(r.get(col_fecha), errors="coerce"))
               and pd.to_datetime(r.get(col_fecha), errors="coerce").date().strftime("%Y-%m-%d") in ultimos_3_dias]
    desde_3d = min(ultimos_3_dias) if ultimos_3_dias else str(hoy)
    hasta_3d = max(ultimos_3_dias) if ultimos_3_dias else str(hoy)

    log.info(f"  Registros último año: {len(regs_anio):,}  |  Últimos 3 días registrados: {sorted(ultimos_3_dias)}")

    # Tabla de materia seca por insumo (nombre exacto → % MS)
    # Valores actualizados por Nicolás · 2026-08 (Dirección) — v15.64
    MS_PCT = {
        "GLUTEN DE MAIZ":       51.0,   # 53 → 51 (Dir 2026-08)
        "MAIZ GRANO":           86.0,   # 89 → 86 (Dir 2026-08)
        "SILO DE MAIZ":         47.0,   # 58 → 47 (Dir 2026-08)
        "HARINA GERMEN":        99.0,   # sin cambio
        "NUCLEO CONC 5% LDB":   95.5,   # 98 → 95.5 (Dir 2026-08)
    }
    def get_ms(desc):
        return MS_PCT.get(desc.strip().upper(), None)

    # ── Totales anuales por insumo ──
    anual_por_ins = {}   # desc -> {cod, kg}
    total_anual   = 0.0
    for r in regs_anio:
        desc = str(r.get(col_desc) or "Sin descripción").strip() if col_desc else "Sin descripción"
        cod  = str(r.get(col_cod)  or "").strip()               if col_cod  else ""
        kg   = to_num(r.get(col_kg, 0))
        total_anual += kg
        if desc not in anual_por_ins:
            anual_por_ins[desc] = {"cod": cod, "kg": 0.0}
        anual_por_ins[desc]["kg"] += kg

    # Redondear y ordenar por kg desc
    por_insumo_anual = sorted(
        [{"desc": d, "cod": v["cod"], "kg": round(v["kg"], 1),
          "ms_pct": get_ms(d),
          "kg_ms":  round(v["kg"] * get_ms(d) / 100, 1) if get_ms(d) is not None else None}
         for d, v in anual_por_ins.items()],
        key=lambda x: -x["kg"]
    )
    total_anual_ms = round(sum(r["kg_ms"] for r in por_insumo_anual if r["kg_ms"] is not None), 1)
    log.info(f"  Total anual: {total_anual:,.0f} kg  |  Insumos distintos: {len(por_insumo_anual)}")
    for ins in por_insumo_anual[:5]:
        log.info(f"    {ins['desc']:<30} {ins['kg']:>12,.1f} kg")

    # ── Promedio diario últimos 3 días registrados ──
    semanal_por_ins = {}
    dias_con_datos  = set()
    total_3d = 0.0
    for r in regs_3d:
        desc = str(r.get(col_desc) or "Sin descripción").strip() if col_desc else "Sin descripción"
        cod  = str(r.get(col_cod)  or "").strip()               if col_cod  else ""
        kg   = to_num(r.get(col_kg, 0))
        total_3d += kg
        if desc not in semanal_por_ins:
            semanal_por_ins[desc] = {"cod": cod, "kg": 0.0, "dias": set()}
        semanal_por_ins[desc]["kg"] += kg
        try:
            fd = pd.to_datetime(r.get(col_fecha), errors="coerce")
            if fd and not pd.isnull(fd):
                dia_str = fd.strftime("%Y-%m-%d")
                dias_con_datos.add(dia_str)
                semanal_por_ins[desc]["dias"].add(dia_str)
        except: pass

    # Divisor = días únicos con registros en los últimos 3 (mínimo 1)
    n_dias = max(len(dias_con_datos), 1)
    log.info(f"  Últimos 3 días con registros: {n_dias} ({sorted(dias_con_datos)})")
    # ── DEBUG: listar TODOS los insumos detectados en los 3 últimos días ──
    log.info(f"  ╔═══ DEBUG · INSUMOS DETECTADOS EN 3 DÍAS ═══╗")
    log.info(f"  ║  Total kg suma: {total_3d:,.1f}")
    log.info(f"  ║  Insumos distintos: {len(semanal_por_ins)}")
    for desc, vd in sorted(semanal_por_ins.items(), key=lambda x: -x[1]['kg']):
        ms_marca = f"MS={get_ms(desc)}%" if get_ms(desc) is not None else "MS=??? (no en MS_PCT)"
        log.info(f"  ║   • {desc:<35} {vd['kg']:>12,.1f} kg en {len(vd['dias'])} dias [{ms_marca}]")
    log.info(f"  ╚════════════════════════════════════════════╝")

    por_insumo_3d = sorted(
        [{"desc": d, "cod": v["cod"],
          "kg_3d": round(v["kg"], 1),
          "dias_registrados": len(v["dias"]),
          "promedio_diario":    round(v["kg"] / max(len(v["dias"]), 1), 1),
          "ms_pct":             get_ms(d),
          "promedio_diario_ms": round(v["kg"] / max(len(v["dias"]), 1) * get_ms(d) / 100, 1)
                                if get_ms(d) is not None else None}
         for d, v in semanal_por_ins.items()],
        key=lambda x: -x["kg_3d"]
    )
    prom_diario_total    = round(total_3d / n_dias, 1)
    prom_diario_total_ms = round(sum(
        r["promedio_diario_ms"] for r in por_insumo_3d if r["promedio_diario_ms"] is not None
    ), 1)
    # % MS global = kg MS / kg TC × 100
    pct_ms_global = round(prom_diario_total_ms / prom_diario_total * 100, 1) if prom_diario_total > 0 else 0.0
    log.info(f"  Total 3d: {total_3d:,.0f} kg  |  Días: {n_dias}  |  Prom diario TC: {prom_diario_total:,.1f} kg/día  |  MS: {prom_diario_total_ms:,.1f} kg/día  |  %MS: {pct_ms_global:.1f}%")

    return {
        "meta": {
            "generado":    datetime.now().isoformat(),
            "periodo":     periodo,
            "tabla":       "v_PB_ConsumoDetallado",
            "col_kg":      col_kg,
            "desde_anual": str(hace_anio),
            "hasta":       str(hoy),
            "registros_anio": len(regs_anio),
            "registros_3d":   len(regs_3d),
        },
        "anual": {
            "total_kg":    round(total_anual, 1),
            "total_kg_ms": total_anual_ms,
            "por_insumo":  por_insumo_anual,
        },
        "semanal": {
            "desde":                 desde_3d,
            "hasta":                 hasta_3d,
            "dias_registrados":      n_dias,
            "dias_detalle":          sorted(dias_con_datos),
            "total_kg_3d":           round(total_3d, 1),
            "promedio_diario_kg":    prom_diario_total,
            "promedio_diario_kg_ms": prom_diario_total_ms,
            "pct_ms_global":         pct_ms_global,
            "por_insumo":            por_insumo_3d,
        },
    }

# ═══════════════════════════════════════════════════════════
#  MÓDULO 10 · VALUACIÓN EN PESOS
#  Scraping de precios externos:
#   · MAG  → Índice Arrendamiento ($/kg hacienda) por mes
#   · BCR  → Precio pizarra promedio Maíz y Soja ($/ton) por mes
#  Calcula valuación total mensual: Hacienda + Insumos + Financiero + USD
# ═══════════════════════════════════════════════════════════

def _ar_num(s):
    """Convierte número formato argentino '1.234.567,89' a float.

    v15.49: se limpia también el signo '$'. BCR agregó el símbolo de moneda a
    la celda de promedio de la pizarra alrededor de 2026-05 ('$272.348,05'), y
    float('$272.348,05') tiraba ValueError → _scrap_bcr_precio devolvía None →
    forward-fill silencioso de precios viejos durante 3 meses. La página seguía
    respondiendo 200 con el dato: el bug era de parseo, no de transporte.
    """
    if s is None:
        return None
    s = str(s).strip().replace('\xa0', '').replace(' ', '').replace('$', '')
    if not s or s == '-' or s == '—':
        return None
    try:
        return float(s.replace('.', '').replace(',', '.'))
    except ValueError:
        return None


def _mes_rango(periodo_str):
    """
    Dado 'YYYY-MM' devuelve (primer_dia, ultimo_dia) como datetime.date.
    """
    import calendar
    from datetime import date as _date
    year, month = int(periodo_str[:4]), int(periodo_str[5:7])
    ultimo = calendar.monthrange(year, month)[1]
    return _date(year, month, 1), _date(year, month, ultimo)


def _html_tabla(html_bytes, encoding='latin-1'):
    """
    Extrae filas de la primera tabla HTML encontrada.
    Devuelve lista de listas de strings (texto de cada celda).
    """
    from html.parser import HTMLParser

    class _TP(HTMLParser):
        def __init__(self):
            super().__init__()
            self.rows, self._row, self._cell, self._in = [], [], [], False
        def handle_starttag(self, tag, attrs):
            if tag in ('td', 'th'):
                self._in = True; self._cell = []
            elif tag == 'tr':
                self._row = []
        def handle_endtag(self, tag):
            if tag in ('td', 'th'):
                self._row.append(''.join(self._cell).strip()); self._in = False
            elif tag == 'tr':
                if self._row:
                    self.rows.append(self._row)
        def handle_data(self, data):
            if self._in:
                self._cell.append(data)

    for enc in (encoding, 'utf-8', 'latin-1', 'cp1252'):
        try:
            txt = html_bytes.decode(enc, errors='replace')
            break
        except Exception:
            continue

    p = _TP()
    p.feed(txt)
    return p.rows


def _scrap_mag_indice(periodo_str):
    """
    Scraping MAG: devuelve Índice Arrendamiento promedio del mes ($/kg hacienda)
    para el período 'YYYY-MM'.  Retorna float o None si no hay datos.
    """
    import urllib.request, urllib.parse

    primer, ultimo = _mes_rango(periodo_str)
    fi = primer.strftime('%d/%m/%Y')
    ff = ultimo.strftime('%d/%m/%Y')

    payload = urllib.parse.urlencode({
        'ID': '', 'CP': '', 'FLASH': '',
        'USUARIO': 'SIN IDENTIFICAR',
        'OPCIONMENU': '', 'OPCIONSUBMENU': '',
        'txtFechaIni': fi,
        'txtFechaFin': ff,
    }).encode('utf-8')

    url = 'https://www.mercadoagroganadero.com.ar/dll/hacienda2.dll/haciinfo000013'
    try:
        req = urllib.request.Request(url, data=payload, method='POST')
        req.add_header('User-Agent', 'Mozilla/5.0 (compatible; PEGSA-Bot/1.0)')
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read()
    except Exception as e:
        log.warning(f'    MAG request error ({periodo_str}): {e}')
        return None

    filas = _html_tabla(raw, encoding='latin-1')
    # Buscar fila "Totales" → columna 3 = Índice Arrendamiento
    # Encabezado: Fecha | Cab. ingresadas | Importe | Índice Arrendamiento | Variación
    for fila in filas:
        if fila and 'total' in fila[0].lower():
            if len(fila) >= 4:
                val = _ar_num(fila[3])
                if val:
                    log.info(f'    MAG {periodo_str}: índice={val:,.3f} $/kg')
                    return val
    log.warning(f'    MAG {periodo_str}: sin datos en la tabla')
    return None


def _scrap_bcr_precio(product_id, nombre, periodo_str):
    """
    Scraping BCR Cámara Arbitral: precio pizarra promedio mensual ($/ton).
    product_id: 3=Maíz, 13=Soja
    Retorna float ($/ton) o None.
    """
    import urllib.request

    primer, ultimo = _mes_rango(periodo_str)
    ds = primer.strftime('%Y-%m-%d')
    de = ultimo.strftime('%Y-%m-%d')

    url = (f'https://www.cac.bcr.com.ar/es/precios-de-pizarra/consultas'
           f'?product={product_id}&type=average&date_start={ds}&date_end={de}')
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'Mozilla/5.0 (compatible; PEGSA-Bot/1.0)')
        with urllib.request.urlopen(req, timeout=20) as resp:
            raw = resp.read()
    except Exception as e:
        # v15.49: loguear el TIPO de excepción, no solo el mensaje — permite
        # distinguir timeout / 403 / problema de certificado de un vistazo.
        log.warning(f'    BCR {nombre} request error ({periodo_str}): {type(e).__name__}: {e}')
        return None

    filas = _html_tabla(raw, encoding='utf-8')
    # Estructura: [nombre], [Fecha Desde, Fecha Hasta, Promedio], [val, val, PRECIO]
    for fila in filas:
        if len(fila) >= 3 and fila[0].startswith('0') and '/' in fila[0]:
            # Fila de datos: primera celda es una fecha
            val = _ar_num(fila[2])
            if val:
                log.info(f'    BCR {nombre} {periodo_str}: {val:,.2f} $/ton')
                return val
    # v15.49: dumpear lo que se recibió para que el próximo diagnóstico no
    # arranque de cero (así se detecta al toque si BCR vuelve a cambiar el
    # formato de la tabla, como el signo $ agregado en 2026-05).
    log.warning(f'    BCR {nombre} {periodo_str}: sin datos en la tabla. '
                f'Filas parseadas: {len(filas)} · primeras 3: {filas[:3]}')
    return None


_BCR_EXCEL_CACHE = None   # {'maiz': {'YYYY-MM': prom}, 'soja': {...}}

def _cargar_precios_bcr_excel(carpeta):
    """
    v15.49 — Fallback local para precios de pizarra BCR.

    Lee TODOS los .xlsx de datos/precios_bcr/ y arma promedios mensuales por
    producto. Red de contención por si el scraping vuelve a fallar (BCR ya
    cambió el acceso/formato dos veces: v14.2 y el signo $ de v15.49).

    El producto se identifica por el CONTENIDO de la celda A4 ('Maíz'/'Soja'),
    NO por el nombre del archivo — el usuario los baja de BCR con nombres
    arbitrarios. Verificado: el promedio mensual de estos diarios coincide al
    centavo con el promedio mensual que publica BCR (julio 2026 maíz 272.348,05).

    Returns: {'maiz': {'2026-07': 272348.05, ...}, 'soja': {...}}
    """
    global _BCR_EXCEL_CACHE
    if _BCR_EXCEL_CACHE is not None:
        return _BCR_EXCEL_CACHE

    import unicodedata, fnmatch
    from collections import defaultdict

    def _norm(s):
        s = unicodedata.normalize('NFKD', str(s or ''))
        return ''.join(c for c in s if not unicodedata.combining(c)).strip().lower()

    base = Path(carpeta) / 'precios_bcr'
    acum = {'maiz': defaultdict(list), 'soja': defaultdict(list)}

    if not base.exists():
        log.info('    precios_bcr/ no existe — sin fallback de Excel')
        _BCR_EXCEL_CACHE = {'maiz': {}, 'soja': {}}
        return _BCR_EXCEL_CACHE

    # v15.44: matching case-insensitive (Path.glob es case-sensitive en Python
    # aunque estés en Windows)
    archivos = [p for p in base.iterdir()
                if p.is_file() and fnmatch.fnmatch(p.name.lower(), '*.xlsx')
                and not p.name.startswith('~$')]

    import openpyxl as _oxl
    for ruta in archivos:
        try:
            wb = _oxl.load_workbook(str(ruta), read_only=True, data_only=True)
            ws = wb.active

            # Identificar el producto: buscar 'maiz'/'soja' en las primeras 10
            # filas de la col A (normalmente A4, pero no lo damos por sentado)
            prod = None
            for r in range(1, 11):
                v = _norm(ws.cell(r, 1).value)
                if v == 'maiz': prod = 'maiz'; break
                if v == 'soja': prod = 'soja'; break
            if not prod:
                log.warning(f'    precios_bcr: {ruta.name} — no se pudo identificar el producto, se omite')
                wb.close(); continue

            n = 0
            for row in ws.iter_rows(min_row=1, max_col=2, values_only=True):
                fecha, precio = row[0], row[1]
                if not hasattr(fecha, 'strftime'):
                    continue                      # header o fila vacía
                try:
                    precio = float(precio)
                except (TypeError, ValueError):
                    continue
                if precio <= 0:
                    continue
                acum[prod][fecha.strftime('%Y-%m')].append(precio)
                n += 1
            wb.close()
            log.info(f'    precios_bcr: {ruta.name} -> {prod}, {n} precios diarios')
        except Exception as e:
            log.warning(f'    precios_bcr: error leyendo {ruta.name}: {type(e).__name__}: {e}')

    out = {}
    for prod, meses in acum.items():
        out[prod] = {m: round(sum(v)/len(v), 2) for m, v in meses.items() if v}
        if out[prod]:
            ks = sorted(out[prod])
            log.info(f'    precios_bcr {prod}: {len(ks)} meses ({ks[0]} a {ks[-1]})')

    _BCR_EXCEL_CACHE = out
    return out


def _scrap_bna_tc_historico(periodo):
    """
    Consulta el TC dólar Billete Venta del BNA para el último día hábil del mes.
    Fuente: https://www.bna.com.ar/Cotizador/HistoricasMonedas (POST form)
    periodo: "YYYY-MM"
    Retorna float o None si no hay datos disponibles.
    """
    import urllib.request, urllib.parse, calendar
    from datetime import date, timedelta

    try:
        year, month = int(periodo[:4]), int(periodo[5:7])
    except Exception:
        return None

    # Último día del mes
    last_day = calendar.monthrange(year, month)[1]
    fecha_fin = date(year, month, last_day)

    # Buscar desde 10 días antes (para cubrir fines de semana / feriados)
    fecha_ini = fecha_fin - timedelta(days=10)

    payload = urllib.parse.urlencode({
        'moneda':      '2',          # USD
        'cotizacion':  '1',          # Billete
        'fechaDesde':  fecha_ini.strftime('%d/%m/%Y'),
        'fechaHasta':  fecha_fin.strftime('%d/%m/%Y'),
    }).encode('utf-8')

    url = 'https://www.bna.com.ar/Cotizador/HistoricasMonedas'
    try:
        req = urllib.request.Request(url, data=payload, method='POST')
        req.add_header('User-Agent', 'Mozilla/5.0 (compatible; PEGSA-Bot/1.0)')
        req.add_header('Content-Type', 'application/x-www-form-urlencoded')
        req.add_header('Referer', 'https://www.bna.com.ar/Cotizador/HistoricasMonedas')
        with urllib.request.urlopen(req, timeout=15) as resp:
            raw = resp.read()
    except Exception as e:
        log.warning(f'    BNA TC histórico {periodo}: request error: {e}')
        return None

    try:
        txt = raw.decode('utf-8', errors='replace')
    except Exception:
        txt = raw.decode('latin-1', errors='replace')

    # Parsear tabla de resultados: columnas Fecha | Compra | Venta
    import re
    # Buscar filas <tr> con celdas de fecha DD/MM/YYYY y valores numéricos
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', txt, re.S | re.I)
    last_venta = None
    for row in rows:
        cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.S | re.I)
        if len(cells) >= 3:
            fecha_txt = re.sub(r'<[^>]+>', '', cells[0]).strip()
            venta_txt = re.sub(r'<[^>]+>', '', cells[2]).strip()
            # Verificar que sea una fila de fecha válida
            if re.match(r'\d{2}/\d{2}/\d{4}', fecha_txt):
                v = _ar_num(venta_txt)
                if v and v > 100:   # TC razonable > $100
                    last_venta = v

    if last_venta:
        log.info(f'    BNA TC histórico {periodo}: ${last_venta:,.2f}/USD')
    else:
        log.warning(f'    BNA TC histórico {periodo}: sin datos en respuesta')
    return last_venta


def _scrap_mep_tc_ambito(periodo):
    """
    Obtiene el dólar MEP promedio mensual desde la API de Ambito.
    Las empresas argentinas usan dólar MEP como referencia de valuación.
    periodo: "YYYY-MM"
    Retorna float (promedio mensual) o None si falla.
    Fuente: https://mercados.ambito.com/dolar/mep/historico-general/{d}/{m}/{a}/{d}/{m}/{a}
    """
    import urllib.request, calendar
    from datetime import date

    try:
        year, month = int(periodo[:4]), int(periodo[5:7])
    except Exception:
        return None

    last_day = calendar.monthrange(year, month)[1]
    d_ini = f'01/{month:02d}/{year}'
    d_fin = f'{last_day:02d}/{month:02d}/{year}'
    url = (f'https://mercados.ambito.com/dolar/mep/historico-general/'
           f'01/{month:02d}/{year}/{last_day:02d}/{month:02d}/{year}')
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'Mozilla/5.0 (compatible; PEGSA-Bot/1.0)')
        req.add_header('Referer', 'https://www.ambito.com/contenidos/dolar-mep-historico.html')
        req.add_header('Accept', 'application/json, text/plain, */*')
        with urllib.request.urlopen(req, timeout=15) as resp:
            raw = resp.read()
    except Exception as e:
        log.debug(f'    MEP TC Ambito {periodo}: request error: {e}')
        return None

    try:
        import json as _json
        obj = _json.loads(raw)
        # Respuesta: lista de [["fecha","valor",...], ...] con cabecera en [0]
        rows = obj if isinstance(obj, list) else obj.get('data', [])
        valores = []
        for row in rows:
            if not isinstance(row, list) or len(row) < 2:
                continue
            cell = str(row[1]).replace(',', '.').strip()
            try:
                v = float(cell)
                if v > 100:
                    valores.append(v)
            except Exception:
                pass
        if valores:
            promedio = round(sum(valores) / len(valores), 2)
            log.info(f'    MEP TC Ambito {periodo}: ${promedio:,.2f}/USD (promedio {len(valores)} días)')
            return promedio
        log.debug(f'    MEP TC Ambito {periodo}: sin valores válidos en respuesta')
        return None
    except Exception as e:
        log.debug(f'    MEP TC Ambito {periodo}: parse error: {e}')
        return None


def _scrap_bna_tc():
    """
    Scraping BNA: devuelve el tipo de cambio dólar Billete Venta del día actual.
    Fuente: https://www.bna.com.ar/Personas
    Tabla 0 = Billete: fila "Dolar U.S.A" → celda[2] = Venta
    Retorna float (ARS por USD) o None si falla.
    """
    import urllib.request

    url = 'https://www.bna.com.ar/Personas'
    try:
        req = urllib.request.Request(url)
        req.add_header('User-Agent', 'Mozilla/5.0 (compatible; PEGSA-Bot/1.0)')
        with urllib.request.urlopen(req, timeout=15) as resp:
            raw = resp.read()
    except Exception as e:
        log.warning(f'    BNA TC request error: {e}')
        return None

    # Parsear todas las tablas; tabla[0] = Billete, tabla[1] = Divisa
    from html.parser import HTMLParser

    class _AllTables(HTMLParser):
        def __init__(self):
            super().__init__()
            self.tables = []
            self._cur_table = None
            self._cur_row   = None
            self._cur_cell  = None
            self._depth     = 0

        def handle_starttag(self, tag, attrs):
            tag = tag.lower()
            if tag == 'table':
                self._cur_table = []
                self._depth += 1
            elif tag in ('tr',):
                if self._cur_table is not None and self._depth == 1:
                    self._cur_row = []
            elif tag in ('td', 'th'):
                if self._cur_row is not None:
                    self._cur_cell = []

        def handle_endtag(self, tag):
            tag = tag.lower()
            if tag == 'table':
                if self._cur_table is not None:
                    self.tables.append(self._cur_table)
                self._cur_table = None
                self._depth -= 1
            elif tag == 'tr':
                if self._cur_row is not None and self._cur_table is not None:
                    self._cur_table.append(self._cur_row)
                self._cur_row = None
            elif tag in ('td', 'th'):
                if self._cur_cell is not None and self._cur_row is not None:
                    self._cur_row.append(' '.join(self._cur_cell).strip())
                self._cur_cell = None

        def handle_data(self, data):
            if self._cur_cell is not None:
                self._cur_cell.append(data)

    try:
        txt = raw.decode('utf-8', errors='replace')
    except Exception:
        txt = raw.decode('latin-1', errors='replace')

    parser = _AllTables()
    parser.feed(txt)

    if not parser.tables:
        log.warning('    BNA TC: no se encontraron tablas en la página')
        return None

    # Tabla 0 = Billete
    tabla_billete = parser.tables[0]
    for fila in tabla_billete:
        if fila and 'dolar' in fila[0].lower():
            # celda[2] = Venta
            if len(fila) >= 3:
                val = _ar_num(fila[2])
                if val:
                    log.info(f'    BNA TC Billete Venta: ${val:,.2f}/USD')
                    return val
    log.warning('    BNA TC: no se encontró fila "Dolar U.S.A" en tabla Billete')
    return None


def actualizar_valuacion(carpeta, snaps_historico):
    """
    Calcula la valuación mensual en pesos para cada snapshot histórico.
    Componentes:
      · Hacienda PEGSA = kg_pegsa × indice_MAG ($/kg)
      · Insumos        = kg_maiz × precio_maiz/ton + kg_soja × precio_soja/ton
      · Financiero     = disponible + cartera - emitidos + cobrar_hac - pagar_hac + lcg + tercio - darwash   # v15.23
      #   (darwash = suma col B de la hoja 'cuenta  corriente con darwash' del propio archivo financiero)
      · USD            = usd_ars (ya convertido al TC del mes)
    Cachea precios scrapeados para no re-consultar períodos ya guardados.
    """
    val_path = Path(carpeta) / 'valuacion_historica.json'

    # v15.23: la posición Darwash ahora viene DENTRO de cada snapshot financiero
    # (campo 'darwash_pos', calculado en _parse_financiero_nuevo desde la hoja
    # 'cuenta  corriente con darwash' del propio archivo). Se eliminó el cargador
    # de tesoreria_darwash_historico.json de v15.22 (daba cifra parcial: $2.297M
    # en vez de $2.430M al 27-may, porque leía otra fuente). procesar_tesoreria_
    # darwash sigue generando ese JSON para el módulo Tesorería — no se toca.

    # ── TC Dólar MEP promedio mensual — Fuente: Ambito historico ──
    # Usados como fallback cuando scraping en tiempo real falla.
    # Valores = promedio de cotizaciones diarias del mes. Fuente: ambito.com/dolar-mep-historico
    _TC_APROX = {
        '2024-12': 1111.0,
        '2025-01': 1165.0,
        '2025-02': 1198.0,
        '2025-03': 1262.0,
        '2025-04': 1245.0,   # banda cambiaria abr-2025
        '2025-05': 1160.0,
        '2025-06': 1189.0,
        '2025-07': 1276.0,
        '2025-08': 1269.0,
        '2025-09': 1435.0,
        '2025-10': 1498.0,
        '2025-11': 1466.0,
        '2025-12': 1483.0,
        '2026-01': 1478.0,
        '2026-02': 1393.0,
        '2026-03': 1422.0,
    }

    # ── Cargar caché existente ──
    # Solo cachear TC que son "reales" (scraping o implícito), NO aproximados.
    cache = {}   # {periodo_str: {mag, bcr_maiz, bcr_soja, bna_tc_venta}}
    val_snaps_prev = {}
    if val_path.exists():
        try:
            with open(val_path, encoding='utf-8') as _f:
                _old = json.load(_f)
            for _s in _old.get('snapshots', []):
                p = _s.get('periodo', '')
                val_snaps_prev[p] = _s
                pr = _s.get('precios', {})
                # Cachear precios de commodities siempre
                if pr.get('mag_indice') or pr.get('bcr_maiz_ton') or pr.get('bcr_soja_ton'):
                    cache.setdefault(p, {}).update({
                        k: pr[k] for k in ('mag_indice','bcr_maiz_ton','bcr_soja_ton') if pr.get(k)
                    })
                # Cachear TC solo si fue obtenido por scraping real (no aproximado)
                tc_guardado  = pr.get('bna_tc_venta')
                tc_es_aprox  = (tc_guardado is not None and tc_guardado == _TC_APROX.get(p))
                tc_es_real   = (tc_guardado is not None and not tc_es_aprox)
                if tc_es_real:
                    cache.setdefault(p, {})['bna_tc_venta'] = tc_guardado
        except Exception:
            pass

    # ── TC implícito desde snapshots con usd_ars + usd_cant conocidos ──
    _tc_implicito_ref = None
    for snap in snaps_historico:
        fin_ = snap.get('financiero', {})
        ars_ = float(fin_.get('usd_ars')  or 0)
        cnt_ = float(fin_.get('usd_cant') or 0)
        if ars_ > 0 and cnt_ > 0:
            tc_imp = round(ars_ / cnt_, 2)
            per_   = snap.get('periodo', '')
            # Este TC es real — guardarlo en caché con prioridad
            cache.setdefault(per_, {})['bna_tc_venta'] = tc_imp
            log.info(f'    TC implícito {per_}: ${tc_imp:,.2f}/USD (usd_ars/usd_cant)')
            _tc_implicito_ref = tc_imp

    # ── Scraping BNA TC: actual (para período corriente) + histórico (para meses anteriores) ──
    from datetime import date as _dtoday
    _periodo_hoy = _dtoday.today().strftime('%Y-%m')
    log.info('  Consultando BNA TC actual...')
    _bna_tc_hoy = _scrap_bna_tc()   # TC del día de hoy (None si falla)

    nuevos_snaps = []

    # v15.21: Forward-fill de precios cuando el scraping devuelve None.
    # Probable cuando la BCR aún no publicó el precio mensual del último mes
    # cerrado (típico desfasaje). El forward-fill se aplica SOLO al cálculo de
    # componentes; los campos en 'precios' mantienen el valor scrapeado real
    # (None si falló) para que cuando se publique el precio real después, el
    # scraping/cache lo capture solo y se actualice sin intervención manual.
    _ult_mag      = None
    _ult_bcr_maiz = None
    _ult_bcr_soja = None
    _ult_bna_tc   = None

    # Garantizar orden ASC por periodo para que el forward-fill funcione.
    snaps_historico = sorted(snaps_historico, key=lambda s: s.get('periodo', ''))

    for snap in snaps_historico:
        periodo  = snap.get('periodo', '')
        hm       = snap.get('hacienda_masa', {})
        fin      = snap.get('financiero', {})
        ins      = snap.get('insumos', {})
        pegsa    = hm.get('pegsa', {})

        if not periodo:
            continue

        log.info(f'  · Valuación {periodo}')

        # ── 1. Obtener precios (con caché) ──
        cached = cache.get(periodo, {})

        mag_indice = cached.get('mag_indice')
        if mag_indice is None:
            mag_indice = _scrap_mag_indice(periodo)

        # v15.49: cadena de fuentes — cache → scraping BCR → Excel local. El
        # scraping va PRIMERO para que el sistema se auto-sane cuando BCR vuelva,
        # sin depender de que el usuario mantenga los archivos. El Excel es red de
        # contención. El forward-fill (más abajo) queda como último recurso pero
        # ahora casi nunca debería activarse. Un valor que viene del Excel es el
        # dato real del mes (no forward-fill): se captura antes de bcr_*_scraped,
        # así heredado queda en False.
        _bcr_xls = _cargar_precios_bcr_excel(carpeta)

        bcr_maiz = cached.get('bcr_maiz_ton')
        if bcr_maiz is None:
            bcr_maiz = _scrap_bcr_precio(3, 'Maíz', periodo)
        if bcr_maiz is None:
            bcr_maiz = _bcr_xls.get('maiz', {}).get(periodo)
            if bcr_maiz:
                log.info(f'    BCR maíz {periodo}: ${bcr_maiz:,.2f}/ton (Excel local)')

        bcr_soja = cached.get('bcr_soja_ton')
        if bcr_soja is None:
            bcr_soja = _scrap_bcr_precio(13, 'Soja', periodo)
        if bcr_soja is None:
            bcr_soja = _bcr_xls.get('soja', {}).get(periodo)
            if bcr_soja:
                log.info(f'    BCR soja {periodo}: ${bcr_soja:,.2f}/ton (Excel local)')

        # ── 1b. TC Dólar MEP: caché real → Ambito MEP → BNA histórico → aprox tabla ──
        # Las empresas argentinas usan dólar MEP como referencia de valuación.
        bna_tc = cached.get('bna_tc_venta')   # solo TC reales en caché
        if bna_tc is None:
            if periodo == _periodo_hoy:
                bna_tc = _bna_tc_hoy
            else:
                # 1) Intentar MEP promedio mensual desde Ambito (fuente preferida)
                bna_tc = _scrap_mep_tc_ambito(periodo)
                # 2) Fallback: BNA histórico (último día hábil del mes)
                if bna_tc is None:
                    bna_tc = _scrap_bna_tc_historico(periodo)
                    if bna_tc:
                        log.info(f'    BNA TC {periodo}: ${bna_tc:,.2f} (scraping BNA histórico)')
            # Fallback: TC aproximado por mes (promedios MEP Ambito conocidos)
            if bna_tc is None:
                bna_tc = _TC_APROX.get(periodo)
                if bna_tc:
                    log.info(f'    MEP TC {periodo}: ${bna_tc:,.0f} (tabla mensual MEP — aproximado)')
            # Último fallback: TC actual del día
            if bna_tc is None and _bna_tc_hoy:
                log.info(f'    TC {periodo}: ${_bna_tc_hoy:,.0f} (TC actual como fallback)')
                bna_tc = _bna_tc_hoy

        # ── 1c. v15.21 · Forward-fill. Si scraping/cache devolvió None, usar el
        # último mes con precio conocido. Las variables _ult_* se actualizan SOLO
        # con valores reales (no heredados) para no propagar herencias en cadena
        # cuando varios meses seguidos fallen el scraping.
        # IMPORTANTE: estos valores heredados se usan en el cálculo de componentes
        # (maiz_pesos, soja_pesos, hacienda_pesos, total_pesos), pero NO se
        # persisten en 'precios' — esos quedan con el valor scrapeado real (None
        # si falló) para que el próximo tick re-scrapee y obtenga el precio real
        # cuando la BCR lo publique.
        mag_indice_scraped = mag_indice
        bcr_maiz_scraped   = bcr_maiz
        bcr_soja_scraped   = bcr_soja
        bna_tc_scraped     = bna_tc

        if mag_indice is None and _ult_mag is not None:
            mag_indice = _ult_mag
            log.info(f'    MAG {periodo}: {mag_indice:,.2f} (forward-fill del mes anterior)')
        elif mag_indice is not None:
            _ult_mag = mag_indice

        if bcr_maiz is None and _ult_bcr_maiz is not None:
            bcr_maiz = _ult_bcr_maiz
            log.info(f'    BCR maíz {periodo}: ${bcr_maiz:,.0f}/ton (forward-fill del mes anterior)')
        elif bcr_maiz is not None:
            _ult_bcr_maiz = bcr_maiz

        if bcr_soja is None and _ult_bcr_soja is not None:
            bcr_soja = _ult_bcr_soja
            log.info(f'    BCR soja {periodo}: ${bcr_soja:,.0f}/ton (forward-fill del mes anterior)')
        elif bcr_soja is not None:
            _ult_bcr_soja = bcr_soja

        if bna_tc is None and _ult_bna_tc is not None:
            bna_tc = _ult_bna_tc
            log.info(f'    BNA TC {periodo}: ${bna_tc:,.2f}/USD (forward-fill del mes anterior)')
        elif bna_tc is not None:
            _ult_bna_tc = bna_tc

        # ── 2. Hacienda PEGSA en pesos ──
        kg_pegsa       = float(pegsa.get('kg_proyectado') or 0)
        hacienda_pesos = round(kg_pegsa * mag_indice) if mag_indice else None

        # ── 3. Insumos en pesos (solo Maíz y Soja) ──
        # items puede ser dict {nombre: kg} o lista [{nombre, stock_kg}]
        kg_maiz = kg_soja = 0.0
        if ins and ins.get('items'):
            _items = ins['items']
            if isinstance(_items, dict):
                # Formato real: {"MAIZ GRANO (KG)": 1575000, ...}
                for _nom, _kg in _items.items():
                    _n = str(_nom or '').upper()
                    _v = float(_kg or 0)
                    if 'MAIZ' in _n or 'MAÍZ' in _n:
                        kg_maiz += _v
                    elif 'SOJA' in _n:
                        kg_soja += _v
            else:
                # Formato alternativo: [{nombre, stock_kg}]
                for it in _items:
                    nom = str(it.get('nombre', '') or '').upper()
                    kg  = float(it.get('stock_kg') or 0)
                    if 'MAIZ' in nom or 'MAÍZ' in nom:
                        kg_maiz += kg
                    elif 'SOJA' in nom:
                        kg_soja += kg

        maiz_pesos = round(kg_maiz * bcr_maiz / 1000) if bcr_maiz else None
        soja_pesos = round(kg_soja * bcr_soja / 1000) if bcr_soja else None
        insumos_pesos = (
            (maiz_pesos or 0) + (soja_pesos or 0)
            if (maiz_pesos is not None or soja_pesos is not None) else None
        )

        # ── 4. Posición financiera en pesos ──
        disp    = float(fin.get('disponible')       or 0)
        cartera = float(fin.get('cheques_cartera')  or 0)
        emit    = float(fin.get('cheques_emitidos') or 0)
        cobrar  = float(fin.get('cobrar_hacienda')  or 0)
        pagar   = float(fin.get('pagar_hacienda')   or 0)
        lcg     = float(fin.get('lcg')              or 0)
        tercio  = float(fin.get('tercio_bravo')     or 0)
        # v15.23: posición Darwash del PROPIO financiero (hoja 'cuenta corriente
        # con darwash'), pasivo PEGSA → restar. Reemplaza el cruce con
        # tesoreria_darwash_historico.json de v15.22.
        darwash_pos    = float(fin.get('darwash_pos') or 0)
        darwash_origen = fin.get('darwash_origen')
        fin_pesos = (round(disp + cartera - emit + cobrar - pagar + lcg + tercio - darwash_pos)
                     if any([disp, cartera, cobrar]) else None)

        # ── 5. Dólares en pesos ──
        # Prioridad: usd_ars (ya convertido en el Excel); fallback: usd_cant × TC actual
        _usd_ars  = float(fin.get('usd_ars')  or 0)
        _usd_cant = float(fin.get('usd_cant') or 0)
        if _usd_ars > 0:
            usd_pesos = round(_usd_ars)
        elif _usd_cant > 0 and bna_tc:
            usd_pesos = round(_usd_cant * bna_tc)
            log.info(f'    USD fallback: {_usd_cant:,.0f} USD × ${bna_tc:,.0f} = ${usd_pesos:,.0f}')
        else:
            usd_pesos = None

        # ── 6. Total pesos ──
        componentes = [hacienda_pesos, insumos_pesos, fin_pesos, usd_pesos]
        total_pesos = round(sum(c for c in componentes if c is not None)) if any(c is not None for c in componentes) else None

        # ── 7. Total USD (usando BNA TC) ──
        total_usd = round(total_pesos / bna_tc, 0) if (total_pesos is not None and bna_tc) else None

        s = {
            'periodo':  periodo,
            'fecha':    snap.get('fecha', ''),
            'precios': {
                # v15.21: persistimos lo scrapeado (puede ser None) para que el
                # cache no fije el valor heredado. La próxima corrida re-intenta
                # scraping y, si la BCR publica el precio real, lo agarra.
                'mag_indice':    mag_indice_scraped,
                'bcr_maiz_ton':  bcr_maiz_scraped,
                'bcr_soja_ton':  bcr_soja_scraped,
                'bna_tc_venta':  bna_tc_scraped,
            },
            'precios_efectivos': {
                # v15.21: valores realmente usados en el cálculo de componentes
                # (heredados del mes anterior si el scraped fue None), con flag
                # por campo para auditoría.
                'mag_indice':    mag_indice,
                'bcr_maiz_ton':  bcr_maiz,
                'bcr_soja_ton':  bcr_soja,
                'bna_tc_venta':  bna_tc,
                'heredado': {
                    'mag_indice':    mag_indice_scraped is None and mag_indice is not None,
                    'bcr_maiz_ton':  bcr_maiz_scraped is None and bcr_maiz is not None,
                    'bcr_soja_ton':  bcr_soja_scraped is None and bcr_soja is not None,
                    'bna_tc_venta':  bna_tc_scraped is None and bna_tc is not None,
                },
            },
            'componentes': {
                'hacienda_kg_pegsa': round(kg_pegsa),
                'hacienda_pesos':    hacienda_pesos,
                'maiz_kg':           round(kg_maiz),
                'maiz_pesos':        maiz_pesos,
                'soja_kg':           round(kg_soja),
                'soja_pesos':        soja_pesos,
                'insumos_pesos':     insumos_pesos,
                'financiero_pesos':  fin_pesos,
                'darwash_pos':       round(darwash_pos),   # v15.23 (pasivo restado)
                'darwash_origen':    darwash_origen,       # v15.23 (hoja del propio financiero)
                'usd_cant':          round(_usd_cant) if _usd_cant else None,
                'usd_pesos':         usd_pesos,
                'total_pesos':       total_pesos,
                'total_usd':         total_usd,
            }
        }
        nuevos_snaps.append(s)
        tc_str = f' | TC ${bna_tc:,.0f}' if bna_tc else ''
        usd_str = f' = U$S {total_usd:,.0f}' if total_usd else ''
        log.info(f'    Total {periodo}: {("${:,.0f}".format(total_pesos)) if total_pesos else "—"}{tc_str}{usd_str}')

    nuevos_snaps.sort(key=lambda x: x.get('periodo', ''))

    # v15.49: alertar cuando los precios de commodities quedan heredados
    # (forward-fill). El scraping de BCR estuvo roto de 2026-05 a 2026-08 sin que
    # nadie lo notara porque el forward-fill lo tapaba. Miramos los últimos 2
    # meses; si maíz o soja vinieron heredados, se avisa en el log y en el JSON.
    _stale = []
    for _s in nuevos_snaps[-2:]:
        _her = (_s.get('precios_efectivos', {}) or {}).get('heredado', {}) or {}
        if _her.get('bcr_maiz_ton') or _her.get('bcr_soja_ton'):
            _stale.append(_s.get('periodo'))
    if _stale:
        log.warning(f'  ⚠ Precios BCR heredados en: {", ".join(_stale)} — '
                    f'revisar scraping o actualizar datos/precios_bcr/')

    resultado = {
        'generado':  datetime.now().isoformat(),
        'metodo':    'scraping_mag_bcr_bna',
        'snapshots': nuevos_snaps,
        'precios_stale': _stale,
    }
    guardar(resultado, carpeta, 'valuacion_historica.json')
    log.info(f'  ✓ valuacion_historica.json — {len(nuevos_snaps)} períodos')
    return resultado


# ═══════════════════════════════════════════════════════════
#  RUNNING BALANCE · STOCK DIARIO HISTÓRICO
#  Recalcula cada día desde movimientos reales en lugar de
#  acumular snapshots diarios (que quedan desactualizados
#  cuando se cargan compras/ventas con fecha retroactiva).
# ═══════════════════════════════════════════════════════════

def recalcular_stock_diario_desde_movimientos(
        regs_stock, cols_stock,
        regs_ing,   cols_ing,
        regs_egr,   cols_egr,
        carpeta,    periodo,
        dias=90):
    """
    Recalcula el stock diario histórico usando running balance.
    stock(D) = stock(D+1) - ingresos(D+1) + egresos(D+1)
    Baseline = V_STOCK_HACIENDA actual (estado definitivo de hoy).
    Retiene solo los últimos `dias` días (90 por defecto) para controlar
    el tamaño del archivo. Las entradas más antiguas se descartan.
    El resultado reemplaza completamente stock_diario.json en cada ejecución,
    incorporando automáticamente cualquier carga retroactiva de movimientos.
    """
    from datetime import date as _date, timedelta as _td

    hoy = _date.today()

    # ── 1. Baseline: stock de hoy ──────────────────────────────
    kpis_hoy      = calcular_kpis(regs_stock or [], cols_stock or [])
    total_cab_hoy = int(kpis_hoy.get("total_cabezas", 0))
    total_kg_hoy  = kpis_hoy.get("total_kg_estimado_hoy", 0) or 0
    avg_kg_hoy    = total_kg_hoy / max(total_cab_hoy, 1)

    # Snapshots de propietario (Hotelero) para hoy
    prop_hoy = {
        p: {"cabezas": int(v.get("cabezas", 0)),
            "kg_estimado": int(v.get("kg_estimado", 0))}
        for p, v in kpis_hoy.get("por_propietario", {}).items()
    }

    # v15.51: mismo tratamiento que por_propietario para establecimiento y
    # categoría. Antes se escribían siempre {} y el desglose del panel Diario
    # nunca tuvo datos.
    est_hoy = {
        e: {"cabezas": int(v.get("cabezas", 0)),
            "kg_estimado": int(v.get("kg_estimado", 0))}
        for e, v in kpis_hoy.get("por_establecimiento", {}).items()
    }
    cat_hoy = {
        c: {"cabezas": int(v.get("cabezas", 0)),
            "kg_estimado": int(v.get("kg_estimado", 0))}
        for c, v in kpis_hoy.get("por_categoria", {}).items()
    }

    log.info(f"  Baseline hoy ({hoy}): {total_cab_hoy:,} cab · {total_kg_hoy/1000:,.0f} t")
    log.info(f"  Propietarios baseline: {list(prop_hoy.keys())}")
    log.info(f"  Establecimientos baseline: {list(est_hoy.keys())}")
    log.info(f"  Categorías baseline: {list(cat_hoy.keys())}")

    # ── 2. Resolución de columnas ─────────────────────────────
    def _fc(cols, *keys):
        """Busca la primera columna cuyo nombre exacto (o en minúsculas) coincide."""
        for k in keys:
            for c in (cols or []):
                if c == k or c.lower() == k.lower():
                    return c
        return None

    col_fi = _fc(cols_ing, "FechaIngreso", "fecha_ingreso", "FECHA_INGRESO", "fecha")
    col_ci = _fc(cols_ing, "CantidadIngreso", "cantidadingreso", "CANTIDADINGRESO", "cantidad", "Cantidad", "cabezas")
    col_hi = _fc(cols_ing, "Hotelero", "hotelero", "HOTELERO", "propietario", "Propietario")

    col_fe = _fc(cols_egr, "FechaSalida", "fecha_salida", "FECHA_SALIDA", "fecha")
    col_ce = _fc(cols_egr, "CantidadEgreso", "cantidadegreso", "CANTIDADEGRESO", "cantidad", "Cantidad", "cabezas")
    col_he = _fc(cols_egr, "Hotelero", "hotelero", "HOTELERO", "propietario", "Propietario")

    log.info(f"  Ingresos cols → fecha={col_fi} cant={col_ci} hotelero={col_hi}")
    log.info(f"  Egresos cols  → fecha={col_fe} cant={col_ce} hotelero={col_he}")

    # ── 3. Función fecha → YYYY-MM-DD ─────────────────────────
    def _fs(val):
        if val is None:
            return None
        if isinstance(val, _date):
            return val.strftime("%Y-%m-%d")
        if hasattr(val, "strftime"):
            return val.strftime("%Y-%m-%d")
        try:
            from datetime import datetime as _dt
            return _dt.fromisoformat(str(val)[:10]).strftime("%Y-%m-%d")
        except Exception:
            return None

    # ── 4. Acumular movimientos por fecha ─────────────────────
    ing_total = {}  # {fecha_str: cabezas_int}
    ing_prop  = {}  # {fecha_str: {hotelero: cabezas_int}}
    egr_total = {}
    egr_prop  = {}

    for r in (regs_ing or []):
        fs = _fs(r.get(col_fi)) if col_fi else None
        if not fs:
            continue
        try:
            cant = int(round(float(r.get(col_ci, 0) or 0)))
        except Exception:
            cant = 0
        if cant <= 0:
            continue
        prop = str(r.get(col_hi) or "").strip() if col_hi else ""
        ing_total[fs] = ing_total.get(fs, 0) + cant
        if prop:
            ing_prop.setdefault(fs, {})
            ing_prop[fs][prop] = ing_prop[fs].get(prop, 0) + cant

    for r in (regs_egr or []):
        fs = _fs(r.get(col_fe)) if col_fe else None
        if not fs:
            continue
        try:
            cant = int(round(float(r.get(col_ce, 0) or 0)))
        except Exception:
            cant = 0
        if cant <= 0:
            continue
        prop = str(r.get(col_he) or "").strip() if col_he else ""
        egr_total[fs] = egr_total.get(fs, 0) + cant
        if prop:
            egr_prop.setdefault(fs, {})
            egr_prop[fs][prop] = egr_prop[fs].get(prop, 0) + cant

    # Diagnóstico: primeras fechas con movimientos
    _sample_ing = sorted(ing_total.items())[-5:]
    _sample_egr = sorted(egr_total.items())[-5:]
    log.info(f"  Ingresos últimas 5 fechas: {_sample_ing}")
    log.info(f"  Egresos  últimas 5 fechas: {_sample_egr}")

    # ── 5. Cargar historial acumulado de los tres desgloses ──────
    # v15.51: además de por_propietario (v15.21) ahora se acumulan también
    # por_establecimiento y por_categoria, con el mismo patrón. Se acumulan día
    # a día desde ejecuciones anteriores; NO se reconstruyen hacia atrás (el
    # running balance opera sobre totales, no sobre desgloses). Estrategia:
    # conservar histórico guardado; hoy se sobreescribe con la vista actual.
    _diario_path = Path(carpeta) / "stock_diario.json"
    _hist = {"por_propietario": {}, "por_establecimiento": {}, "por_categoria": {}}
    if _diario_path.exists():
        try:
            with open(_diario_path, encoding="utf-8") as _fh:
                _old = json.load(_fh)
            for _s in _old.get("snapshots", []):
                _fs2 = _s.get("fecha", "")
                if not _fs2:
                    continue
                _h = _s.get("hacienda") or {}
                for _k in _hist:
                    _v = _h.get(_k)
                    if _v:
                        _hist[_k][_fs2] = _v
        except Exception:
            pass
    # Hoy siempre desde la vista actual (dato fidedigno)
    _hoy_str = hoy.strftime("%Y-%m-%d")
    _hist["por_propietario"][_hoy_str]     = prop_hoy
    _hist["por_establecimiento"][_hoy_str] = est_hoy
    _hist["por_categoria"][_hoy_str]       = cat_hoy
    for _k, _v in _hist.items():
        log.info(f"  {_k} acumulado: {len(_v)} fechas con datos")

    # ── 6. Running balance hacia atrás desde hoy ──────────────
    snapshots = []
    cab_d = total_cab_hoy

    for i in range(dias + 1):
        dia = hoy - _td(days=i)
        fs  = dia.strftime("%Y-%m-%d")

        if i > 0:
            # stock(D) = stock(D+1) - ingresos(D+1) + egresos(D+1)
            fs_next = (dia + _td(days=1)).strftime("%Y-%m-%d")
            ing_n   = ing_total.get(fs_next, 0)
            egr_n   = egr_total.get(fs_next, 0)
            cab_d   = max(0, cab_d - ing_n + egr_n)

        kg_d = int(cab_d * avg_kg_hoy)

        snapshots.append({
            "fecha": fs,
            "hacienda": {
                "total_cabezas":       int(cab_d),
                "total_kg_estimado":   kg_d,
                "por_propietario":     _hist["por_propietario"].get(fs, {}),
                "por_establecimiento": _hist["por_establecimiento"].get(fs, {}),
                "por_categoria":       _hist["por_categoria"].get(fs, {})
            }
        })

    # Ordenar ascendente (más antiguo primero)
    snapshots.sort(key=lambda s: s["fecha"])

    # ── 7. Forward-fill defensivo de los desgloses ──────────────
    # v15.21: bug detectado 2026-06-18 — ciertos snapshots tienen total_cabezas
    # correcto pero el desglose vacío ({}) o con todos en 0 (la vista WinCampo
    # del día devolvió ceros y se arrastró, o nunca se acumuló). Eso rompe el
    # gráfico Evolución Diaria (stacked) → cae a 0 visual aunque el total esté OK.
    # Defense-in-depth: pasada ASC heredando el desglose del último día sano; los
    # totales se mantienen como están.
    # v15.51: la misma defensa se aplica a los tres desgloses. Como
    # por_establecimiento/por_categoria recién arrancan a poblarse desde hoy
    # hacia adelante, los días previos al primer día sano quedan vacíos (no hay
    # de dónde copiar) — es esperado; para el histórico completo está la vista
    # mensual.
    for _key in ("por_propietario", "por_establecimiento", "por_categoria"):
        _ult_sano = None
        _n_filled = 0
        for _snap in snapshots:
            _h = _snap.get("hacienda", {})
            _d = _h.get(_key) or {}
            _sano = bool(_d) and any((v or {}).get("cabezas", 0) > 0 for v in _d.values())
            if _sano:
                _ult_sano = _d
            elif _ult_sano:
                _h[_key] = _ult_sano
                _n_filled += 1
        if _n_filled:
            log.info(f"  v15.51 forward-fill {_key}: {_n_filled} día(s) heredados")

    diario = {
        "generado":  datetime.now().isoformat(),
        "periodo":   periodo,
        "metodo":    "running_balance",
        "dias":      len(snapshots),
        "snapshots": snapshots,
    }
    guardar(diario, carpeta, "stock_diario.json")
    log.info(f"  ✓ stock_diario.json — {len(snapshots)} días · running balance")
    return diario


def _xlsx_es_placeholder_onedrive(path):
    """v15.32: detecta archivos OneDrive en estado placeholder (cloud-only).
    OneDrive reporta size > 0 en metadata pero el contenido no está en disco
    local → pd.read_excel falla con [Errno 22]. Leemos los primeros 4 bytes:
    si NO son la signature ZIP/XLSX (PK\\x03\\x04), es placeholder o corrupto.
    Returns: (es_placeholder: bool, motivo: str|None)
    """
    try:
        if not os.path.exists(path):
            return True, "no_existe"
        size = os.path.getsize(path)
        if size < 100:
            return True, f"tamano_invalido ({size} bytes)"
        with open(path, "rb") as f:
            head = f.read(4)
        if head != b"PK\x03\x04":
            return True, f"header_no_xlsx ({head!r})"
        return False, None
    except OSError as e:
        # Errno 22 acá típicamente significa "online-only file" no hidratado.
        return True, f"oserror_{e.errno}: {e}"


def _intentar_warming_onedrive(path, max_intentos=2):
    """v15.32: fuerza a OneDrive a hidratar el archivo leyendo unos KB.
    Retorna True si tras el warming pasa el check de placeholder, False si
    sigue inaccesible.
    """
    import time
    for _ in range(max_intentos):
        try:
            with open(path, "rb") as f:
                f.read(8192)   # forzar a OneDrive a bajar al menos 8 KB
            time.sleep(0.5)
            es_ph, _motivo = _xlsx_es_placeholder_onedrive(path)
            if not es_ph:
                return True
        except OSError:
            time.sleep(1.0)
    return False


def main():
    separador("PEGSA & BULLTRADE - Actualizador de Datos")
    log.info(f"  Inicio: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    separador()

    cfg     = cargar_config()
    periodo = cfg["OPCIONES"]["periodo"]
    carpeta = resolver_carpeta_salida(cfg["ONEDRIVE"].get("carpeta", ""))
    log.info(f"  Periodo : {periodo}")
    log.info(f"  Destino : {carpeta}")
    log.info("")

    # v15.10: migración completa. Las 5 tablas (Stock, Egresos, Ingresos,
    # Muertes, Stock Insumos) vienen 100% de WinCampo Web; Consumo del Mixer
    # Dropbox. Ya no hay SQL ni conectar(): el adapter es la única fuente.
    try:
        from wincampo_source import WinCampoAPI
        wcampo = WinCampoAPI()
        log.info("  + WinCampoAPI conectado (modo wincampo_web)")
    except Exception as e:
        log.error(f"  x Falla conectar WinCampoAPI: {e}")
        raise

    resumen = {"generado": datetime.now().isoformat(), "periodo": periodo, "modulos": {}}

    # ── v15.13.2: Pre-step · Egresos + procesar_productivo para ADP dinámico ──
    # Debe correr ANTES de Stock: Stock usa _ADP_CAL_RUNTIME (vía calc_engorde en
    # extraer) y acá lo actualizamos con el adp_calibrado real per categoría.
    # egresos_data/regs_egr/cols_egr/prod_data quedan en scope para reusar abajo.
    separador("Pre-step · Egresos + ADP dinámico")
    import pandas as pd
    from datetime import date, timedelta
    # v15.67: el fetch se amplía hasta el 1/1 del año base más viejo, porque
    # procesar_productivo calcula los teóricos de ADP sobre esos años. El resto
    # del pipeline (stock diario, movimientos) sigue viendo 730 días exactos:
    # ampliar SOLO egresos descalzaría la serie diaria contra ingresos.
    _hoy_egr   = date.today()
    _base_egr  = adp_base_anios(_hoy_egr)
    _dias_base = max(730, (_hoy_egr - date(min(_base_egr), 1, 1)).days)
    fd_egr = (_hoy_egr - timedelta(days=_dias_base)).isoformat()
    fh_egr = _hoy_egr.isoformat()
    egresos_data = wcampo.fetch_egresos(fecha_desde=fd_egr, fecha_hasta=fh_egr)
    df_egr = pd.DataFrame(egresos_data)
    log.info(f"  + WinCampo Web devolvio {len(df_egr):,} egresos ({_dias_base}d, base ADP {_base_egr})")
    # Recorte estándar 730d — lo que consume el resto del pipeline (sin cambios)
    regs_egr, cols_egr = extraer("v_PB_Egresos", fecha_col="FechaSalida", dias=730, df_override=df_egr)
    # Recorte ancho — solo para los teóricos de ADP
    regs_prod, _cols_prod = extraer("v_PB_Egresos", fecha_col="FechaSalida", dias=_dias_base, df_override=df_egr)
    log.info(f"  + Productivo usa {len(regs_prod):,} egresos ({_dias_base}d) | resto del pipeline: {len(regs_egr):,} (730d)")
    prod_data = procesar_productivo(regs_prod, cols_egr, periodo)
    # Actualizar _ADP_CAL_RUNTIME (in-place) con el adp_calibrado dinámico per cat
    pc90 = prod_data.get("por_categoria_90d", {})
    for _cat, _info in pc90.items():
        _cal = _info.get("adp_calibrado")
        if _cal is not None and _cal > 0:
            _ADP_CAL_RUNTIME[_cat] = _cal
    log.info(f"  + ADP_CAL_RUNTIME (dinámico, clamp ±25%): {_ADP_CAL_RUNTIME}")
    log.info(f"    Fallback si categoría sin observado: {ADP_CAL_FALLBACK}")

    separador("Stock de Hacienda")
    tabla      = cfg["TABLAS"].get("stock_hacienda", "V_STOCK_HACIENDA")
    # v15.10: única fuente WinCampo Web. extraer() aplica las transformaciones
    # (DIAS_EN_FEEDLOT, CLASIFICACION, etc.) sobre el DataFrame del adapter.
    import pandas as pd
    # v15.45: degradación elegante. WinCampo migró el stock a una cola asincrónica
    # y puede fallar (cola rota, API caída, timeout). Si el fetch falla NO matamos
    # todo el pipeline — los módulos independientes (tesorería, mixer, trazabilidad,
    # precios) siguen; los JSONs de stock del run anterior quedan como
    # last-known-good. Defaults vacíos para que los módulos intermedios que leen
    # regs/kpis/kpis_haras (muertes, indicadores) degraden sin NameError. Patrón v15.32.
    stock_data = None
    regs, cols = [], []
    kpis = {}
    kpis_haras = {}
    _regs_stock_hoy = None
    _cols_stock_hoy = None
    try:
        stock_data = wcampo.fetch_stock_hacienda()
        df_stock = pd.DataFrame(stock_data)
        log.info(f"  + WinCampo Web devolvio {len(df_stock):,} cabezas")
        regs, cols = extraer(tabla, df_override=df_stock)
        # Guardar referencia al stock actual para el recálculo diario posterior
        _regs_stock_hoy = regs
        _cols_stock_hoy = cols
    except Exception as e:
        log.error(f"  x Stock de Hacienda falló: {type(e).__name__}: {e}")
        log.error("  x Continuando con los módulos independientes (stock queda con datos del run anterior)")
        resumen["modulos"]["stock_hacienda"] = {"ok": False, "error": f"{type(e).__name__}: {str(e)[:200]}"}

    if regs:
        kpis = calcular_kpis(regs, cols)
        meta = {
            "generado":  datetime.now().isoformat(),
            "periodo":   periodo,
            "tabla":     tabla,
            "registros": len(regs),
            "columnas":  cols,
        }

        # ── 1. JSON liviano: solo KPIs (carga inicial del portal) ──
        guardar({"meta": meta, "kpis": kpis}, carpeta, f"stock_kpis_{periodo}.json")
        log.info(f"  ✓ stock_kpis_{periodo}.json")

        # ── 2. JSON detalle completo (pestaña Detalle Completo) ──
        guardar({"meta": meta, "detalle": regs}, carpeta, f"stock_detalle_{periodo}.json")
        log.info(f"  ✓ stock_detalle_{periodo}.json")

        # ── 3. JSON por propietario (HOTELERO) ──
        col_prop = "HOTELERO"
        if col_prop in (regs[0] if regs else {}):
            propietarios = {}
            for r in regs:
                p = str(r.get(col_prop) or "Sin datos").strip()
                if p not in propietarios:
                    propietarios[p] = []
                propietarios[p].append(r)
            for prop, rows_prop in propietarios.items():
                nombre_archivo = re.sub(r'[^a-zA-Z0-9_]', '_', prop)
                kpis_prop = calcular_kpis(rows_prop, cols)
                guardar({"propietario": prop, "kpis": kpis_prop},
                        carpeta, f"stock_prop_{nombre_archivo}_{periodo}.json")
            log.info(f"  ✓ {len(propietarios)} archivos stock_prop_*_{periodo}.json")

        # ── 4. JSON por establecimiento (NOMBRE_CORRAL) ──
        col_est   = "NOMBRE_CORRAL"
        kpis_haras = {}   # se llena si existe "El Haras"
        if col_est in (regs[0] if regs else {}):
            establecimientos = {}
            for r in regs:
                e = str(r.get(col_est) or "Sin asignar").strip()
                if e not in establecimientos:
                    establecimientos[e] = []
                establecimientos[e].append(r)
            for est, rows_est in establecimientos.items():
                nombre_archivo = re.sub(r'[^a-zA-Z0-9_]', '_', est)
                kpis_est = calcular_kpis(rows_est, cols)
                guardar({"establecimiento": est, "kpis": kpis_est},
                        carpeta, f"stock_est_{nombre_archivo}_{periodo}.json")
                if est.strip().upper() == "EL HARAS":
                    kpis_haras = kpis_est
                    log.info(f"  ↳ El Haras → {kpis_haras.get('total_cabezas',0):,} cab · {kpis_haras.get('total_kg_estimado_hoy',0):,.0f} kg")
            log.info(f"  ✓ {len(establecimientos)} archivos stock_est_*_{periodo}.json")

        resumen["modulos"]["stock_hacienda"] = {
            "ok": True, "registros": len(regs), "cabezas": kpis.get("total_cabezas")
        }

        log.info("")
        log.info("  RESUMEN GENERAL:")
        log.info(f"  Cabezas totales       : {kpis.get('total_cabezas', 0):,}")
        log.info(f"  Kg totales estimado   : {kpis.get('total_kg_estimado_hoy', 0):,} kg")
        log.info(f"  Ton. estimado hoy     : {kpis.get('total_ton_estimado_hoy', 0):,} t")
        log.info(f"  Kg promedio estimado  : {kpis.get('kg_promedio_estimado', 0)} kg")
        log.info(f"  Dias prom. feedlot    : {kpis.get('dias_promedio_feedlot', 0)}")

        if kpis.get("por_propietario"):
            log.info("")
            log.info("  Por propietario (HOTELERO):")
            for g, d in sorted(kpis["por_propietario"].items(), key=lambda x: -x[1]["cabezas"]):
                log.info(f"    {g:<22} {int(d['cabezas']):>7,} cab  /  {d['ton_estimado']:>7,.1f} t  /  {d['kg_promedio']} kg prom.")

        if kpis.get("por_establecimiento"):
            log.info("")
            log.info("  Por establecimiento (NOMBRE_CORRAL):")
            for g, d in sorted(kpis["por_establecimiento"].items(), key=lambda x: -x[1]["cabezas"]):
                log.info(f"    {g:<22} {int(d['cabezas']):>7,} cab  /  {d['ton_estimado']:>7,.1f} t  /  {d['kg_promedio']} kg prom.")

        if kpis.get("por_clasificacion"):
            log.info("")
            log.info("  Por clasificacion:")
            for g, d in sorted(kpis["por_clasificacion"].items(), key=lambda x: -x[1]["cabezas"]):
                log.info(f"    {g:<12} {int(d['cabezas']):>7,} cab  /  {d['ton_estimado']:>7,.1f} t")

        if kpis.get("por_categoria_final"):
            log.info("")
            log.info("  Por categoria final:")
            for g, d in sorted(kpis["por_categoria_final"].items(), key=lambda x: -x[1]["cabezas"]):
                log.info(f"    {g:<25} {int(d['cabezas']):>7,} cab  /  {d['kg_promedio']} kg prom.")

    # ── 5. JSON Stock de Insumos ──
    separador("Stock de Insumos")
    tabla_ins = cfg["TABLAS"].get("stock_insumos", "v_PB_StockInsumos")
    # v15.10: WinCampo Web. El adapter renombra STOCK_ACTUAL -> STOCK; el filtro
    # de 7 insumos (INSUMOS_INCLUIDOS) se aplica abajo.
    import pandas as pd
    insumos_data = wcampo.fetch_stock_insumos()
    df_ins = pd.DataFrame(insumos_data)
    log.info(f"  + WinCampo Web devolvio {len(df_ins):,} insumos (stock actual)")
    regs_ins, cols_ins = extraer(tabla_ins, df_override=df_ins)

    INSUMOS_INCLUIDOS = {
        2: "MAIZ GRANO",
        9: "SOJA",
        8: "NUCLEO CONC 5% LDB",
        99: "DIESEL",
        6: "HARINA GERMEN",
        7: "GLUTEN DE MAIZ",
        3: "SILO DE MAIZ",
    }

    col_nombre = "DESC_INSUMO"
    col_stock  = "STOCK"
    col_cod    = "COD_INSUMO"

    insumos = []
    total_kg = 0
    for r in regs_ins:
        try:
            cod = int(float(r.get(col_cod) or -1))
        except:
            cod = -1
        if cod not in INSUMOS_INCLUIDOS:
            continue
        nombre_raw = str(r.get(col_nombre) or "").strip()
        raw_stock  = r.get(col_stock) if col_stock else 0
        stock_kg   = round(float(raw_stock or 0), 2)
        insumos.append({
            "nombre":   nombre_raw,
            "stock_kg": stock_kg,
        })
        total_kg += stock_kg

    insumos.sort(key=lambda x: -x["stock_kg"])

    # ── Cruzar con consumo diario para calcular días restantes ──
    # Nota: consumo_data aún no está disponible acá (se procesa en módulo 9)
    # Se calcula al final en módulo 10 y se enriquece el JSON
    meta_ins = {
        "generado": datetime.now().isoformat(),
        "periodo":  periodo,
        "tabla":    tabla_ins,
        "registros": len(insumos),
    }
    guardar({
        "meta": meta_ins,
        "insumos": insumos,
        "total_kg": round(total_kg, 2),
    }, carpeta, f"stock_insumos_{periodo}.json")
    log.info(f"  ✓ stock_insumos_{periodo}.json  ({len(insumos)} insumos)")
    for ins in insumos:
        log.info(f"    {ins['nombre']:<28} {ins['stock_kg']:>12,.1f} kg")
        resumen["modulos"]["stock_insumos"] = {"ok": True, "registros": len(insumos), "total_kg": round(total_kg,2)}
    else:
        log.warning("  ⚠ Sin datos de insumos")

    # ── Actualizar STOCK DE INSUMOS.xlsx en carpeta stock mensuales ──
    if insumos:
        try:
            from pathlib import Path as _Path
            _carpeta_sm = _Path(carpeta).parent / "stock mensuales"
            _today_str  = datetime.now().strftime('%Y-%m-%d')
            log.info(f"  → Actualizando STOCK DE INSUMOS.xlsx ({_today_str})...")
            actualizar_stock_insumos_excel(insumos, str(_carpeta_sm), _today_str)
        except Exception as _e:
            log.warning(f"  ⚠ Error actualizando STOCK DE INSUMOS.xlsx: {_e}")

    # ── 6. JSON Movimientos Productivos (Ingresos + Egresos) ──
    separador("Movimientos Productivos")
    tabla_ing = cfg["TABLAS"].get("movimientos_ingresos", "v_PB_Ingresos")

    # v15.6/v15.10: Ingresos desde WinCampo Web (fetch_ingresos). Endpoint sin cap.
    # v15.13.2: egresos (regs_egr/cols_egr) ya cargados en el pre-step — NO se
    # vuelven a fetchear. Reusa fd_egr/fh_egr del pre-step.
    ingresos_data = wcampo.fetch_ingresos(fecha_desde=fd_egr, fecha_hasta=fh_egr)
    df_ing = pd.DataFrame(ingresos_data)
    log.info(f"  + WinCampo Web devolvio {len(df_ing):,} sub-grupos de ingresos (730d)")
    regs_ing, cols_ing = extraer(tabla_ing, fecha_col="FechaIngreso", dias=730, df_override=df_ing)

    mov_data = procesar_movimientos(regs_ing, cols_ing, regs_egr, cols_egr, periodo)

    guardar(mov_data, carpeta, f"movimientos_{periodo}.json")
    log.info(f"  ✓ movimientos_{periodo}.json")
    m = mov_data.get("resumen", {})
    log.info(f"  Ingresos  :  {m.get('total_cabezas_ingresadas',0):>8,} cab  /  {m.get('total_kg_ingresado',0):>12,.0f} kg")
    log.info(f"  Egresos   :  {m.get('total_cabezas_egresadas',0):>8,} cab  /  {m.get('total_kg_egresado',0):>12,.0f} kg")
    log.info(f"  Saldo neto:  {m.get('saldo_cabezas',0):>+8,} cab  /  {m.get('saldo_kg',0):>+12,.0f} kg")
    resumen["modulos"]["movimientos"] = {"ok": True, **m}

    # ── 7. JSON Muertes + Tasa de Mortandad ──
    separador("Muertes & Tasa de Mortandad")
    tabla_muertes = cfg["TABLAS"].get("muertes", "V_MUERTES")
    # v15.7/v15.10: Muertes desde WinCampo Web (fetch_muertes, wrapper sobre
    # egresos MOTIVO=M con remap a las columnas de V_MUERTES), rango 730d.
    import pandas as pd
    from datetime import date, timedelta
    fd = (date.today() - timedelta(days=730)).isoformat()
    fh = date.today().isoformat()
    muertes_raw = wcampo.fetch_muertes(fecha_desde=fd, fecha_hasta=fh)
    df_m = pd.DataFrame(muertes_raw)
    log.info(f"  + WinCampo Web devolvio {len(df_m):,} muertes MOTIVO=M (730d)")
    regs_m, cols_m = extraer(tabla_muertes, fecha_col="FECHA_MUERTE", dias=730, df_override=df_m)

    # Reusar regs_ing/cols_ing (ya cargados en módulo 6) y regs/cols de stock hacienda
    # regs_ing ya fue cargado arriba; regs (stock) también — los pasamos directamente
    muertes_data = procesar_muertes(
        regs_m,   cols_m,
        regs_ing, cols_ing,
        regs,     cols,      # V_STOCK_HACIENDA cargado en módulo 1
        periodo
    )
    guardar(muertes_data, carpeta, f"muertes_{periodo}.json")
    log.info(f"  ✓ muertes_{periodo}.json")
    mort = muertes_data.get("mortandad", {})
    resumen["modulos"]["muertes"] = {
        "ok":               True,
        "total_anio":       muertes_data["anio"].get("total_muertes", 0),
        "total_mes_ant":    muertes_data["mes_anterior"].get("total_muertes", 0),
        "tasa_mensual_pct": mort.get("tasa_mensual_pct"),
    }

    # ── 7b. JSON Muertes 30 días ──
    separador("Muertes & Tasa — Últimos 30 días")
    muertes_30d_data = procesar_muertes_30d(
        regs_m,   cols_m,
        regs_ing, cols_ing,
        regs,     cols,
        periodo
    )
    guardar(muertes_30d_data, carpeta, f"muertes_30d_{periodo}.json")
    log.info(f"  ✓ muertes_30d_{periodo}.json")
    m30 = muertes_30d_data.get("mortandad", {})
    log.info(f"    Muertes 30d: {m30.get('muertes_30d',0)} | Tasa: {m30.get('tasa_mensual_pct')}%")
    resumen["modulos"]["muertes_30d"] = {
        "ok":               True,
        "total_30d":        muertes_30d_data["detalle"].get("total_muertes", 0),
        "tasa_mensual_pct": m30.get("tasa_mensual_pct"),
    }

    # ── 8. JSON Parámetros Productivos (ADP + Estadía) ──
    separador("Parámetros Productivos")
    # v15.13.2: prod_data ya calculado en el pre-step (antes de Stock), solo se guarda.
    guardar(prod_data, carpeta, f"productivo_{periodo}.json")
    log.info(f"  ✓ productivo_{periodo}.json")
    g = prod_data.get("general", {})
    log.info(f"  ADP promedio  : {g.get('adp_promedio')} kg/día")
    log.info(f"  Estadía prom  : {g.get('estadia_promedio')} días")
    log.info(f"  Cabezas       : {g.get('cabezas',0):,}")
    resumen["modulos"]["productivo"] = {
        "ok":               True,
        "adp_promedio":     g.get("adp_promedio"),
        "estadia_promedio": g.get("estadia_promedio"),
        "cabezas":          g.get("cabezas", 0),
    }

    # ── 9. JSON Consumo de Alimento ──
    # FUENTE PRIMARIA: base Access del Mixer (Dropbox) — kilos reales cargados.
    # FUENTE FALLBACK: vista SQL v_PB_ConsumoDetallado (deprecated, pierde ~25%).
    separador("Consumo de Alimento (Mixer Dropbox)")
    consumo_data = None
    try:
        from consumo_mixer import procesar_consumo_mixer, DEFAULT_MIXER_DB
        mixer_path = cfg["RUTAS"].get("mixer_db", DEFAULT_MIXER_DB) if cfg.has_section("RUTAS") else DEFAULT_MIXER_DB
        consumo_data = procesar_consumo_mixer(
            mixer_path=mixer_path,
            periodo=periodo,
            dias_diario=30,
            log=log,
        )
        log.info(f"  ✓ Consumo leído del Mixer ({consumo_data['meta']['ultimo_completo']} = último día completo)")
    except Exception as e:
        log.error(f"  x No se pudo leer el Mixer: {e}")
        log.error(f"  -> No hay fallback SQL (v15.10). Saltando modulo Consumo.")
        consumo_data = {"meta": {"error": str(e)}, "anual": {}, "semanal": {}, "diario": []}
    guardar(consumo_data, carpeta, f"consumo_{periodo}.json")
    log.info(f"  ✓ consumo_{periodo}.json")
    ca = consumo_data.get("anual",   {})
    cs = consumo_data.get("semanal", {})
    log.info(f"  Total anual     : {ca.get('total_kg',0):,.0f} kg")
    log.info(f"  Prom. diario 7d : {cs.get('promedio_diario_kg',0):,.1f} kg/día")
    resumen["modulos"]["consumo"] = {
        "ok":                True,
        "total_anual_kg":    ca.get("total_kg", 0),
        "promedio_diario_kg": cs.get("promedio_diario_kg", 0),
    }

    # ── 10. Indicadores cruzados (consumo × stock El Haras × ADP) ──
    separador("Indicadores Productivos")
    try:
        # Denominadores: siempre El Haras (donde se da el alimento)
        kg_stock_haras = kpis_haras.get("total_kg_estimado_hoy", 0) if kpis_haras else kpis.get("total_kg_estimado_hoy", 0)
        cab_haras      = kpis_haras.get("total_cabezas", 0)         if kpis_haras else kpis.get("total_cabezas", 0)
        usando_haras   = bool(kpis_haras)

        prom_diario_ms = cs.get("promedio_diario_kg_ms", 0)
        prom_diario_tc = cs.get("promedio_diario_kg", 0)

        # ADP: usar el último mes cerrado de productivo (más representativo que el promedio histórico)
        _por_mes = prod_data.get("por_mes", {})
        _ultimo_mes = max(_por_mes.keys()) if _por_mes else None
        adp_prom = (_por_mes[_ultimo_mes].get("adp_promedio") or 0) if _ultimo_mes else (g.get("adp_promedio", 0) or 0)
        log.info(f"  ADP conversión: {adp_prom} kg/día (último mes: {_ultimo_mes or 'n/a'})")

        log.info(f"  Denominador: {'El Haras' if usando_haras else 'PEGSA total (Haras no encontrado)'}")
        log.info(f"  Cabezas     : {cab_haras:,}")
        log.info(f"  Kg PV       : {kg_stock_haras:,.0f}")

        # ── 1. % Consumo de Peso Vivo ──
        pct_pv = round(prom_diario_ms / kg_stock_haras * 100, 2) if kg_stock_haras > 0 else None

        # ── 2. Consumo por cabeza (TC y MS) ──
        consumo_cab_tc = round(prom_diario_tc / cab_haras, 2) if cab_haras > 0 else None
        consumo_cab_ms = round(prom_diario_ms / cab_haras, 2) if cab_haras > 0 else None

        # ── 3. Conversión alimenticia ──
        prod_diaria_kg = adp_prom * cab_haras if adp_prom and cab_haras else 0
        conversion     = round(prom_diario_ms / prod_diaria_kg, 2) if prod_diaria_kg > 0 else None

        indicadores = {
            "generado":      datetime.now().isoformat(),
            "denominador":   "El Haras" if usando_haras else "PEGSA total",
            "fuentes": {
                "kg_stock_haras":  kg_stock_haras,
                "kg_stock_total":  kpis.get("total_kg_estimado_hoy", 0),
                "cab_haras":       cab_haras,
                "cab_total":       kpis.get("total_cabezas", 0),
                "prom_diario_ms":  prom_diario_ms,
                "prom_diario_tc":  prom_diario_tc,
                "adp_promedio":    adp_prom,
                "adp_mes":         _ultimo_mes,
                "prod_diaria_kg":  round(prod_diaria_kg, 1),
                "dias_consumo":    cs.get("dias_registrados", 0),
            },
            "indicadores": {
                "pct_peso_vivo": {
                    "valor":       pct_pv,
                    "unidad":      "% PV",
                    "descripcion": "Consumo MS como % del peso vivo — El Haras",
                    # v15.20: rangos de 5 zonas (escala · normal amarillo · óptimo verde)
                    "esc_min":        1.5,
                    "esc_max":        3.5,
                    "ref_normal_min": 2.3,
                    "ref_normal_max": 2.9,
                    "ref_opt_min":    2.4,
                    "ref_opt_max":    2.8,
                    # legacy (compat con consumidores que aún leen ref_min/opt/max):
                    "ref_min":     2.3,
                    "ref_opt":     2.6,
                    "ref_max":     2.9,
                    "formula":     "kg MS/día ÷ kg PV El Haras × 100",
                },
                "consumo_por_cabeza": {
                    "valor_tc":    consumo_cab_tc,
                    "valor_ms":    consumo_cab_ms,
                    "unidad":      "kg/cab/día",
                    "descripcion": "Alimento por animal por día — El Haras",
                    # v15.20: rangos de 5 zonas
                    "esc_min":        8.0,
                    "esc_max":        23.0,
                    "ref_normal_min": 11.0,
                    "ref_normal_max": 20.0,
                    "ref_opt_min":    13.0,
                    "ref_opt_max":    18.0,
                    # legacy:
                    "ref_min":     11.0,
                    "ref_max":     20.0,
                    "formula":     "kg TC/día ÷ cabezas El Haras",
                },
                "conversion_alimenticia": {
                    "valor":       conversion,
                    "unidad":      "kg MS : kg carne",
                    "descripcion": "Kg MS por cab por día dividido ADP — El Haras",
                    # v15.20: rangos de 5 zonas (simétrico)
                    "esc_min":        5.0,
                    "esc_max":        13.0,
                    "ref_normal_min": 6.0,
                    "ref_normal_max": 10.0,
                    "ref_opt_min":    7.0,
                    "ref_opt_max":    9.0,
                    # legacy:
                    "ref_min":     7.0,
                    "ref_max":     9.0,
                    "formula":     "(kg MS/día ÷ cabezas El Haras) ÷ ADP",
                },
            },
        }
        guardar(indicadores, carpeta, f"indicadores_{periodo}.json")
        log.info(f"  ✓ indicadores_{periodo}.json")
        log.info(f"  % Peso Vivo        : {pct_pv}%   (ref óptimo 2.4–2.8% · normal 2.3–2.9 · escala 1.5–3.5)")
        log.info(f"  Consumo/cab (TC)   : {consumo_cab_tc} kg/cab/día   (ref óptimo 13–18 · normal 11–20 · escala 8–23)")
        log.info(f"  Consumo/cab (MS)   : {consumo_cab_ms} kg MS/cab/día")
        log.info(f"  Conversión alim.   : {conversion}:1   (ref óptimo 7–9 · normal 6–10 · escala 5–13)")
        log.info(f"  Producción diaria  : {prod_diaria_kg:,.0f} kg (ADP {adp_prom} × {cab_haras:,} cab Haras)")

        # ── Histórico de eficiencia — snapshot diario ──────────────────────
        try:
            hoy_str   = datetime.now().strftime("%Y-%m-%d")
            hist_path = Path(carpeta) / "eficiencia_historico.json"
            if hist_path.exists():
                with open(hist_path, encoding="utf-8") as _f:
                    hist_ef = json.load(_f)
            else:
                hist_ef = {"registros": []}

            # Reemplazar si ya existe entrada de hoy, sino agregar
            registro_hoy = {
                "fecha":          hoy_str,
                "cabezas":        cab_haras,
                "kg_pv":          round(kg_stock_haras, 0),
                "consumo_ms_cab": consumo_cab_ms,
                "consumo_tc_cab": consumo_cab_tc,
                "pct_pv":         pct_pv,
                "conversion":     conversion,
                "adp":            adp_prom,
                "adp_mes":        _ultimo_mes,
            }
            hist_ef["registros"] = [r for r in hist_ef["registros"] if r.get("fecha") != hoy_str]
            hist_ef["registros"].append(registro_hoy)
            # Ordenar cronológicamente y mantener últimos 365 días
            hist_ef["registros"].sort(key=lambda r: r.get("fecha",""))
            hist_ef["registros"] = hist_ef["registros"][-365:]
            hist_ef["generado"]  = datetime.now().isoformat()

            guardar(hist_ef, carpeta, "eficiencia_historico.json")
            log.info(f"  ✓ eficiencia_historico.json  ({len(hist_ef['registros'])} registros)")
        except Exception as _e:
            log.warning(f"  ⚠ No se pudo actualizar eficiencia_historico.json: {_e}")
        resumen["modulos"]["indicadores"] = {
            "ok":             True,
            "denominador":    "El Haras" if usando_haras else "PEGSA total",
            "pct_peso_vivo":  pct_pv,
            "consumo_cab_tc": consumo_cab_tc,
            "consumo_cab_ms": consumo_cab_ms,
            "conversion":     conversion,
        }
    except Exception as e:
        log.warning(f"  ⚠ No se pudieron calcular indicadores cruzados: {e}")
        resumen["modulos"]["indicadores"] = {"ok": False, "error": str(e)}

    # ── 10b. Precios de Indiferencia diarios (6 categorías-tipo) ──
    separador("Precios de Indiferencia")
    try:
        from calcular_indiferencia import actualizar_indiferencia_historico
        actualizar_indiferencia_historico(carpeta, log=log)
    except Exception as e:
        log.warning(f"  ⚠ No se pudieron calcular precios de indiferencia: {e}")

    # ── 11. Enriquecer stock_insumos con días de consumo restantes ──
    separador("Días de Stock Restantes")
    try:
        # Mapa nombre → promedio_diario TC desde consumo semanal
        consumo_por_nombre = {}
        for ins_c in cs.get("por_insumo", []):
            nombre = ins_c.get("desc", "").strip().upper()
            consumo_por_nombre[nombre] = ins_c.get("promedio_diario", 0)

        # Enriquecer cada insumo del stock
        insumos_enriquecidos = []
        for ins in insumos:
            nombre_up = ins["nombre"].strip().upper()
            prom_tc   = consumo_por_nombre.get(nombre_up, None)
            if prom_tc and prom_tc > 0:
                dias = round(ins["stock_kg"] / prom_tc, 1)
            else:
                dias = None
            ins_enr = dict(ins)
            ins_enr["consumo_diario_tc"] = prom_tc
            ins_enr["dias_restantes"]    = dias
            insumos_enriquecidos.append(ins_enr)
            if dias is not None:
                log.info(f"  {ins['nombre']:<28} stock {ins['stock_kg']:>12,.0f} kg ÷ {prom_tc:>8,.1f} kg/día = {dias:>6.1f} días")
            else:
                log.info(f"  {ins['nombre']:<28} stock {ins['stock_kg']:>12,.0f} kg  (sin consumo registrado)")

        # Reescribir stock_insumos con dias_restantes incluido
        guardar({
            "meta":     meta_ins,
            "insumos":  insumos_enriquecidos,
            "total_kg": round(total_kg, 2),
        }, carpeta, f"stock_insumos_{periodo}.json")
        log.info(f"  ✓ stock_insumos_{periodo}.json actualizado con días restantes")
    except Exception as e:
        log.warning(f"  ⚠ No se pudieron calcular días restantes: {e}")

    # ── 12. JSON Tesorería (Excel YYYY-MM-DD_financiero.xlsx en OneDrive) ──
    separador("Tesorería Financiera")
    try:
        import glob, os as _os

        # Buscar todos los YYYY-MM-DD_financiero.xlsx en subcarpeta financiero/
        subcarpeta_fin = _os.path.join(carpeta, "financiero")
        if not _os.path.exists(subcarpeta_fin):
            _os.makedirs(subcarpeta_fin)
            log.info(f"  ✓ Carpeta creada: {subcarpeta_fin}")
        patron  = _os.path.join(subcarpeta_fin, "[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]_financiero.xlsx")
        archivos = sorted(glob.glob(patron))
        log.info(f"  Archivos financiero encontrados: {len(archivos)}")

        def safe_float(v):
            if v is None: return None
            try:
                f = float(v)
                return f if pd.notna(f) else None
            except: return None

        def proc_financiero(ruta):
            """Procesa un Excel financiero con hoja 'resumen' y retorna dict con todos los datos."""
            nombre    = _os.path.basename(ruta)
            fecha_str = nombre[:10]

            sheets = pd.read_excel(ruta, sheet_name=None, header=None)

            # ── Hoja: resumen ──
            res = sheets.get('resumen')
            if res is None:
                log.warning(f"  ⚠ {nombre}: sin hoja 'resumen'")
                return None

            # Semanas: 17 columnas desde col 3
            N_COLS = 17
            from datetime import date as _date, timedelta
            try:
                base_date = _date.fromisoformat(fecha_str)
            except:
                base_date = _date.today()
            sem_labels = [(base_date + timedelta(weeks=i)).strftime('%d/%m') for i in range(N_COLS)]

            def gres(row, col):
                return safe_float(res.iloc[row, col]) or 0.0

            # Filas clave en hoja 'resumen' (confirmadas con Excel 2026-03-20)
            FILAS_RES = {
                'cheques_cobro':      30,
                'saldo_acum_bancos':  33,
                'venta_hacienda':     46,  # Total ingresos (incluye hacienda)
                'hoteleria':          43,  # Hotelería y alimentación Feed
                'total_ingresos':     46,
                'pagos_feedlot':      53,
                'pagos_admin':        54,
                'pago_impuestos':     55,
                'pago_flete':         56,
                'pago_agricultura':   57,
                'total_egresos':      65,
                'darwash':            67,
                'saldo_semanal':      69,
                'saldo_acumulado':    70,
            }
            series_flujo = {k: [gres(row, c) for c in range(3, 3+N_COLS)]
                            for k, row in FILAS_RES.items()}

            # ── Hoja: posicion hoy ──
            ph = sheets.get('posicion hoy', pd.DataFrame())
            def gph(r, c): return safe_float(ph.iloc[r, c]) if len(ph) > r else None

            bancos_peg  = [{'nombre': n, 'saldo': gph(i,4) or 0}
                           for i,n in [(2,'PECUARIA BNA'),(3,'PECUARIA BANCOR'),
                                       (4,'PECUARIA LA PAMPA'),(5,'PECUARIA GALICIA'),
                                       (6,'PECUARIA SANTANDER')] if gph(i,4)]
            bancos_bull = [{'nombre': n, 'saldo': gph(i,4) or 0}
                           for i,n in [(7,'BULLTRADE BNA'),(8,'BULLTRADE BANCOR'),
                                       (9,'BULLTRADE LA PAMPA'),(10,'BULLTRADE GALICIA'),
                                       (11,'BULLTRADE SANTANDER')] if gph(i,4)]
            efectivo     = gph(13,4) or 0
            becerra      = gph(16,4) or 0
            fima_bull    = gph(17,4) or 0
            fima_peg     = gph(18,4) or 0
            fci          = becerra + fima_bull + fima_peg
            echeq        = gph(21,4) or 0
            saldo_disp   = gph(22,4) or 0
            usd_ars      = gph(25,3) or 0
            usd_cant     = gph(25,1) or 0

            # Fix del bug del Excel: la celda de saldo_semanal de la primera
            # semana (fila 70 col D, fórmula =-D31+D47-D66-D68+B28) incluye
            # el saldo_inicial (B28 = 'posicion hoy'!E23) como arrastre; las
            # semanas 1..n (cols E-T) usan =-X31+X47-X66-X68 sin ese +B28.
            # Lo removemos para que toda la serie sea flujo puro semanal.
            # NO se recalcula saldo_acumulado: la fila 71 del Excel ya es el
            # saldo proyectado correcto, y el invariante
            #   saldo_acumulado[i] == saldo_inicial + cumsum(saldo_semanal)
            # se mantiene solo al corregir el índice 0.
            ss = series_flujo.get('saldo_semanal')
            if ss and len(ss) > 0:
                _orig_ss0 = ss[0]
                ss[0] = ss[0] - saldo_disp
                log.info(f"  saldo_semanal[0] ajustado: {_orig_ss0:,.2f} -> {ss[0]:,.2f} (removido saldo_inicial {saldo_disp:,.2f})")

            # ── Hoja: cheques pendiente ──
            cheq_raw = sheets.get('cheques pendiente', pd.DataFrame())
            from datetime import date as _date2
            hoy_c = _date2.fromisoformat(fecha_str)
            cheq_por_bucket = []
            total_cartera   = 0.0
            if len(cheq_raw) > 4:
                cheq_df = cheq_raw.iloc[4:].copy()
                cheq_df['fecha']   = pd.to_datetime(cheq_df[1], errors='coerce')
                cheq_df['importe'] = pd.to_numeric(cheq_df[5], errors='coerce').fillna(0)
                cheq_df = cheq_df[cheq_df['fecha'].notna() & (cheq_df['importe'] > 0)]
                def buck(f):
                    d = (f.date() - hoy_c).days
                    if d <= 7:   return '0-7 días'
                    if d <= 14:  return '8-14 días'
                    if d <= 30:  return '15-30 días'
                    if d <= 60:  return '31-60 días'
                    return '+60 días'
                cheq_df['bucket'] = cheq_df['fecha'].apply(buck)
                pb = cheq_df.groupby('bucket')['importe'].agg(['sum','count']).reindex(
                    ['0-7 días','8-14 días','15-30 días','31-60 días','+60 días'], fill_value=0).reset_index()
                pb.columns = ['bucket','monto','cantidad']
                cheq_por_bucket = pb.to_dict('records')
                total_cartera   = float(cheq_df['importe'].sum())

            # ── Hoja: vencimientos de hacienda ──
            hac      = sheets.get('vencimientos de hacienda', pd.DataFrame())
            hac_comp, hac_vent = [], []
            if len(hac) > 2:
                # Proveedores compras: fila 1, cols 1..N
                provs_comp = []
                for c in range(1, hac.shape[1]):
                    v = str(hac.iloc[1, c]).strip()
                    if v not in ['nan','NaT','']: provs_comp.append((c, v))
                # Proveedores ventas: fila 21, cols 1..N
                provs_vent = []
                for c in range(1, hac.shape[1]):
                    v = str(hac.iloc[21, c]).strip() if len(hac) > 21 else ''
                    if v not in ['nan','NaT','']: provs_vent.append((c, v))

                for i in range(2, min(19, len(hac))):
                    r = hac.iloc[i]; f = pd.to_datetime(r[0], errors='coerce')
                    if pd.isna(f): continue
                    detalle = []
                    for col, nombre in provs_comp:
                        v = safe_float(r[col]) or 0
                        if v: detalle.append({'empresa': nombre, 'monto': round(v, 2)})
                    total = sum(d['monto'] for d in detalle)
                    if total: hac_comp.append({'fecha': str(f.date()), 'monto': round(total, 2), 'detalle': detalle})

                for i in range(22, min(39, len(hac))):
                    r = hac.iloc[i]; f = pd.to_datetime(r[0], errors='coerce')
                    if pd.isna(f): continue
                    detalle = []
                    for col, nombre in provs_vent:
                        v = safe_float(r[col]) or 0
                        if v: detalle.append({'empresa': nombre, 'monto': round(v, 2)})
                    total = sum(d['monto'] for d in detalle)
                    if total: hac_vent.append({'fecha': str(f.date()), 'monto': round(total, 2), 'detalle': detalle})

            # ── Hoja: gastos varios ──
            gv_sheet = sheets.get('gastos varios', pd.DataFrame())
            gastos, cat = [], ''
            for i in range(1, len(gv_sheet)):
                r = gv_sheet.iloc[i]; concepto = str(r[0]).strip()
                if not concepto or concepto == 'nan': continue
                if concepto.endswith(':'): cat = concepto.rstrip(':'); continue
                mt = safe_float(r[2]) or 0
                mb = safe_float(r[1]) or 0
                freq = str(r[3]).strip() if len(r) > 3 and pd.notna(r[3]) else ''
                if mt or mb:
                    gastos.append({'categoria': cat, 'concepto': concepto,
                                   'monto_bruto': mb, 'monto_total': mt, 'frecuencia': freq})

            # ── Hoja: cuenta corriente con darwash ──
            # La col 2 tiene el tipo ('ingreso'/'egreso') explícito — usarla directamente
            cc_sheet = sheets.get('cuenta  corriente con darwash', pd.DataFrame())
            darwash_secs, s_nom, s_items = [], '', []
            for i in range(len(cc_sheet)):
                r = cc_sheet.iloc[i]; nom = str(r[0]).strip()
                if nom.endswith(':') and nom not in ['nan','']:
                    if s_items: darwash_secs.append({'nombre': s_nom, 'items': s_items})
                    s_nom = nom.rstrip(':'); s_items = []; continue
                f = pd.to_datetime(r[0], errors='coerce')
                m = safe_float(r[1])
                tipo_col = str(r[2]).strip().lower() if len(r) > 2 and pd.notna(r[2]) else ''
                if pd.notna(f) and m is not None and m != 0:
                    # Usar tipo del Excel si está disponible, sino inferir por signo
                    if tipo_col in ('ingreso', 'egreso'):
                        tipo = tipo_col
                        monto_real = abs(m) if tipo == 'ingreso' else -abs(m)
                    else:
                        # fallback: negativo en Excel = ingreso para PEGSA
                        monto_real = -m
                        tipo = 'ingreso' if monto_real > 0 else 'egreso'
                    s_items.append({'fecha': str(f.date()), 'monto': monto_real, 'tipo': tipo})
            if s_items: darwash_secs.append({'nombre': s_nom, 'items': s_items})

            return {
                'archivo':    nombre,
                'fecha_corte': fecha_str,
                'posicion': {
                    'bancos_peg':            bancos_peg,
                    'bancos_bull':           bancos_bull,
                    'efectivo':              efectivo,
                    'becerra':               becerra,
                    'fima_bull':             fima_bull,
                    'fima_peg':              fima_peg,
                    'fci':                   fci,
                    'echeq':                 echeq,
                    'saldo_disponibilidades': saldo_disp,
                    'usd_ars':               usd_ars,
                    'usd_cant':              usd_cant,
                },
                'cheques': {
                    'total_cartera':    total_cartera,
                    'por_vencimiento':  cheq_por_bucket,
                },
                'hacienda':  {'compras': hac_comp, 'ventas': hac_vent},
                'gastos':    gastos,
                'darwash':   darwash_secs,
                'flujo': {
                    'semanas': sem_labels,
                    'saldo_inicial': saldo_disp,
                    'series':  {k: [round(v, 2) for v in vals]
                                for k, vals in series_flujo.items()},
                },
            }

        # Procesar todos los archivos
        # v15.32: robusto ante placeholders OneDrive (cloud-only). Un archivo
        # no hidratado hacía fallar pd.read_excel con [Errno 22] y tumbaba todo
        # el módulo (ok=false → banner stale). Ahora detectamos placeholder,
        # intentamos warming, y si falla skipeamos ESE archivo reportándolo como
        # warning — el resto de los cortes se procesa igual.
        cortes_proc = []
        ultimo_corte = None
        archivos_skipados = []
        for ruta in archivos:
            nombre = _os.path.basename(ruta)
            fecha_str = nombre[:10]
            try: datetime.strptime(fecha_str, "%Y-%m-%d")
            except:
                log.warning(f"  ⚠ Ignorando: {nombre}"); continue

            es_ph, motivo = _xlsx_es_placeholder_onedrive(ruta)
            if es_ph:
                log.warning(f"  ⚠ {nombre}: placeholder OneDrive ({motivo}). Intentando warming...")
                if not _intentar_warming_onedrive(ruta):
                    log.warning(f"  ⚠ {nombre}: skipeado ({motivo})")
                    archivos_skipados.append({"archivo": nombre, "motivo": motivo})
                    continue
                log.info(f"  ✓ {nombre}: hidratado tras warming")

            log.info(f"  Procesando: {nombre}")
            try:
                resultado = proc_financiero(ruta)
            except Exception as _e:
                log.warning(f"  ⚠ {nombre}: error al parsear ({type(_e).__name__}: {_e}). Skipeado.")
                archivos_skipados.append({"archivo": nombre, "motivo": f"parse_error: {type(_e).__name__}"})
                continue
            if resultado:
                cortes_proc.append(resultado)
                ultimo_corte = resultado
                sd = resultado['posicion']['saldo_disponibilidades']
                log.info(f"  ✓ {nombre} · saldo: ${sd:,.0f}")

        if cortes_proc:
            # JSON con todos los cortes (para histórico)
            guardar({'generado': datetime.now().isoformat(), 'cortes': cortes_proc},
                    carpeta, "financiero_historico.json")
            log.info(f"  ✓ financiero_historico.json — {len(cortes_proc)} cortes")

            # JSON del último corte (para módulo 6 del portal)
            guardar(ultimo_corte, carpeta, "tesoreria_ultimo.json")
            log.info(f"  ✓ tesoreria_ultimo.json — corte {ultimo_corte['fecha_corte']}")

            resumen["modulos"]["tesoreria"] = {
                "ok":          True,
                "cortes":      len(cortes_proc),
                "ultimo_corte": ultimo_corte['fecha_corte'],
                "saldo_disp":  ultimo_corte['posicion']['saldo_disponibilidades'],
            }
            # v15.32: archivos skipados como warning (no fatal)
            if archivos_skipados:
                resumen["modulos"]["tesoreria"]["warnings"] = archivos_skipados
                resumen["modulos"]["tesoreria"]["warning_count"] = len(archivos_skipados)
                log.warning(f"  ⚠ Tesorería: {len(archivos_skipados)} archivo(s) skipado(s) "
                            f"(placeholder/parse). Módulo OK con {len(cortes_proc)} cortes.")
        elif archivos_skipados:
            # Había archivos pero NINGUNO se pudo procesar → error real
            log.warning("  ⚠ Tesorería: ningún archivo financiero se pudo procesar")
            resumen["modulos"]["tesoreria"] = {
                "ok":      False,
                "error":   "No se pudo procesar ningún archivo financiero",
                "archivos_skipados": archivos_skipados,
            }
        else:
            log.info("  ℹ Sin archivos YYYY-MM-DD_financiero.xlsx en la carpeta")
            resumen["modulos"]["tesoreria"] = {"ok": True, "cortes": 0}

    except Exception as e:
        log.warning(f"  ⚠ Módulo tesorería falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["tesoreria"] = {"ok": False, "error": str(e)}

    # ── MÓDULO MERCADO Y PRECIOS (web scraping) ───────────────
    separador("MÓDULO 7 · MERCADO Y PRECIOS")
    try:
        _repo_txt = Path(__file__).parent / "repo_github_path.txt"
        _repo_path = _repo_txt.read_text(encoding="utf-8").strip() if _repo_txt.exists() else ""
        if not _repo_path:
            _repo_path = cfg.get("GITHUB", {}).get("repo_path", "")
        actualizar_mercado_precios(carpeta, _repo_path)
    except Exception as e:
        log.warning(f"  ⚠ Módulo mercado falló: {e}")
        import traceback; log.warning(traceback.format_exc())

    # ── MÓDULO 8 · HISTÓRICO MENSUAL (reconstrucción completa) ──
    separador("MÓDULO 8 · HISTÓRICO MENSUAL")
    try:
        import calendar as _cal
        _hist_path = Path(carpeta) / "stock_historico.json"
        _kpis_path = Path(carpeta) / f"stock_kpis_{periodo}.json"
        _ins_path  = Path(carpeta) / f"stock_insumos_{periodo}.json"
        _mov_path  = Path(carpeta) / f"movimientos_{periodo}.json"

        if _kpis_path.exists() and _mov_path.exists():
            with open(_kpis_path, encoding="utf-8") as _f: _kpis_raw = json.load(_f)
            with open(_mov_path,  encoding="utf-8") as _f: _mov_raw  = json.load(_f)
            _ins_raw = json.load(open(_ins_path, encoding="utf-8")) if _ins_path.exists() else {}

            _k      = _kpis_raw.get("kpis", {})
            _anio   = _mov_raw.get("anio", {})
            _ing_mes  = _anio.get("ingresos", {}).get("por_mes", {})
            _egr_mes  = _anio.get("egresos",  {}).get("por_mes", {})
            _ing_prop = _anio.get("ingresos", {}).get("por_propietario", {})
            _egr_prop = _anio.get("egresos",  {}).get("por_propietario", {})
            _ing_cat  = _anio.get("ingresos", {}).get("por_categoria", {})
            _egr_cat  = _anio.get("egresos",  {}).get("por_categoria", {})

            # Totales anuales para distribución proporcional
            _tic = sum(v.get("cabezas",0) or 0 for v in _ing_prop.values()) or 1
            _tec = sum(v.get("cabezas",0) or 0 for v in _egr_prop.values()) or 1
            _tic2= sum(v.get("cabezas",0) or 0 for v in _ing_cat.values())  or 1
            _tec2= sum(v.get("cabezas",0) or 0 for v in _egr_cat.values())  or 1

            # Punto de anclaje: hoy
            _stock_a = _k.get("total_cabezas", 0)
            _kg_a    = _k.get("total_kg_estimado_hoy", 0)
            _prop_a  = {p: v["cabezas"] for p,v in _k.get("por_propietario",{}).items()}
            _cat_a   = {c: v.get("cabezas",0) for c,v in _k.get("por_categoria",{}).items()}

            _hoy_str  = datetime.now().strftime("%Y-%m")
            _all_meses= sorted(set(list(_ing_mes)+list(_egr_mes)))
            _meses    = sorted([m for m in _all_meses if m <= _hoy_str], reverse=True)

            _snaps = []
            for _mes in _meses:
                _ic = _ing_mes.get(_mes,{}).get("cabezas",0) or 0
                _ec = _egr_mes.get(_mes,{}).get("cabezas",0) or 0
                _ikg= _ing_mes.get(_mes,{}).get("kg",0)      or 0
                _ekg= _egr_mes.get(_mes,{}).get("kg",0)      or 0

                _y, _m = int(_mes[:4]), int(_mes[5:7])
                _ld   = _cal.monthrange(_y, _m)[1]
                _fecha= f"{_y:04d}-{_m:02d}-{_ld:02d}"

                _pp = {p: {"cabezas": c, "kg_estimado": round(_kg_a * c / max(_stock_a,1))}
                       for p,c in _prop_a.items()}
                _pc = {c: {"cabezas": cab, "kg_estimado": round(_kg_a * cab / max(_stock_a,1))}
                       for c,cab in _cat_a.items()}

                _snaps.append({
                    "fecha": _fecha, "periodo": _mes,
                    "hacienda": {
                        "total_cabezas":     _stock_a,
                        "total_kg_estimado": max(0, round(_kg_a)),
                        "por_propietario":   _pp,
                        "por_categoria":     _pc,
                    },
                    "insumos": {
                        "total_kg": _ins_raw.get("total_kg", 0),
                        "items": [{"nombre": it["nombre"], "stock_kg": it["stock_kg"]}
                                  for it in _ins_raw.get("insumos", [])]
                    } if _mes == _hoy_str else {"total_kg": 0, "items": []}
                })

                # Retroceder al mes anterior
                _stock_a = max(0, _stock_a + _ec - _ic)
                _kg_a    = max(0, _kg_a    + _ekg- _ikg)
                _prop_a  = {p: max(0, round(c + (_egr_prop.get(p,{}).get("cabezas",0) or 0)/_tec*_ec
                                              - (_ing_prop.get(p,{}).get("cabezas",0) or 0)/_tic*_ic))
                            for p,c in _prop_a.items()}
                _cat_a   = {c: max(0, round(cab + (_egr_cat.get(c,{}).get("cabezas",0) or 0)/_tec2*_ec
                                              - (_ing_cat.get(c,{}).get("cabezas",0) or 0)/_tic2*_ic))
                            for c,cab in _cat_a.items()}

            _snaps.reverse()
            _hist_out = {"generado": datetime.now().isoformat(), "fuente": "reconstruccion_sql",
                         "snapshots": _snaps}
            guardar(_hist_out, carpeta, "stock_historico.json")
            log.info(f"  ✓ stock_historico.json — {len(_snaps)} meses reconstruidos "
                     f"({_snaps[0]['periodo']} → {_snaps[-1]['periodo']})")
            resumen["modulos"]["historico"] = {
                "ok": True, "snapshots": len(_snaps),
                "rango": f"{_snaps[0]['periodo']} → {_snaps[-1]['periodo']}"
            }
        else:
            log.info("  ℹ stock_kpis o movimientos no encontrados — histórico omitido")
            resumen["modulos"]["historico"] = {"ok": True, "snapshots": 0}
    except Exception as e:
        log.warning(f"  ⚠ Snapshot histórico falló: {e}")
        resumen["modulos"]["historico"] = {"ok": False, "error": str(e)}

    # ── MÓDULO 9 · COMPORTAMIENTO HISTÓRICO MENSUAL ───────────
    separador("MÓDULO 9 · COMPORTAMIENTO HISTÓRICO MENSUAL")
    try:
        from pathlib import Path as _Path9
        _carpeta_sm9 = _Path9(carpeta).parent / "stock mensuales"
        _hist9 = actualizar_comportamiento_historico(carpeta, str(_carpeta_sm9))
        _n9 = _hist9.get('total', 0) if _hist9 else 0
        resumen["modulos"]["comportamiento_historico"] = {
            "ok": True, "snapshots": _n9,
        }
    except Exception as e:
        log.warning(f"  ⚠ Módulo 9 falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["comportamiento_historico"] = {"ok": False, "error": str(e)}

    # ── MÓDULO 10 · VALUACIÓN EN PESOS ───────────────────────────
    separador("MÓDULO 10 · VALUACIÓN EN PESOS")
    try:
        _val_path = Path(carpeta) / "comportamiento_historico.json"
        if _val_path.exists():
            with open(_val_path, encoding="utf-8") as _fv:
                _hist9_data = json.load(_fv)
            _snaps_hist9 = _hist9_data.get("snapshots", [])
            if _snaps_hist9:
                _val = actualizar_valuacion(carpeta, _snaps_hist9)
                resumen["modulos"]["valuacion"] = {
                    "ok": True, "periodos": len(_val.get("snapshots", []))
                }
            else:
                log.info("  ℹ comportamiento_historico.json sin snapshots — valuación omitida")
                resumen["modulos"]["valuacion"] = {"ok": True, "periodos": 0}
        else:
            log.info("  ℹ comportamiento_historico.json no encontrado — valuación omitida")
            resumen["modulos"]["valuacion"] = {"ok": True, "periodos": 0}
    except Exception as e:
        log.warning(f"  ⚠ Módulo 10 falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["valuacion"] = {"ok": False, "error": str(e)}

    # ── STOCK DIARIO · RUNNING BALANCE ──────────────────────────
    # Recalcula el historial completo desde movimientos reales,
    # incorporando automáticamente cargas retroactivas de compras/ventas.
    separador("Stock Diario · Running Balance")
    if _regs_stock_hoy is None:
        # v15.45: el stock de hoy no está disponible (fetch falló). NO recalculamos
        # el running balance con baseline 0 — eso pisaría stock_diario.json con una
        # caída a cero. Se conserva el JSON del run anterior (last-known-good).
        log.warning("  ⚠ Stock de hoy no disponible — se conserva stock_diario.json del run anterior")
        resumen["modulos"]["stock_diario"] = {"ok": False, "error": "stock_hacienda no disponible este run"}
    else:
        try:
            _diario = recalcular_stock_diario_desde_movimientos(
                _regs_stock_hoy, _cols_stock_hoy,
                regs_ing,        cols_ing,
                regs_egr,        cols_egr,
                carpeta,         periodo,
                dias=90
            )
            resumen["modulos"]["stock_diario"] = {
                "ok":          True,
                "dias":        _diario["dias"],
                "ultima_fecha": datetime.now().strftime("%Y-%m-%d"),
                "metodo":      "running_balance",
            }
        except Exception as e:
            log.warning(f"  ⚠ Running balance diario falló: {e}")
            import traceback; log.warning(traceback.format_exc())
            resumen["modulos"]["stock_diario"] = {"ok": False, "error": str(e)}

    # ── TESORERÍA DARWASH (v11) ─────────────────────────────
    # Análisis financiero independiente de Darwash. Lee el XLSX más
    # reciente de `datos/financiero DW/` y vuelca tesoreria_darwash.json
    # + tesoreria_darwash_historico.json (acumulado por fecha_corte).
    separador("Tesorería · Darwash")
    try:
        _dw_snap, _dw_hist = procesar_tesoreria_darwash(carpeta, log)
        resumen["modulos"]["tesoreria_darwash"] = {
            "ok":          _dw_snap is not None,
            "fecha_corte": _dw_snap["fecha_corte"] if _dw_snap else None,
            "snapshots":   len(_dw_hist["snapshots"]) if _dw_hist else 0,
        }
    except Exception as e:
        log.warning(f"  ⚠ procesar_tesoreria_darwash falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["tesoreria_darwash"] = {"ok": False, "error": str(e)}

    # ── TRAZABILIDAD · Caravanas declaradas (Google Drive) ─────
    # Lee los Excel colaborativos de G:\Mi unidad\Trazabilidad\ y vuelca
    # trazabilidad_resumen.json (KPIs por hoja + consolidado global).
    separador("Trazabilidad · Caravanas Declaradas")
    try:
        _traz = procesar_trazabilidad(carpeta, log)
        resumen["modulos"]["trazabilidad"] = {
            "ok":      _traz is not None,
            "hojas":   len(_traz["hojas"]) if _traz else 0,
            "activas": _traz["consolidado"]["activas"] if _traz else 0,
        }
    except Exception as e:
        log.warning(f"  ⚠ procesar_trazabilidad falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["trazabilidad"] = {"ok": False, "error": str(e)}

    # ── PRECIOS DE INFERENCIA (v8) ─────────────────────────────
    # Lee el Excel del simulador y vuelca precios_inferencia.json +
    # precios_inferencia_historico.json (acumulado semanal por fecha).
    separador("Precios de Inferencia")
    try:
        _snap, _hist = procesar_precios_inferencia(carpeta, log)
        resumen["modulos"]["precios_inferencia"] = {
            "ok":       _snap is not None,
            "fecha":    _snap["meta"]["fecha"] if _snap else None,
            "semanas":  len(_hist["semanas"]) if _hist else 0,
        }
    except Exception as e:
        log.warning(f"  ⚠ procesar_precios_inferencia falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["precios_inferencia"] = {"ok": False, "error": str(e)}

    # ── COMPRAS REALES (v15.57) ────────────────────────────────
    # Lee el Excel de compras del OneDrive compartido y vuelca
    # precios_compra_real.json — el precio REAL pagado por categoría, que las
    # tarjetas de indiferencia muestran al lado del tope teórico.
    separador("Compras reales · precio pagado")
    try:
        _compras = procesar_compras_reales(carpeta, log)
        resumen["modulos"]["compras_reales"] = {
            "ok":           _compras is not None,
            "filas":        _compras["meta"]["filas_totales"] if _compras else 0,
            "descartadas":  _compras["meta"]["filas_descartadas"] if _compras else 0,
            "categorias":   len(_compras["por_categoria"]) if _compras else 0,
        }
    except Exception as e:
        log.warning(f"  ⚠ procesar_compras_reales falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["compras_reales"] = {"ok": False, "error": str(e)}

    # ── %PV MENSUAL HISTÓRICO (v15.58) ─────────────────────────
    # La serie diaria de pct_pv (eficiencia_historico) solo cubre desde
    # 2026-04-30 y el módulo Resultado por Remito costea animales con hasta
    # 360 días de estadía → esto reconstruye el %PV mes a mes.
    #   - numerador:   consumo_{periodo}.json → por_mes (mixer, días válidos)
    #   - denominador: comportamiento_historico → El Haras kg_proyectado,
    #                  PROMEDIO del mes = (fin mes anterior + fin mes) / 2
    #   - ÷ 0.92: mismo ajuste que el indicador del portal (data.js ~L761)
    # Va acá y NO en la sección 9 porque comportamiento_historico.json se
    # genera en el Módulo 9 (más arriba) y se lee desde DISCO, no de memoria:
    # así funciona igual si algún módulo intermedio falló.
    separador("%PV mensual histórico")
    try:
        _pct = generar_pct_pv_mensual(carpeta, periodo, log)
        resumen["modulos"]["pct_pv_mensual"] = {
            "ok":     _pct is not None,
            "meses":  len(_pct["meses"]) if _pct else 0,
            "desde":  _pct["meta"]["desde"] if _pct else None,
            "hasta":  _pct["meta"]["hasta"] if _pct else None,
        }
    except Exception as e:
        log.warning(f"  ⚠ generar_pct_pv_mensual falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["pct_pv_mensual"] = {"ok": False, "error": str(e)}

    # ── RESULTADO POR REMITO (v15.59) ──────────────────────────
    # Costo completo de cada venta con remito. Va acá, al final, porque necesita
    # pct_pv_mensual.json y precios_compra_real.json ya guardados en disco.
    # Reusa los egresos del pre-step (no vuelve a pegarle a la API).
    separador("Resultado por Remito")
    try:
        _rr = generar_resultado_remitos(carpeta, periodo, egresos_data, log)
        resumen["modulos"]["resultado_remitos"] = {
            "ok":        _rr is not None,
            "remitos":   _rr["meta"]["remitos"] if _rr else 0,
            "cobertura": _rr["meta"]["cobertura_global_pct"] if _rr else None,
        }
    except Exception as e:
        log.warning(f"  ⚠ generar_resultado_remitos falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["resultado_remitos"] = {"ok": False, "error": str(e)}

    # ── CARAVANAS FANTASMA (v15.69) ────────────────────────────
    # Va después del cruce: reusa la cache de Datamars y el stock de este tick.
    separador("Caravanas fantasma")
    try:
        _fa = generar_fantasmas(carpeta, periodo, egresos_data, log, stock_data=regs or None)
        resumen["modulos"]["fantasmas"] = {
            "ok":    _fa is not None,
            "total": _fa["meta"]["total"] if _fa else 0,
        }
    except Exception as e:
        log.warning(f"  ⚠ generar_fantasmas falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["fantasmas"] = {"ok": False, "error": str(e)}

    # ── ANÁLISIS DE COSTOS · denominadores operativos (v15.61) ─
    # Sólo los denominadores de los ratios. El JSON contable del módulo
    # (analisis_costos_datos.json) lo regenera Nicolás a mano con su ejecutable
    # de Physis — el pipeline NUNCA lo escribe.
    separador("Análisis de Costos · operativos")
    try:
        _op = generar_analisis_costos_operativos(carpeta, periodo, log)
        resumen["modulos"]["analisis_costos_operativos"] = {
            "ok":    _op is not None,
            "meses": len(_op) if _op else 0,
        }
    except Exception as e:
        log.warning(f"  ⚠ generar_analisis_costos_operativos falló: {e}")
        import traceback; log.warning(traceback.format_exc())
        resumen["modulos"]["analisis_costos_operativos"] = {"ok": False, "error": str(e)}

    # ── v15.14: Smoke test post-pipeline ──
    # Valida que los JSONs tengan números razonables. Si falla, queda registrado
    # en modulos.smoke_test → el banner stale v15.12 lo detecta como módulo con
    # error y se muestra en el portal. Pipeline NUNCA aborta por esto.
    separador("Smoke Test · Validación post-pipeline")
    try:
        smoke = smoke_test(carpeta, periodo)
        if smoke["ok"]:
            log.info(f"  ✓ Smoke test OK ({smoke['checks_passed']} checks)")
        else:
            log.error(f"  ✗ Smoke test FALLÓ ({smoke['checks_failed']} checks fallidos):")
            for err in smoke["errors"]:
                log.error(f"     - {err}")
        resumen["modulos"]["smoke_test"] = {
            "ok": smoke["ok"],
            "checks_passed": smoke["checks_passed"],
            "checks_failed": smoke["checks_failed"],
            "errors": smoke["errors"][:5],  # max 5 para no explotar el JSON
        }
    except Exception as e:
        log.error(f"  ✗ Smoke test crasheó: {e}")
        resumen["modulos"]["smoke_test"] = {"ok": False, "errors": [f"smoke test crashed: {e}"]}

    separador()
    guardar(resumen, carpeta, "ultima_actualizacion.json")

    # ── AUTO GIT PUSH ─────────────────────────────────────────
    separador("GIT · PUBLICAR EN GITHUB PAGES")
    _repo_txt = Path(__file__).parent / "repo_github_path.txt"
    _repo_path = _repo_txt.read_text(encoding="utf-8").strip() if _repo_txt.exists() else ""
    if not _repo_path:
        _repo_path = cfg.get("GITHUB", {}).get("repo_path", "")
    if _repo_path and Path(_repo_path).is_dir():
        import subprocess, datetime as _dt, shutil as _shutil
        try:
            _ts   = _dt.datetime.now().strftime("%a %d/%m/%Y %H:%M:%S")
            _repo = Path(_repo_path)

            # 1) Copiar JSONs de datos de OneDrive → repo
            #    Excluir archivos con credenciales o configuración sensible
            _EXCLUIR = {"lector-robot", "credential", "secret", "config", "key"}
            _copiados = 0
            for _json in Path(carpeta).glob("*.json"):
                _nombre_lower = _json.name.lower()
                if any(excl in _nombre_lower for excl in _EXCLUIR):
                    log.info(f"  ⚠ Omitido (sensible): {_json.name}")
                    continue
                _dst = _repo / _json.name
                _shutil.copy2(str(_json), str(_dst))
                _copiados += 1
            log.info(f"  ✓ {_copiados} JSON copiados a repo")

            # 2) Commit + push
            subprocess.run(["git", "-C", str(_repo), "add", "-A"],
                           check=True, capture_output=True)
            _res = subprocess.run(
                ["git", "-C", str(_repo), "commit", "-m",
                 f"Actualizacion automatica {_ts}"],
                capture_output=True, text=True
            )
            if "nothing to commit" in _res.stdout or "nothing to commit" in _res.stderr:
                log.info("  ℹ Git: sin cambios nuevos para publicar")
            else:
                log.info(f"  ✓ Git commit: Actualizacion automatica {_ts}")
                # Intento 1: push directo
                push = subprocess.run(
                    ["git", "-C", str(_repo), "push"],
                    capture_output=True, text=True, timeout=60
                )
                if push.returncode == 0:
                    log.info("  ✓ Git push OK → GitHub Pages actualizado")
                else:
                    # Intento 2: si rechazaron por non-fast-forward, hacer pull --rebase y retry
                    err = (push.stderr or "").strip()
                    if "non-fast-forward" in err or "fetch first" in err or "rejected" in err:
                        log.warning("  ⚠ Push rechazado (remoto tiene commits nuevos). Haciendo pull --rebase...")
                        pull = subprocess.run(
                            ["git", "-C", str(_repo), "pull", "--rebase", "--autostash"],
                            capture_output=True, text=True, timeout=60
                        )
                        if pull.returncode == 0:
                            log.info("  ✓ Pull --rebase OK")
                            push2 = subprocess.run(
                                ["git", "-C", str(_repo), "push"],
                                capture_output=True, text=True, timeout=60
                            )
                            if push2.returncode == 0:
                                log.info("  ✓ Git push OK → GitHub Pages actualizado (después de rebase)")
                            else:
                                log.warning(f"  ⚠ Git push falló incluso después de rebase: {push2.stderr.strip()[:200]}")
                        else:
                            log.warning(f"  ⚠ Pull --rebase falló: {pull.stderr.strip()[:200]}")
                            log.warning("    Resolución manual: cd al repo y correr 'git pull --rebase && git push'")
                    else:
                        log.warning(f"  ⚠ Git push falló: {err[:200]}")
        except Exception as _e:
            log.warning(f"  ⚠ Git error: {_e}")
            import traceback; log.warning(traceback.format_exc())
    else:
        log.info("  ℹ Git: repo_github_path.txt no configurado, se omite push")

    separador("FINALIZADO")
    log.info(f"  Archivos guardados en: {carpeta}")
    log.info("  OneDrive sincronizara automaticamente")
    separador()
    print()
    esperar_si_interactivo("  Presiona Enter para cerrar...")


# ══════════════════════════════════════════════════════════════
# MÓDULO 7 — MERCADO Y PRECIOS
# Fuentes:
#   - Hacienda: Mercado de Cañuelas (decampoacampo.com)
#   - Granos:   BCR Cámara Arbitral (cac.bcr.com.ar/es/precios-de-pizarra)
#   - Negocios: Google Sheets (CARGAS + COMPRAS)
# ══════════════════════════════════════════════════════════════

# ID de la planilla de negocios en Google Sheets
GSHEET_ID = "1_N1k3QkNQ8NMfs-uz_FHmpLd8afR067-EsoThkg0RWk"
# URL base para export CSV público (requiere que la hoja esté "Publicada en web")
GSHEET_CSV_BASE = f"https://docs.google.com/spreadsheets/d/{GSHEET_ID}/export?format=csv&sheet="


def _http_get(url, timeout=20):
    """Descarga URL como texto. Devuelve str o None."""
    import urllib.request
    try:
        req = urllib.request.Request(
            url,
            headers={
                "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                "Accept": "text/html,application/json,*/*",
                "Accept-Language": "es-AR,es;q=0.9",
            }
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            raw = resp.read()
            # detectar encoding
            ct = resp.headers.get("Content-Type", "")
            enc = "utf-8"
            if "charset=" in ct:
                enc = ct.split("charset=")[-1].split(";")[0].strip()
            return raw.decode(enc, errors="replace")
    except Exception as e:
        log.debug(f"    _http_get({url[:60]}...): {e}")
        return None


def _parse_ar_num(s):
    """Convierte string numérico argentino/internacional a float."""
    s = str(s or "").strip().replace(" ", "").replace("$", "")
    if not s or s in ("-", "—", ""):
        return None
    if "," in s and "." in s:
        # "1.234,56" → 1234.56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        # puede ser "1234,56" (decimal) o "1.234" (miles)
        partes = s.split(",")
        if len(partes) == 2 and len(partes[1]) <= 2:
            s = s.replace(",", ".")  # decimal: 1234,56
        else:
            s = s.replace(",", "")   # miles: 1,234
    elif "." in s:
        partes = s.split(".")
        if len(partes) == 2 and len(partes[1]) == 3:
            s = s.replace(".", "")  # miles: 1.234
    try:
        return float(s)
    except Exception:
        return None


# ──────────────────────────────────────────────────────────────
# 7_HIST. Histórico de precios — Excel en OneDrive
# ──────────────────────────────────────────────────────────────
def actualizar_historico_excel(hacienda, commodities, carpeta, today):
    """
    Agrega/actualiza una fila en historico_precios.xlsx con los precios del día.
    Si no llegaron precios de alguna columna, replica el último valor conocido (carry-forward).
    Requiere: pandas + openpyxl
    """
    try:
        import pandas as pd
        from pathlib import Path
    except ImportError:
        log.warning("  ⚠ pandas no disponible; se omite historico_precios.xlsx")
        return

    ARCHIVO = Path(carpeta) / "historico_precios.xlsx"

    # Mapeo: clave interna → nombre de columna Excel
    # IMPORTANTE: Las claves deben ser substrings exactos de los nombres de categoría
    # devueltos por la API de Cañuelas (plural: "Novillitos", "Novillos", "Vacas", etc.)
    # Ejemplo de respuesta API: "Novillitos hasta 390 Kg.", "Novillitos 391/430 Kg.", etc.
    COLS_HAC = [
        ("novillitos hasta",    "Novillito <390kg $/kg"),
        ("novillitos 391",      "Novillito 391/430kg $/kg"),
        ("novillos 431",        "Novillo 431/460kg $/kg"),
        ("novillos 461",        "Novillo 461/490kg $/kg"),
        ("vaquillona",          "Vaquillona <390kg $/kg"),
        ("vacas buenas",        "Vaca Buena $/kg"),
        ("vacas regulares",     "Vaca Regular $/kg"),
        ("vacas conserva",      "Vaca Conserva $/kg"),
        ("ternero",             "Ternero $/kg"),
        ("ternera",             "Ternera $/kg"),
    ]
    COLS_COM = [
        ("maíz",   "Maíz $/tn"),   # usar tilde para que coincida con nombre="Maíz"
        ("soja",   "Soja $/tn"),
        ("trigo",  "Trigo $/tn"),
        ("sorgo",  "Sorgo $/tn"),
    ]

    # Construir dict de precios de hacienda de hoy
    def _hac_precio(key_lower):
        for h in hacienda:
            if key_lower in h.get("categoria", "").lower():
                p = h.get("precio", 0)
                if p and p > 500:
                    return p
        return None

    def _com_precio(key_lower):
        for c in commodities:
            if key_lower in c.get("nombre", "").lower():
                p = c.get("precio", 0)
                if p and p > 1000:
                    return p
        return None

    # Leer Excel existente o crear DataFrame vacío
    todas_cols = ["Fecha"] + [c for _, c in COLS_HAC] + [c for _, c in COLS_COM]
    if ARCHIVO.exists():
        try:
            df = pd.read_excel(ARCHIVO, sheet_name="Historico", dtype={"Fecha": str})
            # Asegurar que todas las columnas existen
            for col in todas_cols:
                if col not in df.columns:
                    df[col] = None
        except Exception as e:
            log.warning(f"  ⚠ No se pudo leer {ARCHIVO.name}: {e}. Se crea nuevo.")
            df = pd.DataFrame(columns=todas_cols)
    else:
        df = pd.DataFrame(columns=todas_cols)

    # Obtener última fila como valores de carry-forward
    if len(df) > 0:
        last = df.iloc[-1].to_dict()
    else:
        last = {}

    # Armar la fila de hoy
    fila = {"Fecha": today}
    for key, col in COLS_HAC:
        p = _hac_precio(key)
        if p:
            fila[col] = p
        else:
            # Carry-forward: usar el último valor conocido
            fila[col] = last.get(col, None)

    for key, col in COLS_COM:
        p = _com_precio(key)
        if p:
            fila[col] = p
        else:
            fila[col] = last.get(col, None)

    # Si hoy ya existe, reemplazar; si no, agregar
    if "Fecha" in df.columns and today in df["Fecha"].values:
        df.loc[df["Fecha"] == today, list(fila.keys())] = list(fila.values())
        log.info(f"  ✓ historico_precios.xlsx — fila {today} actualizada")
    else:
        df = pd.concat([df, pd.DataFrame([fila])], ignore_index=True)
        log.info(f"  ✓ historico_precios.xlsx — fila {today} agregada ({len(df)} días totales)")

    # Guardar con formato
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(ARCHIVO, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Historico", index=False)
            ws = writer.sheets["Historico"]

            # Estilo encabezado
            fill_hdr = PatternFill("solid", fgColor="1F4E79")
            font_hdr = Font(bold=True, color="FFFFFF", size=10)
            for cell in ws[1]:
                cell.fill = fill_hdr
                cell.font = font_hdr
                cell.alignment = Alignment(horizontal="center", wrap_text=True)

            # Freeze primera fila
            ws.freeze_panes = "A2"

            # Ancho automático
            for col_idx, col_name in enumerate(todas_cols, start=1):
                max_len = max(len(str(col_name)), 10)
                ws.column_dimensions[get_column_letter(col_idx)].width = max_len * 1.2

            # Filas alternadas
            fill_par  = PatternFill("solid", fgColor="EBF3FB")
            fill_impar = PatternFill("solid", fgColor="FFFFFF")
            for row_idx in range(2, ws.max_row + 1):
                fill = fill_par if row_idx % 2 == 0 else fill_impar
                for cell in ws[row_idx]:
                    cell.fill = fill
                    cell.alignment = Alignment(horizontal="center")

        log.info(f"  ✓ historico_precios.xlsx guardado en {ARCHIVO.parent}")
    except ImportError:
        # openpyxl no disponible, guardar sin formato
        df.to_excel(ARCHIVO, sheet_name="Historico", index=False)
        log.info(f"  ✓ historico_precios.xlsx guardado (sin formato — instalar openpyxl)")
    except Exception as e:
        log.warning(f"  ⚠ Error guardando historico_precios.xlsx: {e}")


# ──────────────────────────────────────────────────────────────
# 7a. Hacienda — Mercado de Cañuelas via deCampoaCampo
# ──────────────────────────────────────────────────────────────
def scrape_canuelas():
    """Devuelve lista de {categoria, precio, variacion, unidad} o [].
    Usa la API JSON interna de deCampoaCampo:
      GET /gh_funciones.php?function=getListadoPreciosGordo
    Respuesta: {"hoy":"25/03/2026", "data":[
      {"categoria":"Novillitos hasta 390 Kg.",
       "precio_semana_1": 5204,
       "variacion_precio_semana_1": -113, ...}, ...]}
    """
    # ── 1) API JSON (fuente primaria) ────────────────────────────
    url_api = "https://www.decampoacampo.com/gh_funciones.php?function=getListadoPreciosGordo"
    text = _http_get(url_api)
    if text:
        try:
            data = json.loads(text)
            items = data.get("data", [])
            hacienda = []
            for item in items:
                cat = str(item.get("categoria", "")).strip()
                precio = item.get("precio_semana_1") or item.get("precio_semana_2") or 0
                var    = item.get("variacion_precio_semana_1") or 0
                if cat and precio and 500 < float(precio) < 30000:
                    hacienda.append({
                        "categoria": cat,
                        "precio":    round(float(precio), 2),
                        "variacion": round(float(var or 0), 2),
                        "unidad":    "$/kg + IVA"
                    })
            if hacienda:
                log.info(f"  ✓ Cañuelas API: {len(hacienda)} categorías — "
                         + " | ".join(f"{h['categoria']} ${h['precio']:,.0f}" for h in hacienda))
                return hacienda
        except (json.JSONDecodeError, TypeError, ValueError) as e:
            log.debug(f"    Cañuelas API JSON error: {e}")

    # ── 2) Fallback: HTML renderizado de la página outside ───────
    import re
    url_html = "https://www.decampoacampo.com/__dcac/outside/canuelas/precios"
    text = _http_get(url_html)
    if not text:
        log.info("  ℹ Cañuelas: sin respuesta de red")
        return []

    hacienda = []
    rows = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL | re.IGNORECASE)
    for row in rows:
        cat_m = re.search(
            r'class=["\']td_precios["\'][^>]*>.*?<h3[^>]*>(.*?)</h3>',
            row, re.DOTALL | re.IGNORECASE
        )
        if not cat_m:
            continue
        categoria = re.sub(r'<[^>]+>', '', cat_m.group(1)).strip()
        if not categoria:
            continue
        precio_m = re.search(
            r'<span[^>]*class=["\']h4["\'][^>]*>([\d.,]+)</span>',
            row, re.IGNORECASE
        )
        if not precio_m:
            continue
        precio = _parse_ar_num(precio_m.group(1))
        if not precio or not (500 < precio < 30000):
            continue
        var_m = re.search(
            r'<span[^>]*class=["\']h4["\'][^>]*>[\d.,]+</span>.*?\(([+-]?[\d.,]+)\)',
            row, re.DOTALL | re.IGNORECASE
        )
        variacion = _parse_ar_num(var_m.group(1)) if var_m else 0
        hacienda.append({
            "categoria": categoria,
            "precio":    round(precio, 2),
            "variacion": round(variacion or 0, 2),
            "unidad":    "$/kg + IVA"
        })

    if hacienda:
        log.info(f"  ✓ Cañuelas HTML: {len(hacienda)} categorías extraídas")
    else:
        log.info("  ℹ Cañuelas: sin precios en respuesta")
    return hacienda


# ──────────────────────────────────────────────────────────────
# 7b. Terneros / Terneras — Entre Surcos y Corrales
# ──────────────────────────────────────────────────────────────
def scrape_entresurcosycorrales():
    """Scrapea precios de terneros y terneras de entresurcosycorralesya.com.
    Fuente: ajax-modulo-ternero.php y ajax-modulo-ternera.php
    Columnas HTML: Categoría | Cantidad | Prom. Kilo | Kilo+ | Kilo- | Prom. Bulto | Bulto+ | Bulto-
    Devuelve lista de {categoria, tipo, precio, precio_max, precio_min, cantidad, unidad} o []
    """
    import re

    ENDPOINTS = [
        ("terneros",  "https://www.entresurcosycorralesya.com/ajax-modulo-ternero.php?desde=&hasta="),
        ("terneras",  "https://www.entresurcosycorralesya.com/ajax-modulo-ternera.php?desde=&hasta="),
    ]

    resultados = []

    for tipo, url in ENDPOINTS:
        text = _http_get(url)
        if not text:
            log.info(f"  ℹ EntreS&C {tipo}: sin respuesta de red")
            continue

        # Extraer filas <tr> con sus <td>
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL | re.IGNORECASE)
        encontradas = 0
        for row in rows:
            cells = re.findall(r'<td[^>]*>(.*?)</td>', row, re.DOTALL | re.IGNORECASE)
            if len(cells) < 5:
                continue
            # Limpiar HTML de cada celda
            clean = [re.sub(r'<[^>]+>', '', c).strip() for c in cells]
            cat      = clean[0]
            cantidad = _parse_ar_num(clean[1]) or 0
            prom_kg  = _parse_ar_num(clean[2])
            kg_max   = _parse_ar_num(clean[3])
            kg_min   = _parse_ar_num(clean[4])

            if not cat or not prom_kg or prom_kg < 100:
                continue

            resultados.append({
                "categoria":  cat,
                "tipo":       tipo,
                "precio":     round(prom_kg, 2),
                "precio_max": round(kg_max, 2) if kg_max else None,
                "precio_min": round(kg_min, 2) if kg_min else None,
                "cantidad":   int(cantidad),
                "unidad":     "$/kg vivo",
            })
            encontradas += 1

        log.info(f"  ✓ EntreS&C {tipo}: {encontradas} categorías")

    if resultados:
        # Log de las categorías clave de interés
        CLAVE = {"Terneros 130-160 Kg.", "Terneros 230-260 Kg.", "Novillitos 330-370 Kg.",
                 "Terneras 130-150 Kg.", "Terneras 150-170 Kg."}
        for r in resultados:
            if r["categoria"] in CLAVE:
                log.info(f"    ★ {r['categoria']}: ${r['precio']:,.2f}/kg ({r['cantidad']:,} cab)")

    return resultados


# ──────────────────────────────────────────────────────────────
# 7c. Granos — BCR Cámara Arbitral Precios de Pizarra
# ──────────────────────────────────────────────────────────────
def scrape_bcr_pizarra():
    """Devuelve {maiz, soja, trigo, sorgo} en $/tn o {} si falla.

    v14.2 (1/6/2026): el sitio de BCR cambió de estructura — ya no usa
    <table><tr><td>. Ahora cada grano vive en su propio bloque:

        <div class="board board-maiz ">
          <div class="board-wrapper">
            <h3>… Maíz …</h3>
            <div class="price"> $254.620,00 </div>
            …
          </div>
        </div>

    La versión vieja del scraper devolvía {} para maíz/trigo/sorgo y
    matcheaba mal soja (898987 en lugar del precio real 465000) porque
    el fallback de free-text se enganchaba con un número equivocado del
    HTML. La nueva regex apunta directamente a board-<grano> + price.

    Si NO se extrae maíz (señal canaria — el grano más estable), se
    guarda una copia del HTML en datos/debug_bcr_last.html para que el
    próximo cambio de estructura se diagnostique sin re-fetch.
    """
    import re
    url = "https://www.cac.bcr.com.ar/es/precios-de-pizarra"
    text = _http_get(url)
    if not text:
        return {}

    granos = {}

    # v14.2: regex específica al nuevo DOM. Captura el <div class="price"> que
    # sigue al header del board (sin saltar a otro board gracias al lookahead
    # negativo (?!<div class="board board-)).
    PATTERN_BOARD = re.compile(
        r'<div class="board board-(\w+)[^"]*">'
        r'(?:(?!<div class="board board-).)*?'
        r'<div class="price">\s*\$?\s*([0-9.,]+)',
        re.DOTALL
    )
    for grano_key, raw in PATTERN_BOARD.findall(text):
        p = _parse_ar_num(raw)
        if p and 10000 < p < 2000000:
            granos[grano_key] = round(p)

    # Fallback histórico (v <14.2): tablas HTML. Por si BCR vuelve a un layout
    # con <table>, sigue funcionando sin tener que tocar este código.
    if not granos:
        GRANOS_BUSCAR = {
            "maiz":  ["maíz", "maiz", "corn"],
            "soja":  ["soja", "soybean"],
            "trigo": ["trigo", "wheat"],
            "sorgo": ["sorgo", "sorghum"],
        }
        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', text, re.DOTALL | re.IGNORECASE)
        for row in rows:
            cells_raw = re.findall(r'<t[dh][^>]*>(.*?)</t[dh]>', row, re.DOTALL | re.IGNORECASE)
            cells = [re.sub(r'<[^>]+>', '', c).strip().lower() for c in cells_raw]
            if not cells:
                continue
            row_text = " ".join(cells)
            for grano_key, aliases in GRANOS_BUSCAR.items():
                if grano_key in granos:
                    continue
                if any(alias in row_text for alias in aliases):
                    for cell in cells[1:]:
                        p = _parse_ar_num(cell)
                        if p and 10000 < p < 2000000:
                            granos[grano_key] = round(p)
                            break

    # v14.2: dump del HTML cuando MAÍZ falla — es el canario más confiable
    # (siempre listado en BCR). Si maíz no aparece, lo más probable es que
    # BCR cambió el HTML de nuevo. El dump deja evidencia para el próximo fix
    # sin tener que re-curlear.
    if 'maiz' not in granos and text:
        try:
            debug_path = Path(__file__).parent / 'debug_bcr_last.html'
            debug_path.write_text(text[:80000], encoding='utf-8')
            log.warning(f"  ⚠ BCR: maíz NO extraído — HTML guardado en {debug_path.name} para inspección")
        except Exception as e:
            log.debug(f"    BCR debug dump falló: {e}")

    if granos:
        for g, p in granos.items():
            log.info(f"  ✓ BCR {g}: ${p:,}/tn")
    else:
        log.info("  ℹ BCR: sin precios extraídos")
    return granos


# ──────────────────────────────────────────────────────────────
# 7c. Negocios — Google Sheets con Service Account
# ──────────────────────────────────────────────────────────────

# Ruta al archivo de credenciales de la Service Account
# (relativa al script; también puede ser ruta absoluta)
GSHEET_CREDENTIALS_FILE = Path(__file__).parent / "lector-robot-credentials.json"


def _leer_hoja_api(sheet_id, nombre_hoja, creds_file):
    """Lee una hoja usando Google Sheets API con Service Account.
    Devuelve lista de dicts (una fila = un dict) o None si falla."""
    try:
        from google.oauth2.service_account import Credentials
        from googleapiclient.discovery import build

        scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
        creds  = Credentials.from_service_account_file(str(creds_file), scopes=scopes)
        service = build("sheets", "v4", credentials=creds, cache_discovery=False)
        result  = service.spreadsheets().values().get(
            spreadsheetId=sheet_id,
            range=nombre_hoja
        ).execute()
        values = result.get("values", [])
        if not values or len(values) < 2:
            return []
        headers = [str(h).strip().lower().replace(" ", "_") for h in values[0]]
        rows = []
        for row in values[1:]:
            # Rellenar celdas vacías al final de la fila
            row_padded = row + [""] * (len(headers) - len(row))
            reg = {headers[i]: str(row_padded[i]).strip() for i in range(len(headers))
                   if str(row_padded[i]).strip()}
            if reg:
                rows.append(reg)
        return rows
    except ImportError:
        return None  # Librería no instalada
    except Exception as e:
        log.debug(f"    Google API error ({nombre_hoja}): {e}")
        return None


def leer_negocios_gsheet():
    """Lee hojas CARGAS y COMPRAS usando Service Account.
    Devuelve dict con listas de registros."""
    import csv, io

    resultado = {"ventas": [], "compras": [], "error": None}

    # ── Intentar con Service Account (API) ─────────────────────
    if GSHEET_CREDENTIALS_FILE.exists():
        log.info(f"  → Usando credenciales: {GSHEET_CREDENTIALS_FILE.name}")
        for hoja, clave in [("CARGAS", "ventas"), ("COMPRAS", "compras")]:
            rows = _leer_hoja_api(GSHEET_ID, hoja, GSHEET_CREDENTIALS_FILE)
            if rows is None:
                # Librería no instalada
                resultado["error"] = (
                    "Instalar librerías: pip install google-auth google-api-python-client"
                )
                log.warning("  ⚠ Librerías de Google no instaladas. Ejecutar: "
                            "pip install google-auth google-api-python-client")
                break
            resultado[clave] = rows
            log.info(f"  ✓ Google Sheets API '{hoja}': {len(rows)} registros")
        return resultado

    # ── Fallback: CSV público (si la hoja fue publicada) ────────
    log.info("  ℹ Credenciales no encontradas, intentando CSV público...")
    for hoja, clave in [("CARGAS", "ventas"), ("COMPRAS", "compras")]:
        url = GSHEET_CSV_BASE + hoja
        text = _http_get(url, timeout=25)
        if not text:
            log.info(f"  ℹ Google Sheets hoja '{hoja}': no accesible")
            resultado["error"] = (
                "Colocar lector-robot-credentials.json en la carpeta de datos, "
                "o publicar la hoja en web como CSV"
            )
            continue

        try:
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
            if not rows:
                log.info(f"  ℹ Google Sheets hoja '{hoja}': vacía")
                continue
            registros = []
            for row in rows:
                reg = {k.strip().lower().replace(" ", "_"): v.strip() for k, v in row.items() if v.strip()}
                if reg:
                    registros.append(reg)
            resultado[clave] = registros
            log.info(f"  ✓ Google Sheets CSV '{hoja}': {len(registros)} registros")
        except Exception as e:
            log.warning(f"  ⚠ Error procesando '{hoja}': {e}")

    return resultado


def procesar_negocios(negocios_raw):
    """Procesa los registros crudos de CARGAS y COMPRAS.
    Busca columnas estándar: fecha, categoria, kg_cab, precio_kg,
    precio_carne, kg_total, frigorífico, etc.
    Devuelve resumen agrupado por categoría y frigorífico."""
    import re

    def buscar_col(row, *candidatos):
        """Busca la primera columna que coincida con algún candidato."""
        for cand in candidatos:
            for k, v in row.items():
                if cand in k.lower() and v:
                    return v
        return ""

    ventas_proc  = []
    compras_proc = []

    for r in negocios_raw.get("ventas", []):
        try:
            fecha    = buscar_col(r, "fecha")
            cat      = buscar_col(r, "categ", "categoria", "tipo")
            kg_cab   = _parse_ar_num(buscar_col(r, "kg_cab", "kg/cab", "peso", "kg_prom")) or 0
            # Fallback: columna con nombre exactamente "kg" (no capturada por buscar_col)
            if not kg_cab:
                for _k, _v in r.items():
                    if _k.strip().lower() == 'kg' and _v:
                        kg_cab = _parse_ar_num(_v) or 0
                        break
            precio   = _parse_ar_num(buscar_col(r, "precio_kg", "precio/kg", "precio_c",
                                                "precio_carne", "precio")) or 0
            precio_p = _parse_ar_num(buscar_col(r, "precio_pie", "precio_vivo", "$/pie")) or 0
            rinde    = _parse_ar_num(buscar_col(r, "rinde", "rendimiento", "rto")) or 0
            frigo    = buscar_col(r, "frigorifico", "frigorífico", "destino", "comprador")
            cabezas  = _parse_ar_num(buscar_col(r, "cabezas", "cantidad", "cab")) or 1

            if fecha or precio or kg_cab:
                ventas_proc.append({
                    "fecha":     fecha,
                    "categoria": cat,
                    "kg_cab":    round(kg_cab, 1) if kg_cab else 0,
                    "precio_carne": round(precio, 2) if precio else 0,
                    "precio_pie":   round(precio_p, 2) if precio_p else 0,
                    "rinde":     round(rinde, 3) if rinde else 0,
                    "frigorífico": frigo,
                    "cabezas":   int(cabezas),
                })
        except Exception:
            pass

    for r in negocios_raw.get("compras", []):
        try:
            fecha    = buscar_col(r, "fecha")
            cat      = buscar_col(r, "categ", "categoria", "tipo")
            kg_cab   = _parse_ar_num(buscar_col(r, "kg_cab", "kg/cab", "peso", "kg_prom")) or 0
            # Fallback: columna con nombre exactamente "kg" (no capturada por buscar_col)
            if not kg_cab:
                for _k, _v in r.items():
                    if _k.strip().lower() == 'kg' and _v:
                        kg_cab = _parse_ar_num(_v) or 0
                        break
            precio   = _parse_ar_num(buscar_col(r, "precio_kg", "precio/kg", "precio_c", "precio")) or 0
            cabezas  = _parse_ar_num(buscar_col(r, "cabezas", "cantidad", "cab")) or 1
            origen   = buscar_col(r, "origen", "vendedor", "proveedor", "campo")

            if fecha or precio or kg_cab:
                compras_proc.append({
                    "fecha":     fecha,
                    "categoria": cat,
                    "kg_cab":    round(kg_cab, 1) if kg_cab else 0,
                    "precio_kg": round(precio, 2) if precio else 0,
                    "cabezas":   int(cabezas),
                    "origen":    origen,
                })
        except Exception:
            pass

    # Resumen por categoría
    resumen_cat = {}
    for v in ventas_proc:
        cat = v["categoria"] or "Sin categoría"
        if cat not in resumen_cat:
            resumen_cat[cat] = {"ventas": 0, "cabezas": 0, "precio_prom": [], "rinde_prom": []}
        resumen_cat[cat]["ventas"]  += 1
        resumen_cat[cat]["cabezas"] += v["cabezas"]
        if v["precio_carne"]: resumen_cat[cat]["precio_prom"].append(v["precio_carne"])
        if v["rinde"]:        resumen_cat[cat]["rinde_prom"].append(v["rinde"])

    for k, v in resumen_cat.items():
        pp = v["precio_prom"]
        rp = v["rinde_prom"]
        v["precio_promedio"] = round(sum(pp)/len(pp), 2) if pp else 0
        v["rinde_promedio"]  = round(sum(rp)/len(rp), 4) if rp else 0
        del v["precio_prom"], v["rinde_prom"]

    # Resumen por frigorífico
    resumen_frigo = {}
    for v in ventas_proc:
        frig = v["frigorífico"] or "Desconocido"
        if frig not in resumen_frigo:
            resumen_frigo[frig] = {"ventas": 0, "cabezas": 0, "precio_prom": []}
        resumen_frigo[frig]["ventas"]  += 1
        resumen_frigo[frig]["cabezas"] += v["cabezas"]
        if v["precio_carne"]: resumen_frigo[frig]["precio_prom"].append(v["precio_carne"])

    for k, v in resumen_frigo.items():
        pp = v["precio_prom"]
        v["precio_promedio"] = round(sum(pp)/len(pp), 2) if pp else 0
        del v["precio_prom"]

    return {
        "ventas":        ventas_proc,
        "compras":       compras_proc,
        "resumen_cat":   resumen_cat,
        "resumen_frigo": resumen_frigo,
        "total_ventas":  len(ventas_proc),
        "total_compras": len(compras_proc),
    }


# ──────────────────────────────────────────────────────────────
# Función principal del módulo
# ──────────────────────────────────────────────────────────────
def actualizar_mercado_precios(carpeta, repo):
    """Actualiza mercado_precios.json y negocios_resumen.json."""
    from datetime import date
    today = date.today().isoformat()
    log.info(f"  Fecha: {today}")

    # ── 1. Cargar JSON existente ────────────────────────────────
    repo_json = Path(repo) / "mercado_precios.json" if repo else None
    existing = {}
    if repo_json and repo_json.exists():
        try:
            with open(repo_json, "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass
    elif (Path(carpeta) / "mercado_precios.json").exists():
        try:
            with open(Path(carpeta) / "mercado_precios.json", "r", encoding="utf-8") as f:
                existing = json.load(f)
        except Exception:
            pass

    historico      = existing.get("historico", [])
    insumos_ant    = existing.get("insumos", {})
    comt_ant       = existing.get("commodities", [])

    def _prev_com(nombre, default):
        for c in comt_ant:
            if c.get("nombre","").lower() == nombre.lower():
                return c.get("precio", default)
        return default

    # ── 2. Hacienda — Cañuelas ──────────────────────────────────
    log.info("  → Scraping Mercado de Cañuelas...")
    hacienda = scrape_canuelas()
    if not hacienda:
        hacienda = existing.get("hacienda", [
            {"categoria": "Novillo especial", "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
            {"categoria": "Novillo",          "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
            {"categoria": "Vaca",             "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
            {"categoria": "Vaquillona",       "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
            {"categoria": "Ternero",          "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
            {"categoria": "Ternera",          "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
            {"categoria": "Novillito",        "precio": 0, "variacion": 0, "unidad": "$/kg en pie"},
        ])
        log.info("  ℹ Usando precios anteriores de hacienda")

    # ── 2b. Terneros / Terneras — Entre Surcos y Corrales ───────
    log.info("  → Scraping Entre Surcos y Corrales (terneros/terneras)...")
    terneros_esyc = scrape_entresurcosycorrales()
    if not terneros_esyc:
        terneros_esyc = existing.get("terneros_esyc", [])
        if terneros_esyc:
            log.info("  ℹ Usando precios anteriores de terneros/terneras")

    # ── 3. Granos — BCR Pizarra ─────────────────────────────────
    log.info("  → Scraping BCR Precios de Pizarra...")
    granos = scrape_bcr_pizarra()
    precio_maiz  = granos.get("maiz",  insumos_ant.get("maiz",  243150))
    precio_soja  = granos.get("soja",  _prev_com("Soja",  390000))
    precio_trigo = granos.get("trigo", _prev_com("Trigo", 230000))
    precio_sorgo = granos.get("sorgo", _prev_com("Sorgo", 180000))

    if not granos:
        log.info("  ℹ Usando precios anteriores de granos")

    # ── 4. Insumos (maíz como ancla, resto relaciones del Excel) ─
    mep_hoy = None  # se asigna más adelante; inicializado aquí para evitar UnboundLocalError en Python 3.12+
    REL = {"gluten": 0.5385, "germen": 1.2227, "nucleo": 1.9342, "hominy": 0.8413}
    insumos = {
        "maiz":       precio_maiz,
        "gluten":     round(precio_maiz * REL["gluten"]),
        "nucleo":     round(precio_maiz * REL["nucleo"]),
        "germen":     round(precio_maiz * REL["germen"]),
        "hominy":     round(precio_maiz * REL["hominy"]),
        "silo":       insumos_ant.get("silo",       155482),
        "rollo":      insumos_ant.get("rollo",      25000),
        "hoteleria":  insumos_ant.get("hoteleria",  310),
        "sanidad":    insumos_ant.get("sanidad",    7500),
        "flete_12tn": insumos_ant.get("flete_12tn", 2750),
        "guias":      insumos_ant.get("guias",      1725),
        "dolar":      round(mep_hoy) if mep_hoy else insumos_ant.get("dolar", 1422),
    }

    commodities = [
        {"nombre": "Maíz",  "precio": precio_maiz,  "unidad": "$/tn", "fuente": "BCR Pizarra"},
        {"nombre": "Soja",  "precio": precio_soja,  "unidad": "$/tn", "fuente": "BCR Pizarra"},
        {"nombre": "Trigo", "precio": precio_trigo, "unidad": "$/tn", "fuente": "BCR Pizarra"},
        {"nombre": "Sorgo", "precio": precio_sorgo, "unidad": "$/tn", "fuente": "BCR Pizarra"},
    ]

    # ── 5. Negocios — Google Sheets ─────────────────────────────
    log.info("  → Leyendo Google Sheets (negocios)...")
    negocios_raw = leer_negocios_gsheet()
    negocios     = procesar_negocios(negocios_raw)
    log.info(f"  ✓ Negocios: {negocios['total_ventas']} ventas · {negocios['total_compras']} compras procesadas")

    # ── 6a. Dólar MEP del día (Ambito) ─────────────────────────
    log.info("  → Scraping Dólar MEP (Ambito)...")
    from datetime import date as _date
    _hoy_date = _date.today()
    _mes_actual = _hoy_date.strftime('%Y-%m')
    mep_hoy = _scrap_mep_tc_ambito(_mes_actual)   # promedio del mes en curso
    if mep_hoy is None:
        # Fallback: usar valor anterior o tabla aproximada
        _prev_mep = next((h.get('tc_mep') for h in reversed(historico) if h.get('tc_mep')), None)
        mep_hoy = _prev_mep or _TC_APROX_MEP_REF.get(_mes_actual) or insumos_ant.get('dolar', 1422)
        log.info(f"  ℹ MEP fallback: ${mep_hoy:,.0f}")
    else:
        log.info(f"  ✓ MEP hoy: ${mep_hoy:,.0f}/USD")

    # ── 6b. Histórico diario ────────────────────────────────────
    nov_precio = next((h["precio"] for h in hacienda
                       if "novillo" in h["categoria"].lower()
                       and "especial" not in h["categoria"].lower()), 0)
    ter_precio = next((h["precio"] for h in hacienda
                       if "ternero" in h["categoria"].lower()), 0)

    def _hprice(*substrings):
        """Primer precio de hacienda cuya categoría contiene todos los substrings (case-insensitive)."""
        for h in hacienda:
            cat = h.get("categoria", "").lower()
            if all(s in cat for s in substrings):
                return h.get("precio", 0) or 0
        return 0

    # Tabla MEP de referencia para retroalimentar historial
    _TC_APROX_MEP_REF = {
        '2024-12': 1111.0, '2025-01': 1165.0, '2025-02': 1198.0, '2025-03': 1262.0,
        '2025-04': 1245.0, '2025-05': 1160.0, '2025-06': 1189.0, '2025-07': 1276.0,
        '2025-08': 1269.0, '2025-09': 1435.0, '2025-10': 1498.0, '2025-11': 1466.0,
        '2025-12': 1483.0, '2026-01': 1478.0, '2026-02': 1393.0, '2026-03': 1422.0,
    }

    # Retroalimentar entradas históricas sin tc_mep
    for _h in historico:
        if _h.get('tc_mep') is None:
            _mes = _h.get('fecha', '')[:7]
            _h['tc_mep'] = _TC_APROX_MEP_REF.get(_mes)

    # Precios clave de Entre Surcos y Corrales para el histórico
    def _esyc(cat):
        for r in terneros_esyc:
            if r.get("categoria","").strip().lower() == cat.strip().lower():
                return r.get("precio") or 0
        return 0

    hoy = {
        "fecha":        today,
        "nov_390":      _hprice("novillito", "390"),
        "nov_430":      _hprice("novillito", "430"),
        "nov_460":      _hprice("460"),
        "nov_490":      _hprice("490"),
        "vaq_390":      _hprice("vaquillon"),
        "vac_buena":    _hprice("buena"),
        "vac_regular":  _hprice("regular"),
        "vac_conserva": _hprice("conserva"),
        "ternero":      ter_precio,
        "maiz":         precio_maiz,
        "soja":         precio_soja,
        "novillo":      nov_precio,
        "tc_mep":       round(mep_hoy) if mep_hoy else None,
        # Entre Surcos y Corrales — categorías de referencia para compra
        # Terneros
        "ter_130_160":  _esyc("Terneros 130-160 Kg."),
        "ter_230_260":  _esyc("Terneros 230-260 Kg."),
        "nov_330_370":  _esyc("Novillitos 330-370 Kg."),
        # Terneras (bandas equivalentes)
        "tera_130_150": _esyc("Terneras 130-150 Kg."),
        "tera_150_170": _esyc("Terneras 150-170 Kg."),
        "vaq_250_290":  _esyc("Vaquillonas 250-290 Kg."),
        "vaq_320_360":  _esyc("Vaquillonas 320-360 Kg."),
    }
    historico = [h for h in historico if h.get("fecha") != today]
    historico.append(hoy)
    historico = sorted(historico, key=lambda x: x.get("fecha", ""))[-365:]

    # ── 7. Histórico Excel en OneDrive ──────────────────────────
    log.info("  → Actualizando historico_precios.xlsx...")
    actualizar_historico_excel(hacienda, commodities, carpeta, today)

    # ── 8. Armar y guardar JSONs ────────────────────────────────
    mercado_json = {
        "fecha":        today,
        "fuente":       "Cañuelas · BCR Cámara Arbitral · Entre Surcos y Corrales",
        "hacienda":     hacienda,
        "terneros_esyc": terneros_esyc,
        "commodities":  commodities,
        "insumos":      insumos,
        "historico":    historico,
    }

    negocios_json = {
        "fecha":          today,
        "sheet_id":       GSHEET_ID,
        "total_ventas":   negocios["total_ventas"],
        "total_compras":  negocios["total_compras"],
        "resumen_cat":    negocios["resumen_cat"],
        "resumen_frigo":  negocios["resumen_frigo"],
        "ventas":         negocios["ventas"],
        "compras":        negocios["compras"],
        "error":          negocios_raw.get("error"),
    }

    # ── Generar snapshots históricos de compras (resultado simulado fijo) ──
    try:
        from simulador_negocios import actualizar_negocios_snapshots
        actualizar_negocios_snapshots(carpeta, log=log)
    except Exception as _e:
        log.warning(f"  ⚠ No se pudo generar negocios_snapshots: {_e}")

    # Cargar el snapshot recién generado para copiarlo al repo
    snap_data = None
    snap_path = Path(carpeta) / "negocios_snapshots.json"
    if snap_path.exists():
        try:
            with open(snap_path, encoding="utf-8") as _f:
                snap_data = json.load(_f)
        except Exception:
            pass

    # Guardar en repo GitHub Pages
    _files_to_save = [("mercado_precios.json", mercado_json),
                      ("negocios_resumen.json", negocios_json)]
    if snap_data:
        _files_to_save.append(("negocios_snapshots.json", snap_data))

    for fname, data in _files_to_save:
        if repo_json:
            dest = Path(repo) / fname
            try:
                with open(dest, "w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
                log.info(f"  ✓ {fname} → repo ({dest})")
            except Exception as e:
                log.warning(f"  ⚠ No se pudo guardar {fname} en repo: {e}")

        guardar(data, carpeta, fname)
        log.info(f"  ✓ {fname} → OneDrive")


# ══════════════════════════════════════════════════════════════
#  MÓDULO 9 — COMPORTAMIENTO HISTÓRICO MENSUAL
#  Combina: Masa de kg (Listado Caravanas XLS), Stock Insumos,
#           y Financiero mensual (formato viejo + nuevo)
# ══════════════════════════════════════════════════════════════

def actualizar_stock_insumos_excel(insumos_list, carpeta_stock_mensuales, today_str):
    """
    Añade columna de hoy al archivo STOCK DE INSUMOS.xlsx.
    Inserta la nueva columna en la posición 5 (después de 'Descripción Insumo').
    Si la columna de hoy ya existe, sobreescribe los valores.
    insumos_list: lista de {"nombre": str, "stock_kg": float}
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from pathlib import Path
        from datetime import datetime as _dt

        ruta = Path(carpeta_stock_mensuales) / "STOCK DE INSUMOS.xlsx"
        if not ruta.exists():
            log.warning(f"  ⚠ No existe {ruta}")
            return

        wb = openpyxl.load_workbook(str(ruta))
        ws = wb.active

        # Construir dict {nombre_lower → stock_kg} desde insumos_list
        stock_dict = {i['nombre'].lower().strip(): i['stock_kg'] for i in insumos_list}

        # Función de coincidencia: substring bidireccional
        def match_stock(desc_excel):
            desc_lower = str(desc_excel or '').lower().strip()
            for k, v in stock_dict.items():
                if k in desc_lower or desc_lower in k:
                    return v
            return None

        INSERT_COL = 5  # columna E (después de A=Depos, B=Rubro, C=Cod, D=Descripción)

        # Verificar si la columna de hoy ya existe
        today_col_idx = None
        for c in range(INSERT_COL, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h is None:
                continue
            if hasattr(h, 'strftime'):
                h_str = h.strftime('%Y-%m-%d')
            else:
                h_str = str(h)[:10]
            if h_str == today_str:
                today_col_idx = c
                break

        if today_col_idx is None:
            # Insertar nueva columna en posición 5
            ws.insert_cols(INSERT_COL)
            today_col_idx = INSERT_COL
            # Header: fecha como datetime para que Excel la reconozca
            try:
                hdr_date = _dt.strptime(today_str, '%Y-%m-%d')
            except Exception:
                hdr_date = today_str
            hdr_cell = ws.cell(1, today_col_idx)
            hdr_cell.value = hdr_date
            hdr_cell.number_format = 'DD/MM/YYYY'
            hdr_cell.fill = PatternFill("solid", fgColor="1F4E79")
            hdr_cell.font = Font(bold=True, color="FFFFFF", size=9)
            hdr_cell.alignment = Alignment(horizontal="center")
            log.info(f"  ✓ STOCK DE INSUMOS.xlsx — insertada columna {today_str} (col {today_col_idx})")
        else:
            log.info(f"  ℹ STOCK DE INSUMOS.xlsx — columna {today_str} ya existe, actualizando valores")

        # Rellenar valores para cada fila de insumo
        filled = 0
        for row in range(2, ws.max_row + 1):
            desc = ws.cell(row, 4).value  # columna D = Descripción Insumo
            if desc is None:
                continue
            valor = match_stock(str(desc))
            if valor is not None:
                cell = ws.cell(row, today_col_idx)
                cell.value = round(valor, 2)
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="center")
                filled += 1

        wb.save(str(ruta))
        log.info(f"  ✓ STOCK DE INSUMOS.xlsx — {filled} insumos actualizados para {today_str}")
    except ImportError:
        log.warning("  ⚠ openpyxl no disponible; se omite actualización de STOCK DE INSUMOS.xlsx")
    except Exception as e:
        log.warning(f"  ⚠ Error actualizando STOCK DE INSUMOS.xlsx: {e}")
        import traceback; log.warning(traceback.format_exc())


def _sanitize_xlsx_nan(ruta_src, ruta_dst):
    """
    v15.18: El sistema nuevo WinCampo Web escribe celdas numéricas con el
    string literal 'NaN'. Eso rompe openpyxl._cast_number (int('NaN') →
    ValueError) DURANTE pd.read_excel — el na_values de pandas no alcanza
    porque el crash ocurre dentro de openpyxl, antes de que pandas pueda
    filtrar. Reescribimos el xlsx (es un zip) quitando '<v>NaN</v>' de las
    hojas de cálculo: la celda queda sin valor → se lee como None/NaN, y el
    pd.to_numeric(errors='coerce') aguas abajo la deja en 0.
    Escribe el resultado saneado en ruta_dst. Devuelve nº de celdas saneadas.
    """
    import zipfile, re as _re2
    _pat = _re2.compile(r'<v>\s*[Nn][Aa][Nn]\s*</v>')
    n_fix = 0
    with zipfile.ZipFile(str(ruta_src), 'r') as zin, \
         zipfile.ZipFile(str(ruta_dst), 'w', zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith('xl/worksheets/') and item.filename.endswith('.xml'):
                txt = data.decode('utf-8', errors='replace')
                txt, n = _pat.subn('', txt)
                n_fix += n
                data = txt.encode('utf-8')
            zout.writestr(item, data)
    return n_fix


def _parse_listado_caravanas_html(ruta):
    """
    Parsea un Listado_Caravanas*.{XLS,xlsx} (Stock detallado de Caravanas).
    Soporta 3 formatos históricos:
      - HTML disfrazado de XLS (SQL viejo, hasta marzo 2026)
      - XLSX nativo con headers en R1 (transición, abril 2026)
      - XLSX nativo con metadata en R1 + headers en R2 (WinCampo Web, mayo 2026+)
    Detecta la fila de headers buscando 'Corral'; las columnas se localizan
    luego por substring (_find), tolerante a los renombres del sistema nuevo
    (Nº Tropa, Kilos Ingreso, etc.). Extrae fecha del nombre del archivo,
    mapea Corral → Campo.
    v15.18 Op 3: la masa de kg NO sale de 'Peso Proyectado' sino que se
    recalcula con la lógica del módulo Stock (calc_engorde: ADP+techo por
    categoría en El Haras, path recría en el resto), usando ADP_CAL_FALLBACK
    para meses cerrados y _ADP_CAL_RUNTIME para el mes corriente.
    Returns: dict {fecha, total_cabezas, total_kg, pegsa, por_hotelero}
    o None si falla.
    """
    import re as _re
    from pathlib import Path

    ruta = Path(ruta)
    nombre = ruta.name  # Listado_Caravanas28-02-2026.XLS

    # Extraer fecha del nombre (DD-MM-YYYY)
    m = _re.search(r'(\d{2})-(\d{2})-(\d{4})', nombre)
    if m:
        d, mo, y = m.groups()
        fecha_str = f"{y}-{mo}-{d}"
    else:
        log.warning(f"  ⚠ No se pudo extraer fecha de {nombre}")
        return None

    try:
        # v12.7: Detectar formato. WinCampo solía exportar HTML disfrazado
        # de .XLS; en mayo 2026 empezó a exportar XLSX real (los primeros 2
        # bytes son "PK" del zip). Soportamos ambos.
        raw_head = ruta.read_bytes()[:4]
        es_xlsx_real = raw_head[:2] == b'PK'

        if es_xlsx_real:
            # Branch XLSX nativo. WinCampo guarda a veces con extensión .XLS
            # aunque el archivo sea XLSX real (zip "PK..."); openpyxl lo
            # rechaza por la extensión, así que escribimos a un temp .xlsx.
            # v15.18: el sistema nuevo (WinCampo Web, mayo 2026+) además
            #   (a) mete metadata en la fila 1 y los headers reales en la 2, y
            #   (b) escribe celdas numéricas con el string literal 'NaN' que
            #       rompe openpyxl._cast_number durante el read.
            # _sanitize_xlsx_nan resuelve (b) al copiar al temp; la detección
            # de header debajo resuelve (a). El formato de transición (abril,
            # headers en R1) cae en _hidx=0 → idéntico al comportamiento previo.
            import tempfile
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as _tf:
                _tmp_xlsx = _tf.name
            try:
                _n_nan = _sanitize_xlsx_nan(ruta, _tmp_xlsx)
                df_raw = pd.read_excel(_tmp_xlsx, engine='openpyxl', sheet_name=0,
                                       dtype=object, header=None)
            finally:
                try: Path(_tmp_xlsx).unlink()
                except Exception: pass
            # Detectar la fila de headers buscando 'Corral' (R1 en transición,
            # R2 en WinCampo Web por la fila de metadata).
            _hidx = 0
            for _i in range(min(6, len(df_raw))):
                _vals = [str(v).strip() if v is not None else '' for v in df_raw.iloc[_i].tolist()]
                if any(c.lower() == 'corral' for c in _vals):
                    _hidx = _i
                    break
            df = df_raw.iloc[_hidx + 1:].copy()
            df.columns = [str(c).strip() for c in df_raw.iloc[_hidx].tolist()]
            df = df.reset_index(drop=True)
            log.info(f"    formato: XLSX nativo · headers R{_hidx+1} · "
                     f"{len(df)} filas · {_n_nan} celdas 'NaN' saneadas")
        else:
            # Branch HTML disfrazado (parser nativo, sin dependencias).
            from html.parser import HTMLParser as _HTMLParser

            class _TblParser(_HTMLParser):
                def __init__(self):
                    super().__init__()
                    self.rows, self._row, self._cell, self._in = [], [], [], False
                def handle_starttag(self, tag, attrs):
                    if tag in ('td','th'): self._in=True; self._cell=[]
                    elif tag=='tr': self._row=[]
                def handle_endtag(self, tag):
                    if tag in ('td','th'):
                        self._row.append(''.join(self._cell).strip()); self._in=False
                    elif tag=='tr':
                        if self._row: self.rows.append(self._row)
                def handle_data(self, data):
                    if self._in: self._cell.append(data)

            raw = ruta.read_bytes()
            for _enc in ('utf-8','latin-1','cp1252'):
                try: html_txt = raw.decode(_enc); break
                except: pass
            else: html_txt = raw.decode('utf-8', errors='replace')

            p = _TblParser(); p.feed(html_txt)
            if not p.rows or len(p.rows) < 2:
                log.warning(f"  ⚠ {nombre}: sin tablas HTML"); return None

            headers = p.rows[0]
            data_rows = p.rows[1:]
            ncols = len(headers)
            data_rows = [r + ['']*(ncols-len(r)) if len(r)<ncols else r[:ncols] for r in data_rows]
            df = pd.DataFrame(data_rows, columns=headers)

            # convertir números (formato argentino: punto=miles, coma=decimal)
            def _to_num(v):
                try:
                    v2 = str(v).replace('.','').replace(',','.')
                    return float(v2)
                except: return v
            for col in df.columns:
                df[col] = df[col].apply(lambda v: _to_num(v) if str(v).replace('.','').replace(',','').replace('-','').strip().isdigit() or (str(v).count(',')<=1 and str(v).replace('.','').replace(',','').replace('-','').strip().replace(' ','').isdigit()) else v)

    except Exception as e:
        log.warning(f"  ⚠ {nombre}: error leyendo archivo: {e}")
        return None

    # Normalizar columnas
    df.columns = [str(c).strip() for c in df.columns]

    # Buscar columnas relevantes (tolerante a variaciones)
    def _find(keywords):
        kw_lower = [k.lower() for k in keywords]
        for col in df.columns:
            cl = col.lower()
            if any(k in cl for k in kw_lower):
                return col
        return None

    col_corral   = _find(['corral'])
    col_hotelero = _find(['hotelero'])
    col_peso_p   = _find(['peso proyectado', 'proyectado'])
    col_categoria= _find(['categor'])
    col_kg_ing   = _find(['kilos ingreso', 'kg ingreso'])         # v15.18 Op 3
    col_fecha_in = _find(['fecha de ingreso', 'fecha ingreso'])   # v15.18 Op 3

    # v15.18 Op 3: ahora basta con Kg Ingreso para recalcular la masa; Peso
    # Proyectado queda como fallback si faltara Kg Ingreso/Fecha de Ingreso.
    if col_corral is None or col_hotelero is None or (col_peso_p is None and col_kg_ing is None):
        log.warning(f"  ⚠ {nombre}: columnas requeridas no encontradas. Cols: {list(df.columns)}")
        return None

    # Normalizar datos
    df['_corral_n'] = pd.to_numeric(df[col_corral], errors='coerce')
    df['_hotelero'] = df[col_hotelero].astype(str).str.strip().str.upper()

    # v15.54 (decisión usuario 2026-08-03): excluir el corral 10000 (virtual de
    # WinCampo, tropa PEG.DES.19/02/26) también del histórico mensual, con el
    # mismo criterio que v15.46 aplicó al stock diario. Sin esto los totales
    # mensuales incluían 179 cabezas que el diario no cuenta y las dos series
    # (PEGSA 7.853 diario vs 8.032 mensual) no eran comparables. Se filtra ANTES
    # de todas las agregaciones para que total/PEGSA/por_campo/por_hotelero
    # queden todos consistentes.
    _n_pre = len(df)
    df = df[df['_corral_n'] != 10000].copy()
    if len(df) < _n_pre:
        log.info(f"    Excluidas {_n_pre - len(df)} cabezas del corral 10000 (virtual)")

    # v15.18 Op 3 (decisión usuario 2026-06-11) · kg consistente con módulo Stock.
    # La masa de kg deja de salir de 'Peso Proyectado' (lo que calcula WinCampo) y
    # se recalcula con la MISMA lógica del Stock vivo (v15.13):
    #   - El Haras (corral 1-199): min(kg_ingreso + dias·ADP, TECHO_KG_POR_CAT[cat])
    #   - Resto de campos: path recría → min(kg_ingreso + dias·ENGORDE_RECRIA, TECHO_KG_RECRIA)
    # dias = fecha_corte_del_mes − Fecha de Ingreso, clampeado a [0, TECHO_DIAS].
    # ADP: mes corriente → _ADP_CAL_RUNTIME (calibrado dinámico, corre antes del
    # módulo 9); meses cerrados → ADP_CAL_FALLBACK (el runtime refleja la
    # productividad de hoy, no la del mes histórico).
    # NOTA: la fórmula del PROMPT original usaba kg crudo para no-Haras; se corrigió
    # al path recría real del Stock (0,5 kg/día, techo 380).
    import numpy as _np
    _peso_proy = (pd.to_numeric(df[col_peso_p], errors='coerce').fillna(0)
                  if col_peso_p else pd.Series(0.0, index=df.index))
    if col_kg_ing is None or col_fecha_in is None:
        log.warning(f"    ⚠ {nombre}: faltan Kg Ingreso/Fecha de Ingreso "
                    f"(kg={col_kg_ing}, fi={col_fecha_in}) — uso Peso Proyectado")
        df['_peso'] = _peso_proy
    else:
        _fecha_corte  = datetime.strptime(fecha_str, '%Y-%m-%d').date()
        _es_corriente = _fecha_corte >= datetime.now().date().replace(day=1)
        _adp_src      = _ADP_CAL_RUNTIME if _es_corriente else ADP_CAL_FALLBACK
        _kg_ing = pd.to_numeric(df[col_kg_ing], errors='coerce').fillna(0.0)
        _fi     = pd.to_datetime(df[col_fecha_in], errors='coerce')
        _dias   = (((pd.Timestamp(_fecha_corte) - _fi).dt.days)
                   .clip(lower=0, upper=TECHO_DIAS).fillna(0).astype(int))
        _cat    = (df[col_categoria].astype(str).str.strip().str.upper()
                   if col_categoria else pd.Series('', index=df.index))
        _es_haras = df['_corral_n'].between(1, 199)
        _adp   = _cat.map(lambda c: float(_adp_src.get(c, ADP_CAL_FALLBACK.get(c, 1.0))))
        _techo = _cat.map(lambda c: float(TECHO_KG_POR_CAT.get(c, 650)))
        _kg_haras  = _np.minimum(_kg_ing + _dias * _adp, _techo)
        _kg_recria = _np.minimum(_kg_ing + _dias * ENGORDE_RECRIA, float(TECHO_KG_RECRIA))
        df['_peso'] = pd.Series(_np.where(_es_haras, _kg_haras, _kg_recria),
                                index=df.index).round(1)
        log.info(f"    kg recalc Op 3 (ADP {'runtime' if _es_corriente else 'fallback'}): "
                 f"Peso Proy {float(_peso_proy.sum()):,.0f} → "
                 f"calc_engorde {float(df['_peso'].sum()):,.0f}")

    # v15.16 · Consolidar BULLTRADE SRL → PEGSA también en el histórico mensual.
    # El Listado_Caravanas viene del SQL viejo / export manual y trae los
    # hoteleros sin consolidar. Aplicamos el mismo mapeo que el adapter
    # WinCampo Web para que comportamiento_historico, valuacion_historica
    # y módulo Histórico muestren un único PEGSA en todos los meses.
    try:
        from wincampo_source import HOTELEROS_CONSOLIDADOS_A_PEGSA
        _h_set = HOTELEROS_CONSOLIDADOS_A_PEGSA
    except Exception:
        _h_set = {"BULLTRADE SRL", "BULLTRADE", "BULL TRADE SRL"}
    df['_hotelero'] = df['_hotelero'].where(~df['_hotelero'].isin(_h_set), 'PEGSA')

    # Mapear corral → campo usando tabla CORRALES global
    def _get_campo(nro):
        try:
            n = int(nro)
        except (TypeError, ValueError):
            return "Desconocido"
        for lo, hi, nom in CORRALES:
            if lo <= n <= hi:
                return nom
        return "Otro"

    df['_campo'] = df['_corral_n'].apply(_get_campo)

    # Total general
    total_cab = len(df)
    total_kg  = round(df['_peso'].sum(), 0)

    # Por hotelero
    por_hotelero = {}
    for hot, grp in df.groupby('_hotelero'):
        if not hot or hot in ('NAN', 'NONE', ''):
            continue
        por_hotelero[hot] = {
            'cabezas':       int(len(grp)),
            'kg_proyectado': round(float(grp['_peso'].sum()), 0),
        }

    # PEGSA solamente
    df_peg = df[df['_hotelero'] == 'PEGSA']
    peg_cab = int(len(df_peg))
    peg_kg  = round(float(df_peg['_peso'].sum()), 0)

    por_campo_pegsa = {}
    for campo, grp in df_peg.groupby('_campo'):
        por_campo_pegsa[campo] = {
            'cabezas':       int(len(grp)),
            'kg_proyectado': round(float(grp['_peso'].sum()), 0),
        }

    # v15.52: desglose por campo del GRUPO COMPLETO + cruce campo × hotelero.
    # El feedlot hotelea a terceros: "cuántas cabezas hay en El Haras" no es lo
    # mismo que "cuántas cabezas de PEGSA hay en El Haras". df['_campo'] ya
    # estaba calculado para todas las filas — antes solo se usaba el subset de
    # PEGSA y el resto se descartaba.
    por_campo_grupo = {}
    for campo, grp in df.groupby('_campo'):
        detalle = {}
        for hot, g2 in grp.groupby('_hotelero'):
            if not hot or hot in ('NAN', 'NONE', ''):
                continue
            detalle[hot] = {
                'cabezas':       int(len(g2)),
                'kg_proyectado': round(float(g2['_peso'].sum()), 0),
            }
        por_campo_grupo[campo] = {
            'cabezas':       int(len(grp)),
            'kg_proyectado': round(float(grp['_peso'].sum()), 0),
            'por_hotelero':  detalle,
        }

    log.info(f"  ✓ {nombre} — total {total_cab:,} cab / {total_kg:,.0f} kg | "
             f"PEGSA {peg_cab:,} cab / {peg_kg:,.0f} kg | "
             f"campos grupo: {len(por_campo_grupo)}")

    return {
        'fecha':          fecha_str,
        'archivo':        nombre,
        'total_cabezas':  total_cab,
        'total_kg':       float(total_kg),
        'por_hotelero':   por_hotelero,
        'por_campo':      por_campo_grupo,   # v15.52 (nivel superior: grupo por campo)
        'pegsa': {
            'cabezas':       peg_cab,
            'kg_proyectado': float(peg_kg),
            'por_campo':     por_campo_pegsa,
        },
    }


def _parse_financiero_viejo(df, fecha_str):
    """
    Parsea financiero en FORMATO VIEJO (hoja única, layout semanal).
    Columnas: col0=Label, col1=Referencia/Monto, col2=Sub-label, col3=Semana0, col4=Semana1...
    Returns dict estandarizado con los campos financieros clave.
    """
    def _sf(v):
        try:
            f = float(v)
            import math
            return f if not math.isnan(f) else None
        except Exception:
            return None

    # ── disponible: "saldo disponibilidades" → col3 ──
    disponible = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'saldo disponibilidades' in label:
            disponible = _sf(row.iloc[3]) or 0.0
            break

    # ── cheques en cartera corrientes: rows entre "cheq cartera ctes" y "compra dolares" → col1 ──
    cheq_ctes = 0.0
    in_ctes = False
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'cheq cartera ctes' in label:
            in_ctes = True
            continue
        if in_ctes:
            if 'compra dolares' in label or 'saldo disponibilidades' in label:
                break
            v = _sf(row.iloc[1]) if len(row) > 1 else None
            if v and v > 0:
                cheq_ctes += v

    # ── cheques en cartera diferidos: "total disponib+chdif" → sum cols4+ positivos ──
    cheq_dif = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'total disponib' in label and 'chdif' in label:
            for c in range(4, len(row)):
                v = _sf(row.iloc[c])
                if v and v > 0:
                    cheq_dif += v
            break

    cheques_cartera = cheq_ctes + cheq_dif

    # ── cheques diferidos emitidos: "total cheques emitidos" → abs(sum cols4+) ──
    cheques_emitidos = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'total cheques emitidos' in label:
            for c in range(3, len(row)):
                v = _sf(row.iloc[c])
                if v and v < 0:
                    cheques_emitidos += abs(v)
            break

    # ── cobrar hacienda: "total vtos x ventas" o "vencimientos a cobrar" → sum cols3+ positivos ──
    cobrar_hacienda = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'total vtos x ventas' in label or 'vencimientos a cobrar' in label:
            for c in range(3, len(row)):
                v = _sf(row.iloc[c])
                if v and v > 0:
                    cobrar_hacienda += v
            break

    # ── pagar hacienda (v15.24): sumar el DETALLE de proveedores ──
    # La fila 'Total Vtos x Compras Hacienda' / 'Vencimientos a Pagar hacienda'
    # mezcla las columnas semanales de proveedores CON las filas 'Darwash - *'
    # (que viven DENTRO del mismo bloque de compras) → sumar abs de sus columnas
    # infla el valor (al feb-26: $2.783M vs $492M real). Solución: sumar las
    # filas de detalle entre el header 'VENCIMIENTOS COMPRAS HACIENDA A PAGAR' y
    # la primera fila 'Darwash -*' / 'Total Vtos' / próxima sección. Cada fila es
    # un proveedor con su importe en la columna-semana de vencimiento.
    # Validado feb-26 = $492.170.660 (matchea con el usuario).
    pagar_hacienda = 0.0
    _ph_start = None
    for i, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'vencimientos' in label and 'compras' in label and 'hacienda' in label and 'pagar' in label:
            _ph_start = i + 1
            break
    if _ph_start is not None:
        for i in range(_ph_start, min(_ph_start + 60, len(df))):
            row   = df.iloc[i]
            label = str(row.iloc[0] if row.iloc[0] is not None else '').strip().lower()
            # Cortar al llegar al total o a la próxima sección.
            if ('total vtos' in label or 'vencimientos a cobrar' in label
                    or 'vta hacienda' in label or 'ventas de hacienda' in label):
                break
            # Saltar las filas 'Darwash -*' (se cuentan aparte en darwash_pos),
            # sin cortar — por si aparecen intercaladas antes de proveedores.
            if label.startswith('darwash'):
                continue
            for c in range(2, len(row)):
                v = _sf(row.iloc[c])
                if v and v > 0:
                    pagar_hacienda += v

    # ── dólares: primera "compra dolares" en sección disponibilidades → col1=qty, col3=ARS ──
    usd_cant = 0.0; usd_ars = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if 'compra dolares' in label:
            usd_cant = _sf(row.iloc[1]) or 0.0
            # ARS puede estar en col3 o ser 0 en archivos viejos
            usd_ars = _sf(row.iloc[3]) or 0.0
            break

    # ── LCG: fila con "lcg" en label → col2 (valor de referencia/activo) ──
    lcg = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if label.startswith('lcg') or ' lcg' in label:
            # El valor está en col2 (col0=label, col1=vacío, col2=monto)
            v = _sf(row.iloc[2]) if len(row) > 2 else None
            if v is None:
                v = _sf(row.iloc[1]) if len(row) > 1 else None
            lcg = abs(v or 0.0)
            if lcg > 0:
                break

    # ── Tercio Bravo: "terciobravo", "tercio bravo", "aporte tercio" → col2 ──
    tercio_bravo = 0.0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').lower()
        if any(k in label for k in ['terciobravo', 'tercio bravo', 'aporte tercio', 'terciob']):
            v = _sf(row.iloc[2]) if len(row) > 2 else None
            if v is None:
                v = _sf(row.iloc[1]) if len(row) > 1 else None
            tercio_bravo = abs(v or 0.0)
            if tercio_bravo > 0:
                break

    # ── v15.24: Cuenta Corriente Darwash en formato viejo ──
    # Filas con label 'Darwash - *' (viven dentro del bloque compras a pagar).
    # Suma CON SIGNO de sus columnas semanales = posición neta. Positiva = pasivo
    # PEGSA → Darwash → se resta en módulo 10 (igual criterio que formato nuevo).
    # Antes (v15.23) estaba hardcodeado a 0. Validado feb-26: 6 filas → +$573.622.408.
    darwash_pos    = 0.0
    darwash_origen = None
    _dw_n = 0
    for _, row in df.iterrows():
        label = str(row.iloc[0] if row.iloc[0] is not None else '').strip().lower()
        if label.startswith('darwash -') or label.startswith('darwash-'):
            _dw_n += 1
            for c in range(1, len(row)):
                v = _sf(row.iloc[c])
                if v:
                    darwash_pos += v
    if _dw_n:
        darwash_origen = f"formato viejo · {_dw_n} filas 'Darwash -*'"

    return {
        'fecha':            fecha_str,
        'formato':          'viejo',
        'disponible':       round(disponible, 2),
        'cheques_cartera':  round(cheques_cartera, 2),
        'cheques_emitidos': round(cheques_emitidos, 2),
        'cobrar_hacienda':  round(cobrar_hacienda, 2),
        'pagar_hacienda':   round(pagar_hacienda, 2),
        'usd_cant':         round(usd_cant, 0),
        'usd_ars':          round(usd_ars, 2),
        'lcg':              round(lcg, 2),
        'tercio_bravo':     round(tercio_bravo, 2),
        'darwash_pos':      round(darwash_pos, 2),   # v15.24 (era 0.0 en v15.23)
        'darwash_origen':   darwash_origen,          # v15.24
    }


def _parse_financiero_nuevo(sheets, fecha_str):
    """
    Parsea financiero en FORMATO NUEVO (multi-hoja: resumen, posicion hoy, etc.).
    Returns dict estandarizado con los campos financieros clave.
    """
    def _sf(v):
        try:
            f = float(v)
            import math
            return f if not math.isnan(f) else None
        except Exception:
            return None

    # ── posicion hoy ──
    ph = sheets.get('posicion hoy', pd.DataFrame())

    def gph(r, c):
        try:
            return _sf(ph.iloc[r, c])
        except Exception:
            return None

    # Row 22 = "saldo Disponibilidades" → col4 = SALDO FINAL
    saldo_disp = gph(22, 4) or 0.0
    # Row 25 = "COMPRA DOLARES" → col1=qty, col3=ARS
    usd_cant = gph(25, 1) or 0.0
    usd_ars  = gph(25, 3) or 0.0

    # ── resumen (LCG y Tercio Bravo) ──
    res = sheets.get('resumen', pd.DataFrame())

    def gres(r, c):
        try:
            return _sf(res.iloc[r, c])
        except Exception:
            return None

    # Row 1 = "LCG - aportes..." → col1; Row 2 = "aporte terciobravo" → col1
    lcg          = abs(gres(1, 1) or 0.0)
    tercio_bravo = abs(gres(2, 1) or 0.0)

    # ── v15.22 · cheques pendiente: distinción EMITIDOS vs CARTERA ──
    # La hoja 'cheques pendiente' lista cheques EMITIDOS (a pagar), NO cheques al
    # cobro. Razonamiento contable (confirmado por usuario 2026-06-18):
    #   · venc <  fecha del archivo: ya vencidos pero aún no presentados al banco
    #     → YA están descontados dentro de 'disponible'. Sumarlos como emitidos
    #     sería doble-conteo (se restarían dos veces).
    #   · venc >= fecha del archivo: diferidos; no afectan el saldo bancario
    #     actual → pasivo cierto futuro = cheques_emitidos.
    # Cartera (cheques al cobro) = 0 en este formato: los marcados 'AL COBRO=SI'
    # son presentaciones del día (no pendientes de cobro). Se loguea su suma por
    # visibilidad, NO se computa como cartera.
    # (Antes de v15.22: se sumaba TODA la columna F como cartera (+$1.954M
    #  ficticios) y los emitidos salían de la row 30 del resumen — ambos mal.)
    cheq_raw         = sheets.get('cheques pendiente', pd.DataFrame())
    total_cartera    = 0.0
    cheques_emitidos = 0.0
    _al_cobro_dia    = 0.0
    if len(cheq_raw) > 4:
        from datetime import datetime as _dtv
        _sub   = cheq_raw.iloc[4:]
        _venc  = pd.to_datetime(_sub.iloc[:, 1], errors='coerce')           # col B
        _imp   = pd.to_numeric(_sub.iloc[:, 5], errors='coerce').fillna(0)  # col F
        _cobro = _sub.iloc[:, 6].astype(str).str.upper().str.strip() if _sub.shape[1] > 6 \
                 else pd.Series(['']*len(_sub), index=_sub.index)           # col G 'AL COBRO'
        try:
            _fecha_archivo = _dtv.strptime(fecha_str, '%Y-%m-%d')
        except Exception:
            _fecha_archivo = None
        if _fecha_archivo is not None:
            cheques_emitidos = float(_imp[(_venc >= _fecha_archivo) & (_imp > 0)].sum())
        else:
            cheques_emitidos = float(_imp[_imp > 0].sum())
        _al_cobro_dia = float(_imp[_cobro.isin(['SI', 'SÍ', 'YES', 'TRUE']) & (_imp > 0)].sum())
    log.info(f"    cheques pendiente → emitidos diferidos (venc>={fecha_str})="
             f"${cheques_emitidos:,.0f} | cartera=$0 | al-cobro-del-día=${_al_cobro_dia:,.0f}")

    # ── vencimientos de hacienda ──
    hac = sheets.get('vencimientos de hacienda', pd.DataFrame())
    cobrar_hacienda = 0.0
    pagar_hacienda  = 0.0
    if len(hac) > 22:
        # Compras hacienda (pagar): filas 2-18
        for i in range(2, min(19, len(hac))):
            r = hac.iloc[i]
            f = pd.to_datetime(r.iloc[0], errors='coerce')
            if pd.isna(f):
                continue
            for c in range(1, hac.shape[1]):
                v = _sf(r.iloc[c])
                if v and v > 0:
                    pagar_hacienda += v
        # Ventas hacienda (cobrar): filas 22-38
        for i in range(22, min(39, len(hac))):
            r = hac.iloc[i]
            f = pd.to_datetime(r.iloc[0], errors='coerce')
            if pd.isna(f):
                continue
            for c in range(1, hac.shape[1]):
                v = _sf(r.iloc[c])
                if v and v > 0:
                    cobrar_hacienda += v

    # ── v15.23 · Posición Cuenta Corriente Darwash (pasivo PEGSA → Darwash) ──
    # FUENTE A (preferida): suma de la col B de la hoja dedicada
    # 'cuenta  corriente con darwash' (¡OJO: DOS espacios entre 'cuenta' y
    # 'corriente', peculiaridad del template!). Cada fila es un movimiento
    # (col A fecha, col B importe ya con signo: egreso +, ingreso −, col C tipo).
    # Robusta: no depende de un rango fijo de filas.
    # FUENTE B (fallback): la fila del resumen rotulada 'Cuenta Corriente
    # Darwash' (~R98), sumando sus columnas — da el mismo total que Fuente A.
    # (Se descartó sumar R100-R104 como proponía el PROMPT: los labels de esas
    #  filas cambian entre archivos y sumar sus columnas duplica el desglose
    #  semanal → da un total erróneo. Verificado empíricamente 2026-06-18.)
    # Validado al 27-may: $2.430.035.525 (matchea exacto con el usuario).
    darwash_pos    = 0.0
    darwash_origen = None
    _dw_name = next((n for n in sheets.keys()
                     if 'darwash' in str(n).lower() and 'corriente' in str(n).lower()), None)
    if _dw_name is not None:
        _dw = sheets[_dw_name]
        if _dw.shape[1] > 1:
            _colb = pd.to_numeric(_dw.iloc[:, 1], errors='coerce').fillna(0)
            darwash_pos    = float(_colb.sum())
            darwash_origen = f"hoja {_dw_name!r} col B"
    if darwash_origen is None or darwash_pos == 0:
        # Fallback: fila 'Cuenta Corriente Darwash' del resumen (por label)
        for _r in range(len(res)):
            _lbl = gres(_r, 0)
            if isinstance(_lbl, str) and 'cuenta corriente darwash' in _lbl.lower():
                _s = 0.0
                for _c in range(1, res.shape[1]):
                    _v = gres(_r, _c)
                    if isinstance(_v, (int, float)):
                        _s += float(_v)
                if _s:
                    darwash_pos    = _s
                    darwash_origen = f"resumen fila {_r} 'Cuenta Corriente Darwash'"
                break
    log.info(f"    Darwash pos: ${darwash_pos:,.0f} (origen: {darwash_origen or 'no encontrado'})")

    return {
        'fecha':            fecha_str,
        'formato':          'nuevo',
        'disponible':       round(saldo_disp, 2),
        'cheques_cartera':  round(total_cartera, 2),
        'cheques_emitidos': round(cheques_emitidos, 2),
        'cobrar_hacienda':  round(cobrar_hacienda, 2),
        'pagar_hacienda':   round(pagar_hacienda, 2),
        'usd_cant':         round(usd_cant, 0),
        'usd_ars':          round(usd_ars, 2),
        'lcg':              round(lcg, 2),
        'tercio_bravo':     round(tercio_bravo, 2),
        'darwash_pos':      round(darwash_pos, 2),   # v15.23
        'darwash_origen':   darwash_origen,          # v15.23 (auditoría)
    }


def parse_financiero_historico(ruta):
    """
    Detecta el formato (viejo=hoja única / nuevo=multi-hoja) y parsea el
    archivo YYYY-MM-DD_financiero.xlsx, retornando un dict estandarizado.
    """
    import os as _os
    nombre    = _os.path.basename(ruta)
    fecha_str = nombre[:10]

    try:
        sheets = pd.read_excel(ruta, sheet_name=None, header=None, engine='openpyxl')
    except Exception as e:
        log.warning(f"  ⚠ {nombre}: error abriendo: {e}")
        return None

    if 'resumen' in sheets:
        # FORMATO NUEVO (2026-03-20 en adelante)
        return _parse_financiero_nuevo(sheets, fecha_str)
    else:
        # FORMATO VIEJO (hasta 2026-02-28)
        # Buscar hoja principal (Hoja1 o la primera disponible)
        hoja = sheets.get('Hoja1')
        if hoja is None:
            hoja = sheets.get('Sheet1')
        if hoja is None:
            hoja = list(sheets.values())[0] if sheets else None
        if hoja is None:
            log.warning(f"  ⚠ {nombre}: no se encontró hoja de datos")
            return None
        return _parse_financiero_viejo(hoja, fecha_str)


def _normalizar_por_campo(por_campo, total_kg_sellado):
    """
    v15.52: escala los kg del desglose para que sumen EXACTO al total sellado.

    Las CABEZAS no se tocan — son un conteo puro del Excel, siempre exacto y
    reproducible. Los kg sí, porque el re-parseo usa los parámetros de ADP
    actuales mientras que el mes sellado se calculó con los de su momento. La
    diferencia es marginal, pero romper la identidad
    suma(desglose) == total_sellado sería peor que un redondeo.
    """
    if not por_campo:
        return por_campo
    suma = sum((v.get('kg_proyectado') or 0) for v in por_campo.values())
    if not suma or not total_kg_sellado:
        return por_campo
    f = float(total_kg_sellado) / float(suma)
    if abs(f - 1.0) < 1e-9:
        return por_campo
    if abs(f - 1.0) > 0.0001:
        log.warning(f"    ⚠ _normalizar_por_campo: factor {f:.5f} (>0.01%) — "
                    f"los parámetros de ADP cambiaron bastante; conviene mirarlo")
    for v in por_campo.values():
        v['kg_proyectado'] = round((v.get('kg_proyectado') or 0) * f, 0)
        for h in (v.get('por_hotelero') or {}).values():
            h['kg_proyectado'] = round((h.get('kg_proyectado') or 0) * f, 0)
    return por_campo


def actualizar_comportamiento_historico(carpeta, carpeta_stock_mensuales):
    """
    MÓDULO 9: construye/actualiza comportamiento_historico.json.

    Para cada mes con Listado_Caravanas disponible:
      - hacienda_masa: del XLS del mes
      - insumos:       columna de esa fecha en STOCK DE INSUMOS.xlsx
      - financiero:    del financiero más próximo a esa fecha

    Salida: comportamiento_historico.json en carpeta datos.
    """
    import glob as _glob
    import os as _os
    from pathlib import Path

    log.info("  Escaneando Listado_Caravanas...")
    # 1. Listar archivos Listado_Caravanas
    # v15.18: matchea .XLS (legacy), .xls, .xlsx, .XLSX (sistema nuevo WinCampo
    # Web exporta .xlsx). Si para un mismo mes coexisten varios archivos (p.ej.
    # el .xlsx ORIGINAL de WinCampo Web + un .XLS pre-procesado a mano), se
    # deduplica por mes PREFIRIENDO el .xlsx nativo. Bajo Op 3 el kg se recalcula
    # desde Kg Ingreso + Fecha de Ingreso, y el pre-procesado puede traer la
    # fecha como serial Excel ('45875') que rompe el cálculo de días → el .xlsx
    # original es la fuente canónica correcta.
    patron_xls = _os.path.join(carpeta_stock_mensuales, "Listado_Caravanas*.[Xx][Ll][Ss]*")
    _todos = sorted(set(_glob.glob(patron_xls)))
    def _pref_ext(_f):
        return 0 if _os.path.basename(_f).rsplit('.', 1)[-1].lower() == 'xlsx' else 1
    _por_mes = {}
    for _f in _todos:
        _key = _os.path.basename(_f).rsplit('.', 1)[0].lower()
        if _key not in _por_mes or _pref_ext(_f) < _pref_ext(_por_mes[_key]):
            _por_mes[_key] = _f
    archivos_cara = sorted(_por_mes.values())
    log.info(f"  Archivos Listado_Caravanas: {len(archivos_cara)}")

    # 2. Listar archivos financieros y parsearlos todos
    log.info("  Escaneando archivos financieros...")
    patron_fin = _os.path.join(carpeta, "financiero",
                               "[0-9][0-9][0-9][0-9]-[0-9][0-9]-[0-9][0-9]_financiero.xlsx")
    archivos_fin = sorted(_glob.glob(patron_fin))
    log.info(f"  Archivos financieros: {len(archivos_fin)}")

    financieros = {}  # fecha_str → dict
    for ruta_f in archivos_fin:
        nombre_f = _os.path.basename(ruta_f)
        fecha_f  = nombre_f[:10]
        # Validar fecha antes de procesar
        try:
            datetime.strptime(fecha_f, '%Y-%m-%d')
        except ValueError:
            log.warning(f"  ⚠ Fecha inválida en nombre '{nombre_f}' — se omite")
            continue
        res = parse_financiero_historico(ruta_f)
        if res:
            financieros[fecha_f] = res
            log.info(f"    ✓ {nombre_f} — disp: ${res['disponible']:,.0f} "
                     f"| cartera: ${res['cheques_cartera']:,.0f} "
                     f"| lcg: ${res['lcg']:,.0f}")
    log.info(f"  Financieros parseados: {len(financieros)}")

    # 3. Leer STOCK DE INSUMOS.xlsx (mapa fecha → col de insumos)
    log.info("  Leyendo STOCK DE INSUMOS.xlsx...")
    insumos_por_fecha = {}  # 'YYYY-MM-DD' → {nombre: kg}
    ruta_ins = Path(carpeta_stock_mensuales) / "STOCK DE INSUMOS.xlsx"
    if ruta_ins.exists():
        try:
            import openpyxl as _oxl
            wb_ins = _oxl.load_workbook(str(ruta_ins), read_only=True, data_only=True)
            ws_ins = wb_ins.active

            # Leer encabezados (fila 1): cols 5+ son fechas
            headers = {}
            for c in range(5, ws_ins.max_column + 1):
                h = ws_ins.cell(1, c).value
                if h is None:
                    continue
                if hasattr(h, 'strftime'):
                    h_str = h.strftime('%Y-%m-%d')
                else:
                    try:
                        h_str = str(h)[:10]
                    except Exception:
                        continue
                if len(h_str) == 10 and h_str[4] == '-':
                    headers[c] = h_str

            # Leer nombres de insumos (columna D = col4)
            nombres_insumos = {}
            for r in range(2, ws_ins.max_row + 1):
                desc = ws_ins.cell(r, 4).value
                if desc:
                    nombres_insumos[r] = str(desc).strip()

            # Para cada columna-fecha, leer los valores
            for col_idx, fecha_col in headers.items():
                items = {}
                total = 0.0
                for r, nombre_ins in nombres_insumos.items():
                    v = ws_ins.cell(r, col_idx).value
                    try:
                        kg = float(v) if v is not None else 0.0
                    except Exception:
                        kg = 0.0
                    items[nombre_ins] = round(kg, 2)
                    total += kg
                insumos_por_fecha[fecha_col] = {
                    'items': items,
                    'total_kg': round(total, 2)
                }
            wb_ins.close()
            log.info(f"  ✓ STOCK DE INSUMOS.xlsx — {len(headers)} fechas leídas")
        except Exception as e:
            log.warning(f"  ⚠ Error leyendo STOCK DE INSUMOS.xlsx: {e}")
    else:
        log.warning(f"  ⚠ No existe {ruta_ins}")

    def _try_parse_date(s):
        """Intenta parsear una fecha YYYY-MM-DD; retorna None si es inválida."""
        try:
            return datetime.strptime(s, '%Y-%m-%d')
        except ValueError:
            return None

    # 4. Función para encontrar el financiero más próximo a una fecha (sin exceder)
    def _financiero_mas_proximo(fecha_target_str):
        """Retorna el dict financiero cuya fecha es la más cercana a fecha_target (≤ fecha_target)."""
        dt_target = _try_parse_date(fecha_target_str)
        if dt_target is None:
            return None
        candidatos = [(f, d) for f, d in financieros.items()
                      if _try_parse_date(f) is not None and f <= fecha_target_str]
        if not candidatos:
            candidatos = [(f, d) for f, d in financieros.items()
                          if _try_parse_date(f) is not None]
        if not candidatos:
            return None
        candidatos.sort(key=lambda x: abs((_try_parse_date(x[0]) - dt_target).days))
        return candidatos[0][1]

    # 5. Función para encontrar los insumos más próximos a una fecha
    def _insumos_mas_proximos(fecha_target_str):
        """Retorna insumos de la fecha más cercana a fecha_target."""
        dt_target = _try_parse_date(fecha_target_str)
        if dt_target is None:
            return None, None
        candidatos = [(f, d) for f, d in insumos_por_fecha.items()
                      if _try_parse_date(f) is not None and f <= fecha_target_str]
        if not candidatos:
            candidatos = [(f, d) for f, d in insumos_por_fecha.items()
                          if _try_parse_date(f) is not None]
        if not candidatos:
            return None, None
        candidatos.sort(key=lambda x: abs((_try_parse_date(x[0]) - dt_target).days))
        return candidatos[0][0], candidatos[0][1]

    # 6. Construir snapshots
    # v15.19: snapshots INMUTABLES. Los meses CERRADOS (periodo < mes corriente)
    # congelan su hacienda_masa: una vez calculado con los parámetros vigentes al
    # cierre, NO se recalcula aunque después cambien ADP_CAL/TECHO/recría. Sólo el
    # mes corriente recalcula cada tick ("stock vivo"). financiero e insumos
    # SIEMPRE se refrescan — el usuario sube el financiero tarde (p.ej. el de mayo
    # lo carga en jun/jul), así que congelar el snapshot entero rompería ese flujo.
    log.info("  Construyendo snapshots mensuales...")

    # Cargar snapshots previos para reutilizar hacienda_masa de meses cerrados.
    import json as _json
    snaps_previos = {}  # periodo → hacienda_masa cacheado (con parametros_calc)
    _ruta_prev = Path(carpeta) / "comportamiento_historico.json"
    if _ruta_prev.exists():
        try:
            with open(_ruta_prev, encoding='utf-8') as _fp:
                _prev = _json.load(_fp)
            for _s in _prev.get('snapshots', []):
                _hm = _s.get('hacienda_masa')
                if _s.get('periodo') and isinstance(_hm, dict) and 'parametros_calc' in _hm:
                    # v15.21: NO sellar snapshots heredados (forward-fill por
                    # parseo vacío). Se excluyen del cache para que el mes vuelva
                    # a reparsear cada tick y se auto-sane cuando el XLS sea
                    # legible (entonces sella con el cálculo real).
                    if _hm.get('parametros_calc', {}).get('origen') == 'heredado':
                        continue
                    snaps_previos[_s['periodo']] = _hm
            log.info(f"  Snapshots previos con parametros_calc: {len(snaps_previos)}")
        except Exception as _e:
            log.warning(f"  ⚠ No se pudo leer comportamiento_historico.json previo: {_e}")

    mes_corriente = datetime.now().strftime('%Y-%m')
    snapshots = []

    for ruta_c in archivos_cara:
        nombre_c = _os.path.basename(ruta_c)

        # Período (YYYY-MM) y fecha (YYYY-MM-DD) desde el nombre del archivo
        m2 = re.search(r'(\d{2})-(\d{2})-(\d{4})', nombre_c)
        if m2:
            d2, mo2, y2 = m2.groups()
            periodo   = f"{y2}-{mo2}"
            fecha_nom = f"{y2}-{mo2}-{d2}"
        else:
            periodo, fecha_nom = None, None

        es_corriente = (periodo == mes_corriente)
        hm_cache = snaps_previos.get(periodo) if periodo else None

        if hm_cache is not None and not es_corriente:
            # Mes cerrado ya sellado → congelar: reutilizar hacienda_masa tal cual.
            masa = hm_cache
            fecha_snap = masa.get('fecha', fecha_nom)

            # v15.54: sacar el corral 10000 (virtual) de los meses sellados que lo
            # tenían baked-in (jun/jul 2026), con el criterio de v15.46. Resta
            # EXACTA usando los valores 'Otro' ya sellados (todo el corral 10000 es
            # PEGSA) — no recomputa parámetros, así que ningún otro valor se mueve.
            # Self-terminante: una vez removido 'Otro' no vuelve a activarse, y el
            # parser (v15.54) ya evita que reingrese en re-parseos/meses futuros.
            _otro = ((masa.get('pegsa') or {}).get('por_campo') or {}).get('Otro')
            if _otro and _otro.get('cabezas'):
                _oc = _otro.get('cabezas') or 0
                _ok = _otro.get('kg_proyectado') or 0
                masa['pegsa']['cabezas']       = (masa['pegsa'].get('cabezas') or 0) - _oc
                masa['pegsa']['kg_proyectado'] = (masa['pegsa'].get('kg_proyectado') or 0) - _ok
                masa['total_cabezas']          = (masa.get('total_cabezas') or 0) - _oc
                masa['total_kg']               = (masa.get('total_kg') or 0) - _ok
                masa['pegsa']['por_campo'].pop('Otro', None)
                _ph = (masa.get('por_hotelero') or {}).get('PEGSA')
                if _ph:
                    _ph['cabezas']       = (_ph.get('cabezas') or 0) - _oc
                    _ph['kg_proyectado'] = (_ph.get('kg_proyectado') or 0) - _ok
                # Forzar re-injerto del por_campo del grupo sobre los totales ya
                # limpios (el bloque v15.52 de abajo lo re-arma y re-normaliza).
                masa.pop('por_campo', None)
                masa.pop('por_campo_origen', None)
                masa['corral10000_excluido'] = 'v15.54'
                log.info(f"    → {periodo}: excluido corral 10000 sellado "
                         f"(−{int(_oc)} cab / −{_ok:,.0f} kg)")

            # v15.52: injerto NO destructivo del desglose por campo del grupo.
            # Los meses sellados por v15.19 no lo tienen. Se re-parsea el
            # Listado_Caravanas SOLO para extraer ese bloque nuevo; ningún valor
            # preexistente se modifica. Los kg se normalizan al total sellado
            # para conservar la identidad suma(desglose) == total. Corre una sola
            # vez por mes (después 'por_campo' ya está y no se re-parsea).
            if 'por_campo' not in masa:
                try:
                    _fresh = _parse_listado_caravanas_html(ruta_c)
                    if _fresh and _fresh.get('por_campo'):
                        masa['por_campo'] = _normalizar_por_campo(
                            _fresh['por_campo'], masa.get('total_kg'))
                        masa['por_campo_origen'] = 'injertado_v15.52'
                        log.info(f"    → {periodo}: por_campo injertado "
                                 f"({len(masa['por_campo'])} campos, totales sellados intactos)")
                    else:
                        log.warning(f"    ⚠ {periodo}: no se pudo injertar por_campo "
                                    f"(re-parseo sin datos)")
                except Exception as _e:
                    log.warning(f"    ⚠ {periodo}: error injertando por_campo: "
                                f"{type(_e).__name__}: {_e}")

            log.info(f"  → {nombre_c} · {periodo} CERRADO — hacienda_masa congelado "
                     f"(parametros_calc {masa.get('parametros_calc', {}).get('version', '?')})")
        else:
            # Mes corriente, o cerrado sin sellar (1ª corrida post-v15.19) →
            # (re)calcular y sellar el bloque parametros_calc dentro de hacienda_masa.
            log.info(f"  → Procesando {nombre_c}" + (" (mes corriente)" if es_corriente else ""))
            masa = _parse_listado_caravanas_html(ruta_c)
            if masa is None:
                log.warning(f"    ⚠ Skipping {nombre_c} — error en parseo")
                continue
            fecha_snap = masa['fecha']
            masa['parametros_calc'] = {
                'version':          'v15.19',
                'adp_cal_por_cat':  dict(_ADP_CAL_RUNTIME if es_corriente else ADP_CAL_FALLBACK),
                'techo_kg_por_cat': dict(TECHO_KG_POR_CAT),
                'engorde_recria':   ENGORDE_RECRIA,
                'techo_kg_recria':  TECHO_KG_RECRIA,
                'fecha_calculado':  datetime.now().isoformat(),
            }

        # financiero + insumos SIEMPRE se refrescan (NO se congelan)
        fin = _financiero_mas_proximo(fecha_snap)
        fin_log = (f"financiero: {fin['fecha']} (${fin['disponible']:,.0f})"
                   if fin else "financiero: no disponible")
        fecha_ins, ins_data = _insumos_mas_proximos(fecha_snap)
        ins_log = (f"insumos: {fecha_ins} ({ins_data['total_kg']:,.0f} kg)"
                   if ins_data else "insumos: no disponibles")
        log.info(f"    {fin_log} | {ins_log}")

        snap = {
            'fecha':   fecha_snap,
            'periodo': periodo if periodo else fecha_snap[:7],
            'hacienda_masa': masa,
            'insumos': {
                'fecha_col':   fecha_ins,
                'items':       ins_data['items']       if ins_data else {},
                'total_kg':    ins_data['total_kg']    if ins_data else 0.0,
            },
            'financiero': fin if fin else {
                'fecha': None, 'disponible': 0, 'cheques_cartera': 0,
                'cheques_emitidos': 0, 'cobrar_hacienda': 0, 'pagar_hacienda': 0,
                'usd_cant': 0, 'usd_ars': 0, 'lcg': 0, 'tercio_bravo': 0,
            },
        }
        snapshots.append(snap)

    # Ordenar por fecha
    snapshots.sort(key=lambda s: s['fecha'])

    # v15.21 · Blindaje (preventivo): forward-fill de cab/kg si un mes parseó
    # vacío (total_cabezas==0). Hereda total_cabezas/total_kg/por_hotelero/pegsa
    # del último mes sano y marca parametros_calc.origen='heredado' para que el
    # sellado v15.19 NO lo congele (reparsea cada tick hasta auto-sanar).
    # NOTA: el modo "el mes desaparece" (parseo→None→snapshot omitido) ya está
    # mitigado por v15.19 (meses cerrados se sirven del cache, inmunes al lock).
    _ult_masa_sana = None
    _n_masa_fill = 0
    for _snap9 in snapshots:
        _masa9 = _snap9.get('hacienda_masa', {})
        if (_masa9.get('total_cabezas', 0) or 0) == 0:
            if _ult_masa_sana is not None:
                _masa9['total_cabezas'] = _ult_masa_sana.get('total_cabezas', 0)
                _masa9['total_kg']      = _ult_masa_sana.get('total_kg', 0)
                _masa9['por_hotelero']  = {h: dict(d) for h, d in _ult_masa_sana.get('por_hotelero', {}).items()}
                _masa9['pegsa']         = dict(_ult_masa_sana.get('pegsa', {}))
                _pc = _masa9.setdefault('parametros_calc', {})
                _pc['origen']           = 'heredado'
                _pc['heredado_de']      = _ult_masa_sana.get('fecha', '?')
                _n_masa_fill += 1
                log.info(f"    Mes {_snap9.get('periodo')}: cab/kg heredados de "
                         f"{_ult_masa_sana.get('fecha','?')} (parseo vacío)")
        else:
            _ult_masa_sana = _masa9
    if _n_masa_fill:
        log.info(f"  Módulo 9: {_n_masa_fill} mes(es) con cab/kg heredados (forward-fill v15.21)")

    output = {
        'generado':  datetime.now().isoformat(),
        'snapshots': snapshots,
        'total':     len(snapshots),
    }

    guardar(output, carpeta, "comportamiento_historico.json")
    n = len(snapshots)
    if snapshots:
        log.info(f"  ✓ comportamiento_historico.json — {n} meses "
                 f"({snapshots[0]['fecha']} → {snapshots[-1]['fecha']})")
    else:
        log.info(f"  ✓ comportamiento_historico.json — 0 meses (sin Listado_Caravanas)")
    return output


def procesar_precios_inferencia(carpeta_out, log=None):
    """v8 / v15.34: lee el Excel "referencia precios de mercado simulador.xlsx"
    del subdir simulador/simulador/ y vuelca DOS archivos:

      precios_inferencia.json           — snapshot actual (sobrescribe)
      precios_inferencia_historico.json — acumulado por fecha (upsert)

    Layout fijo del Excel (Hoja1):
      A1  "Fecha:"               B1 <datetime>
      A2  "Categoria:"           B2..E2  4 categorías
      A3  "Kg Compra:"           B3..E3
      A4  "kg Venta:"            B4..E4
      A5  "Precio venta:"        B5..E5
      A6  "Rinde:"               B6..E6  (decimal 0..1)
      A7  "cost kg prod:"        B7..E7
      A8  "Dias Feed:"           B8..E8
      A9  "Engorde:"             B9..E9    ← v15.34 (kg/día)
      A10 "Eficiencia Engorde:"  B10..E10  ← v15.34 (decimal, puede ser negativo)
      A11 "Precio comp:"         B11..E11  ← KPI principal

    El Excel está en OneDrive y a veces queda bloqueado por el sync —
    copiamos a tempfile antes de leer.

    Devuelve (snapshot_dict, hist_dict) o (None, None) si no hay archivo
    o si hay error. NO levanta excepción — sólo loguea warning.
    """
    if log is None:
        log = logging.getLogger("inferencia")

    base = Path(__file__).resolve().parent.parent
    xl = base / "simulador" / "simulador" / "referencia precios de mercado simulador.xlsx"
    if not xl.exists():
        log.warning(f"  ⚠ Excel de inferencia no encontrado en {xl}, saltando")
        return None, None

    # Copiar a temp (OneDrive a veces bloquea la lectura directa)
    import shutil, tempfile
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tf:
            tmp_path = _tf.name
        shutil.copy2(str(xl), tmp_path)
    except Exception as e:
        log.warning(f"  ⚠ no pude copiar Excel a temp: {e}")
        return None, None

    try:
        try:
            import openpyxl
        except ImportError:
            log.warning("  ⚠ openpyxl no instalado; saltando precios de inferencia")
            return None, None
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
        except Exception as e:
            log.warning(f"  ⚠ no pude abrir Excel: {e}")
            return None, None
        ws = wb[wb.sheetnames[0]]

        # v15.66: categorías DINÁMICAS — se leen las columnas B.. mientras la
        # fila 2 (Categoria) tenga valor. Los ids históricos se preservan por
        # nombre; una categoría nueva genera su id como slug del nombre.
        import unicodedata as _ud
        def _slug(n):
            s = _ud.normalize("NFKD", str(n).strip().lower())
            s = "".join(ch for ch in s if not _ud.combining(ch))
            return "_".join(s.split())
        IDS_POR_NOMBRE = {
            "vaca 100 dias": "vaca_100",
            "vaca 60 dias":  "vaca_60",
            "novillo":       "novillo",
            "vaquillona":    "vaquillona",
        }

        # B1: fecha del snapshot. Aceptar datetime o string ISO.
        b1 = ws.cell(row=1, column=2).value
        if isinstance(b1, datetime):
            fecha_iso = b1.strftime("%Y-%m-%d")
        elif isinstance(b1, str) and b1.strip():
            # Intentar parsear varios formatos
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
                try:
                    fecha_iso = datetime.strptime(b1.strip(), fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            else:
                fecha_iso = datetime.now().strftime("%Y-%m-%d")
        else:
            fecha_iso = datetime.now().strftime("%Y-%m-%d")

        def cell(r, c):
            v = ws.cell(row=r, column=c).value
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return None

        items = []
        col = 1
        while True:
            col += 1  # arranca en B=2
            nombre = ws.cell(row=2, column=col).value
            if nombre is None or not str(nombre).strip():
                break
            nombre = str(nombre).strip()
            item_id = IDS_POR_NOMBRE.get(_slug(nombre).replace("_", " "), _slug(nombre))
            # Capitalizar primera letra de cada palabra para presentación uniforme
            nombre_disp = " ".join(w.capitalize() if w.lower() != "días" else "días" for w in nombre.split())
            items.append({
                "id":                 item_id,
                "nombre":             nombre_disp,
                "nombre_raw":         nombre,
                "kg_compra":          cell(3, col),
                "kg_venta":           cell(4, col),
                "precio_venta":       cell(5, col),
                "rinde":              cell(6, col),
                "cost_kg_prod":       cell(7, col),
                "dias_feed":          cell(8, col),
                # v15.34: 2 datos nuevos del Excel (filas A9 y A10)
                "engorde":            cell(9, col),    # kg/día (ej 1.14)
                "eficiencia_engorde": cell(10, col),   # decimal con signo (ej -0.25 = -25%)
                "precio_comp":        cell(11, col),
            })

        snapshot = {
            "meta": {
                "fecha":    fecha_iso,
                "generado": datetime.now().isoformat(),
                "archivo":  xl.name,
            },
            "items": items,
        }

        # 1) snapshot actual
        out_actual = Path(carpeta_out) / "precios_inferencia.json"
        with out_actual.open("w", encoding="utf-8") as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)

        # 2) histórico (upsert por fecha)
        hist_path = Path(carpeta_out) / "precios_inferencia_historico.json"
        if hist_path.exists():
            try:
                with hist_path.open("r", encoding="utf-8") as f:
                    hist = json.load(f)
                if not isinstance(hist, dict) or "semanas" not in hist:
                    hist = {"meta": {}, "semanas": []}
            except Exception:
                hist = {"meta": {}, "semanas": []}
        else:
            hist = {"meta": {}, "semanas": []}

        nueva_semana = {"fecha": fecha_iso, "items": items}
        # Si la fecha ya existe, reemplazar; sino, push y ordenar.
        idx = next((i for i, s in enumerate(hist["semanas"]) if s.get("fecha") == fecha_iso), None)
        if idx is not None:
            hist["semanas"][idx] = nueva_semana
        else:
            hist["semanas"].append(nueva_semana)
            hist["semanas"].sort(key=lambda s: s.get("fecha", ""))

        hist["meta"] = {
            "generado":  datetime.now().isoformat(),
            "n_semanas": len(hist["semanas"]),
        }
        with hist_path.open("w", encoding="utf-8") as f:
            json.dump(hist, f, ensure_ascii=False, indent=2, default=str)

        fecha_disp = "/".join(reversed(fecha_iso.split("-")))
        log.info(f"  ✓ Precios de inferencia: snapshot {fecha_disp} · histórico {len(hist['semanas'])} semanas")
        return snapshot, hist
    finally:
        try: Path(tmp_path).unlink()
        except Exception: pass


# v15.57: el Excel de compras se carga a mano y las categorías vienen con
# mayúsculas/minúsculas mezcladas y variantes ('Ternero overo', 'Vaca Directo',
# 'vaca holando'). 16 grafías distintas para 7 categorías reales. Se normaliza
# sin acentos y en minúscula, con mapa explícito — NO por coincidencia parcial:
# 'ternera' y 'ternero' comparten prefijo y un startswith/in las mezclaría.
_CAT_COMPRAS = {
    'vaca': 'Vaca', 'vaca directo': 'Vaca', 'vaca holando': 'Vaca',
    'vaquillona': 'Vaquillona',
    'novillo': 'Novillo',
    'novillito': 'Novillito',
    'toro': 'Toro',
    'ternero': 'Ternero', 'ternero overo': 'Ternero',
    'ternera': 'Ternera', 'ternera overa': 'Ternera',
}

COMPRAS_VENTANA_DIAS = 90


def _norm_cat_compra(s):
    """Minúscula, sin acentos, espacios colapsados."""
    import unicodedata
    s = str(s).strip().lower()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    return ' '.join(s.split())


def _norm_tropa(s):
    """v15.59: clave de match de tropa entre WinCampo y el Excel de compras.

    Los formatos difieren: la API trae 'PEG-MYA-05-03-2026', 'PECMAR 11/03/2026',
    'BULLHUIN15/05/2026'; el Excel 'PEG.MYA.05/03/26'. Se saca todo lo que no sea
    alfanumérico, se pasa a mayúscula y se homogeneiza el año final a 2 dígitos.
    """
    import re
    k = re.sub(r'[^A-Za-z0-9]', '', str(s or '')).upper()
    return re.sub(r'20(\d{2})$', r'\1', k)


def procesar_precios_racion(carpeta_out, log=None):
    """v15.59: lee 'preico de racion feelot.xlsx' (sí, 'preico' — el nombre del
    archivo del usuario está así y NO se corrige) y devuelve los precios
    mensuales del feedlot:

        {'2026-07': {'tc': 237.37, 'ms': 0.6676, 'dia': 450, 'san': 7500}, ...}

    tc  = $/kg tal cual · ms = % materia seca (decimal) · dia = $/día-animal de
    estructura · san = $/cabeza de sanidad al ingreso.
    El $/kg de MS se calcula como tc / ms (jul-26: 237,37 / 0,6676 = 355,56).
    ⚠ El 'Costo Alimentación' de WinCampo NO sirve (da $3,6/kg MS contra $355).

    Devuelve {} si no encuentra el archivo — el caller degrada.
    """
    if log is None:
        log = logging.getLogger("racion")

    SUBDIR = Path("archivos de pecuaria compartidos") / "haras"
    ruta_dir = None
    cand = Path(carpeta_out).resolve()
    for base in [cand] + list(cand.parents):
        if (base / SUBDIR).is_dir():
            ruta_dir = base / SUBDIR
            break
    if ruta_dir is None:
        log.warning(f"  ⚠ Carpeta de haras no encontrada desde {carpeta_out}, sin precios de ración")
        return {}

    import fnmatch
    cands = [p for p in Path(ruta_dir).iterdir()
             if p.is_file() and fnmatch.fnmatch(p.name.lower(), '*racion*feelot*.xlsx')]
    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    if not cands:
        log.warning(f"  ⚠ Excel de ración no encontrado en {ruta_dir}, saltando")
        return {}
    xl = cands[0]

    try:
        import openpyxl
    except ImportError:
        log.warning("  ⚠ openpyxl no instalado; sin precios de ración")
        return {}

    import shutil, tempfile
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tf:
            tmp_path = _tf.name
        shutil.copy2(str(xl), tmp_path)
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        filas = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
        try: wb.close()
        except Exception: pass
    except Exception as e:
        log.warning(f"  ⚠ no pude leer {xl.name}: {e}")
        return {}
    finally:
        if tmp_path:
            try: Path(tmp_path).unlink()
            except Exception: pass

    if len(filas) < 2:
        log.warning(f"  ⚠ {xl.name} vacío, sin precios de ración")
        return {}

    # Headers por nombre (mes · TC precio · % ms · costo dia animal · costo ingreso sanidad)
    hdr = [_norm_cat_compra(c) if c is not None else "" for c in filas[0]]
    def _idx(*nombres):
        for n in nombres:
            if n in hdr:
                return hdr.index(n)
        return None
    i_mes = _idx("mes")
    i_tc  = _idx("tc precio", "tc")
    i_ms  = _idx("% ms", "ms")
    i_dia = _idx("costo dia animal", "dia")
    i_san = _idx("costo ingreso sanidad", "sanidad")
    if None in (i_mes, i_tc, i_ms, i_dia, i_san):
        log.warning(f"  ⚠ headers inesperados en {xl.name}: {hdr}, saltando ración")
        return {}

    out = {}
    for r in filas[1:]:
        f = r[i_mes]
        if f is None:
            continue
        if isinstance(f, datetime):
            mes = f"{f.year:04d}-{f.month:02d}"
        else:
            mes = str(f)[:7].replace("/", "-")
        try:
            out[mes] = {
                "tc":  float(r[i_tc]),
                "ms":  float(r[i_ms]),
                "dia": float(r[i_dia]),
                "san": float(r[i_san]),
            }
        except (TypeError, ValueError):
            log.warning(f"  ⚠ fila de ración no numérica en {mes}, salteada")
    if out:
        _u = max(out)
        log.info(f"  ✓ Precios de ración: {len(out)} meses ({min(out)} → {_u}) · "
                 f"$/kg MS {_u} = {out[_u]['tc']/out[_u]['ms']:,.2f}")
    return out


def procesar_compras_reales(carpeta_out, log=None):
    """v15.57: lee 'compras de hacienda.xlsx' (hoja OK) y genera
    precios_compra_real.json con el precio REALMENTE pagado por categoría,
    para contrastarlo contra el tope de indiferencia del simulador.

    El archivo lo mantiene el usuario a mano y vive FUERA de la carpeta de
    datos, en el OneDrive compartido:
      <OneDrive>\\archivos de pecuaria compartidos\\haras\\compras de hacienda.xlsx

    Estructura de la hoja OK (se lee POR NOMBRE de header, no por posición —
    las columnas Cab WinCampo/Dif cantidad se corrieron cuando agregaron
    'comision' y una columna vacía en el medio):
      tropa · Fecha · Categoria · Origen · Cantidad · Precio Compra ($/kg) ·
      kg (kg/cabeza) · comision · <vacía> · Cab WinCampo · Dif cantidad
    Hay filas separadoras vacías entre tropas; Cab WinCampo/Dif cantidad sólo
    vienen en la primera fila de cada tropa y NO se usan.

    La hoja 'VER WIN' (conciliación contra WinCampo) queda fuera de alcance.

    Devuelve el dict volcado, o None si no hay archivo / error.
    NO levanta excepción — sólo loguea warning (igual que trazabilidad).
    """
    if log is None:
        log = logging.getLogger("compras")
    from datetime import date, timedelta

    # ── Resolución de la ruta ────────────────────────────────
    # carpeta_out es <OneDrive>\PEGSA_Portal\datos (config 'auto'), pero el
    # archivo cuelga de la RAÍZ del OneDrive. En vez de contar .parent a ciegas
    # (que se rompe si carpeta_out fuese PEGSA_Portal, o si corre desde el repo)
    # subimos buscando la carpeta compartida. Override por config.ini.
    SUBDIR = Path("archivos de pecuaria compartidos") / "haras"
    ruta_dir = None
    try:
        _cfg_path = Path(__file__).parent / "config.ini"
        if _cfg_path.exists():
            _cfg = configparser.ConfigParser()
            _cfg.read(_cfg_path, encoding="utf-8")
            if _cfg.has_section("RUTAS") and _cfg["RUTAS"].get("compras_dir"):
                ruta_dir = Path(_cfg["RUTAS"].get("compras_dir"))
    except Exception:
        pass
    if ruta_dir is None:
        cand = Path(carpeta_out).resolve()
        for base in [cand] + list(cand.parents):
            if (base / SUBDIR).is_dir():
                ruta_dir = base / SUBDIR
                break
    if ruta_dir is None or not Path(ruta_dir).is_dir():
        log.warning(f"  ⚠ Carpeta de compras no encontrada desde {carpeta_out}, saltando")
        return None

    # v15.44: matching case-insensitive — Path.glob() es case-sensitive aun en
    # Windows y el usuario puede renombrar el archivo con otra caja.
    import fnmatch
    cands = [p for p in Path(ruta_dir).iterdir()
             if p.is_file() and fnmatch.fnmatch(p.name.lower(), 'compras de hacienda*.xlsx')]
    cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
    xl = cands[0] if cands else None
    log.info(f"  compras: {ruta_dir} · archivo={xl.name if xl else None}")
    if xl is None:
        log.warning(f"  ⚠ 'compras de hacienda.xlsx' no está en {ruta_dir}, saltando")
        return None

    try:
        import openpyxl
    except ImportError:
        log.warning("  ⚠ openpyxl no instalado; saltando compras reales")
        return None

    # Copiar a temp — OneDrive a veces bloquea la lectura directa.
    import shutil, tempfile
    tmp_path = None
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tf:
            tmp_path = _tf.name
        shutil.copy2(str(xl), tmp_path)
    except Exception as e:
        log.warning(f"  ⚠ no pude copiar {xl.name} a temp: {e}")
        return None

    try:
        wb = openpyxl.load_workbook(tmp_path, data_only=True, read_only=True)
        if "OK" not in wb.sheetnames:
            log.warning(f"  ⚠ {xl.name} no tiene hoja 'OK' (tiene {wb.sheetnames}), saltando")
            return None
        filas = list(wb["OK"].iter_rows(values_only=True))
        try: wb.close()
        except Exception: pass
        if len(filas) < 2:
            log.warning(f"  ⚠ hoja OK vacía en {xl.name}, saltando")
            return None

        hdr = [_norm_cat_compra(c) if c is not None else "" for c in filas[0]]
        try:
            i_tropa = hdr.index("tropa")
            i_fecha = hdr.index("fecha")
            i_cat   = hdr.index("categoria")
            i_cant  = hdr.index("cantidad")
            i_prec  = hdr.index("precio compra")
            i_kg    = hdr.index("kg")
            # v15.59: comisión fila por fila (no un % fijo). Puede no existir.
            i_com   = hdr.index("comision") if "comision" in hdr else None
        except ValueError as e:
            log.warning(f"  ⚠ headers inesperados en hoja OK ({hdr}): {e}, saltando")
            return None

        hoy   = date.today()
        desde = hoy - timedelta(days=COMPRAS_VENTANA_DIAS)
        acc_v, acc_t = {}, {}          # ventana / histórico completo
        acc_tropa    = {}              # v15.59: índice por tropa normalizada
        acc_cat_mes  = {}              # v15.59: categoría → mes → kg/plata
        n_filas = 0
        desc = {}
        grafias_desconocidas = {}

        def _add(d, cat, cab, kg_cab, precio):
            a = d.setdefault(cat, {"cabezas": 0, "kg": 0.0, "plata": 0.0, "operaciones": 0})
            a["cabezas"]     += cab
            a["kg"]          += cab * kg_cab      # kilos totales de la operación
            a["plata"]       += cab * kg_cab * precio
            a["operaciones"] += 1

        def _descartar(motivo, detalle):
            desc[motivo] = desc.get(motivo, 0) + 1
            log.warning(f"  ⚠ compra descartada ({motivo}): {detalle}")

        for r in filas[1:]:
            # filas separadoras entre tropas
            if r[i_cat] is None or r[i_fecha] is None:
                continue
            n_filas += 1
            tropa = r[i_tropa] or "?"
            f = r[i_fecha]
            f = f.date() if isinstance(f, datetime) else f
            if not isinstance(f, date):
                _descartar("fecha_invalida", f"{tropa} · fecha={r[i_fecha]!r}")
                continue
            try:
                cab    = int(float(r[i_cant]))
                precio = float(r[i_prec])
                kg_cab = float(r[i_kg])
            except (TypeError, ValueError):
                _descartar("fila_incompleta", f"{tropa} · cant/precio/kg no numéricos")
                continue

            # v15.57: filtros defensivos y genéricos (no hardcodean filas).
            # Dos filas malas confirmadas en el archivo al 2026-08-12:
            #  1) 'BUL.HVG.02/01/2026' cargada con fecha 2026-12-29 (el nombre
            #     dice 02/01). Rompería la ventana de 90 días.
            #  2) 1 cabeza a $13.333/kg con 60 kg/cab — un ternero de 60 kg no
            #     existe; es error de carga.
            if f > hoy:
                _descartar("fecha_futura", f"{tropa} · {f}")
                continue
            if precio > 10000 or precio <= 0:
                _descartar("precio_fuera_rango", f"{tropa} · ${precio:,.0f}/kg · {kg_cab:.0f} kg")
                continue
            if kg_cab < 100 or kg_cab > 800:
                _descartar("peso_implausible", f"{tropa} · {kg_cab:.0f} kg/cab")
                continue
            if cab <= 0:
                _descartar("cantidad_invalida", f"{tropa} · {cab} cab")
                continue

            cat = _CAT_COMPRAS.get(_norm_cat_compra(r[i_cat]))
            if cat is None:
                g = str(r[i_cat]).strip()
                grafias_desconocidas[g] = grafias_desconocidas.get(g, 0) + 1
                _descartar("categoria_desconocida", f"{tropa} · {g!r}")
                continue

            # ── v15.59 · índice POR TROPA (para Resultado por Remito) ──
            # ⚠ split por '+': hay filas con
            #   'PEG.VIL.25/07/2026 + PEG.VIL.24/07/2026 + …' (84 de 221).
            # Sin indexar CADA tropa por separado la cobertura de precios cae
            # de ~90% a ~50% — fue el bug más caro del prototipo.
            _com = None
            if i_com is not None:
                try:
                    _com = float(r[i_com])
                except (TypeError, ValueError):
                    _com = None
            for _t in str(tropa).split("+"):
                _t = _t.strip()
                if not _t:
                    continue
                _k = _norm_tropa(_t)
                if not _k:
                    continue
                a = acc_tropa.setdefault(_k, {
                    "tropa": _t, "categorias": {}, "kg": 0.0, "plata": 0.0,
                    "comision": _com, "cabezas": 0, "filas": 0,
                })
                a["kg"]     += cab * kg_cab
                a["plata"]  += cab * kg_cab * precio
                a["cabezas"] += cab
                a["filas"]  += 1
                a["comision"] = _com          # última fila de la tropa manda
                _c = a["categorias"].setdefault(cat, {"kg": 0.0, "plata": 0.0,
                                                      "ultimo": None, "filas": 0,
                                                      "precios": set()})
                _c["kg"]    += cab * kg_cab
                _c["plata"] += cab * kg_cab * precio
                _c["filas"] += 1
                _c["precios"].add(round(precio, 2))
                # ⚠ v15.59: una misma tropa+categoria puede tener VARIAS filas con
                # precios distintos (lotes distintos dentro de la misma compra):
                # PECHUIN17/04/2026 tiene vaca a $2.851,93 y vaca a $2.004,69.
                # Se toma la ULTIMA para reproducir exacto el prototipo validado,
                # pero se guarda tambien el ponderado por kilos (precio_kg_pond),
                # que es lo que realmente se pago. Ver bitacora v15.59.
                _c["ultimo"] = precio

            # v15.59: por categoría y MES de compra — lo usa el precio de
            # reposición (promedio ponderado del último mes con compras de la
            # categoría).
            _m = acc_cat_mes.setdefault(cat, {}).setdefault(f"{f.year:04d}-{f.month:02d}",
                                                            {"kg": 0.0, "plata": 0.0, "cabezas": 0})
            _m["kg"]     += cab * kg_cab
            _m["plata"]  += cab * kg_cab * precio
            _m["cabezas"] += cab

            _add(acc_t, cat, cab, kg_cab, precio)
            if f >= desde:
                _add(acc_v, cat, cab, kg_cab, precio)

        def _cerrar(d):
            """Promedio PONDERADO POR KILOS, no simple: una compra de 1 cabeza
            no puede pesar lo mismo que una de 115."""
            out = {}
            for cat, a in sorted(d.items(), key=lambda x: -x[1]["cabezas"]):
                if a["cabezas"] <= 0 or a["kg"] <= 0:
                    continue
                out[cat] = {
                    "cabezas":     a["cabezas"],
                    "kg_cab":      round(a["kg"] / a["cabezas"], 1),
                    "precio_kg":   round(a["plata"] / a["kg"]),
                    "operaciones": a["operaciones"],
                }
            return out

        n_desc = sum(desc.values())
        salida = {
            "meta": {
                "generado":         datetime.now().isoformat(),
                "archivo":          xl.name,
                "ventana_dias":     COMPRAS_VENTANA_DIAS,
                "desde":            desde.isoformat(),
                "hasta":            hoy.isoformat(),
                "filas_totales":    n_filas,
                "filas_descartadas": n_desc,
                "motivos_descarte": desc,
            },
            "por_categoria":       _cerrar(acc_v),
            "por_categoria_total": _cerrar(acc_t),
            # v15.59: índice por tropa para el módulo Resultado por Remito.
            # NO cambia nada de lo que consumen las tarjetas de indiferencia.
            "por_tropa": {
                k: {
                    "tropa":         a["tropa"],
                    "precio_kg":     round(a["plata"] / a["kg"], 2) if a["kg"] > 0 else None,
                    "kg":            round(a["kg"], 1),
                    "cabezas":       a["cabezas"],
                    "comision":      a["comision"],
                    "filas":         a["filas"],
                    "por_categoria": {
                        c: {
                            "precio_kg":      round(v["ultimo"], 2),
                            "precio_kg_pond": round(v["plata"] / v["kg"], 2),
                            "kg":             round(v["kg"], 1),
                            "filas":          v["filas"],
                            "multiprecio":    len(v["precios"]) > 1,
                        }
                        for c, v in a["categorias"].items() if v["kg"] > 0
                    },
                }
                for k, a in sorted(acc_tropa.items()) if a["kg"] > 0
            },
            # v15.59: categoría → mes → precio ponderado (precio de reposición).
            "por_categoria_mes": {
                c: {
                    m: {"precio_kg": round(v["plata"] / v["kg"], 2),
                        "kg": round(v["kg"], 1), "cabezas": v["cabezas"]}
                    for m, v in sorted(meses.items()) if v["kg"] > 0
                }
                for c, meses in sorted(acc_cat_mes.items())
            },
        }
        if grafias_desconocidas:
            salida["meta"]["categorias_desconocidas"] = grafias_desconocidas

        out_path = Path(carpeta_out) / "precios_compra_real.json"
        with out_path.open("w", encoding="utf-8") as f:
            json.dump(salida, f, ensure_ascii=False, indent=2, default=str)

        # Si mañana el archivo trae 30 descartes en vez de 2, hay que enterarse.
        log.info(f"  ✓ Compras reales: {n_filas} filas · {n_desc} descartadas {desc or ''} · "
                 f"{len(salida['por_categoria'])} categorías en {COMPRAS_VENTANA_DIAS}d "
                 f"(desde {desde.isoformat()})")
        for cat, v in salida["por_categoria"].items():
            log.info(f"      {cat:<12} {v['cabezas']:>5} cab · {v['kg_cab']:>6.1f} kg/cab · $ {v['precio_kg']:,}/kg")
        return salida
    finally:
        if tmp_path:
            try: Path(tmp_path).unlink()
            except Exception: pass


# v15.58: mismo ajuste que aplica el portal en el indicador de %PV
# (data.js ~L761: pct_peso_vivo.valor / 0.92). Se guardan crudo y ajustado.
AJUSTE_MS_PCT_PV = 0.92


def generar_pct_pv_mensual(carpeta_out, periodo, log=None):
    """v15.58: cruza el consumo mensual del mixer con los kg PV de El Haras y
    vuelca pct_pv_mensual.json — el %PV mes a mes de toda la historia.

    Lo consume el módulo Resultado por Remito, que hoy usa un 2,63% de
    referencia (un objetivo anual, no el consumo real) que sobreestima el
    alimento 30-40%.

        pct_pv_crudo    = (kg_ms_mes / dias_con_registro) / kg_pv_promedio * 100
        pct_pv_ajustado = pct_pv_crudo / 0.92

    Denominador = PROMEDIO del mes (fin mes anterior + fin mes) / 2, no fin de
    mes: el snapshot es del último día y el stock se mueve mucho adentro del mes
    (jun-26 arrancó en 3,57M y terminó en 2,99M kg → fin de mes inflaría el %PV
    ~9%).

    Devuelve el dict volcado, o None si falta alguna fuente.
    NO levanta excepción — sólo loguea warning y no pisa el JSON existente.
    """
    if log is None:
        log = logging.getLogger("pctpv")

    base = Path(carpeta_out)

    # ── Fuentes (ambas desde disco) ──────────────────────────
    try:
        with (base / f"consumo_{periodo}.json").open(encoding="utf-8") as f:
            _consumo = json.load(f)
    except Exception as e:
        log.warning(f"  ⚠ no pude leer consumo_{periodo}.json ({e}), saltando %PV mensual")
        return None
    _meses_consumo = ((_consumo.get("por_mes") or {}).get("meses")) or {}
    if not _meses_consumo:
        log.warning("  ⚠ consumo sin bloque por_mes (mixer caído?), saltando %PV mensual")
        return None

    try:
        with (base / "comportamiento_historico.json").open(encoding="utf-8") as f:
            _comp = json.load(f)
    except Exception as e:
        log.warning(f"  ⚠ no pude leer comportamiento_historico.json ({e}), saltando %PV mensual")
        return None
    _snaps = _comp.get("snapshots") or []
    if not _snaps:
        log.warning("  ⚠ comportamiento_historico sin snapshots, saltando %PV mensual")
        return None

    # kg PV de El Haras a fin de cada mes (el snapshot es del último día).
    # v15.58.1: el mixer alimenta a TODOS los animales de El Haras, propios y de
    # hoteleros (Bulltrade, Darwash, Las Taperas, Tercio Bravo, Saguaipé, UGMA…) —
    # confirmado por el usuario. El denominador es El Haras pegsa + hoteleros.
    # Con pegsa solo, 2025 daba 3,1-3,9% de %PV, imposible: en 2025 los hoteleros
    # pesaban 0,8-1,3M kg (30-50% extra), en 2026 bajaron a 0,3-0,4M.
    # ⚠ por_hotelero trae una clave 'PEGSA' que duplica la hacienda propia (y además
    # es de TODOS los campos, no solo El Haras): hay que excluirla.
    # Limitación conocida: por_hotelero no abre por campo. Si algún hotelero tuviera
    # hacienda fuera de El Haras, esto sobreestima un poco el denominador (contra la
    # serie diaria: ±1% en abr/may-26, +5/+9% en jun/jul-26).
    kg_fin_mes = {}
    kg_hoteleros_fin_mes = {}
    for s in _snaps:
        per = s.get("periodo")
        _hm = s.get("hacienda_masa") or {}
        kg = ((_hm.get("pegsa") or {})
              .get("por_campo", {}).get("El Haras", {}).get("kg_proyectado"))
        if not (per and kg):
            continue
        _hot = sum(
            float((v or {}).get("kg_proyectado") or 0)
            for k, v in (_hm.get("por_hotelero") or {}).items()
            if str(k).strip().upper() != "PEGSA"
        )
        kg_fin_mes[per]           = float(kg) + _hot
        kg_hoteleros_fin_mes[per] = _hot

    def _mes_anterior(mes):
        y, m = int(mes[:4]), int(mes[5:7])
        return f"{y-1:04d}-12" if m == 1 else f"{y:04d}-{m-1:02d}"

    meses_out         = {}
    meses_sin_kg_pv   = []
    meses_sin_consumo = []
    fuera_de_rango    = []

    for mes in sorted(_meses_consumo):
        c = _meses_consumo[mes]
        kg_ms_mes = c.get("kg_ms_total")
        n_dias    = c.get("dias_con_registro") or 0
        if not kg_ms_mes or n_dias <= 0:
            continue

        prev = kg_fin_mes.get(_mes_anterior(mes))
        fin  = kg_fin_mes.get(mes)
        if fin is not None and prev is not None:
            kg_pv, fuente = (prev + fin) / 2, "promedio_snapshots"
        elif fin is not None:
            # Primer mes de la serie de snapshots: no hay mes anterior.
            kg_pv, fuente = fin, "fin_mes"
        elif prev is not None:
            # Mes en curso: todavía no hay snapshot propio.
            kg_pv, fuente = prev, "fin_mes_anterior"
        else:
            meses_sin_kg_pv.append(mes)
            continue

        kg_ms_dia = kg_ms_mes / n_dias
        crudo     = kg_ms_dia / kg_pv * 100
        meses_out[mes] = {
            "kg_ms_mes":         round(kg_ms_mes, 1),
            "kg_ms_dia":         round(kg_ms_dia, 1),
            "dias_calendario":   c.get("dias_calendario"),
            "dias_con_registro": n_dias,
            "kg_pv_fin_mes":     round(fin) if fin is not None else None,
            # v15.58.1: cuánto de kg_pv_fin_mes son hoteleros (para auditar el aporte).
            "kg_pv_hoteleros_fin_mes": (round(kg_hoteleros_fin_mes[mes])
                                        if mes in kg_hoteleros_fin_mes else None),
            "kg_pv_haras":       round(kg_pv),
            "fuente_kg_pv":      fuente,
            "pct_pv_crudo":      round(crudo, 2),
            "pct_pv_ajustado":   round(crudo / AJUSTE_MS_PCT_PV, 2),
            "parcial":           bool(c.get("parcial")),
        }
        if not (1.0 <= crudo <= 3.5):
            fuera_de_rango.append((mes, round(crudo, 2), round(kg_ms_dia), round(kg_pv)))

    for mes in sorted(kg_fin_mes):
        if mes not in _meses_consumo:
            meses_sin_consumo.append(mes)

    if not meses_out:
        log.warning("  ⚠ ningún mes con consumo Y kg PV, no se genera pct_pv_mensual.json")
        return None

    salida = {
        "meta": {
            "generado":        datetime.now().isoformat(),
            "formula":         "(kg_ms_mes / dias_con_registro) / kg_pv_promedio_mes / 0.92 * 100",
            "ajuste_ms":       AJUSTE_MS_PCT_PV,
            "fuente_consumo":  f"consumo_{periodo}.json -> por_mes (mixer Dropbox, dias validos)",
            "fuente_kg_pv":    ("comportamiento_historico.json -> El Haras pegsa + hoteleros "
                                "(kg_proyectado), promedio (fin mes ant + fin mes)/2"),
            "desde":           min(meses_out),
            "hasta":           max(meses_out),
            "meses":           len(meses_out),
            "meses_sin_kg_pv":   meses_sin_kg_pv,
            "meses_sin_consumo": meses_sin_consumo,
        },
        "meses": meses_out,
    }
    guardar(salida, carpeta_out, "pct_pv_mensual.json")

    log.info(f"  ✓ %PV mensual: {len(meses_out)} meses ({min(meses_out)} → {max(meses_out)})")
    if meses_sin_kg_pv:
        log.info(f"    {len(meses_sin_kg_pv)} mes(es) con consumo pero sin kg PV "
                 f"(mixer más viejo que los snapshots): {meses_sin_kg_pv[0]} → {meses_sin_kg_pv[-1]}")
    for mes, crudo, kgms, kgpv in fuera_de_rango:
        log.warning(f"  ⚠ %PV fuera de rango [1,0-3,5] en {mes}: {crudo}% "
                    f"(kg_ms_dia={kgms:,} · kg_pv={kgpv:,})")
    return salida


# ═══════════════════════════════════════════════════════════
#  v15.59 · RESULTADO POR REMITO
# ═══════════════════════════════════════════════════════════
# Port 1:1 del motor calc() del prototipo v2.5 validado por el usuario
# (Claude_Outputs\Scripts_Auxiliares\modulo_resultado_remito\).
RR_DESDE          = "2026-07-01"   # alcance: ventas con remito desde acá
RR_PV_MIN         = 2.0            # límites de negocio del % consumo MS…
RR_PV_MAX         = 3.0            # …los meses fuera se acotan al límite
RR_FACTOR_VACA    = 1.30           # la vaca come +30% (decisión de prudencia)
RR_COMISION_DEF   = 0.03           # fallback si el Excel no trae comisión
# Categoría → grupo de mortandad de muertes_2025.json
RR_GRUPO_MORT = {
    "Vaca": "Vacas", "Toro": "Machos", "Novillo": "Machos",
    "Novillito": "Machos", "Ternero": "Machos",
    "Vaquillona": "Hembras", "Ternera": "Hembras",
}
# ⚠ La API devuelve la categoría como código de 2 letras (v15.5.1) y el motivo
# como 1 letra (V=venta, T=traslado, M=muerte) — NO como los strings largos del
# SQL viejo. Se mapea al nombre largo del Excel de compras (_CAT_COMPRAS).
RR_CAT_CODE = {
    "TM": "Ternero", "TH": "Ternera", "NT": "Novillito", "NV": "Novillo",
    "VQ": "Vaquillona", "VA": "Vaca", "TO": "Toro",
}


def _rr_cat(c):
    k = str(c or "").strip().upper()
    if k in RR_CAT_CODE:
        return RR_CAT_CODE[k]
    return _CAT_COMPRAS.get(_norm_cat_compra(k), k.title())


# ── v15.68.2 · Cruce con Datamars por CARAVANA (la lectura real del bastón) ──
# Cuando la caravana no lee o no coincide, el cargador de WinCampo le asigna al
# egreso un animal cualquiera del stock: el remito hereda kg de ingreso, precio,
# fecha y estadía de OTRO animal (remito 2274: 50 días de estadía en WinCampo
# contra 11 reales).
#
# v15.68.2 — Criterio de Nicolás, apoyado en dos hechos del negocio: las
# caravanas NO se repiten y un animal que salió no vuelve a salir. Entonces no
# hace falta adivinar qué sesión de balanza corresponde a cada remito: alcanza
# con preguntar si esa caravana fue leída ALGUNA VEZ en los últimos 30 días,
# en cualquier sesión (ingresos incluidos). El match remito↔sesión de la primera
# versión era demasiado estricto para la realidad de la balanza (remitos
# cruzados o combinados en una sesión, sesiones de ingreso el mismo día) y
# dejaba 38 de 60 remitos con el control en rojo.
RR_DM_VENTANA_DIAS  = 30     # se busca la caravana en [fe − 30, fe + 1]
RR_DM_DIAS_ADELANTE = 1
RR_DM_DIAS_SALIDA   = 3      # lectura a ±3 días del egreso = la pesada de salida

# v15.68.4 · El filtro por sexo se DESESTIMÓ (decisión de Nicolás). Datamars no
# carga `Sexo` (vacío en 6.558 de 6.559 pesadas) y la Categoria viene nula en el
# 30 % de las pesadas: sobre esa base el filtro no descartaba nada (sc_por_sexo
# dio 0 en los 64 remitos) y solo agregaba una vía de falsos negativos. La
# confirmación es por PRESENCIA de la caravana en la ventana, y nada más.
#
# Tampoco hay lógica de caravana reutilizada: desde 2026 la caravana es la
# electrónica oficial y el EID es único de por vida.


def _rr_rfid(v):
    """RFID / EID → solo dígitos. Vacío → None."""
    d = re.sub(r"\D", "", str(v or ""))
    return d or None


# v15.68.5 · Regla de Nicolás: SOLO la caravana electrónica se cruza contra la
# balanza. El código visual y el placeholder que pone el cargador cuando no hay
# caravana son "sin caravana" directo, sin consultar Datamars.
#
#   032010036311450  → electrónica (15 dígitos, sin letras)   → se busca
#   HI215A238        → visual (tiene letras)                  → SC directo
#   13052026-001     → placeholder del cargador (ddmmyyyy-NNN)→ SC directo
#   (vacío)          → SC directo
#
# Todos los EID de Datamars son de 15 dígitos (verificado sobre 5.404 distintos),
# así que las no electrónicas no matchearían igual: el desglose es informativo y
# no mueve el conteo de SC.
_RR_PLACEHOLDER = re.compile(r"^\d{6,8}-\d+[A-Za-z]*$")


def _rr_caravana_tipo(raw):
    """Tipo de caravana de WinCampo. Solo 'electronica' se cruza con Datamars."""
    t = str(raw or "").strip()
    if not t:
        return "vacia"
    if _RR_PLACEHOLDER.match(t):
        return "placeholder"
    if re.search(r"[A-Za-z]", t):
        return "visual"
    if len(re.sub(r"\D", "", t)) == 15:
        return "electronica"
    # ni 15 dígitos ni letras ni placeholder: 3 casos en el histórico
    # (un '9820004543431100' de 16 dígitos y dos de 8). Tampoco se cruzan.
    return "otra"


def _rr_moda(vals):
    """Valor más repetido, determinista ante empate (orden por str)."""
    vals = [v for v in vals if v is not None]
    if not vals:
        return None
    return sorted(set(vals), key=lambda v: (-vals.count(v), str(v)))[0]


def _rr_datamars(carpeta_datos, filas_rem, por_tropa, log):
    """v15.68.2 · Marca e imputa los animales "sin caravana" de cada remito.

    Para cada fila de egreso se busca su RFID en TODAS las pesadas de Datamars
    con fecha en [fecha_egreso − RR_DM_VENTANA_DIAS, fecha_egreso + 1], sin
    importar de qué sesión vengan (ingresos incluidos). Si aparece, la fila está
    CONFIRMADA — presencia y nada más. Si no aparece es SIN CARAVANA (SC) y se
    le imputa el origen de la tropa mayoritaria confirmada del remito.

    De cada confirmada se anota, informativo, si la lectura elegida es la pesada
    de SALIDA (a ±RR_DM_DIAS_SALIDA días del egreso) o una lectura ANTERIOR
    (típicamente su ingreso), y si esa caravana tuvo más de una lectura en la
    ventana.

    MUTA `filas_rem` in place. El kg de EGRESO nunca se toca: es el real
    (Datamars y WinCampo coinciden en el peso).

    Criterio de Nicolás: si no hubo balanza en la ventana, vale lo cargado en
    WinCampo — cero cambios, marcado "sin verificar".

    Devuelve (verificacion_por_remito, meta_datamars). Nunca levanta.
    """
    verif = {rem: {"estado": "sin_datamars", "sc": 0} for rem in filas_rem}
    meta = {"activo": False, "motivo": None, "ventana_dias": RR_DM_VENTANA_DIAS,
            "sesiones_cache": 0, "ultima_sesion": None,
            "remitos_verificados": 0, "remitos_sin_datamars": 0,
            "confirmadas_total": 0, "sc_total": 0,
            "sc_sin_electronica_total": 0, "sc_no_leida_total": 0,
            "confirmadas_salida_total": 0, "confirmadas_anterior_total": 0,
            "dobles_lectura_total": 0,
            "cab_total": sum(len(f) for f in filas_rem.values())}
    if not filas_rem:
        return verif, meta

    try:
        import datamars_source as _DM
    except Exception as e:
        meta["motivo"] = f"import falló: {type(e).__name__}"
        log.warning(f"  ⚠ datamars_source no importable ({e}) — remitos sin verificar")
        return verif, meta

    info = {}
    for rem, filas in filas_rem.items():
        info[rem] = {"fe": max(x["fe"] for x in filas)}

    try:
        sesiones, m = _DM.sincronizar(
            carpeta_datos, [i["fe"] for i in info.values()], log=log,
            dias_atras=RR_DM_VENTANA_DIAS, dias_adelante=RR_DM_DIAS_ADELANTE)
    except Exception as e:
        meta["motivo"] = f"sincronizar falló: {type(e).__name__}"
        log.warning(f"  ⚠ Datamars: {e} — remitos sin verificar")
        return verif, meta

    meta["motivo"] = m.get("motivo")
    meta["sesiones_cache"] = m.get("sesiones_cache", 0)
    meta["ultima_sesion"] = m.get("ultima_sesion")
    if not sesiones:
        meta["remitos_sin_datamars"] = len(filas_rem)
        return verif, meta
    meta["activo"] = True

    # ── Índice global de lecturas: EID → [(fecha, pesada)] ──
    # Una sola pasada por toda la cache; después cada fila es un lookup.
    lecturas, sin_eid_fechas = {}, []
    for s in sesiones:
        for p in (s.get("pesadas") or []):
            f = _DM._fecha(p.get("fecha")) or s["fecha"]
            e = p.get("eid")
            if e:
                lecturas.setdefault(e, []).append((f, p))
            else:
                sin_eid_fechas.append(f)
    # segundo índice sin ceros a la izquierda, por si WinCampo o Datamars los recortan
    lecturas_slz = {}
    for e, v in lecturas.items():
        k = e.lstrip("0")
        if k != e:
            lecturas_slz.setdefault(k, []).extend(v)
    todas_fechas = sorted({f for v in lecturas.values() for f, _ in v} | set(sin_eid_fechas))

    for rem, filas in filas_rem.items():
        fe = info[rem]["fe"]
        desde = fe - _td(days=RR_DM_VENTANA_DIAS)
        hasta = fe + _td(days=RR_DM_DIAS_ADELANTE)
        if not any(desde <= f <= hasta for f in todas_fechas):
            # no hubo balanza en la ventana → el remito vale como WinCampo
            verif[rem] = {"estado": "sin_datamars", "sc": 0,
                          "ventana_dias": RR_DM_VENTANA_DIAS,
                          "desde": desde.isoformat(), "hasta": hasta.isoformat()}
            meta["remitos_sin_datamars"] += 1
            continue

        conf, sc = [], []
        dias_dm = []
        n_salida = n_anterior = n_doble = 0
        n_sin_elec = n_no_leida = 0
        for x in filas:
            # visual, placeholder o vacía: sin caravana directo, sin consultar
            # Datamars. Solo la electrónica se cruza.
            if x["car_tipo"] != "electronica":
                x["sc_tipo"] = "sin_electronica"
                n_sin_elec += 1
                sc.append(x)
                continue
            r = x["rfid"]
            cand = (lecturas.get(r) or []) if r else []
            if r and not cand:
                cand = lecturas_slz.get(r.lstrip("0")) or []
            enventana = [(f, p) for f, p in cand if desde <= f <= hasta]
            if not enventana:
                x["sc_tipo"] = "no_leida"
                n_no_leida += 1
                sc.append(x)
                continue
            # la lectura más cercana a la fecha de egreso
            f, p = min(enventana, key=lambda t: abs((t[0] - fe).days))
            x["dm_fecha"] = f
            x["dm_dias"] = p.get("dias_datamars")
            x["dm_gpv"] = p.get("gpv_datamars")
            # ¿es la pesada de salida o una lectura anterior (su ingreso)?
            x["dm_tipo"] = ("salida" if abs((f - fe).days) <= RR_DM_DIAS_SALIDA
                            else "anterior")
            if x["dm_tipo"] == "salida":
                n_salida += 1
            else:
                n_anterior += 1
            # más de un día con lectura en la ventana. No es un error: con
            # estadías cortas el ingreso y la salida caen las dos adentro.
            x["dm_doble"] = len({d for d, _ in enventana}) > 1
            if x["dm_doble"]:
                n_doble += 1
            if p.get("dias_datamars"):
                dias_dm.append(p["dias_datamars"])
            conf.append(x)

        kge_tot = sum(x["kge"] for x in filas) or 0.0
        kge_sc = sum(x["kge"] for x in sc)
        n_sc = len(sc)
        bloque = {
            "estado": "verificado",
            "ventana_dias": RR_DM_VENTANA_DIAS,
            "desde": desde.isoformat(), "hasta": hasta.isoformat(),
            "confirmadas": len(conf), "sc": n_sc,
            # desglose de los sin caravana: los que WinCampo no tiene con
            # caravana electrónica, y los que sí la tienen pero el bastón no la
            # leyó. Suman `sc`.
            "sc_sin_electronica": n_sin_elec, "sc_no_leida": n_no_leida,
            "sc_pct_cab": round(n_sc / len(filas) * 100, 1) if filas else None,
            "sc_pct_kg": round(kge_sc / kge_tot * 100, 1) if kge_tot else None,
            # informativos: de las confirmadas, cuántas por su pesada de salida
            # y cuántas por una lectura anterior; y cuántas caravanas tuvieron
            # más de un día de lectura en la ventana.
            "confirmadas_salida": n_salida, "confirmadas_anterior": n_anterior,
            "dobles_lectura": n_doble,
            "tropa_imputada": None, "imputado": None,
            "pesadas_sin_eid_ventana": sum(1 for f in sin_eid_fechas if desde <= f <= hasta),
            "dias_datamars_prom": (round(sum(dias_dm) / len(dias_dm), 1) if dias_dm else None),
        }
        meta["remitos_verificados"] += 1
        meta["confirmadas_total"] += len(conf)
        meta["sc_total"] += n_sc
        meta["sc_sin_electronica_total"] += n_sin_elec
        meta["sc_no_leida_total"] += n_no_leida
        meta["confirmadas_salida_total"] += n_salida
        meta["confirmadas_anterior_total"] += n_anterior
        meta["dobles_lectura_total"] += n_doble

        for x in conf:
            x["origen"] = "confirmado"
        if not n_sc:
            verif[rem] = bloque
            continue
        if not conf:
            # ninguna caravana leída: no hay de dónde imputar, vale WinCampo.
            # El frontend lo pinta en rojo para que Nicolás cargue el origen.
            verif[rem] = bloque
            continue

        # tropa mayoritaria entre las CONFIRMADAS; empate → la que pesa parecido
        por_t = {}
        for x in conf:
            por_t.setdefault(str(x["tropa"] or ""), []).append(x)
        maxc = max(len(v) for v in por_t.values())
        cands = sorted([t for t, v in por_t.items() if len(v) == maxc])
        if len(cands) > 1:
            kg_sc = kge_sc / n_sc
            cands.sort(key=lambda t: abs(sum(y["kge"] for y in por_t[t]) / len(por_t[t]) - kg_sc))
        ref = por_t[cands[0]]

        cat_imp = _rr_moda([y["cat"] for y in ref])
        fi_imp = _rr_moda([y["fi"] for y in ref])
        corral_imp = _rr_moda([y["corral"] for y in ref])
        hot_imp = _rr_moda([y["hotelero"] for y in ref])
        tropa_imp = ref[0]["tropa"]

        # kg de ingreso por cabeza: promedio de las CONFIRMADAS de esa tropa y
        # categoría en este mismo remito.
        #
        # ⚠ El prompt pedía el promedio de la tropa completa del Excel de
        # compras (kg ÷ cabezas), pero ese número mezcla categorías: en el
        # remito 2274 la tropa PEG.UTE.23/07/2026 tiene 154 cabezas y 40.819 kg
        # entre novillos, vaquillonas y toros → 265 kg/cab, contra los 669 kg
        # reales de los 16 toros confirmados. Con 265 el ADP del remito daba
        # 17,6 kg/día. El propio ejemplo del prompt (kg_ingreso_cab = 668,9) es
        # el promedio de las confirmadas, así que se toma ese y el promedio del
        # Excel queda de respaldo para cuando no hay confirmadas con kg.
        t_ex = por_tropa.get(_norm_tropa(tropa_imp)) or {}
        _k = [y["kgi"] for y in ref if y["kgi"] and y["cat"] == cat_imp]
        if not _k:
            _k = [y["kgi"] for y in ref if y["kgi"]]
        kg_cab = (sum(_k) / len(_k)) if _k else None
        if not kg_cab and t_ex.get("kg") and t_ex.get("cabezas"):
            kg_cab = t_ex["kg"] / t_ex["cabezas"]

        for x in sc:
            x["tropa"] = tropa_imp
            if cat_imp:
                x["cat"] = cat_imp
            if fi_imp:
                x["fi"] = fi_imp
            x["corral"] = corral_imp
            x["hotelero"] = hot_imp
            if kg_cab:
                x["kgi"] = kg_cab
            x["origen"] = "imputado"

        _pc = (t_ex.get("por_categoria") or {}).get(cat_imp) or {}
        bloque["tropa_imputada"] = tropa_imp
        bloque["imputado"] = {
            "kg_ingreso_cab": round(kg_cab, 1) if kg_cab else None,
            "precio_kg": _pc.get("precio_kg") or t_ex.get("precio_kg"),
            "fecha_ingreso": fi_imp.isoformat() if fi_imp else None,
            "dias": ((fe - fi_imp).days if fi_imp else None),
        }
        verif[rem] = bloque

    return verif, meta


# ── v15.69 · CARAVANAS FANTASMA ─────────────────────────────────────
# El otro lado del problema de las caravanas inventadas. Cuando una vaca sale y
# su caravana no lee, el cargador pone en el remito otra del mismo dueño y
# categoría — eso está bien y así queda. Pero la que salió DE VERDAD sigue
# figurando en el stock de WinCampo con su caravana electrónica, y el bastón sí
# la leyó ese día. Esa es la "fantasma": está en el campo solo en los papeles.
#
# Para que pueda salir en la próxima venta como "sin caravana", el cargador
# tiene que renombrarla a un placeholder (ddmmyy-N). Este informe le dice
# exactamente cuáles son y qué placeholder ponerles. Cuando las renombra, al
# tick siguiente ya no están en stock con ese EID y desaparecen solas de la
# lista: no hay nada que marcar a mano.
RR_FANT_MIN_DIAS = 7   # si ingresó hace <= 7 días, la lectura es su pesada de ingreso


def generar_fantasmas(carpeta_out, periodo, egresos_data, log=None, stock_data=None):
    """v15.69: caravanas leídas en una salida que siguen en el stock de WinCampo.

    Una caravana es fantasma cuando se cumplen las cuatro:
      1. es un EID de 15 dígitos leído en una sesión de Datamars cuya fecha
         coincide con la fecha de egreso de algún remito de venta;
      2. sigue en el stock actual de WinCampo;
      3. su FECHA_INGRESO es anterior a la lectura por más de RR_FANT_MIN_DIAS
         (así la pesada de ingreso del día no cuenta);
      4. no está en ningún remito de venta de ese día (si está, es una
         confirmada normal).

    Devuelve el dict volcado o None. NO levanta excepción.
    """
    if log is None:
        log = logging.getLogger("fantasmas")
    base = Path(carpeta_out)

    def _d(v):
        if not v:
            return None
        t = str(v)[:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(t, fmt).date()
            except ValueError:
                continue
        return None

    # ── stock actual ──
    stock_gen = None
    filas_stock = stock_data
    if not filas_stock:
        try:
            with (base / f"stock_detalle_{periodo}.json").open(encoding="utf-8") as f:
                _sd = json.load(f)
            filas_stock = _sd.get("detalle") or []
            stock_gen = (_sd.get("meta") or {}).get("generado")
        except Exception as e:
            log.warning(f"  ⚠ Fantasmas: sin stock_detalle ({e}), se saltea")
            return None
    if not filas_stock:
        log.warning("  ⚠ Fantasmas: stock vacío, se saltea")
        return None

    por_rfid = {}
    for r in filas_stock:
        d = _rr_rfid(r.get("RFID"))
        if d:
            por_rfid.setdefault(d, r)
            por_rfid.setdefault(d.lstrip("0"), r)

    # ── remitos de venta por fecha (los mismos que el módulo) ──
    desde = _d(RR_DESDE)
    rem_dia, rfid_dia = {}, {}
    for e in (egresos_data or []):
        motivo = str(e.get("MotivoSalida") or "").strip().upper()
        if not (motivo == "V" or "VENTA" in motivo):
            continue
        fe = _d(e.get("FechaSalida"))
        if fe is None or fe < desde:
            continue
        rem = str(e.get("NRO_TRANSACCION") or "").strip()
        if not rem:
            continue
        rem_dia.setdefault(fe, {}).setdefault(rem, e.get("Destino") or e.get("Consignatario"))
        d = _rr_rfid(e.get("RFID"))
        if d:
            rfid_dia.setdefault(fe, set()).add(d)
            rfid_dia[fe].add(d.lstrip("0"))
    if not rem_dia:
        log.info("  Fantasmas: sin remitos de venta, nada que revisar")
        return None

    # ── sesiones de la cache cuya fecha cae en un día de remito ──
    try:
        import datamars_source as _DM
    except Exception as e:
        log.warning(f"  ⚠ Fantasmas: datamars_source no importable ({e})")
        return None
    index = _DM.cargar_index(carpeta_out)
    if not index:
        log.info("  Fantasmas: cache de Datamars vacía, se saltea")
        return None

    fant = []
    for sid, ses in index.items():
        f = _d(ses.get("fecha"))
        if f is None or f not in rem_dia:
            continue
        pesadas = _DM.cargar_sesion(carpeta_out, sid)
        if not pesadas:
            continue
        en_remito = rfid_dia.get(f) or set()
        for p in pesadas:
            eid = p.get("eid")
            # solo la caravana electrónica: los EID de Datamars son de 15 dígitos
            if not eid or len(eid) != 15:
                continue
            if eid in en_remito or eid.lstrip("0") in en_remito:
                continue                      # salió y está en el remito: normal
            st = por_rfid.get(eid) or por_rfid.get(eid.lstrip("0"))
            if st is None:
                continue                      # ya no está en stock: nada que hacer
            fi = _d(st.get("FECHA_INGRESO"))
            if fi is None or (f - fi).days <= RR_FANT_MIN_DIAS:
                continue                      # es su propia pesada de ingreso
            fant.append({
                "eid": eid,
                "hotelero": st.get("HOTELERO"),
                "tropa": st.get("NRO_TROPA"),
                "categoria": st.get("CLASIFICACION") or _rr_cat(st.get("CATEGORIA")),
                "corral": st.get("NRO_CORRAL"),
                "caravana_visual": st.get("NRO_CARAVANA"),
                "kg_ingreso": st.get("KG_INGRESO"),
                "fecha_ingreso": fi.isoformat(),
                "dias_en_stock": (f - fi).days,
                "fecha_lectura": f.isoformat(),
                "sesion_id": int(sid),
                "sesion_nombre": ses.get("nombre"),
                "peso_lectura": p.get("peso"),
                "remitos_del_dia": [{"remito": r, "comprador": c}
                                    for r, c in sorted(rem_dia[f].items())],
            })

    # orden pedido: fecha de lectura desc, hotelero, tropa
    fant.sort(key=lambda x: (x["fecha_lectura"], str(x["hotelero"] or ""),
                             str(x["tropa"] or ""), x["eid"]), reverse=False)
    fant.sort(key=lambda x: x["fecha_lectura"], reverse=True)

    # placeholder sugerido: ddmmyy-N, correlativo dentro del día, en el mismo
    # orden en que se muestran (así el cargador los va poniendo de arriba abajo)
    _n = {}
    for x in fant:
        f = _d(x["fecha_lectura"])
        _n[x["fecha_lectura"]] = _n.get(x["fecha_lectura"], 0) + 1
        x["placeholder_sugerido"] = f"{f.strftime('%d%m%y')}-{_n[x['fecha_lectura']]}"

    por_hot, por_cat, por_fecha = {}, {}, {}
    for x in fant:
        por_hot[x["hotelero"] or "—"] = por_hot.get(x["hotelero"] or "—", 0) + 1
        por_cat[x["categoria"] or "—"] = por_cat.get(x["categoria"] or "—", 0) + 1
        por_fecha[x["fecha_lectura"]] = por_fecha.get(x["fecha_lectura"], 0) + 1

    salida = {
        "meta": {
            "generado": datetime.now().isoformat(),
            "total": len(fant),
            "dias_min": RR_FANT_MIN_DIAS,
            "por_hotelero": dict(sorted(por_hot.items(), key=lambda kv: -kv[1])),
            "por_categoria": dict(sorted(por_cat.items(), key=lambda kv: -kv[1])),
            "por_fecha": dict(sorted(por_fecha.items(), reverse=True)),
            "stock_generado": stock_gen,
            "criterio": ("EID de 15 dígitos leído por el bastón el mismo día que un remito de "
                         "venta, que sigue en el stock de WinCampo, ingresó hace más de "
                         f"{RR_FANT_MIN_DIAS} días y no figura en ningún remito de ese día"),
        },
        "fantasmas": fant,
    }
    guardar(salida, carpeta_out, "fantasmas.json")
    _det = " · ".join(f"{h} {n}" for h, n in list(salida["meta"]["por_hotelero"].items())[:4])
    log.info(f"  ✓ Fantasmas: {len(fant)} caravanas leídas en salida que siguen en stock ({_det})")
    return salida


def generar_resultado_remitos(carpeta_out, periodo, egresos_data, log=None):
    """v15.59: resultado económico por remito de venta.

    Costo = compra (Excel de compras, por tropa+categoría) + comisión (fila por
    fila) + alimento (%PV mensual real × peso interpolado × $/kg MS del mes) +
    estructura ($/día-animal) + sanidad (mes de ingreso) + mortandad (tasa del
    portal × costo de compra). La VENTA no se calcula acá: se carga en el portal.

    Lee todo de disco salvo los egresos, que ya vienen del pre-step de main
    (no se re-consulta la API).

    Devuelve el dict volcado o None. NO levanta excepción.
    """
    if log is None:
        log = logging.getLogger("remitos")
    from datetime import date as _date, timedelta as _td

    base = Path(carpeta_out)

    def _load(nombre):
        try:
            with (base / nombre).open(encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log.warning(f"  ⚠ no pude leer {nombre}: {e}")
            return None

    _compras = _load("precios_compra_real.json") or {}
    por_tropa = _compras.get("por_tropa") or {}
    cat_mes   = _compras.get("por_categoria_mes") or {}
    if not por_tropa:
        log.warning("  ⚠ sin índice de compras por tropa, saltando resultado por remito")
        return None

    _pctpv = _load("pct_pv_mensual.json") or {}
    PCTPV  = {m: v.get("pct_pv_ajustado") for m, v in (_pctpv.get("meses") or {}).items()
              if v.get("pct_pv_ajustado")}
    if not PCTPV:
        log.warning("  ⚠ sin pct_pv_mensual, saltando resultado por remito")
        return None
    PV_FALLBACK = round(sum(PCTPV.values()) / len(PCTPV), 2)

    PRECIOS = procesar_precios_racion(carpeta_out, log)
    if not PRECIOS:
        log.warning("  ⚠ sin precios de ración, saltando resultado por remito")
        return None
    _meses_prec = sorted(PRECIOS)

    # Tasas de mortandad del portal (las mismas que muestra el módulo).
    _muertes = _load(f"muertes_{periodo}.json") or {}
    _grupos  = ((_muertes.get("mortandad") or {}).get("por_grupo")) or {}
    MORT_PCT = {c: (_grupos.get(g, {}).get("tasa_mensual_pct") or 0.0)
                for c, g in RR_GRUPO_MORT.items()}

    def _mk(ym):
        """Mes de precios más cercano disponible (acota a los extremos)."""
        if ym in PRECIOS:
            return ym
        return _meses_prec[0] if ym < _meses_prec[0] else _meses_prec[-1]

    def _d(s):
        """Fecha ISO o dd/mm/yyyy → date."""
        if not s:
            return None
        s = str(s)[:10]
        for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    # ── Filtrar egresos: solo VENTA, con remito, desde RR_DESDE ──
    # v15.68: primero se juntan las filas INDIVIDUALES (1 por animal, con su
    # RFID). El agrupado por (remito, tropa, cat, corral, fi, fe) se hace
    # DESPUÉS del cruce con Datamars, porque la imputación de los "sin caravana"
    # les cambia tropa, categoría y fecha de ingreso.
    desde = _d(RR_DESDE)
    filas_rem = {}
    n_venta = 0
    for e in (egresos_data or []):
        motivo = str(e.get("MotivoSalida") or "").strip().upper()
        # La API manda 'V'; el SQL viejo mandaba 'VENTA ...'. Se toleran los dos
        # igual que filtrar_solo_venta() de procesar_productivo.
        if not (motivo == "V" or "VENTA" in motivo):
            continue          # traslados (T) y muertes (M) quedan afuera
        fe = _d(e.get("FechaSalida"))
        if fe is None or fe < desde:
            continue
        rem = str(e.get("NRO_TRANSACCION") or "").strip()
        if not rem:
            continue
        n_venta += 1
        filas_rem.setdefault(rem, []).append({
            "tropa": e.get("NRO_TROPA"), "cat": _rr_cat(e.get("Categoria")),
            "corral": e.get("NRO_CORRAL"), "hotelero": e.get("HOTELERO"),
            "comprador": e.get("Destino") or e.get("Consignatario"),
            "fi": _d(e.get("FechaIngreso")), "fe": fe,
            "kgi": float(e.get("KgIngreso") or 0),
            "kge": float(e.get("KgEgreso") or 0),
            "estadia": e.get("Estadia"), "rfid": _rr_rfid(e.get("RFID")),
            "car_tipo": _rr_caravana_tipo(e.get("RFID")),
            "origen": "sin_verificar",
        })

    # ── v15.68 · Cruce con la lectura real del bastón ──
    # Sin credenciales / sin sesión del día, `verificacion` queda en
    # sin_datamars / sin_sesion y NADA cambia: el remito vale como WinCampo.
    verif, meta_dm = _rr_datamars(base, filas_rem, por_tropa, log)

    # ── Agrupado (el origen entra en la clave: las filas imputadas se ven
    #    separadas de las confirmadas de la misma tropa) ──
    grupos = {}
    for rem, _filas in filas_rem.items():
        for x in _filas:
            k = (rem, str(x["tropa"] or ""), x["cat"], str(x["corral"] or ""),
                 x["fi"], x["fe"], x["origen"] == "imputado")
            g = grupos.setdefault(k, {
                "remito": rem, "tropa": x["tropa"], "cat": x["cat"],
                "corral": x["corral"], "fi": x["fi"], "fe": x["fe"],
                "hotelero": x["hotelero"], "comprador": x["comprador"],
                "origen": x["origen"], "imputado": x["origen"] == "imputado",
                "cab": 0, "kgi": 0.0, "kge": 0.0, "dias": x["estadia"],
                # v15.68.2: contraste independiente de WinCampo — estadía y
                # ganancia que calcula Datamars desde la primera pesada del EID.
                "dm_dias_l": [], "dm_gpv_l": [], "dm_fecha": None,
                "dm_tipos": set(), "dm_dobles": 0,
                "car_tipos": set(), "sc_tipos": set(),
            })
            g["cab"] += 1
            g["kgi"] += x["kgi"]
            g["kge"] += x["kge"]
            if x.get("dm_dias") is not None:
                g["dm_dias_l"].append(x["dm_dias"])
            if x.get("dm_gpv") is not None:
                g["dm_gpv_l"].append(x["dm_gpv"])
            if x.get("dm_fecha") and (g["dm_fecha"] is None or x["dm_fecha"] > g["dm_fecha"]):
                g["dm_fecha"] = x["dm_fecha"]
            if x.get("dm_tipo"):
                g["dm_tipos"].add(x["dm_tipo"])
            if x.get("dm_doble"):
                g["dm_dobles"] += 1
            g["car_tipos"].add(x["car_tipo"])
            if x.get("sc_tipo"):
                g["sc_tipos"].add(x["sc_tipo"])

    if not grupos:
        log.warning(f"  ⚠ sin egresos de venta con remito desde {RR_DESDE}")
        return None

    # ── Precio de reposición por categoría: promedio ponderado del ÚLTIMO mes
    #    con compras de esa categoría (del Excel de compras).
    repo_cat = {}
    for c, meses in cat_mes.items():
        if meses:
            _u = max(meses)
            repo_cat[c] = {"precio_kg": meses[_u]["precio_kg"], "mes": _u}
    _ultp = _meses_prec[-1]
    repo_ms = PRECIOS[_ultp]["tc"] / PRECIOS[_ultp]["ms"]

    # ── Cálculo por remito ────────────────────────────────────
    remitos_out = {}
    sin_precio_global, con_precio_global = set(), set()
    for rem in sorted({g["remito"] for g in grupos.values()}):
        filas_g = [g for g in grupos.values() if g["remito"] == rem]

        # Precio de cada fila: tropa+categoría; si no está, promedio ponderado
        # de las compañeras del MISMO remito (y se marca estimado).
        def _precio(g):
            t = por_tropa.get(_norm_tropa(g["tropa"]))
            if not t:
                return None, None
            pc = (t.get("por_categoria") or {}).get(g["cat"])
            com = t.get("comision")
            if pc and pc.get("precio_kg"):
                return pc["precio_kg"], com
            # tropa conocida pero sin esa categoría → promedio de la tropa
            return t.get("precio_kg"), com

        kg_con, imp_con = 0.0, 0.0
        for g in filas_g:
            p, _ = _precio(g)
            if p:
                kg_con += g["kgi"]
                imp_con += g["kgi"] * p
        prom = imp_con / kg_con if kg_con else 0.0

        compra = comision = ali = est = san = mort = 0.0
        cab = kgi = kge = ms_tot = pv_den = a_dias = kg_sin = kgi_mort = 0.0
        n_sin = 0
        clamped, sin_pv = {}, 0
        filas_out, tropas_sin = [], []

        for g in sorted(filas_g, key=lambda x: (-x["kgi"],)):
            p_real, com_tropa = _precio(g)
            estimado = p_real is None
            p = p_real if p_real else prom
            com_pct = com_tropa if com_tropa is not None else RR_COMISION_DEF
            c = g["kgi"] * p
            kgp = (g["kgi"] + g["kge"]) / 2
            fcat = RR_FACTOR_VACA if str(g["cat"]).strip().lower() == "vaca" else 1.0
            mpct = (MORT_PCT.get(g["cat"], 0.0)) / 100

            fi, fe = g["fi"], g["fe"]
            total_d = max((fe - fi).days, 1)
            a = s = ms = 0.0
            dtot = elap = 0
            lim = False
            cur = fi
            while cur < fe:
                ym = f"{cur.year:04d}-{cur.month:02d}"
                pp = PRECIOS[_mk(ym)]
                pv_raw = PCTPV.get(ym)
                if pv_raw is None:
                    pv_raw = PV_FALLBACK
                    sin_pv += 1
                pv_c = min(RR_PV_MAX, max(RR_PV_MIN, pv_raw))
                if pv_c != pv_raw:
                    clamped[ym] = pv_raw
                    lim = True
                pv = pv_c / 100 * fcat          # el factor va DESPUÉS del límite
                nx = _date(cur.year + (cur.month // 12), (cur.month % 12) + 1, 1)
                if nx > fe:
                    nx = fe
                dd = (nx - cur).days
                # peso interpolado: el animal va de kgi a kge linealmente y cada
                # tramo usa el peso en su punto medio
                kgm = g["kgi"] + (g["kge"] - g["kgi"]) * ((elap + dd / 2) / total_d)
                m = kgm * pv * dd
                ms += m
                a  += m * (pp["tc"] / pp["ms"])
                s  += g["cab"] * dd * pp["dia"]
                dtot += dd
                elap += dd
                cur = nx

            # sanidad: ÚNICA por cabeza, al precio del MES DE INGRESO
            sa = g["cab"] * PRECIOS[_mk(f"{fi.year:04d}-{fi.month:02d}")]["san"]

            compra += c
            comision += c * com_pct
            ali += a
            est += s
            san += sa
            mort += c * mpct
            kgi_mort += g["kgi"] * mpct
            cab += g["cab"]
            kgi += g["kgi"]
            kge += g["kge"]
            ms_tot += ms
            pv_den += kgp * dtot
            a_dias += g["cab"] * dtot
            if estimado:
                kg_sin += g["kgi"]
                n_sin += 1
                tropas_sin.append({
                    "tropa": g["tropa"], "categoria": g["cat"], "cabezas": g["cab"],
                    "kg_ingreso": round(g["kgi"], 1),
                    "fecha_ingreso": fi.isoformat(), "estimado_a": round(p, 2),
                })
                sin_precio_global.add(g["tropa"])
            else:
                con_precio_global.add(g["tropa"])

            filas_out.append({
                # v15.68: 'confirmado' = el bastón leyó su caravana; 'imputado' =
                # sin caravana, origen imputado a la tropa mayoritaria del
                # remito; 'sin_verificar' = no hubo sesión de Datamars.
                "origen": g["origen"], "imputado": g["imputado"],
                "fecha_lectura_datamars": (g["dm_fecha"].isoformat() if g["dm_fecha"] else None),
                "tipo_lectura": (None if not g["dm_tipos"]
                                 else (list(g["dm_tipos"])[0] if len(g["dm_tipos"]) == 1 else "mixto")),
                "caravana_tipo": (None if not g["car_tipos"]
                                  else (list(g["car_tipos"])[0] if len(g["car_tipos"]) == 1 else "mixto")),
                "sc_tipo": (None if not g["sc_tipos"]
                            else (list(g["sc_tipos"])[0] if len(g["sc_tipos"]) == 1 else "mixto")),
                "dobles_lectura": g["dm_dobles"],
                "dias_datamars": (round(sum(g["dm_dias_l"]) / len(g["dm_dias_l"]), 1)
                                  if g["dm_dias_l"] else None),
                "gpv_datamars": (round(sum(g["dm_gpv_l"]) / len(g["dm_gpv_l"]), 1)
                                 if g["dm_gpv_l"] else None),
                "tropa": g["tropa"], "categoria": g["cat"], "corral": g["corral"],
                "hotelero": g["hotelero"], "comprador": g["comprador"],
                "cabezas": g["cab"],
                "fecha_ingreso": fi.isoformat(), "fecha_egreso": fe.isoformat(),
                "dias": dtot,
                "kg_ingreso": round(g["kgi"], 1), "kg_egreso": round(g["kge"], 1),
                "precio_kg": round(p, 2), "estimado": estimado,
                "comision_pct": round(com_pct * 100, 2),
                "costo_compra": round(c, 2), "kg_ms": round(ms, 1),
                "pct_ms": round(ms / (kgp * dtot) * 100, 2) if dtot and kgp else None,
                "acotado": lim,
                "alimento": round(a, 2), "estructura": round(s, 2), "sanidad": round(sa, 2),
                "mortandad": round(c * mpct, 2),
            })

        costo = compra + comision + ali + est + san + mort
        kg_prod = kge - kgi

        # Reposición: mismos kg de entrada y mismos kg MS a precio de hoy.
        cats = {f["categoria"] for f in filas_out}
        if len(cats) == 1 and list(cats)[0] in repo_cat:
            _rc = repo_cat[list(cats)[0]]
            rp, rp_lbl = _rc["precio_kg"], f"prom. {list(cats)[0]} {_rc['mes']}"
        else:
            rp, rp_lbl = prom, "prom. histórico del remito"
        compra_repo = kgi * rp
        com_repo = compra_repo * RR_COMISION_DEF
        ali_repo = ms_tot * repo_ms
        mort_repo = kgi_mort * rp
        costo_repo = compra_repo + com_repo + ali_repo + mort_repo + est + san

        remitos_out[rem] = {
            "filas": filas_out,
            "cabezas": int(cab), "tropas": len(filas_out),
            "kg_ingreso": round(kgi, 1), "kg_egreso": round(kge, 1),
            "kg_producidos": round(kg_prod, 1), "kg_ms": round(ms_tot, 1),
            "fecha_egreso": max(f["fecha_egreso"] for f in filas_out),
            "comprador": next((f["comprador"] for f in filas_out if f["comprador"]), None),
            "costos": {
                "compra": round(compra, 2), "comision": round(comision, 2),
                "alimento": round(ali, 2), "estructura": round(est, 2),
                "sanidad": round(san, 2), "mortandad": round(mort, 2),
                "total": round(costo, 2),
                "por_kg_vendido": round(costo / kge, 2) if kge else None,
            },
            "indicadores": {
                "kg_prom_ingreso": round(kgi / cab, 1) if cab else None,
                "kg_prom_salida": round(kge / cab, 1) if cab else None,
                "estadia_prom": round(a_dias / cab) if cab else None,
                "adp": round(kg_prod / a_dias, 3) if a_dias else None,
                "pct_ms": round(ms_tot / pv_den * 100, 2) if pv_den else None,
                "conversion_ms": round(ms_tot / kg_prod, 2) if kg_prod > 0 else None,
                "costo_kg_producido": round((ali + est + san) / kg_prod, 2) if kg_prod > 0 else None,
                "precio_prom_pagado": round(compra / kgi, 2) if kgi else None,
            },
            "reposicion": {
                "precio_kg": round(rp, 2), "fuente_precio": rp_lbl,
                "precio_kg_ms": round(repo_ms, 2), "mes_ms": _ultp,
                "compra": round(compra_repo, 2), "comision": round(com_repo, 2),
                "alimento": round(ali_repo, 2), "mortandad": round(mort_repo, 2),
                "total": round(costo_repo, 2),
                "por_kg_vendido": round(costo_repo / kge, 2) if kge else None,
            },
            "verificacion": verif.get(rem) or {"estado": "sin_datamars", "sc": 0},
            "cobertura_pct": round((kgi - kg_sin) / kgi * 100, 1) if kgi else None,
            "tropas_sin_precio": tropas_sin,
            "precio_estimado": round(prom, 2),
            "meses_pv_acotados": {m: round(v, 2) for m, v in sorted(clamped.items())},
            "dias_sin_pv": sin_pv,
        }

    _kg_tot = sum(r["kg_ingreso"] for r in remitos_out.values())
    _kg_sin = sum(sum(t["kg_ingreso"] for t in r["tropas_sin_precio"])
                  for r in remitos_out.values())
    salida = {
        "meta": {
            "generado": datetime.now().isoformat(),
            "desde": RR_DESDE,
            "remitos": len(remitos_out),
            "egresos_venta": n_venta,
            "cobertura_global_pct": round((_kg_tot - _kg_sin) / _kg_tot * 100, 1) if _kg_tot else None,
            "pv_min": RR_PV_MIN, "pv_max": RR_PV_MAX,
            "factor_vaca": RR_FACTOR_VACA,
            "comision_default": RR_COMISION_DEF,
            "pv_fallback": PV_FALLBACK,
            "tasas_mortandad": MORT_PCT,
            "datamars": meta_dm,
            "fuentes": {
                "egresos": "WinCampo lst_egresos_hacienda (MOTIVO=VENTA, NRO_TRANSACCION=remito)",
                "compras": "compras de hacienda.xlsx -> precios_compra_real.json (por_tropa)",
                "racion":  "preico de racion feelot.xlsx",
                "pct_pv":  "pct_pv_mensual.json (pct_pv_ajustado, limites 2-3%)",
                "mortandad": f"muertes_{periodo}.json (tasa por grupo)",
                "caravanas": ("Datamars Livestock /odata/WeightRecords (EID leido por el "
                              "baston); sin sesion vale lo cargado en WinCampo"),
            },
        },
        "remitos": remitos_out,
    }
    guardar(salida, carpeta_out, "resultado_remitos.json")
    log.info(f"  ✓ Resultado por remito: {len(remitos_out)} remitos desde {RR_DESDE} · "
             f"{n_venta} egresos de venta · cobertura {salida['meta']['cobertura_global_pct']}%")
    if meta_dm.get("activo"):
        log.info(f"    Datamars (ventana {meta_dm['ventana_dias']}d): "
                 f"{meta_dm['remitos_verificados']} remitos verificados · "
                 f"{meta_dm['remitos_sin_datamars']} sin lecturas · "
                 f"{meta_dm['confirmadas_total']} caravanas leídas · "
                 f"{meta_dm['sc_total']} sin caravana de {meta_dm['cab_total']} cab · "
                 f"{meta_dm['sc_sin_electronica_total']} sin caravana electrónica + "
                 f"{meta_dm['sc_no_leida_total']} electrónicas no leídas · "
                 f"{meta_dm['confirmadas_salida_total']} confirmadas por su pesada de salida, "
                 f"{meta_dm['confirmadas_anterior_total']} por una lectura anterior")
    else:
        log.info(f"    Datamars inactivo ({meta_dm.get('motivo')}) — "
                 f"todos los remitos valen como WinCampo")
    for rem, r in sorted(remitos_out.items()):
        _v = r.get("verificacion") or {}
        _sc = (f" · {_v['sc']} sin caravana ({_v.get('sc_pct_cab')}%)" if _v.get("sc") else "")
        log.info(f"      {rem}: {r['cabezas']:>3} cab · costo $ {r['costos']['total']:,.0f} "
                 f"· $ {r['costos']['por_kg_vendido']:,.0f}/kg · cobertura {r['cobertura_pct']}%"
                 f"{_sc}")
        if _v.get("estado") == "verificado" and not _v.get("confirmadas"):
            log.warning(f"      ⚠ {rem}: ninguna de sus {r['cabezas']} caravanas se leyó en "
                        f"los últimos {_v.get('ventana_dias')} días — vale WinCampo, "
                        f"hay que cargar el origen a mano")
    return salida


def generar_analisis_costos_operativos(carpeta_out, periodo, log=None):
    """v15.61: denominadores mensuales de los ratios del módulo Análisis de Costos
    (Physis) → analisis_costos_operativos.json.

        {"2025-01": {"alim": 3330933, "cab": 8767}, ...}

    alim = kg de alimento TAL CUAL dados por el mixer en el mes.
    cab  = cabezas PROMEDIO del feedlot (El Haras, TODOS los propietarios).

    El contable (analisis_costos_datos.json) lo regenera Nicolás a mano con su
    ejecutable de Physis; esto es lo único de ese módulo que se refresca solo.

    Devuelve el dict volcado o None. NO levanta excepción.
    """
    if log is None:
        log = logging.getLogger("costosop")
    from datetime import date as _date

    base = Path(carpeta_out)

    def _load(nombre):
        try:
            with (base / nombre).open(encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            log.warning(f"  ⚠ no pude leer {nombre}: {e}")
            return None

    _consumo = _load(f"consumo_{periodo}.json") or {}
    meses_mix = ((_consumo.get("por_mes") or {}).get("meses")) or {}
    if not meses_mix:
        log.warning("  ⚠ consumo sin por_mes, saltando operativos de costos")
        return None

    _efi = _load("eficiencia_historico.json") or {}
    regs = _efi.get("registros") or []
    _comp = _load("comportamiento_historico.json") or {}
    snaps = _comp.get("snapshots") or []

    # ── Cabezas por mes desde la serie DIARIA (la medición buena) ──
    diario = {}
    for r in regs:
        f = str(r.get("fecha") or "")[:7]
        c = r.get("cabezas")
        if f and c:
            diario.setdefault(f, []).append(float(c))

    # ── Cabezas a fin de mes desde los snapshots (meses viejos) ──
    # El Haras pegsa + hoteleros, excluyendo la clave 'PEGSA' de por_hotelero
    # (duplica la propia y además es de todos los campos) — mismo criterio que
    # el denominador de %PV de v15.58.1. Limitación conocida: por_hotelero no
    # abre por campo, así que los meses viejos pueden sobreestimar un poco.
    fin_mes = {}
    for s in snaps:
        per = s.get("periodo")
        hm = s.get("hacienda_masa") or {}
        eh = ((hm.get("pegsa") or {}).get("por_campo", {}).get("El Haras", {}) or {}).get("cabezas")
        if not (per and eh):
            continue
        hot = sum(float((v or {}).get("cabezas") or 0)
                  for k, v in (hm.get("por_hotelero") or {}).items()
                  if str(k).strip().upper() != "PEGSA")
        fin_mes[per] = float(eh) + hot

    def _mes_ant(m):
        y, mm = int(m[:4]), int(m[5:7])
        return f"{y-1:04d}-12" if mm == 1 else f"{y:04d}-{mm-1:02d}"

    hoy = _date.today()
    mes_actual = f"{hoy.year:04d}-{hoy.month:02d}"
    salida, fuentes = {}, {}
    for m in sorted(meses_mix):
        if m < "2025-01" or m > mes_actual:
            continue
        d = meses_mix[m]
        kg = d.get("kg_total") or 0
        n_reg = d.get("dias_con_registro") or 0
        n_desc = len(d.get("dias_descartados") or [])
        if kg <= 0 or n_reg <= 0:
            continue
        # v15.61: los días descartados del corte mensual son días en que un mixer
        # no subió datos — el alimento SE repartió igual. Para un TOTAL mensual
        # (no un promedio) hay que reponerlos al promedio de los días válidos.
        alim = kg + n_desc * kg / n_reg

        cab, fuente = None, None
        if len(diario.get(m, [])) >= 25:
            cab = sum(diario[m]) / len(diario[m])
            fuente = "serie_diaria"
        else:
            ant, act = fin_mes.get(_mes_ant(m)), fin_mes.get(m)
            if ant is not None and act is not None:
                cab, fuente = (ant + act) / 2, "snapshots_prom"
            elif act is not None:
                cab, fuente = act, "fin_mes"
            elif ant is not None:
                cab, fuente = ant, "fin_mes_anterior"
        if cab is None:
            continue
        salida[m] = {"alim": int(round(alim)), "cab": int(round(cab))}
        fuentes[m] = fuente

    if not salida:
        log.warning("  ⚠ ningún mes con alimento Y cabezas, no se genera operativos")
        return None

    guardar(salida, carpeta_out, "analisis_costos_operativos.json")
    log.info(f"  ✓ Análisis de Costos · operativos: {len(salida)} meses "
             f"({min(salida)} → {max(salida)})")
    for m in sorted(salida)[-4:]:
        log.info(f"      {m}: alim {salida[m]['alim']:>10,} kg · cab {salida[m]['cab']:>6,} ({fuentes[m]})")
    return salida


def procesar_tesoreria_darwash(carpeta_out, log=None):
    """v11: lee el XLSX más reciente de `datos/financiero DW/` (subido
    semanalmente por el usuario) y lo convierte al shape canónico de
    tesoreria_ultimo.json (posicion + fecha_corte + flujo {saldo_inicial,
    semanas, series.saldo_semanal, series.saldo_acumulado}).

    Layout del XLSX (validado contra "2026_05_27_financiero darwash.xlsx"):
      R1 col B  → FINAL DE FINANCIERO (= posición final)
      R5 col B  → Capital corriente darwash
      R7 col E..W → fechas de semanas (datetime); col E es la primera
                    de cierre, F adelante son las proyectadas
      R22 col D → Saldo Disponibilidades (= saldo inicial)
      R82 col E..W → Saldo Acumulado con el total de movimientos
                     (es la serie que el panel/módulo grafica)

    Vuelca dos archivos:
      tesoreria_darwash.json           — snapshot actual (sobrescribe)
      tesoreria_darwash_historico.json — acumulado por fecha_corte (upsert)

    Tolerante: si la carpeta no existe o está vacía → warning + None.
    """
    if log is None:
        log = logging.getLogger("dw")
    base = Path(__file__).resolve().parent.parent
    carpeta_dw = base / "datos" / "financiero DW"
    if not carpeta_dw.is_dir():
        log.warning(f"  ⚠ Carpeta {carpeta_dw} no existe, saltando")
        return None, None
    xlsxs = sorted(carpeta_dw.glob("*.xlsx"), key=lambda p: p.stat().st_mtime, reverse=True)
    if not xlsxs:
        log.warning(f"  ⚠ Sin archivos XLSX en {carpeta_dw}, saltando")
        return None, None

    # v15.32: elegir el más reciente HIDRATADO. Si el más nuevo es placeholder
    # OneDrive (cloud-only), intentar warming y, si falla, caer al siguiente.
    xl = None
    for _cand in xlsxs:
        _es_ph, _motivo = _xlsx_es_placeholder_onedrive(str(_cand))
        if _es_ph and not _intentar_warming_onedrive(str(_cand)):
            log.warning(f"  ⚠ DW {_cand.name}: placeholder OneDrive ({_motivo}), probando anterior...")
            continue
        xl = _cand
        break
    if xl is None:
        log.warning(f"  ⚠ Todos los XLSX DW son placeholders inaccesibles, saltando")
        return None, None
    # Copia a temp para evitar lock de OneDrive
    import shutil, tempfile
    try:
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tf:
            tmp_path = _tf.name
        shutil.copy2(str(xl), tmp_path)
    except Exception as e:
        log.warning(f"  ⚠ no pude copiar XLSX DW a temp: {e}")
        return None, None

    try:
        try:
            import openpyxl
        except ImportError:
            log.warning("  ⚠ openpyxl no instalado; saltando tesorería DW")
            return None, None
        try:
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
        except Exception as e:
            log.warning(f"  ⚠ no pude abrir XLSX DW: {e}")
            return None, None
        ws = wb[wb.sheetnames[0]]

        def cell(r, c):
            v = ws.cell(row=r, column=c).value
            try:
                return float(v) if v is not None else None
            except (TypeError, ValueError):
                return v if isinstance(v, datetime) else None

        # Localizar filas por etiqueta de col A (robusto contra desplazamientos
        # menores del layout). Match case-insensitive substring.
        def find_row(*kws):
            for r in range(1, min(ws.max_row + 1, 100)):
                a = ws.cell(row=r, column=1).value
                if isinstance(a, str):
                    s = a.lower().strip()
                    if all(kw.lower() in s for kw in kws):
                        return r
            return None

        r_final_fin = find_row("final", "financiero") or 1
        r_capital   = find_row("capital", "corriente")
        r_saldo_disp = find_row("saldo", "disponibilidades")  # R22
        r_total_disp = find_row("total", "disponib")          # R28
        r_saldo_acum = find_row("saldo", "acumulado", "total", "movimientos")  # R82
        if r_saldo_acum is None:
            # Fallback: usar el primer "Saldo Acumulado" (R77 o R82)
            r_saldo_acum = find_row("saldo", "acumulado")

        # Header semanal (típicamente R7). Buscamos la fila con varias
        # datetime en cols E..W.
        r_header = None
        for r in range(5, 15):
            datetimes_in_row = sum(
                1 for c in range(5, 24)
                if isinstance(ws.cell(row=r, column=c).value, datetime)
            )
            if datetimes_in_row >= 5:
                r_header = r
                break
        if r_header is None:
            log.warning("  ⚠ No encontré fila de fechas semanales en el XLSX DW, saltando")
            return None, None

        # Extraer fechas y saldos acumulados (col E=5 hasta donde haya fechas)
        semanas_labels = []
        semanas_fechas = []   # ISO YYYY-MM-DD
        saldo_acum = []
        for c in range(5, 24):
            f = ws.cell(row=r_header, column=c).value
            if not isinstance(f, datetime):
                continue
            v = ws.cell(row=r_saldo_acum, column=c).value if r_saldo_acum else None
            try:
                v = float(v) if v is not None else 0.0
            except (TypeError, ValueError):
                v = 0.0
            semanas_labels.append(f"{f.day:02d}/{f.month:02d}")
            semanas_fechas.append(f.strftime("%Y-%m-%d"))
            saldo_acum.append(v)

        if not semanas_labels:
            log.warning("  ⚠ No pude extraer semanas del XLSX DW")
            return None, None

        # saldo_semanal = delta vs semana anterior (saldo_acum[i] - saldo_acum[i-1])
        # La primera semana usa el saldo_inicial como base.
        saldo_inicial = (cell(r_saldo_disp, 4) if r_saldo_disp else None) or 0.0
        prev = saldo_inicial
        saldo_sem = []
        for v in saldo_acum:
            saldo_sem.append(v - prev)
            prev = v

        posicion = cell(r_final_fin, 2)
        capital_corriente = cell(r_capital, 2) if r_capital else None
        fecha_corte = semanas_fechas[0]  # primer cierre del bloque semanal

        snapshot = {
            "meta": {
                "fuente":   "xlsx_financiero_dw",
                "archivo":  xl.name,
                "generado": datetime.now().isoformat(),
            },
            "fecha_corte": fecha_corte,
            "posicion":    posicion,
            "capital_corriente": capital_corriente,
            "flujo": {
                "saldo_inicial": saldo_inicial,
                "semanas":       semanas_labels,
                "series": {
                    "saldo_semanal":   saldo_sem,
                    "saldo_acumulado": saldo_acum,
                },
            },
        }

        # 1) snapshot actual
        out_actual = Path(carpeta_out) / "tesoreria_darwash.json"
        with out_actual.open("w", encoding="utf-8") as f:
            json.dump(snapshot, f, ensure_ascii=False, indent=2, default=str)

        # 2) histórico — upsert por fecha_corte
        hist_path = Path(carpeta_out) / "tesoreria_darwash_historico.json"
        if hist_path.exists():
            try:
                with hist_path.open("r", encoding="utf-8") as f:
                    hist = json.load(f)
                if not isinstance(hist, dict) or "snapshots" not in hist:
                    hist = {"meta": {}, "snapshots": []}
            except Exception:
                hist = {"meta": {}, "snapshots": []}
        else:
            hist = {"meta": {}, "snapshots": []}

        nueva = {
            "fecha_corte": fecha_corte,
            "posicion":    posicion,
            "saldo_inicial": saldo_inicial,
            "semanas":       semanas_labels,
            "saldo_acumulado": saldo_acum,
            "archivo":     xl.name,
        }
        idx = next((i for i, s in enumerate(hist["snapshots"]) if s.get("fecha_corte") == fecha_corte), None)
        if idx is not None:
            hist["snapshots"][idx] = nueva
        else:
            hist["snapshots"].append(nueva)
            hist["snapshots"].sort(key=lambda s: s.get("fecha_corte", ""))
        hist["meta"] = {
            "generado":    datetime.now().isoformat(),
            "n_snapshots": len(hist["snapshots"]),
        }
        with hist_path.open("w", encoding="utf-8") as f:
            json.dump(hist, f, ensure_ascii=False, indent=2, default=str)

        fecha_disp = "/".join(reversed(fecha_corte.split("-")))
        log.info(f"  ✓ Tesorería DW: snapshot {fecha_disp} · "
                 f"{len(semanas_labels)} semanas · histórico {len(hist['snapshots'])} snapshots")
        return snapshot, hist
    finally:
        try: Path(tmp_path).unlink()
        except Exception: pass


# ═══════════════════════════════════════════════════════════
#  TRAZABILIDAD · Caravanas declaradas (Google Drive)  (v15.x)
# ═══════════════════════════════════════════════════════════
# Fuente: carpeta colaborativa de Google Drive sincronizada localmente vía
# "Google Drive for Desktop" en G:\Mi unidad\Trazabilidad\. Dos Excel que el
# usuario y sus compañeros mantienen a mano:
#   · CARAVANAS DECLARADAS.xlsx  → hojas EL HARAS POST BLANQUEO + LAS TAPERAS SIN BLANQUEO
#   · hilton el descanso *.xlsx   → hoja HILTON  (fecha en el nombre → glob)
# Cada tick del bot los re-lee, así los cambios de los compañeros se reflejan
# en el portal cada hora. Genera trazabilidad_resumen.json.

TRAZABILIDAD_DIR_DEFAULT = r"G:\Mi unidad\Trazabilidad"

# Hojas auxiliares que NUNCA se procesan (columnas de apoyo, no caravanas).
_TRAZ_HOJAS_AUX = ("COINCIDIR", "DECLA", "DECLARACION")

# Definición declarativa de las hojas a exponer.
_TRAZ_HOJAS_CFG = [
    {"clave": "el_haras",    "titulo": "El Haras · Post Blanqueo",
     "archivo": "CARAVANAS DECLARADAS.xlsx",              "match": ("HARAS",)},
    {"clave": "las_taperas", "titulo": "Las Taperas · Sin Blanqueo",
     "archivo": "CARAVANAS DECLARADAS.xlsx",              "match": ("TAPERAS",)},
    {"clave": "hilton",      "titulo": "Hilton · El Descanso",
     "glob": "hilton*el*descanso*.xlsx",                  "match": ("HILTON",)},
]


def _traz_norm(s):
    """Normaliza texto: sin tildes, espacios colapsados, MAYÚSCULAS."""
    import unicodedata, re as _re
    if s is None:
        return ""
    s = str(s).strip()
    s = "".join(c for c in unicodedata.normalize("NFD", s)
                if unicodedata.category(c) != "Mn")
    return _re.sub(r"\s+", " ", s).upper()


def _traz_map_headers(hdr):
    """Mapea la fila de encabezados → índices de columna canónicos.
    Robusto a typos (FEHCA→FECHA), a "feedlot" como sinónimo de BOLSA y a
    "100 DIAS"/"90 DIAS" como umbral largo. Primera coincidencia gana (algunas
    hojas repiten headers en columnas auxiliares a la derecha)."""
    m = {}
    def setk(k, i):
        if k not in m:
            m[k] = i
    for idx, h in enumerate(hdr):
        n = _traz_norm(h).replace("FEHCA", "FECHA")
        if n == "CARAVANA":                                  setk("caravana", idx)
        elif n == "PROPIETARIO":                             setk("propietario", idx)
        # v15.44: "PRESENTE" aparece en HILTON-EL DESCANSO.xlsx rearmado
        # (el usuario renombra la columna "BOLSA" a "PRESENTE"). Aceptamos ambos
        # nombres para evitar regresiones si el equipo alterna entre convenciones.
        elif n in ("BOLSA", "FEEDLOT", "PRESENTE"):          setk("bolsa", idx)
        elif n == "CATEGORIA":                               setk("categoria", idx)
        elif "FECHA 40" in n:                                setk("f40", idx)
        elif ("FECHA 90" in n or "FECHA 100" in n or
              "90 DIAS" in n or "100 DIAS" in n):            setk("f90", idx)
        elif "FECHA SALIDA" in n:                            setk("fsalida", idx)
    return m


def _traz_as_date(v):
    from datetime import datetime as _dt, date as _date
    if isinstance(v, _dt):  return v.date()
    if isinstance(v, _date): return v
    return None


def _traz_analizar_hoja(rows, hoy):
    """v15.43: además de los KPIs por hoja y por propietario, agrega el
    sub-desglose por_categoria dentro de por_propietario (ej:
    por_propietario['PEGSA']['por_categoria']['VACA']) con sus propias
    proximas_40d/90d — necesario para la vista drill-down."""
    if not rows:
        return None
    m = _traz_map_headers(rows[0])
    if "caravana" not in m:
        return None

    HORIZONTE_DIAS = 90   # días hacia adelante para el timeline (frontend puede recortar a 30/60)

    activas = 0; sin_usar = 0
    categorias = {}; estado = {"CLASIFICAR": 0, "CON BOLSA": 0}
    c40 = c90 = 0
    estado_40d = {"CLASIFICAR": 0, "CON BOLSA": 0}
    estado_90d = {"CLASIFICAR": 0, "CON BOLSA": 0}
    categorias_40d = {}; categorias_90d = {}
    por_prop = {}
    prox_40 = {}; prox_90 = {}

    def _get_or_new_prop(p):
        if p not in por_prop:
            por_prop[p] = {
                "activas": 0, "sin_usar": 0,
                "categorias": {}, "estado": {"CLASIFICAR": 0, "CON BOLSA": 0},
                "cumple_40d": 0, "cumple_90d": 0,
                "estado_40d": {"CLASIFICAR": 0, "CON BOLSA": 0},
                "estado_90d": {"CLASIFICAR": 0, "CON BOLSA": 0},
                "categorias_40d": {}, "categorias_90d": {},
                # v15.43 — sub-desglose por categoría con proximas
                "por_categoria": {},   # {"VACA": {activas, c40, c90, prox_40d, prox_90d, estado}}
            }
        return por_prop[p]

    def _get_or_new_cat(p_dict, cat):
        if cat not in p_dict["por_categoria"]:
            p_dict["por_categoria"][cat] = {
                "activas": 0,
                "estado": {"CLASIFICAR": 0, "CON BOLSA": 0},
                "cumple_40d": 0, "cumple_90d": 0,
                "_prox_40d": {},   # dict interno {date: count}
                "_prox_90d": {},
            }
        return p_dict["por_categoria"][cat]

    for r in rows[1:]:
        if not any(x is not None for x in r):
            continue
        if m["caravana"] >= len(r) or r[m["caravana"]] is None:
            continue
        prop_raw = r[m["propietario"]] if "propietario" in m and m["propietario"] < len(r) else None
        if not prop_raw or str(prop_raw).strip() == "":
            continue
        fsal = r[m["fsalida"]] if "fsalida" in m and m["fsalida"] < len(r) else None
        if fsal not in (None, ""):
            continue

        prop = _traz_norm(prop_raw)
        bn = _traz_norm(r[m["bolsa"]]) if "bolsa" in m and m["bolsa"] < len(r) else ""

        # sin usar = tarjeta informativa (no cuenta como activa)
        if bn == "":
            sin_usar += 1
            _get_or_new_prop(prop)["sin_usar"] += 1
            continue

        activas += 1
        p = _get_or_new_prop(prop)
        p["activas"] += 1

        cat = _traz_norm(r[m["categoria"]]) if "categoria" in m and m["categoria"] < len(r) else ""
        cat = cat or "SIN CATEGORIA"
        categorias[cat] = categorias.get(cat, 0) + 1
        p["categorias"][cat] = p["categorias"].get(cat, 0) + 1

        # v15.43 — abrir sub-nivel de categoría dentro del propietario
        pc = _get_or_new_cat(p, cat)
        pc["activas"] += 1

        est_key = "CLASIFICAR" if bn == "CLASIFICAR" else "CON BOLSA"
        estado[est_key] += 1
        p["estado"][est_key] += 1
        pc["estado"][est_key] += 1

        d40 = _traz_as_date(r[m["f40"]]) if "f40" in m and m["f40"] < len(r) else None
        d90 = _traz_as_date(r[m["f90"]]) if "f90" in m and m["f90"] < len(r) else None

        if d40 is not None:
            if d40 <= hoy:
                c40 += 1
                estado_40d[est_key] += 1
                categorias_40d[cat] = categorias_40d.get(cat, 0) + 1
                p["cumple_40d"] += 1
                p["estado_40d"][est_key] += 1
                p["categorias_40d"][cat] = p["categorias_40d"].get(cat, 0) + 1
                pc["cumple_40d"] += 1
            elif (d40 - hoy).days <= HORIZONTE_DIAS:
                prox_40[d40] = prox_40.get(d40, 0) + 1
                pc["_prox_40d"][d40] = pc["_prox_40d"].get(d40, 0) + 1

        if d90 is not None:
            if d90 <= hoy:
                c90 += 1
                estado_90d[est_key] += 1
                categorias_90d[cat] = categorias_90d.get(cat, 0) + 1
                p["cumple_90d"] += 1
                p["estado_90d"][est_key] += 1
                p["categorias_90d"][cat] = p["categorias_90d"].get(cat, 0) + 1
                pc["cumple_90d"] += 1
            elif (d90 - hoy).days <= HORIZONTE_DIAS:
                prox_90[d90] = prox_90.get(d90, 0) + 1
                pc["_prox_90d"][d90] = pc["_prox_90d"].get(d90, 0) + 1

    pct = lambda x, tot: round(100.0 * x / tot, 1) if tot else 0.0

    # Ordenar propietarios por activas desc; propietarios internos ordenados
    por_prop_out = {}
    for p_name, p in sorted(por_prop.items(), key=lambda kv: -kv[1]["activas"]):
        p["categorias"]     = dict(sorted(p["categorias"].items(),     key=lambda kv: -kv[1]))
        p["categorias_40d"] = dict(sorted(p["categorias_40d"].items(), key=lambda kv: -kv[1]))
        p["categorias_90d"] = dict(sorted(p["categorias_90d"].items(), key=lambda kv: -kv[1]))
        p["pct_40d"] = pct(p["cumple_40d"], p["activas"])
        p["pct_90d"] = pct(p["cumple_90d"], p["activas"])

        # v15.43 — cerrar cada sub-categoría con sus proximas ordenadas
        por_cat_out = {}
        for c_name, c in sorted(p["por_categoria"].items(), key=lambda kv: -kv[1]["activas"]):
            c["proximas_40d"] = [
                {"fecha": f.isoformat(), "cabezas": n, "dias_hasta": (f - hoy).days}
                for f, n in sorted(c["_prox_40d"].items())
            ]
            c["proximas_90d"] = [
                {"fecha": f.isoformat(), "cabezas": n, "dias_hasta": (f - hoy).days}
                for f, n in sorted(c["_prox_90d"].items())
            ]
            c["pct_40d"] = pct(c["cumple_40d"], c["activas"])
            c["pct_90d"] = pct(c["cumple_90d"], c["activas"])
            del c["_prox_40d"]; del c["_prox_90d"]   # limpiar dicts internos
            por_cat_out[c_name] = c
        p["por_categoria"] = por_cat_out
        por_prop_out[p_name] = p

    # Timeline de próximas activaciones (ordenado cronológico)
    proximas_40d = [
        {"fecha": f.isoformat(), "cabezas": n, "dias_hasta": (f - hoy).days}
        for f, n in sorted(prox_40.items())
    ]
    proximas_90d = [
        {"fecha": f.isoformat(), "cabezas": n, "dias_hasta": (f - hoy).days}
        for f, n in sorted(prox_90.items())
    ]

    return {
        "activas": activas,
        "sin_usar": sin_usar,
        "categorias": dict(sorted(categorias.items(), key=lambda kv: -kv[1])),
        "estado": estado,
        "cumple_40d": c40,  "pct_40d": pct(c40, activas),
        "cumple_90d": c90,  "pct_90d": pct(c90, activas),
        "estado_40d": estado_40d,
        "estado_90d": estado_90d,
        "categorias_40d": dict(sorted(categorias_40d.items(), key=lambda kv: -kv[1])),
        "categorias_90d": dict(sorted(categorias_90d.items(), key=lambda kv: -kv[1])),
        # v15.41
        "por_propietario": por_prop_out,
        "proximas_40d":    proximas_40d,
        "proximas_90d":    proximas_90d,
    }


def procesar_trazabilidad(carpeta_out, log=None):
    """Lee los Excel de trazabilidad de G:\\Mi unidad\\Trazabilidad\\ y genera
    trazabilidad_resumen.json con KPIs por hoja + consolidado global.
    Tolerante: si la carpeta/archivos no existen → warning + None (no tumba el tick)."""
    if log is None:
        log = logging.getLogger("traz")
    from datetime import date

    # Ruta: config [RUTAS] trazabilidad_dir, si no el default de Google Drive.
    # Lee config.ini de forma tolerante (no usa cargar_config() porque ese hace
    # sys.exit si el archivo no existe, p.ej. corriendo desde el repo).
    ruta_dir = TRAZABILIDAD_DIR_DEFAULT
    try:
        _cfg_path = Path(__file__).parent / "config.ini"
        if _cfg_path.exists():
            _cfg = configparser.ConfigParser()
            _cfg.read(_cfg_path, encoding="utf-8")
            if _cfg.has_section("RUTAS") and _cfg["RUTAS"].get("trazabilidad_dir"):
                ruta_dir = _cfg["RUTAS"].get("trazabilidad_dir")
    except Exception:
        pass
    base = Path(ruta_dir)
    if not base.is_dir():
        log.warning(f"  ⚠ Carpeta trazabilidad no existe: {base}, saltando")
        return None

    try:
        import openpyxl
    except ImportError:
        log.warning("  ⚠ openpyxl no instalado; saltando trazabilidad")
        return None

    import shutil, tempfile
    hoy = date.today()

    def _resolver_archivo(cfg):
        if "glob" in cfg:
            # v15.44: matching case-insensitive. Path.glob() es case-sensitive
            # aunque estés en Windows — falla si el usuario renombra el archivo
            # a MAYÚSCULAS (ej: 'HILTON-EL DESCANSO.xlsx' vs glob
            # 'hilton*el*descanso*.xlsx'). Iteramos la carpeta y matcheamos
            # con fnmatch sobre nombres en lowercase.
            import fnmatch
            pattern = cfg["glob"].lower()
            cands = [
                p for p in base.iterdir()
                if p.is_file() and fnmatch.fnmatch(p.name.lower(), pattern)
            ]
            cands.sort(key=lambda p: p.stat().st_mtime, reverse=True)
            return cands[0] if cands else None
        p = base / cfg["archivo"]
        return p if p.exists() else None

    def _cargar_rows(ruta_xlsx, match_kws):
        """Copia a temp (evita lock de Drive) y devuelve las filas de la hoja
        cuyo nombre normalizado contiene alguno de match_kws (ignora auxiliares)."""
        try:
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as _tf:
                tmp = _tf.name
            shutil.copy2(str(ruta_xlsx), tmp)
        except Exception as e:
            log.warning(f"  ⚠ no pude copiar {ruta_xlsx.name} a temp: {e}")
            return None
        try:
            wb = openpyxl.load_workbook(tmp, data_only=True)
            hoja = None
            for sn in wb.sheetnames:
                nn = _traz_norm(sn)
                if any(aux in nn for aux in _TRAZ_HOJAS_AUX):
                    continue
                if any(kw in nn for kw in match_kws):
                    hoja = sn
                    break
            if hoja is None:
                wb.close()
                return None
            rows = list(wb[hoja].iter_rows(values_only=True))
            wb.close()
            return {"hoja": hoja, "rows": rows}
        except Exception as e:
            log.warning(f"  ⚠ no pude abrir {ruta_xlsx.name}: {e}")
            return None
        finally:
            try: Path(tmp).unlink()
            except Exception: pass

    hojas = []
    for cfg in _TRAZ_HOJAS_CFG:
        ruta = _resolver_archivo(cfg)
        if ruta is None:
            log.warning(f"  ⚠ {cfg['titulo']}: archivo no encontrado en {base}, saltando hoja")
            continue
        cargado = _cargar_rows(ruta, cfg["match"])
        if cargado is None:
            log.warning(f"  ⚠ {cfg['titulo']}: hoja no encontrada en {ruta.name}, saltando")
            continue
        kpis = _traz_analizar_hoja(cargado["rows"], hoy)
        if kpis is None:
            log.warning(f"  ⚠ {cfg['titulo']}: sin datos parseables, saltando")
            continue
        hojas.append({
            "clave":   cfg["clave"],
            "titulo":  cfg["titulo"],
            "archivo": ruta.name,
            "hoja":    cargado["hoja"],
            **kpis,
        })
        log.info(f"  ↳ {cfg['titulo']}: {kpis['activas']:,} activas · "
                 f"40d {kpis['cumple_40d']:,} · 90d {kpis['cumple_90d']:,}")

    if not hojas:
        log.warning("  ⚠ Trazabilidad: ninguna hoja procesada, no se genera JSON")
        return None

    # Consolidado global (suma de todas las hojas) — v15.41
    cons = {
        "activas": 0, "sin_usar": 0,
        "categorias": {}, "estado": {"CLASIFICAR": 0, "CON BOLSA": 0},
        "cumple_40d": 0, "cumple_90d": 0,
        "estado_40d": {"CLASIFICAR": 0, "CON BOLSA": 0},
        "estado_90d": {"CLASIFICAR": 0, "CON BOLSA": 0},
        "categorias_40d": {}, "categorias_90d": {},
        # v15.41
        "por_propietario": {},   # merge de todas las hojas
        "proximas_40d": {},      # dict fecha_iso → count (para deduplicar/mergear)
        "proximas_90d": {},
    }

    for h in hojas:
        cons["activas"]    += h["activas"]
        cons["sin_usar"]   += h["sin_usar"]
        cons["cumple_40d"] += h["cumple_40d"]
        cons["cumple_90d"] += h["cumple_90d"]
        for k, v in h["categorias"].items():
            cons["categorias"][k] = cons["categorias"].get(k, 0) + v
        for k, v in h["estado"].items():
            cons["estado"][k] = cons["estado"].get(k, 0) + v
        for k, v in h["estado_40d"].items():
            cons["estado_40d"][k] = cons["estado_40d"].get(k, 0) + v
        for k, v in h["estado_90d"].items():
            cons["estado_90d"][k] = cons["estado_90d"].get(k, 0) + v
        for k, v in h["categorias_40d"].items():
            cons["categorias_40d"][k] = cons["categorias_40d"].get(k, 0) + v
        for k, v in h["categorias_90d"].items():
            cons["categorias_90d"][k] = cons["categorias_90d"].get(k, 0) + v

        # v15.41: merge de por_propietario (suma agregados)
        for p_name, p_data in h["por_propietario"].items():
            if p_name not in cons["por_propietario"]:
                cons["por_propietario"][p_name] = {
                    "activas": 0, "sin_usar": 0,
                    "categorias": {}, "estado": {"CLASIFICAR": 0, "CON BOLSA": 0},
                    "cumple_40d": 0, "cumple_90d": 0,
                    "estado_40d": {"CLASIFICAR": 0, "CON BOLSA": 0},
                    "estado_90d": {"CLASIFICAR": 0, "CON BOLSA": 0},
                    "categorias_40d": {}, "categorias_90d": {},
                    # v15.43 — sub-desglose por categoría en el consolidado
                    "por_categoria": {},
                }
            p_cons = cons["por_propietario"][p_name]
            p_cons["activas"]    += p_data["activas"]
            p_cons["sin_usar"]   += p_data["sin_usar"]
            p_cons["cumple_40d"] += p_data["cumple_40d"]
            p_cons["cumple_90d"] += p_data["cumple_90d"]
            for k, v in p_data["categorias"].items():
                p_cons["categorias"][k] = p_cons["categorias"].get(k, 0) + v
            for k, v in p_data["estado"].items():
                p_cons["estado"][k] = p_cons["estado"].get(k, 0) + v
            for k, v in p_data["estado_40d"].items():
                p_cons["estado_40d"][k] = p_cons["estado_40d"].get(k, 0) + v
            for k, v in p_data["estado_90d"].items():
                p_cons["estado_90d"][k] = p_cons["estado_90d"].get(k, 0) + v
            for k, v in p_data["categorias_40d"].items():
                p_cons["categorias_40d"][k] = p_cons["categorias_40d"].get(k, 0) + v
            for k, v in p_data["categorias_90d"].items():
                p_cons["categorias_90d"][k] = p_cons["categorias_90d"].get(k, 0) + v

            # v15.43: merge de por_categoria dentro de cada propietario
            for c_name, c_data in p_data.get("por_categoria", {}).items():
                if c_name not in p_cons["por_categoria"]:
                    p_cons["por_categoria"][c_name] = {
                        "activas": 0,
                        "estado": {"CLASIFICAR": 0, "CON BOLSA": 0},
                        "cumple_40d": 0, "cumple_90d": 0,
                        "_prox_40d": {}, "_prox_90d": {},
                    }
                c_cons = p_cons["por_categoria"][c_name]
                c_cons["activas"]    += c_data["activas"]
                c_cons["cumple_40d"] += c_data["cumple_40d"]
                c_cons["cumple_90d"] += c_data["cumple_90d"]
                for k, v in c_data["estado"].items():
                    c_cons["estado"][k] = c_cons["estado"].get(k, 0) + v
                for x in c_data.get("proximas_40d", []):
                    c_cons["_prox_40d"][x["fecha"]] = c_cons["_prox_40d"].get(x["fecha"], 0) + x["cabezas"]
                for x in c_data.get("proximas_90d", []):
                    c_cons["_prox_90d"][x["fecha"]] = c_cons["_prox_90d"].get(x["fecha"], 0) + x["cabezas"]

        # v15.41: merge de próximas activaciones (por fecha)
        for item in h["proximas_40d"]:
            cons["proximas_40d"][item["fecha"]] = cons["proximas_40d"].get(item["fecha"], 0) + item["cabezas"]
        for item in h["proximas_90d"]:
            cons["proximas_90d"][item["fecha"]] = cons["proximas_90d"].get(item["fecha"], 0) + item["cabezas"]

    # Ordenar propietarios y sus dicts internos
    por_prop_cons_out = {}
    for p_name, p_data in sorted(cons["por_propietario"].items(), key=lambda kv: -kv[1]["activas"]):
        p_data["categorias"]     = dict(sorted(p_data["categorias"].items(),     key=lambda kv: -kv[1]))
        p_data["categorias_40d"] = dict(sorted(p_data["categorias_40d"].items(), key=lambda kv: -kv[1]))
        p_data["categorias_90d"] = dict(sorted(p_data["categorias_90d"].items(), key=lambda kv: -kv[1]))
        tot_p = p_data["activas"] or 1
        p_data["pct_40d"] = round(100.0 * p_data["cumple_40d"] / tot_p, 1)
        p_data["pct_90d"] = round(100.0 * p_data["cumple_90d"] / tot_p, 1)
        por_prop_cons_out[p_name] = p_data
    cons["por_propietario"] = por_prop_cons_out

    # v15.43 — cerrar cada categoría del consolidado (proximas ordenadas + pct)
    for p_name in por_prop_cons_out:
        por_cat_out = {}
        for c_name, c in sorted(por_prop_cons_out[p_name]["por_categoria"].items(),
                                 key=lambda kv: -kv[1]["activas"]):
            c["proximas_40d"] = [
                {"fecha": f, "cabezas": n,
                 "dias_hasta": (date.fromisoformat(f) - hoy).days}
                for f, n in sorted(c["_prox_40d"].items())
            ]
            c["proximas_90d"] = [
                {"fecha": f, "cabezas": n,
                 "dias_hasta": (date.fromisoformat(f) - hoy).days}
                for f, n in sorted(c["_prox_90d"].items())
            ]
            tot = c["activas"] or 1
            c["pct_40d"] = round(100.0 * c["cumple_40d"] / tot, 1)
            c["pct_90d"] = round(100.0 * c["cumple_90d"] / tot, 1)
            del c["_prox_40d"]; del c["_prox_90d"]
            por_cat_out[c_name] = c
        por_prop_cons_out[p_name]["por_categoria"] = por_cat_out

    # Convertir dicts de proximas → lista ordenada cronológica
    def _sort_prox(d):
        return [
            {"fecha": f, "cabezas": n, "dias_hasta": (date.fromisoformat(f) - hoy).days}
            for f, n in sorted(d.items())
        ]
    cons["proximas_40d"] = _sort_prox(cons["proximas_40d"])
    cons["proximas_90d"] = _sort_prox(cons["proximas_90d"])

    cons["categorias"]     = dict(sorted(cons["categorias"].items(),     key=lambda kv: -kv[1]))
    cons["categorias_40d"] = dict(sorted(cons["categorias_40d"].items(), key=lambda kv: -kv[1]))
    cons["categorias_90d"] = dict(sorted(cons["categorias_90d"].items(), key=lambda kv: -kv[1]))
    _tot = cons["activas"] or 1
    cons["pct_40d"] = round(100.0 * cons["cumple_40d"] / _tot, 1)
    cons["pct_90d"] = round(100.0 * cons["cumple_90d"] / _tot, 1)

    data = {
        "meta": {
            "generado": datetime.now().isoformat(),
            "hoy":      hoy.isoformat(),
            "fuente":   str(base),
            "archivos": sorted({h["archivo"] for h in hojas}),
            "ok":       True,
        },
        "hojas":       hojas,
        "consolidado": cons,
    }
    guardar(data, carpeta_out, "trazabilidad_resumen.json")
    log.info(f"  ✓ trazabilidad_resumen.json · {len(hojas)} hojas · "
             f"{cons['activas']:,} caravanas activas")
    return data


if __name__ == "__main__":
    main()
